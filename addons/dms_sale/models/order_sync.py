"""
OrderProcessor 自動匯入 Scheduled Action 邏輯。

掃描 /mnt/order_backup/ 下各資料夾的 result.json，
建立草稿 dms.sale.order 並紀錄 dms.sync.log。
"""
import json
import logging
import os
import re
import time

from odoo import models

_logger = logging.getLogger(__name__)

BACKUP_DIR = '/mnt/order_backup'
MTIME_MIN_SECONDS = 60   # result.json 寫入後需靜置 60 秒才視為完整
_FOLDER_STAMP_RE = re.compile(r'_\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2}_\d{4}$')


def _folder_signature(folder_name):
    return _FOLDER_STAMP_RE.sub('', folder_name or '')


class OrderSyncAction(models.AbstractModel):
    _name = 'dms.order.sync'
    _description = 'OrderProcessor 自動匯入'

    # ── 公開方法（Scheduled Action 呼叫）─────────────────
    def run_sync(self):
        """每次由 ir.cron 呼叫，靜默掃描並匯入。"""
        if not os.path.isdir(BACKUP_DIR):
            _logger.warning('[OrderSync] 備份目錄不存在：%s', BACKUP_DIR)
            return

        # 已處理資料夾（去重）
        existing = set(
            self.env['dms.sync.log'].search([]).mapped('folder_name')
        )

        for folder_name in sorted(os.listdir(BACKUP_DIR)):
            folder_path = os.path.join(BACKUP_DIR, folder_name)
            if not os.path.isdir(folder_path):
                continue
            if folder_name in existing:
                continue
            try:
                self._process_folder(folder_path, folder_name)
            except Exception:
                _logger.exception('[OrderSync] 資料夾處理失敗：%s', folder_name)

    # ── 歷史資料一次性標記──────────────────────────────
    def mark_all_existing_ignored(self):
        """將目前所有未處理的資料夾標記為 ignored，不建立訂單。"""
        if not os.path.isdir(BACKUP_DIR):
            return 0
        existing = set(
            self.env['dms.sync.log'].search([]).mapped('folder_name')
        )
        count = 0
        for folder_name in os.listdir(BACKUP_DIR):
            folder_path = os.path.join(BACKUP_DIR, folder_name)
            if not os.path.isdir(folder_path):
                continue
            if folder_name in existing:
                continue
            self.env['dms.sync.log'].create({
                'folder_name': folder_name,
                'state': 'ignored',
                'error_msg': '手動標記為歷史資料，不匯入',
            })
            count += 1
        return count

    def _process_folder_by_name(self, folder_name):
        """以資料夾名稱重新處理（供 UI 重新同步按鈕使用）。

        會繞過「已處理過則跳過」與 mtime 保護。
        """
        if not os.path.isdir(BACKUP_DIR):
            self._write_log(folder_name, 'fail',
                            error_msg=f'備份目錄不存在：{BACKUP_DIR}')
            return
        folder_path = os.path.join(BACKUP_DIR, folder_name)
        if not os.path.isdir(folder_path):
            self._write_log(folder_name, 'fail',
                            error_msg=f'資料夾不存在：{folder_name}')
            return
        try:
            self._process_folder(folder_path, folder_name, skip_mtime=True)
        except Exception as e:
            self._write_log(folder_name, 'fail',
                            error_msg=f'重新同步失敗：{e}')

    # ── 內部：處理單一資料夾 ──────────────────────────
    def _process_folder(self, folder_path, folder_name, skip_mtime=False):
        result_path = os.path.join(folder_path, 'result.json')
        if not os.path.isfile(result_path):
            self._write_log(folder_name, 'skip', error_msg='找不到 result.json')
            return

        # mtime 保護
        if (not skip_mtime
                and (time.time() - os.path.getmtime(result_path))
                < MTIME_MIN_SECONDS):
            # 尚未穩定，不寫 log → 下次再試
            return

        try:
            with open(result_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except Exception as e:
            self._write_log(folder_name, 'fail', error_msg=f'JSON 解析失敗：{e}')
            return

        try:
            vals = self._build_order_vals(data, folder_name)
        except Exception as e:
            self._write_log(folder_name, 'fail', error_msg=f'欄位解析失敗：{e}')
            return

        # xlsx「原始資料」為穩定後備來源：即使身分證已辨識到姓名，仍需補讀缺漏的交易欄位。
        try:
            xlsx_data = self._read_xlsx_fallback(folder_path)
            if xlsx_data:
                vals = self._merge_order_vals(
                    vals,
                    self._build_order_vals(xlsx_data, folder_name),
                )
        except Exception:
            _logger.exception(
                '[OrderSync] xlsx fallback 失敗：%s', folder_name)

        try:
            order = self._find_existing_order(folder_name, vals)
            if order:
                write_vals = self.env['dms.sale.order']._prepare_import_update_vals(order, vals)
                order.write(write_vals)
            else:
                order = self.env['dms.sale.order'].create(vals)
            self._write_log(folder_name, 'success', order_id=order.id)
        except Exception as e:
            self._write_log(folder_name, 'fail', error_msg=f'同步訂單失敗：{e}')

    def _merge_order_vals(self, base_vals, fallback_vals):
        merged = dict(base_vals)

        for key, value in fallback_vals.items():
            if key == 'sale_origin':
                continue
            if merged.get(key) in (False, None, '') and value not in (False, None, ''):
                merged[key] = value

        if not merged.get('product_id') and fallback_vals.get('product_id'):
            merged['product_id'] = fallback_vals['product_id']
        if not merged.get('source_product_name') and fallback_vals.get('source_product_name'):
            merged['source_product_name'] = fallback_vals['source_product_name']
        if not merged.get('color_id') and fallback_vals.get('color_id'):
            merged['color_id'] = fallback_vals['color_id']
        if not merged.get('dealer_id') and fallback_vals.get('dealer_id'):
            merged['dealer_id'] = fallback_vals['dealer_id']
            merged['sale_type'] = fallback_vals.get('sale_type') or 'dealer'
        if merged.get('sale_type') == 'store' and fallback_vals.get('dealer_id'):
            merged['sale_type'] = fallback_vals.get('sale_type') or 'dealer'
        if merged.get('payment_method') == 'cash' and fallback_vals.get('payment_method') == 'installment':
            merged['payment_method'] = 'installment'
        if not merged.get('installment_periods') and fallback_vals.get('installment_periods'):
            merged['installment_periods'] = fallback_vals['installment_periods']
        if not merged.get('finance_company') and fallback_vals.get('finance_company'):
            merged['finance_company'] = fallback_vals['finance_company']
        if not merged.get('finance_company_other') and fallback_vals.get('finance_company_other'):
            merged['finance_company_other'] = fallback_vals['finance_company_other']
        if not merged.get('is_trade_in') and fallback_vals.get('is_trade_in'):
            merged['is_trade_in'] = True

        return merged

    def _find_existing_order(self, folder_name, vals):
        SaleOrder = self.env['dms.sale.order']
        exact = SaleOrder.search([
            ('sale_origin', '=', 'order_processor'),
            ('source_folder', '=', folder_name),
            ('active', '=', True),
            ('state', '!=', 'cancel'),
        ], limit=1)
        if exact:
            return exact

        signature = _folder_signature(folder_name)
        customer_name = (vals.get('customer_name') or '').strip()
        if not signature or not customer_name:
            return SaleOrder.browse()

        candidates = SaleOrder.search([
            ('sale_origin', '=', 'order_processor'),
            ('customer_name', '=', customer_name),
            ('source_folder', '!=', False),
            ('active', '=', True),
            ('state', '!=', 'cancel'),
        ], order='create_date desc, id desc')
        id_number = (vals.get('id_number') or '').strip()
        if id_number:
            candidates = candidates.filtered(
                lambda order: (order.id_number or '').strip() == id_number)

        candidates = candidates.filtered(
            lambda order: _folder_signature(order.source_folder) == signature
            and not order.product_id
            and not (order.source_product_name or '').strip()
        )
        if candidates:
            return candidates[:1]

        return SaleOrder._find_cross_source_import_order(
            vals,
            incoming_origin='order_processor',
        )

    # ── 內部：解析 result.json → 訂單欄位 ────────────
    def _normalize_data(self, data):
        """將 OrderProcessor 多種 result.json 格式統一成 (text_map, front, back)。

        支援：
          (A) 舊格式：{'text_map':{...}, 'front':{...}, 'back':{...}}
              或 {'docx':{'text_map':{...}}, 'front':{...}, 'back':{...}}
          (B) 新格式：{
                '<檔名>.docx': {'text_content': '車主電話：xxx\n車輛型號：...'},
                '身分證正面.jpg': {'辨識面':'front', '擷取欄位':{...}},
                '身分證反面.jpg': {'辨識面':'back',  '擷取欄位':{...}},
              }
        回傳統一過後的 (text_map, front, back) 三個 dict。
        """
        if not isinstance(data, dict):
            return {}, {}, {}

        # (A) 舊格式
        text_map = (
            data.get('text_map')
            or (data.get('docx') or {}).get('text_map')
            or {}
        )
        front = data.get('front') or {}
        back = data.get('back') or {}
        if text_map or front or back:
            return text_map or {}, front, back

        # (B) 新格式：掃描所有 key
        text_map = {}
        for key, val in data.items():
            if not isinstance(val, dict):
                continue
            # docx：把 text_content 多行解析成 key:value
            if key.endswith('.docx') or val.get('type') == 'docx':
                content = val.get('text_content') or ''
                for line in content.splitlines():
                    line = line.strip()
                    # 同時支援全形冒號「：」與半形「:」
                    m = re.match(r'^([^：:]{1,30})\s*[：:]\s*(.*)$', line)
                    if not m:
                        continue
                    k = m.group(1).strip()
                    v = m.group(2).strip()
                    if v:
                        text_map.setdefault(k, v)
            # 身分證正反面
            face = val.get('辨識面')
            fields_map = val.get('擷取欄位') or {}
            if face == 'front':
                front = fields_map
            elif face == 'back':
                back = fields_map
        return text_map, front, back

    # ── 內部：xlsx 原始資料 fallback ─────────────────
    def _read_xlsx_fallback(self, folder_path):
        """從資料夾內任一 xlsx 的「原始資料」sheet 讀出 row → 統一格式。

        用於 result.json 解析後資料缺漏時的後備來源。
        回傳：{'text_map': {...}, 'front': {...}, 'back': {...}}；找不到回傳 None。
        """
        try:
            import openpyxl
        except ImportError:
            return None
        if not os.path.isdir(folder_path):
            return None
        for fname in sorted(os.listdir(folder_path)):
            if not fname.lower().endswith('.xlsx'):
                continue
            fpath = os.path.join(folder_path, fname)
            try:
                wb = openpyxl.load_workbook(
                    fpath, data_only=True, read_only=True)
            except Exception:
                continue
            if '原始資料' not in wb.sheetnames:
                wb.close()
                continue
            ws = wb['原始資料']
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            if len(rows) < 2 or not rows[0]:
                continue
            headers = [(h or '').strip() if isinstance(h, str) else h
                       for h in rows[0]]
            data_row = rows[1]
            row_dict = {}
            for h, v in zip(headers, data_row):
                if not h or v in (None, ''):
                    continue
                row_dict[h] = (v.strip() if isinstance(v, str)
                               else str(v).strip())
            if not row_dict.get('姓名'):
                continue
            text_map = {
                '車輛型號': row_dict.get('機種') or row_dict.get('型號') or '',
                '車輛顏色': row_dict.get('顏色') or '',
                '車行名稱': row_dict.get('車行') or '',
                '車主電話': row_dict.get('手機') or row_dict.get('車主電話') or '',
                '車主Email': row_dict.get('車主Email') or row_dict.get('Email') or '',
                '是否有汰舊': row_dict.get('是否有汰舊') or row_dict.get('汰舊') or '',
                '是否有分期': row_dict.get('分期公司') or '',
                '備註': row_dict.get('備註') or '',
                '配件': row_dict.get('配件') or '',
            }
            text_map = {k: v for k, v in text_map.items() if v}
            front = {}
            if row_dict.get('姓名'):
                front['姓名'] = row_dict['姓名']
            if row_dict.get('生日'):
                front['出生年月日'] = row_dict['生日']
            if row_dict.get('身分證'):
                front['身分證字號'] = row_dict['身分證']
            back = {}
            if row_dict.get('戶籍'):
                back['住址'] = row_dict['戶籍']
            return {'text_map': text_map, 'front': front, 'back': back}
        return None

    def _build_order_vals(self, data, folder_name):
        text_map, front, back = self._normalize_data(data)

        # ── 客戶身分 ──────────────────────────────────
        customer_name = (front.get('姓名') or '').strip() or '（未知）'
        id_number = (front.get('身分證字號') or '').strip()
        birthday_raw = (front.get('出生年月日') or '').strip()
        address_registered = (back.get('住址') or '').strip()

        # 生日：民國 → AD Date
        birthday_ad = None
        bd_match = re.match(r'(\d{2,3})\s*[./年]\s*(\d{1,2})\s*[./月]\s*(\d{1,2})', birthday_raw)
        if bd_match:
            try:
                roc_y, m, d = int(bd_match.group(1)), int(bd_match.group(2)), int(bd_match.group(3))
                birthday_ad = f'{roc_y + 1911}-{m:02d}-{d:02d}'
            except Exception:
                pass

        # ── 聯絡 ──────────────────────────────────────
        customer_phone = (
            text_map.get('車主電話') or text_map.get('聯絡電話') or ''
        ).strip()
        customer_email = (
            text_map.get('車主Email') or text_map.get('Email') or ''
        ).strip()

        # ── 車款 ──────────────────────────────────────
        model_raw = (text_map.get('車輛型號') or '').strip()
        source_product_name = model_raw
        product_id = False
        color_id = False

        if model_raw:
            sku_match = re.search(r'\(([A-Z0-9\-]+)\)', model_raw)
            if sku_match:
                sku = sku_match.group(1)
                product = self.env['dms.product'].search(
                    [('model', '=', sku)], limit=1)
                if not product:
                    product = self.env['dms.product'].search(
                        [('model', 'ilike', sku)], limit=1)
                if product:
                    product_id = product.id

        # ── 顏色 ──────────────────────────────────────
        color_raw = (text_map.get('車輛顏色') or '').strip()
        if color_raw:
            color_name = re.sub(r'\s*\(.*?\)', '', color_raw).strip()
            domain = [('name', 'ilike', color_name)]
            if product_id:
                domain.append(('product_id', '=', product_id))
            color = self.env['dms.product.color'].search(domain, limit=1)
            if color:
                color_id = color.id

        # ── 車行 ──────────────────────────────────────
        dealer_raw = (text_map.get('車行名稱') or '').strip()
        dealer_id = False
        sale_type = 'store'

        if dealer_raw:
            dealer = self.env['dms.dealer'].search(
                [('name', '=', dealer_raw)], limit=1)
            if not dealer:
                dealer = self.env['dms.dealer'].search(
                    [('name', 'ilike', dealer_raw)], limit=1)
            if not dealer:
                # 去除中間空白再比對（如「昌 億」→「昌億」）
                stripped = re.sub(r'\s+', '', dealer_raw)
                if stripped and stripped != dealer_raw:
                    dealer = self.env['dms.dealer'].search(
                        [('name', '=', stripped)], limit=1)
                    if not dealer:
                        dealer = self.env['dms.dealer'].search(
                            [('name', 'ilike', stripped)], limit=1)
            if dealer:
                dealer_id = dealer.id
                sale_type = ('online'
                             if dealer.store_type_id.name == '網路平台'
                             else 'dealer')
            else:
                # 找不到車行 → 嘗試帶入「馭盛」
                fallback = self.env['dms.dealer'].search(
                    [('name', 'ilike', '馭盛')], limit=1)
                if fallback:
                    dealer_id = fallback.id
                    sale_type = 'dealer'
                # 否則維持 store

        # ── 分期 ──────────────────────────────────────
        payment_method = 'cash'
        finance_company = False
        finance_company_other = False
        installment_periods = 0
        installment_raw = (text_map.get('是否有分期') or '').strip()
        finance_raw = (text_map.get('分期公司') or '').strip()

        if any(raw and raw not in ('無', '否', '') for raw in (installment_raw, finance_raw)):
            payment_method = 'installment'

        for raw in (installment_raw, finance_raw):
            match = re.search(r'(\d+)\s*期', raw or '')
            if match:
                installment_periods = int(match.group(1))
                break

        finance_text = finance_raw or installment_raw
        finance_token = re.sub(r'\s+', '', finance_text)
        if '和潤' in finance_token:
            finance_company = '和潤'
        elif '遠信' in finance_token:
            finance_company = '遠信'
        elif '仲信' in finance_token:
            finance_company = '仲信'
        elif (finance_token
              and finance_token not in ('有', '是', '無', '否')
              and not re.fullmatch(r'\d+期', finance_token)):
            finance_company = 'other'
            finance_company_other = finance_text

        # ── 汰舊 ──────────────────────────────────────
        trade_in_raw = (text_map.get('是否有汰舊') or '').strip()
        is_trade_in = bool(trade_in_raw and trade_in_raw not in ('否', '無', ''))

        # ── 配件/備註 ─────────────────────────────────
        extra_other = (text_map.get('配件') or '').strip()
        extra_note = (text_map.get('備註') or '').strip()

        vals = {
            'sale_origin': 'order_processor',
            'source_folder': folder_name,
            'source_product_name': source_product_name,
            'customer_name': customer_name,
            'id_number': id_number,
            'address_registered': address_registered,
            'customer_phone': customer_phone,
            'customer_email': customer_email,
            'product_id': product_id,
            'color_id': color_id,
            'dealer_id': dealer_id,
            'sale_type': sale_type,
            'payment_method': payment_method,
            'finance_company': finance_company,
            'finance_company_other': finance_company_other,
            'installment_periods': installment_periods,
            'is_trade_in': is_trade_in,
            'extra_other': extra_other or False,
            'extra_note': extra_note or False,
        }
        if birthday_ad:
            vals['birthday_ad'] = birthday_ad

        return vals

    # ── 內部：寫 log ──────────────────────────────────
    def _write_log(self, folder_name, state, order_id=None, error_msg=None):
        vals = {
            'folder_name': folder_name,
            'state': state,
        }
        if order_id:
            vals['order_id'] = order_id
        if error_msg:
            vals['error_msg'] = error_msg
        self.env['dms.sync.log'].create(vals)

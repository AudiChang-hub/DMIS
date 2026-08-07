from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.contrib.auth.models import AnonymousUser
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.test import RequestFactory, TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook
from PIL import Image
from pypdf import PdfReader, PdfWriter

from sales.forms import (
    AccessoryFormSet,
    AccessoryLineForm,
    OtherFeeFormSet,
    OtherFeeLineForm,
    PaymentRecordForm,
    SalesOrderForm,
)
from sales.models import (
    AccessoryProduct,
    AccessoryLine,
    OrderChange,
    OrderEvent,
    OrderOperationsProfile,
    OtherFeeLine,
    RegistrationDocument,
    PaymentRecord,
    SalesOrder,
    SalesSource,
    Store,
    SubsidyDocument,
    VehicleColor,
    VehicleInventory,
    VehicleInventoryHistory,
    VehicleIncentiveInstallmentRate,
    VehicleIncentiveRule,
    VehicleModel,
    VehicleSettlementCostRule,
)
from sales.services.secret_fields import decrypt_secret


def uploaded_test_pdf(filename="document.pdf"):
    stream = BytesIO()
    writer = PdfWriter()
    writer.add_blank_page(width=300, height=200)
    writer.write(stream)
    return SimpleUploadedFile(
        filename, stream.getvalue(), content_type="application/pdf"
    )


def uploaded_test_jpeg(filename="photo.jpg"):
    stream = BytesIO()
    Image.new("RGB", (32, 24), "white").save(stream, format="JPEG")
    return SimpleUploadedFile(filename, stream.getvalue(), content_type="image/jpeg")


def uploaded_test_xlsx(filename="document.xlsx"):
    stream = BytesIO()
    workbook = Workbook()
    workbook.active["A1"] = "測試"
    workbook.save(stream)
    return SimpleUploadedFile(
        filename,
        stream.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )


class OrderFlowTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username="tester", password="test-pass-123"
        )
        self.store_a = Store.objects.create(name="總店", code="HQ")
        self.store_b = Store.objects.create(name="二店", code="B02")
        self.model = VehicleModel.objects.create(
            brand="測試廠牌", name="通勤 125", energy_type=VehicleModel.EnergyType.GAS
        )
        self.color = VehicleColor.objects.create(
            vehicle_model=self.model, name="白"
        )
        self.settlement_rule = VehicleSettlementCostRule.objects.create(
            vehicle_model=self.model,
            amount=Decimal("60000"),
            effective_from=date(2026, 1, 1),
        )
        self.vehicle = VehicleInventory.objects.create(
            vehicle_model=self.model,
            color=self.color,
            engine_number="ENG-001",
            ownership_store=self.store_b,
            location_store=self.store_b,
        )

    def make_order(self, signed=False):
        order = SalesOrder(
            owner_name="王小明",
            owner_phone="0912345678",
            owner_address="新北市測試區",
            owner_id_number="A123456789",
            vehicle_model=self.model,
            color=self.color,
            vehicle_price=Decimal("79800"),
            deposit_amount=Decimal("4600"),
            actual_balance=Decimal("75200"),
            id_verified=True,
            status=SalesOrder.Status.ALLOCATION_PENDING,
        )
        if signed:
            order.signed_contract = SimpleUploadedFile(
                "signed.pdf", b"signed contract", content_type="application/pdf"
            )
            order.status = SalesOrder.Status.ALLOCATION_PENDING
        order.save()
        order.calculated_balance = order.calculate_balance()
        order.save(update_fields=["calculated_balance", "updated_at"])
        return order

    def test_unsigned_contract_can_allocate(self):
        order = self.make_order(signed=False)

        order.allocate(self.vehicle)

        self.vehicle.refresh_from_db()
        order.refresh_from_db()
        self.assertEqual(self.vehicle.status, VehicleInventory.Status.RESERVED)
        self.assertEqual(order.status, SalesOrder.Status.ALLOCATED)

    def test_edit_page_available_until_delivery(self):
        order = self.make_order()
        self.client.force_login(self.user)

        response = self.client.get(reverse("order_edit", args=[order.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "編輯訂單")
        self.assertContains(response, "變更原因")
        self.assertContains(response, "取消修改")
        self.assertContains(response, "data-cancel-edit")
        self.assertContains(response, "beforeunload")
        self.assertContains(response, "尚有未儲存的修改")
        self.assertContains(response, 'name="registration_date"')
        self.assertContains(response, 'type="date"')

        order.status = SalesOrder.Status.DELIVERED_DOCS_PENDING
        order.save(update_fields=["status", "updated_at"])
        locked = self.client.get(reverse("order_edit", args=[order.pk]))
        self.assertRedirects(locked, reverse("order_detail", args=[order.pk]))

    def test_order_detail_uses_consistent_section_spacing(self):
        css = (
            __import__("pathlib").Path("static/css/app.css").read_text(
                encoding="utf-8"
            )
        )
        self.assertIn(
            ".detail-grid { display: grid; grid-template-columns: 1fr 1fr; margin-bottom: 20px; gap: 20px; }",
            css,
        )

    def test_desktop_and_mobile_use_distinct_page_shells(self):
        from pathlib import Path

        css = Path("static/css/app.css").read_text(encoding="utf-8")
        base = Path("templates/base.html").read_text(encoding="utf-8")
        detail = Path("templates/sales/order_detail.html").read_text(encoding="utf-8")
        form = Path("templates/sales/order_form.html").read_text(encoding="utf-8")

        self.assertIn("--shell-wide: 1680px", css)
        self.assertIn(".page-shell--wide { max-width: var(--shell-wide); }", css)
        self.assertIn("width: min(100% - 24px, 700px)", css)
        self.assertIn('class="app-header-inner"', base)
        self.assertIn("app.css' %}?v={{ app_version }}", base)
        self.assertIn("app-update.js' %}?v={{ app_version }}", base)
        self.assertIn("page-shell--wide", detail)
        self.assertIn("page-shell--form", form)
        self.assertIn(
            "grid-template-columns: repeat(6, minmax(0, 1fr));",
            css,
        )
        self.assertIn("overflow-x: auto;", css)

    def test_primary_pages_have_consistent_context_navigation(self):
        self.client.force_login(self.user)
        order = self.make_order()

        for route in (
            reverse("order_list"),
            reverse("inventory_list"),
            reverse("inventory_create"),
            reverse("order_create"),
            reverse("order_detail", args=[order.pk]),
            reverse("order_edit", args=[order.pk]),
        ):
            with self.subTest(route=route):
                response = self.client.get(route)
                self.assertEqual(response.status_code, 200)
                self.assertContains(response, 'class="page-context-nav"')
                self.assertContains(response, "data-smart-back")

        dashboard = self.client.get(reverse("dashboard"))
        self.assertNotContains(dashboard, 'class="page-context-nav"')

    def inventory_payload(self, vehicle=None, **overrides):
        vehicle = vehicle or self.vehicle
        payload = {
            "vehicle_model": vehicle.vehicle_model_id,
            "color": vehicle.color_id,
            "engine_number": vehicle.engine_number or "",
            "frame_number": vehicle.frame_number or "",
            "ownership_store": vehicle.ownership_store_id,
            "location_store": vehicle.location_store_id,
            "received_on": vehicle.received_on.isoformat(),
            "condition_note": vehicle.condition_note,
            "condition_resolution": vehicle.condition_resolution,
        }
        payload.update(overrides)
        return payload

    def test_available_inventory_can_be_edited(self):
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("inventory_edit", args=[self.vehicle.pk]),
            self.inventory_payload(
                engine_number="eng-updated",
                location_store=self.store_a.pk,
                condition_note="到店檢查正常",
            ),
        )

        self.assertRedirects(response, reverse("inventory_list"))
        self.vehicle.refresh_from_db()
        self.assertEqual(self.vehicle.engine_number, "ENG-UPDATED")
        self.assertEqual(self.vehicle.location_store, self.store_a)
        self.assertEqual(self.vehicle.condition_note, "到店檢查正常")
        history = self.vehicle.history_entries.get()
        self.assertEqual(
            history.event_type,
            VehicleInventoryHistory.EventType.TRANSFERRED,
        )
        self.assertEqual(history.from_location, self.store_b)
        self.assertEqual(history.to_location, self.store_a)
        self.assertEqual(history.actor_name, "tester")
        self.assertEqual(history.reason, "")
        self.assertEqual(
            history.condition_note_snapshot,
            "到店檢查正常",
        )

    def test_inventory_edit_records_optional_reason_and_friendly_changes(self):
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("inventory_edit", args=[self.vehicle.pk]),
            self.inventory_payload(
                condition_note="右側車殼刮傷",
                condition_resolution="已安排補漆",
                change_reason="到店複檢",
            ),
        )

        self.assertRedirects(response, reverse("inventory_list"))
        history = self.vehicle.history_entries.get()
        self.assertEqual(
            history.event_type,
            VehicleInventoryHistory.EventType.UPDATED,
        )
        self.assertEqual(history.reason, "到店複檢")
        self.assertEqual(history.changes["condition_note"]["label"], "車況說明")
        self.assertEqual(history.changes["condition_note"]["after"], "右側車殼刮傷")

    def test_vehicle_model_edit_includes_price_version_and_motor_fields(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse("vehicle_model_edit", args=[self.model.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "售價版本")
        self.assertContains(response, 'name="motor_power_kw"')
        self.assertContains(response, 'name="horsepower_hp"')
        self.assertContains(response, 'name="price-suggested_retail_price"')

    def test_quick_inventory_entry_includes_manufactured_year_month(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse("inventory_quick_create"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "出廠年月")
        self.assertContains(response, 'name="vehicles-0-manufactured_year_month"')

    def test_inventory_create_hides_ownership_and_sets_internal_compatibility_value(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("inventory_create"),
            {
                "vehicle_model": self.model.pk,
                "color": self.color.pk,
                "engine_number": "ENG-NEW",
                "frame_number": "",
                "location_store": self.store_a.pk,
                "received_on": "2026-07-30",
                "condition_note": "",
                "condition_resolution": "",
            },
        )

        self.assertRedirects(response, reverse("inventory_list"))
        vehicle = VehicleInventory.objects.get(engine_number="ENG-NEW")
        self.assertEqual(vehicle.ownership_store, self.store_a)
        self.assertEqual(
            vehicle.history_entries.get().event_type,
            VehicleInventoryHistory.EventType.CREATED,
        )

    def quick_inventory_payload(self, rows):
        payload = {
            "vehicles-TOTAL_FORMS": str(max(5, len(rows))),
            "vehicles-INITIAL_FORMS": "0",
            "vehicles-MIN_NUM_FORMS": "0",
            "vehicles-MAX_NUM_FORMS": "100",
        }
        for index in range(max(5, len(rows))):
            row = rows[index] if index < len(rows) else {}
            for field in (
                "vehicle_model",
                "color",
                "identifier",
                "received_on",
                "condition_note",
            ):
                payload[f"vehicles-{index}-{field}"] = row.get(field, "")
        return payload

    def test_quick_inventory_entry_creates_gas_and_electric_rows_atomically(self):
        electric_model = VehicleModel.objects.create(
            brand="測試廠牌",
            name="電動車",
            energy_type=VehicleModel.EnergyType.ELECTRIC,
        )
        electric_color = VehicleColor.objects.create(
            vehicle_model=electric_model,
            name="銀",
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("inventory_quick_create"),
            self.quick_inventory_payload(
                [
                    {
                        "vehicle_model": self.model.pk,
                        "color": self.color.pk,
                        "identifier": " quick-eng-01 ",
                        "received_on": "2026-07-30",
                        "condition_note": "外觀正常",
                    },
                    {
                        "vehicle_model": electric_model.pk,
                        "color": electric_color.pk,
                        "identifier": "frame-quick-02",
                        "received_on": "2026-07-30",
                        "condition_note": "",
                    },
                ]
            ),
        )

        self.assertRedirects(response, reverse("inventory_list"))
        gas = VehicleInventory.objects.get(engine_number="QUICK-ENG-01")
        electric = VehicleInventory.objects.get(frame_number="FRAME-QUICK-02")
        self.assertEqual(gas.location_store, self.store_a)
        self.assertEqual(gas.ownership_store, self.store_a)
        self.assertEqual(gas.condition_note, "外觀正常")
        self.assertIsNone(electric.engine_number)
        self.assertEqual(gas.history_entries.count(), 1)
        self.assertEqual(electric.history_entries.count(), 1)

    def test_quick_inventory_entry_marks_duplicate_row_and_creates_nothing(self):
        self.client.force_login(self.user)
        before_count = VehicleInventory.objects.count()

        response = self.client.post(
            reverse("inventory_quick_create"),
            self.quick_inventory_payload(
                [
                    {
                        "vehicle_model": self.model.pk,
                        "color": self.color.pk,
                        "identifier": "BATCH-DUPLICATE",
                        "received_on": "2026-07-30",
                    },
                    {
                        "vehicle_model": self.model.pk,
                        "color": self.color.pk,
                        "identifier": "batch-duplicate",
                        "received_on": "2026-07-30",
                    },
                ]
            ),
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "與第 1 列重複")
        self.assertEqual(VehicleInventory.objects.count(), before_count)

    def test_quick_inventory_entry_rejects_existing_identifier(self):
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("inventory_quick_create"),
            self.quick_inventory_payload(
                [
                    {
                        "vehicle_model": self.model.pk,
                        "color": self.color.pk,
                        "identifier": self.vehicle.engine_number,
                        "received_on": "2026-07-30",
                    }
                ]
            ),
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "此號碼已存在於庫存資料")
        self.assertEqual(VehicleInventory.objects.count(), 1)

    def test_quick_inventory_page_is_list_based_and_hides_location(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse("inventory_quick_create"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "快速進車")
        self.assertContains(response, 'id="quick-entry-rows"')
        self.assertContains(response, "車況說明（選填）")
        self.assertContains(response, "刪除此列")
        self.assertNotContains(response, 'name="location_store"')
        self.assertNotContains(response, "複製")

    def vehicle_model_master_payload(self, **overrides):
        payload = {
            "brand": "SUZUKI",
            "energy_type": VehicleModel.EnergyType.GAS,
            "name": "SUI 125",
            "model_number": "SUI125-ABS",
            "model_year": "2026",
            "model_code": VehicleModel.ModelType.CBS_DISC,
            "displacement_cc": "125",
            "suggested_price": "79800",
            "active": "on",
            "colors-TOTAL_FORMS": "2",
            "colors-INITIAL_FORMS": "0",
            "colors-MIN_NUM_FORMS": "0",
            "colors-MAX_NUM_FORMS": "1000",
            "colors-0-name": "白",
            "colors-0-active": "on",
            "colors-1-name": "灰",
            "colors-1-active": "on",
        }
        payload.update(overrides)
        return payload

    def test_vehicle_model_color_rows_default_to_one_and_edit_has_no_blank_extra(self):
        self.client.force_login(self.user)

        create_response = self.client.get(reverse("vehicle_model_create"))
        edit_response = self.client.get(
            reverse("vehicle_model_edit", args=[self.model.pk])
        )

        self.assertEqual(
            create_response.context["color_formset"].total_form_count(),
            1,
        )
        self.assertEqual(
            edit_response.context["color_formset"].total_form_count(),
            self.model.colors.count(),
        )

    def test_vehicle_model_master_create_and_list_preserve_shared_records(self):
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("vehicle_model_create"),
            self.vehicle_model_master_payload(),
        )

        self.assertRedirects(response, reverse("vehicle_model_list"))
        model = VehicleModel.objects.get(
            brand="SUZUKI",
            name="SUI 125",
            model_year=2026,
            model_number="SUI125-ABS",
            model_code=VehicleModel.ModelType.CBS_DISC,
        )
        self.assertEqual(model.suggested_price, Decimal("79800"))
        self.assertEqual(model.model_number, "SUI125-ABS")
        self.assertEqual(
            str(model),
            "SUZUKI／SUI 125／SUI125-ABS／2026／CBS碟",
        )
        self.assertEqual(
            list(model.colors.order_by("name").values_list("name", flat=True)),
            ["灰", "白"],
        )
        self.vehicle.refresh_from_db()
        self.assertEqual(self.vehicle.vehicle_model, self.model)

        list_response = self.client.get(
            reverse("vehicle_model_list"),
            {"q": "SUI125-ABS", "energy_type": VehicleModel.EnergyType.GAS},
        )
        self.assertContains(list_response, "SUI 125")
        self.assertContains(list_response, "SUI125-ABS")
        self.assertContains(list_response, "2 種")

    def test_vehicle_model_number_is_required_when_maintaining_master(self):
        self.client.force_login(self.user)
        payload = self.vehicle_model_master_payload()
        payload["model_number"] = ""

        response = self.client.post(reverse("vehicle_model_create"), payload)

        self.assertEqual(response.status_code, 200)
        self.assertIn("model_number", response.context["form"].errors)
        self.assertFalse(
            VehicleModel.objects.filter(brand="SUZUKI", name="SUI 125").exists()
        )

    def test_vehicle_model_with_inventory_cannot_change_energy_type(self):
        self.client.force_login(self.user)
        color = self.model.colors.get()
        response = self.client.post(
            reverse("vehicle_model_edit", args=[self.model.pk]),
            {
                "brand": self.model.brand,
                "energy_type": VehicleModel.EnergyType.ELECTRIC,
                "name": self.model.name,
                "model_number": "TEST-125",
                "model_year": "2026",
                "model_code": VehicleModel.ModelType.ABS_DISC,
                "displacement_cc": "",
                "suggested_price": "",
                "active": "on",
                "colors-TOTAL_FORMS": "1",
                "colors-INITIAL_FORMS": "1",
                "colors-MIN_NUM_FORMS": "0",
                "colors-MAX_NUM_FORMS": "1000",
                "colors-0-id": color.pk,
                "colors-0-name": color.name,
                "colors-0-active": "on",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "不能變更能源別")
        self.model.refresh_from_db()
        self.assertEqual(self.model.energy_type, VehicleModel.EnergyType.GAS)

    def test_used_vehicle_color_cannot_be_deleted_but_can_be_deactivated(self):
        self.client.force_login(self.user)
        color = self.color
        base = {
            "brand": self.model.brand,
            "energy_type": self.model.energy_type,
            "name": self.model.name,
            "model_number": "TEST-125",
            "model_year": "2026",
            "model_code": VehicleModel.ModelType.FRONT_DISC_REAR_DRUM,
            "displacement_cc": "125",
            "suggested_price": "",
            "active": "on",
            "colors-TOTAL_FORMS": "1",
            "colors-INITIAL_FORMS": "1",
            "colors-MIN_NUM_FORMS": "0",
            "colors-MAX_NUM_FORMS": "1000",
            "colors-0-id": color.pk,
            "colors-0-name": color.name,
            "colors-0-active": "on",
            "colors-0-DELETE": "on",
        }
        response = self.client.post(
            reverse("vehicle_model_edit", args=[self.model.pk]),
            base,
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "不能刪除")
        self.assertTrue(VehicleColor.objects.filter(pk=color.pk).exists())

        base.pop("colors-0-DELETE")
        base.pop("colors-0-active")
        response = self.client.post(
            reverse("vehicle_model_edit", args=[self.model.pk]),
            base,
        )
        self.assertRedirects(response, reverse("vehicle_model_list"))
        color.refresh_from_db()
        self.assertFalse(color.active)

    def test_reserved_inventory_locks_core_fields_but_allows_condition_updates(self):
        self.vehicle.status = VehicleInventory.Status.RESERVED
        self.vehicle.save(update_fields=["status", "updated_at"])
        other_model = VehicleModel.objects.create(
            brand="其他", name="不可換車型", energy_type=VehicleModel.EnergyType.GAS
        )
        other_color = VehicleColor.objects.create(
            vehicle_model=other_model, name="黑"
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("inventory_edit", args=[self.vehicle.pk]),
            self.inventory_payload(
                vehicle_model=other_model.pk,
                color=other_color.pk,
                engine_number="TAMPERED",
                ownership_store=self.store_a.pk,
                location_store=self.store_a.pk,
                condition_note="調車前發現刮痕",
                condition_resolution="待確認",
            ),
        )

        self.assertRedirects(response, reverse("inventory_list"))
        self.vehicle.refresh_from_db()
        self.assertEqual(self.vehicle.vehicle_model, self.model)
        self.assertEqual(self.vehicle.color, self.color)
        self.assertEqual(self.vehicle.engine_number, "ENG-001")
        self.assertEqual(self.vehicle.ownership_store, self.store_b)
        self.assertEqual(self.vehicle.location_store, self.store_a)
        self.assertEqual(self.vehicle.condition_note, "調車前發現刮痕")

    def test_inventory_list_is_filterable_sortable_table(self):
        second = VehicleInventory.objects.create(
            vehicle_model=self.model,
            color=self.color,
            engine_number="ENG-999",
            ownership_store=self.store_a,
            location_store=self.store_a,
            received_on=date(2026, 1, 1),
            status=VehicleInventory.Status.CONDITION_ISSUE,
            condition_note="測試異常",
        )
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("inventory_list"),
            {
                "q": "ENG-999",
                "status": VehicleInventory.Status.CONDITION_ISSUE,
                "ownership_store": self.store_a.pk,
                "sort": "identifier",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(list(response.context["vehicles"]), [second])
        self.assertContains(response, 'class="inventory-table"')
        self.assertContains(response, "查看／編輯")
        self.assertNotContains(response, 'class="inventory-card"')
        self.assertContains(response, 'name="vehicle_model"')
        self.assertContains(response, 'name="color"')
        self.assertContains(response, 'name="location_store"')
        self.assertContains(response, 'name="sort"')
        self.assertNotContains(response, "庫存歸屬")
        self.assertNotContains(response, 'name="ownership_store"')

    def test_inventory_edit_page_explains_locked_fields(self):
        self.vehicle.status = VehicleInventory.Status.DELIVERY_PENDING
        self.vehicle.save(update_fields=["status", "updated_at"])
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("inventory_edit", args=[self.vehicle.pk])
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "編輯庫存")
        self.assertContains(response, "已進入配車或交付流程")
        self.assertContains(response, "儲存修改")
        self.assertContains(response, 'name="engine_number"')
        self.assertContains(response, "disabled")

    def test_delivered_inventory_locks_location_but_allows_resolution(self):
        self.vehicle.status = VehicleInventory.Status.DELIVERED
        self.vehicle.save(update_fields=["status", "updated_at"])
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("inventory_edit", args=[self.vehicle.pk]),
            self.inventory_payload(
                location_store=self.store_a.pk,
                condition_resolution="交付後補登檢查結果",
            ),
        )

        self.assertRedirects(response, reverse("inventory_list"))
        self.vehicle.refresh_from_db()
        self.assertEqual(self.vehicle.location_store, self.store_b)
        self.assertEqual(
            self.vehicle.condition_resolution,
            "交付後補登檢查結果",
        )

    def test_inventory_list_ignores_malformed_filter_ids(self):
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("inventory_list"),
            {
                "vehicle_model": "not-a-number",
                "color": "bad",
                "ownership_store": "invalid",
                "location_store": "invalid",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.vehicle.identifier)

    def test_navigation_refactor_preserves_critical_order_controls(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse("order_create"))

        self.assertContains(response, 'data-photo-preview="id_front"')
        self.assertContains(response, 'data-photo-preview="id_back"')
        self.assertContains(response, 'data-clear-file="id_id_front"')
        self.assertContains(response, 'data-clear-file="id_id_back"')
        self.assertContains(response, "刪除此筆配件")
        self.assertContains(response, "刪除此筆費用")
        self.assertContains(response, 'id="add-accessory"')
        self.assertContains(response, 'id="add-other-fee"')

    def test_dashboard_searches_all_business_and_related_fields(self):
        order = self.make_order()
        order.owner_email = "owner@example.com"
        order.delivery_destination = "汐止區康寧街測試倉庫"
        order.old_vehicle_tax = Decimal("4321")
        order.actual_balance += Decimal("4321")
        order.save()
        AccessoryLine.objects.create(
            order=order,
            name="專用後靠背",
            quantity=1,
            line_type=AccessoryLine.LineType.PURCHASE,
            amount=Decimal("876"),
            note="客戶指定黑色",
        )
        OtherFeeLine.objects.create(
            order=order, name="特殊文件代辦", amount=Decimal("345")
        )
        OrderEvent.objects.create(
            order=order,
            event_type="manual_note",
            description="已電話確認週六交車",
            actor_name="Sylvia",
        )
        RegistrationDocument.objects.create(
            order=order,
            document_type=RegistrationDocument.DocumentType.OTHER_INSURANCE,
            name="第三人責任險",
            file=SimpleUploadedFile("liability.pdf", b"pdf"),
            uploaded_by="admin",
        )
        self.client.force_login(self.user)

        cases = (
            ("owner@example.com", "Email", "owner@example.com"),
            ("康寧街測試倉庫", "送達地點／託運目的地", "汐止區康寧街測試倉庫"),
            ("4321", "舊車未繳稅金", "4321"),
            ("專用後靠背", "配件名稱", "專用後靠背"),
            ("特殊文件代辦", "其他費用項目", "特殊文件代辦"),
            ("週六交車", "處理紀錄內容", "已電話確認週六交車"),
            ("第三人責任險", "領牌文件名稱", "第三人責任險"),
            ("待配車", "訂單狀態", "待配車"),
        )
        for query, label, value in cases:
            with self.subTest(query=query):
                response = self.client.get(reverse("dashboard"), {"q": query})
                self.assertEqual(response.status_code, 200)
                self.assertContains(response, order.owner_name)
                self.assertContains(response, label)
                self.assertContains(response, value)

    def test_dashboard_search_masks_sensitive_match_values(self):
        order = self.make_order()
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("dashboard"), {"q": order.owner_id_number}
        )

        self.assertContains(response, "證件號碼／統一編號")
        self.assertContains(response, order.masked_id_number)
        self.assertNotContains(response, f"<mark>{order.owner_id_number}</mark>")

    def test_dashboard_search_paginates_without_hiding_total_matches(self):
        SalesOrder.objects.bulk_create(
            [
                SalesOrder(
                    number=f"SO-SEARCH-{index:03d}",
                    owner_name=f"批次搜尋車主 {index}",
                    owner_phone=f"0900{index:06d}",
                    owner_address="測試地址",
                    owner_id_number=f"T{index:09d}",
                    vehicle_model=self.model,
                    color=self.color,
                    status=SalesOrder.Status.ALLOCATION_PENDING,
                )
                for index in range(51)
            ]
        )
        # bulk_create 不觸發 Django signals；大量匯入後必須執行正式的索引重建程序。
        call_command("rebuild_order_search_index", verbosity=0)
        self.client.force_login(self.user)

        first_page = self.client.get(
            reverse("dashboard"), {"q": "批次搜尋車主"}
        )
        second_page = self.client.get(
            reverse("dashboard"), {"q": "批次搜尋車主", "page": 2}
        )

        self.assertEqual(first_page.context["search_result_count"], 51)
        self.assertEqual(len(first_page.context["search_results"]), 50)
        self.assertContains(first_page, "共 51 筆")
        self.assertContains(first_page, "下一頁")
        self.assertEqual(len(second_page.context["search_results"]), 1)
        self.assertContains(second_page, "上一頁")

    def test_forms_provide_field_specific_mobile_keyboard_hints(self):
        order_form = SalesOrderForm()
        accessory_form = AccessoryLineForm()
        fee_form = OtherFeeLineForm()

        self.assertEqual(
            order_form.fields["owner_name"].widget.attrs["lang"], "zh-Hant"
        )
        self.assertEqual(
            order_form.fields["owner_phone"].widget.attrs["inputmode"], "tel"
        )
        self.assertEqual(
            order_form.fields["owner_email"].widget.attrs["inputmode"], "email"
        )
        self.assertEqual(
            order_form.fields["owner_id_number"].widget.attrs["lang"], "en"
        )
        self.assertEqual(
            order_form.fields["vehicle_price"].widget.attrs["inputmode"],
            "decimal",
        )
        self.assertEqual(
            accessory_form.fields["quantity"].widget.attrs["inputmode"],
            "numeric",
        )
        self.assertEqual(
            accessory_form.fields["amount"].widget.attrs["readonly"], True
        )
        self.assertEqual(fee_form.fields["name"].widget.attrs["lang"], "zh-Hant")

    def test_cash_order_clears_installment_details_and_excludes_opening_fee(self):
        order = self.make_order()
        order.payment_type = SalesOrder.PaymentType.CASH
        order.installment_company = "和潤"
        order.installment_amount = Decimal("75000")
        order.installment_periods = 24
        order.installment_opening_fee = Decimal("2500")
        order.installment_monthly = Decimal("3125")

        order.save()
        order.refresh_from_db()

        self.assertEqual(order.installment_company, "")
        self.assertEqual(order.installment_amount, 0)
        self.assertEqual(order.installment_periods, 0)
        self.assertEqual(order.installment_opening_fee, 0)
        self.assertEqual(order.installment_monthly, 0)
        self.assertEqual(order.calculated_balance, Decimal("75200"))

    def test_installment_order_includes_opening_fee(self):
        order = self.make_order()
        order.payment_type = SalesOrder.PaymentType.INSTALLMENT
        order.installment_company = "和潤"
        order.installment_periods = 24
        order.installment_opening_fee = Decimal("2500")
        order.installment_monthly = Decimal("3125")
        order.actual_balance = Decimal("77700")

        order.save()
        order.refresh_from_db()

        self.assertEqual(order.installment_opening_fee, Decimal("2500"))
        self.assertEqual(order.calculated_balance, Decimal("77700"))

    def test_payment_type_change_warns_before_clearing_installment_fields(self):
        template = (
            __import__("pathlib").Path("templates/sales/order_form.html").read_text(
                encoding="utf-8"
            )
        )

        self.assertIn("將清除分期公司、期數、開辦費與每期金額", template)
        self.assertIn('paymentType.value !== "installment"', template)
        self.assertIn('input.value = ""', template)

    def test_edit_order_records_reason_and_before_after_values(self):
        order = self.make_order()
        order.owner_type = SalesOrder.OwnerType.COMPANY
        order.id_front = uploaded_test_jpeg("old-front.jpg")
        order.save(update_fields=["owner_type", "id_front", "updated_at"])
        old_identity_path = Path(order.id_front.path)
        self.assertTrue(old_identity_path.exists())
        self.client.force_login(self.user)
        data = {
            "_order_revision": str(order.revision),
            "source_type": "store",
            "source": "",
            "owner_type": "company",
            "owner_name": "王小明有限公司",
            "owner_name_en": "",
            "owner_phone": order.owner_phone,
            "owner_email": "",
            "owner_birth_date": "",
            "owner_nationality": "",
            "owner_address": order.owner_address,
            "owner_id_number": order.owner_id_number,
            "residence_expiry": "",
            "id_verified": "on",
            "id_front-clear": "on",
            "vehicle_model": str(self.model.pk),
            "color": str(self.color.pk),
            "vehicle_category": "new",
            "payment_type": "cash",
            "vehicle_price": "80000",
            "plate_insurance_fee": "0",
            "installment_opening_fee": "0",
            "deposit_amount": "4600",
            "deposit_date": str(timezone.localdate()),
            "deposit_method": "",
            "installment_company": "",
            "installment_periods": "0",
            "installment_monthly": "0",
            "is_trade_in_subsidy": "",
            "trade_in_plate": "",
            "old_owner_name": "",
            "subsidy_type": "",
            "old_vehicle_valuation": "0",
            "old_vehicle_tax": "0",
            "plate_choice": "none",
            "watched_numbers": "",
            "plate_preference_note": "",
            "delivery_method": "store_pickup",
            "delivery_destination": "",
            "note": "",
            "change_reason": "客戶要求更正公司名稱",
            "accessories-TOTAL_FORMS": "1",
            "accessories-INITIAL_FORMS": "0",
            "accessories-MIN_NUM_FORMS": "0",
            "accessories-MAX_NUM_FORMS": "1000",
            "accessories-0-name": "",
            "accessories-0-quantity": "",
            "accessories-0-line_type": "",
            "accessories-0-amount": "",
            "other_fees-TOTAL_FORMS": "1",
            "other_fees-INITIAL_FORMS": "0",
            "other_fees-MIN_NUM_FORMS": "0",
            "other_fees-MAX_NUM_FORMS": "1000",
            "other_fees-0-name": "",
            "other_fees-0-amount": "",
        }

        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(reverse("order_edit", args=[order.pk]), data)

        self.assertRedirects(response, reverse("order_detail", args=[order.pk]))
        order.refresh_from_db()
        self.assertFalse(order.id_front)
        self.assertFalse(old_identity_path.exists())
        self.assertEqual(order.owner_name, "王小明有限公司")
        self.assertEqual(order.revision, 2)
        self.assertEqual(order.actual_balance, Decimal("75400"))
        self.assertEqual(order.balance_adjustment_reason, "客戶要求更正公司名稱")
        change = order.changes.get()
        self.assertEqual(change.reason, "客戶要求更正公司名稱")
        self.assertEqual(
            change.changes["車主姓名／公司名稱"],
            {"before": "王小明", "after": "王小明有限公司"},
        )

        detail = self.client.get(reverse("order_detail", args=[order.pk]))
        self.assertContains(detail, "tester")
        self.assertContains(detail, "修改了")
        self.assertContains(detail, "客戶要求更正公司名稱")
        self.assertContains(detail, "王小明有限公司")
        self.assertNotContains(
            detail,
            '<strong class="change-label">owner_name</strong>',
            html=True,
        )

    def test_change_history_translates_codes_money_and_line_items(self):
        order = self.make_order()
        OrderChange.objects.create(
            order=order,
            actor_name="王行政",
            reason="依客戶確認內容調整",
            changes={
                "主要付款方式": {"before": "installment", "after": "cash"},
                "訂金": {"before": "5000", "after": "8000"},
                "選號方式": {"before": "none", "after": "preference"},
                "配件": {
                    "before": [],
                    "after": [
                        {
                            "名稱": "手機架",
                            "數量": 1,
                            "類型": "加購",
                            "金額": "850",
                            "安裝日期": "",
                            "備註": "",
                        }
                    ],
                },
            },
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse("order_detail", args=[order.pk]))

        self.assertContains(response, "王行政")
        self.assertContains(response, "依客戶確認內容調整")
        self.assertContains(response, "分期")
        self.assertContains(response, "現金")
        self.assertContains(response, "$5,000")
        self.assertContains(response, "$8,000")
        self.assertContains(response, "一般領牌偏好")
        self.assertContains(response, "手機架 × 1")
        self.assertNotContains(response, "installment")
        self.assertNotContains(response, "preference")
        self.assertNotContains(response, "&#x27;名稱&#x27;")

    def test_signed_contract_can_allocate_and_locks_vehicle(self):
        order = self.make_order(signed=True)

        order.allocate(self.vehicle)

        order.refresh_from_db()
        self.vehicle.refresh_from_db()
        self.assertEqual(self.vehicle.status, VehicleInventory.Status.RESERVED)
        self.assertEqual(order.allocated_vehicle, self.vehicle)
        self.assertEqual(order.status, SalesOrder.Status.ALLOCATED)

    def test_reallocation_atomically_releases_original_and_reserves_new_vehicle(self):
        order = self.make_order()
        order.allocate(self.vehicle)
        replacement = VehicleInventory.objects.create(
            vehicle_model=self.model,
            color=self.color,
            engine_number="ENG-002",
            ownership_store=self.store_a,
            location_store=self.store_a,
        )
        self.client.force_login(self.user)

        detail = self.client.get(reverse("order_detail", args=[order.pk]))
        self.assertContains(detail, "更換配車")
        response = self.client.post(
            reverse("reallocate_vehicle", args=[order.pk]),
            {
                "vehicle": replacement.pk,
                "reason": "原車車況異常",
            },
        )

        self.assertRedirects(
            response,
            f"{reverse('order_detail', args=[order.pk])}?tab=allocation",
        )
        order.refresh_from_db()
        self.vehicle.refresh_from_db()
        replacement.refresh_from_db()
        self.assertEqual(order.allocated_vehicle, replacement)
        self.assertEqual(
            self.vehicle.status, VehicleInventory.Status.AVAILABLE
        )
        self.assertEqual(
            replacement.status, VehicleInventory.Status.RESERVED
        )
        change = order.changes.latest("created_at")
        self.assertEqual(change.reason, "原車車況異常")
        self.assertEqual(
            change.changes["已配車輛"]["before"], str(self.vehicle)
        )
        self.assertEqual(
            change.changes["已配車輛"]["after"], str(replacement)
        )

    def test_reallocation_is_blocked_after_registration_has_started(self):
        order = self.make_order()
        order.allocate(self.vehicle)
        order.registration_date = timezone.localdate()
        order.final_plate_number = "ABC-1234"
        order.save(
            update_fields=[
                "registration_date",
                "final_plate_number",
                "updated_at",
            ]
        )
        replacement = VehicleInventory.objects.create(
            vehicle_model=self.model,
            color=self.color,
            engine_number="ENG-003",
            ownership_store=self.store_a,
            location_store=self.store_a,
        )
        self.client.force_login(self.user)

        detail = self.client.get(reverse("order_detail", args=[order.pk]))
        self.assertContains(detail, "目前無法改配")
        response = self.client.post(
            reverse("reallocate_vehicle", args=[order.pk]),
            {
                "vehicle": replacement.pk,
                "reason": "嘗試改配",
            },
        )

        self.assertRedirects(
            response,
            f"{reverse('order_detail', args=[order.pk])}?tab=allocation",
        )
        order.refresh_from_db()
        self.vehicle.refresh_from_db()
        replacement.refresh_from_db()
        self.assertEqual(order.allocated_vehicle, self.vehicle)
        self.assertEqual(
            self.vehicle.status, VehicleInventory.Status.RESERVED
        )
        self.assertEqual(
            replacement.status, VehicleInventory.Status.AVAILABLE
        )

    def test_registration_date_alone_does_not_block_reallocation(self):
        order = self.make_order()
        order.allocate(self.vehicle)
        order.registration_date = timezone.localdate()
        order.save(update_fields=["registration_date", "updated_at"])
        replacement = VehicleInventory.objects.create(
            vehicle_model=self.model,
            color=self.color,
            engine_number="ENG-004",
            ownership_store=self.store_a,
            location_store=self.store_a,
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("reallocate_vehicle", args=[order.pk]),
            {
                "vehicle": replacement.pk,
                "reason": "僅有舊版預計領牌日期",
            },
        )

        self.assertRedirects(
            response,
            f"{reverse('order_detail', args=[order.pk])}?tab=allocation",
        )
        order.refresh_from_db()
        self.assertEqual(order.allocated_vehicle, replacement)

    def test_order_detail_uses_flexible_work_tabs(self):
        from pathlib import Path

        order = self.make_order()
        self.client.force_login(self.user)

        response = self.client.get(reverse("order_detail", args=[order.pk]))

        self.assertContains(response, 'role="tablist"')
        for tab_name in (
            "訂單資料",
            "配車",
            "補助",
            "領牌",
            "交付",
            "處理紀錄",
        ):
            self.assertContains(response, f"<span>{tab_name}</span>", html=True)
        self.assertNotContains(response, 'data-tab="documents"')
        self.assertNotContains(response, 'data-tab-panel="documents"')
        self.assertContains(response, 'data-tab-panel="order"')
        self.assertContains(response, 'data-tab-panel="registration"')
        self.assertContains(response, 'id="signed-documents"')
        self.assertContains(response, "簽署文件留存")
        self.assertContains(response, "data-document-upload", count=2)
        self.assertContains(response, "data-auto-upload", count=2)
        self.assertContains(response, reverse("contract_upload", args=[order.pk]))
        self.assertContains(
            response, reverse("privacy_consent_upload", args=[order.pk])
        )
        self.assertContains(response, "領牌資料與文件")
        self.assertContains(response, "車輛交付")
        self.assertContains(response, "配車後開放")
        self.assertNotContains(response, "workflow-strip")

        script = Path("static/js/order-detail-tabs.js").read_text(encoding="utf-8")
        self.assertIn('searchParams.get("tab")', script)
        self.assertIn("window.localStorage", script)
        self.assertIn('event.key === "ArrowRight"', script)

    def test_legacy_documents_tab_redirects_to_order_signed_documents(self):
        order = self.make_order()
        self.client.force_login(self.user)

        response = self.client.get(
            f"{reverse('order_detail', args=[order.pk])}?tab=documents"
        )

        self.assertRedirects(
            response,
            f"{reverse('order_detail', args=[order.pk])}?tab=order#signed-documents",
        )

        response_with_created_dialog = self.client.get(
            f"{reverse('order_detail', args=[order.pk])}"
            "?tab=documents&created=1"
        )
        self.assertRedirects(
            response_with_created_dialog,
            f"{reverse('order_detail', args=[order.pk])}"
            "?tab=order&created=1#signed-documents",
        )

    def test_subsidy_documents_are_available_before_allocation(self):
        order = self.make_order()
        SalesOrder.objects.filter(pk=order.pk).update(
            is_trade_in_subsidy=True,
            trade_in_plate="ABC-1234",
            subsidy_type="汰舊換新",
        )
        order.refresh_from_db()
        self.client.force_login(self.user)

        detail = self.client.get(reverse("order_detail", args=[order.pk]))
        self.assertContains(detail, "可與配車同時進行")
        self.assertContains(detail, "舊車主身分證正面")
        self.assertContains(detail, "2／8 已備妥")
        self.assertContains(detail, "新車主存摺封面")
        self.assertNotContains(detail, "舊車主存摺封面")

        response = self.client.post(
            reverse("subsidy_document_upload", args=[order.pk]),
            {
                "document_type": SubsidyDocument.DocumentType.OLD_OWNER_ID_FRONT,
                "file": uploaded_test_jpeg("old-owner-front.jpg"),
            },
        )
        self.assertRedirects(response, reverse("order_detail", args=[order.pk]))
        self.assertIsNone(order.allocated_vehicle)
        self.assertTrue(
            order.subsidy_documents.filter(
                document_type=SubsidyDocument.DocumentType.OLD_OWNER_ID_FRONT
            ).exists()
        )

    def test_subsidy_tab_updates_data_balance_and_change_history(self):
        order = self.make_order()
        self.client.force_login(self.user)

        detail = self.client.get(reverse("order_detail", args=[order.pk]))
        self.assertContains(detail, 'data-subsidy-form')
        self.assertContains(detail, "儲存補助資料")
        self.assertNotContains(detail, "修改補助基本資料")

        response = self.client.post(
            reverse("subsidy_data_update", args=[order.pk]),
            {
                "_order_revision": order.revision,
                "is_trade_in_subsidy": "on",
                "trade_in_plate": "abc-1234",
                "old_owner_name": "陳大華",
                "old_owner_id_number": "b123456789",
                "subsidy_type": "汰舊換新",
                "old_vehicle_valuation": "10000",
                "old_vehicle_tax": "500",
                "change_reason": "客戶提供舊車資料",
            },
        )

        self.assertRedirects(
            response,
            f"{reverse('order_detail', args=[order.pk])}?tab=subsidy",
        )
        order.refresh_from_db()
        self.assertTrue(order.is_trade_in_subsidy)
        self.assertEqual(order.trade_in_plate, "ABC-1234")
        self.assertEqual(order.old_owner_id_number, "B123456789")
        self.assertEqual(order.actual_balance, Decimal("65700"))
        self.assertEqual(order.calculated_balance, Decimal("65700"))
        self.assertEqual(order.revision, 2)
        change = order.changes.latest("created_at")
        self.assertEqual(change.actor_name, "tester")
        self.assertEqual(change.reason, "客戶提供舊車資料")
        self.assertIn("舊車車牌", change.changes)
        self.assertIn("舊車主身分證字號", change.changes)
        self.assertIn("舊車估價", change.changes)

    def test_completed_order_can_enable_subsidy_and_upload_documents(self):
        order = self.make_order()
        SalesOrder.objects.filter(pk=order.pk).update(
            status=SalesOrder.Status.COMPLETED,
            delivered_at=timezone.now(),
            delivered_by="tester",
        )
        order.refresh_from_db()
        self.assertFalse(order.is_editable)
        self.assertTrue(order.can_manage_subsidy)
        self.client.force_login(self.user)

        detail = self.client.get(reverse("order_detail", args=[order.pk]))
        self.assertContains(detail, 'data-subsidy-form')
        self.assertContains(detail, "申請汰舊／政府補助")

        response = self.client.post(
            reverse("subsidy_data_update", args=[order.pk]),
            {
                "_order_revision": order.revision,
                "is_trade_in_subsidy": "on",
                "old_owner_same_as_owner": "on",
                "trade_in_plate": "old-1234",
                "old_owner_name": "",
                "old_owner_id_number": "",
                "subsidy_type": "汰舊換新",
                "old_vehicle_valuation": "0",
                "old_vehicle_tax": "0",
                "change_reason": "交付後開始辦理補助",
            },
        )
        self.assertRedirects(
            response,
            f"{reverse('order_detail', args=[order.pk])}?tab=subsidy",
        )
        order.refresh_from_db()
        self.assertTrue(order.is_trade_in_subsidy)
        self.assertEqual(order.trade_in_plate, "OLD-1234")
        self.assertEqual(order.status, SalesOrder.Status.COMPLETED)

        response = self.client.post(
            reverse("subsidy_document_upload", args=[order.pk]),
            {
                "document_type": SubsidyDocument.DocumentType.RECYCLING_RECEIPT,
                "file": uploaded_test_pdf("recycling-receipt.pdf"),
            },
        )
        self.assertRedirects(response, reverse("order_detail", args=[order.pk]))
        self.assertTrue(
            order.subsidy_documents.filter(
                document_type=SubsidyDocument.DocumentType.RECYCLING_RECEIPT
            ).exists()
        )

    def test_cancelled_order_keeps_subsidy_locked(self):
        order = self.make_order()
        SalesOrder.objects.filter(pk=order.pk).update(
            status=SalesOrder.Status.CANCELLED,
        )
        order.refresh_from_db()
        self.assertFalse(order.can_manage_subsidy)
        self.client.force_login(self.user)

        detail = self.client.get(reverse("order_detail", args=[order.pk]))
        self.assertNotContains(detail, 'data-subsidy-form')

        response = self.client.post(
            reverse("subsidy_data_update", args=[order.pk]),
            {
                "_order_revision": order.revision,
                "is_trade_in_subsidy": "on",
                "trade_in_plate": "OLD-1234",
                "subsidy_type": "汰舊換新",
                "old_vehicle_valuation": "0",
                "old_vehicle_tax": "0",
                "change_reason": "已取消訂單不應修改",
            },
        )
        self.assertEqual(response.status_code, 302)
        order.refresh_from_db()
        self.assertFalse(order.is_trade_in_subsidy)

        response = self.client.post(
            reverse("subsidy_document_upload", args=[order.pk]),
            {
                "document_type": SubsidyDocument.DocumentType.RECYCLING_RECEIPT,
                "file": uploaded_test_pdf("blocked.pdf"),
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertFalse(order.subsidy_documents.exists())

    def test_subsidy_update_preserves_manually_adjusted_balance(self):
        order = self.make_order()
        SalesOrder.objects.filter(pk=order.pk).update(
            actual_balance=Decimal("76000"),
            balance_adjustment_reason="既有人工調整",
        )
        order.refresh_from_db()
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("subsidy_data_update", args=[order.pk]),
            {
                "_order_revision": order.revision,
                "is_trade_in_subsidy": "on",
                "trade_in_plate": "OLD-123",
                "old_owner_name": "",
                "subsidy_type": "汰舊換新",
                "old_vehicle_valuation": "5000",
                "old_vehicle_tax": "0",
                "change_reason": "新增舊車估價",
            },
        )

        self.assertEqual(response.status_code, 302)
        order.refresh_from_db()
        self.assertEqual(order.actual_balance, Decimal("76000"))
        self.assertEqual(order.calculated_balance, Decimal("70200"))
        self.assertEqual(order.balance_adjustment_reason, "新增舊車估價")

    def test_disabling_subsidy_preserves_uploaded_documents(self):
        order = self.make_order()
        SalesOrder.objects.filter(pk=order.pk).update(
            is_trade_in_subsidy=True,
            old_owner_same_as_owner=False,
            trade_in_plate="OLD-123",
            subsidy_type="汰舊換新",
        )
        order.refresh_from_db()
        document = SubsidyDocument.objects.create(
            order=order,
            document_type=SubsidyDocument.DocumentType.OLD_VEHICLE_REGISTRATION,
            file=SimpleUploadedFile(
                "old-license.pdf", b"license", content_type="application/pdf"
            ),
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("subsidy_data_update", args=[order.pk]),
            {
                "_order_revision": order.revision,
                "trade_in_plate": "OLD-123",
                "old_owner_name": "",
                "subsidy_type": "汰舊換新",
                "old_vehicle_valuation": "0",
                "old_vehicle_tax": "0",
                "change_reason": "客戶暫停補助申請",
            },
        )

        self.assertEqual(response.status_code, 302)
        order.refresh_from_db()
        self.assertFalse(order.is_trade_in_subsidy)
        self.assertTrue(SubsidyDocument.objects.filter(pk=document.pk).exists())

    def test_subsidy_update_rejects_stale_revision(self):
        order = self.make_order()
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("subsidy_data_update", args=[order.pk]),
            {
                "_order_revision": order.revision - 1,
                "is_trade_in_subsidy": "on",
                "trade_in_plate": "OLD-123",
                "old_owner_name": "",
                "subsidy_type": "汰舊換新",
                "old_vehicle_valuation": "0",
                "old_vehicle_tax": "0",
                "change_reason": "測試過期版本",
            },
        )

        self.assertEqual(response.status_code, 302)
        order.refresh_from_db()
        self.assertFalse(order.is_trade_in_subsidy)

    def test_subsidy_requirements_add_declaration_for_different_owner(self):
        order = self.make_order()
        SalesOrder.objects.filter(pk=order.pk).update(
            is_trade_in_subsidy=True,
            old_owner_same_as_owner=False,
            trade_in_plate="OLD-123",
            subsidy_type="汰舊換新",
            old_owner_name="陳大華",
        )
        order.refresh_from_db()

        missing = order.missing_subsidy_requirements()

        self.assertIn("新舊車主不同人聲明書", missing)
        self.assertIn("舊車主身分證字號", missing)
        self.assertIn("新車主存摺封面", missing)
        self.assertIn("舊車主存摺封面", missing)
        self.assertEqual(order.subsidy_required_count, 12)

        self.client.force_login(self.user)
        detail = self.client.get(reverse("order_detail", args=[order.pk]))
        self.assertContains(detail, "新車主存摺封面")
        self.assertContains(detail, "舊車主存摺封面")
        self.assertContains(detail, "舊車主身分證字號")

    def test_subsidy_fixed_document_can_be_replaced_and_downloaded(self):
        order = self.make_order()
        SalesOrder.objects.filter(pk=order.pk).update(
            is_trade_in_subsidy=True,
            trade_in_plate="ABC-1234",
            subsidy_type="汰舊換新",
        )
        order.refresh_from_db()
        self.client.force_login(self.user)
        upload_url = reverse("subsidy_document_upload", args=[order.pk])

        response = self.client.post(
            upload_url,
            {
                "document_type": SubsidyDocument.DocumentType.SCRAP_CERTIFICATE,
                "file": uploaded_test_pdf("first.pdf"),
            },
        )
        self.assertRedirects(response, reverse("order_detail", args=[order.pk]))
        first_document = order.subsidy_documents.get(
            document_type=SubsidyDocument.DocumentType.SCRAP_CERTIFICATE
        )
        first_path = Path(first_document.file.path)
        self.assertTrue(first_path.exists())

        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(
                upload_url,
                {
                    "document_type": SubsidyDocument.DocumentType.SCRAP_CERTIFICATE,
                    "file": uploaded_test_pdf("replacement.pdf"),
                },
            )
        self.assertRedirects(response, reverse("order_detail", args=[order.pk]))
        self.assertFalse(first_path.exists())

        documents = order.subsidy_documents.filter(
            document_type=SubsidyDocument.DocumentType.SCRAP_CERTIFICATE
        )
        self.assertEqual(documents.count(), 1)
        document = documents.get()
        response = self.client.get(
            reverse("subsidy_document_file", args=[document.pk])
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("private", response["Cache-Control"])
        self.assertIn("no-store", response["Cache-Control"])
        self.assertEqual(response["Pragma"], "no-cache")
        self.assertEqual(response["X-Content-Type-Options"], "nosniff")

    def test_order_edit_is_exclusive_to_first_session(self):
        order = self.make_order()
        self.client.force_login(self.user)
        first = self.client.get(reverse("order_edit", args=[order.pk]))
        self.assertEqual(first.status_code, 200)

        second_user = get_user_model().objects.create_user(
            username="second-editor", password="test-pass-123"
        )
        second_client = self.client_class()
        second_client.force_login(second_user)
        blocked = second_client.get(reverse("order_edit", args=[order.pk]))

        self.assertRedirects(
            blocked, reverse("order_detail", args=[order.pk])
        )
        order.refresh_from_db()
        self.assertEqual(order.editing_by, self.user.username)

    def test_multiple_other_subsidy_documents_accept_safe_file_types(self):
        order = self.make_order()
        SalesOrder.objects.filter(pk=order.pk).update(
            is_trade_in_subsidy=True,
            trade_in_plate="ABC-1234",
            subsidy_type="汰舊換新",
        )
        self.client.force_login(self.user)
        upload_url = reverse("subsidy_document_upload", args=[order.pk])

        uploads = [
            (
                "補助切結書",
                "客戶補簽",
                uploaded_test_pdf("declaration.pdf"),
            ),
            (
                "補助試算表",
                "",
                uploaded_test_xlsx("calculation.xlsx"),
            ),
        ]
        for name, note, upload in uploads:
            response = self.client.post(
                upload_url,
                {
                    "document_type": SubsidyDocument.DocumentType.OTHER,
                    "name": name,
                    "note": note,
                    "file": upload,
                },
            )
            self.assertRedirects(
                response, reverse("order_detail", args=[order.pk])
            )

        documents = order.subsidy_documents.filter(
            document_type=SubsidyDocument.DocumentType.OTHER
        )
        self.assertEqual(documents.count(), 2)
        self.assertEqual(
            set(documents.values_list("name", flat=True)),
            {"補助切結書", "補助試算表"},
        )
        detail = self.client.get(reverse("order_detail", args=[order.pk]))
        self.assertContains(detail, "其他補助文件")
        self.assertContains(detail, "補助切結書")
        self.assertContains(detail, "客戶補簽")

    def test_other_subsidy_document_rejects_missing_name_and_executable(self):
        order = self.make_order()
        SalesOrder.objects.filter(pk=order.pk).update(
            is_trade_in_subsidy=True
        )
        self.client.force_login(self.user)
        upload_url = reverse("subsidy_document_upload", args=[order.pk])

        for name, filename in (("", "missing-name.pdf"), ("惡意檔案", "bad.exe")):
            response = self.client.post(
                upload_url,
                {
                    "document_type": SubsidyDocument.DocumentType.OTHER,
                    "name": name,
                    "file": SimpleUploadedFile(
                        filename,
                        b"unsafe",
                        content_type="application/octet-stream",
                    ),
                },
            )
            self.assertRedirects(
                response, reverse("order_detail", args=[order.pk])
            )

        self.assertFalse(
            order.subsidy_documents.filter(
                document_type=SubsidyDocument.DocumentType.OTHER
            ).exists()
        )

    @patch("sales.views.recognize_id_card")
    def test_old_owner_id_upload_automatically_fills_empty_fields(self, recognize):
        recognize.return_value = {
            "fields": {
                "name": "陳大華",
                "id_number": "B123456789",
                "id_number_valid": True,
            },
            "warnings": [],
        }
        order = self.make_order()
        SalesOrder.objects.filter(pk=order.pk).update(
            is_trade_in_subsidy=True,
            old_owner_same_as_owner=False,
        )
        self.client.force_login(self.user)
        upload_url = reverse("subsidy_document_upload", args=[order.pk])

        for document_type, filename in (
            (SubsidyDocument.DocumentType.OLD_OWNER_ID_FRONT, "front.jpg"),
            (SubsidyDocument.DocumentType.OLD_OWNER_ID_BACK, "back.jpg"),
        ):
            response = self.client.post(
                upload_url,
                {
                    "document_type": document_type,
                    "file": uploaded_test_jpeg(filename),
                },
            )
            self.assertRedirects(
                response, reverse("order_detail", args=[order.pk])
            )

        order.refresh_from_db()
        self.assertEqual(order.old_owner_name, "陳大華")
        self.assertEqual(order.old_owner_id_number, "B123456789")
        self.assertEqual(order.old_owner_ocr_name, "")
        self.assertEqual(order.old_owner_ocr_id_number, "")
        recognize.assert_called_once()

    @patch("sales.views.recognize_id_card")
    def test_old_owner_ocr_conflict_requires_user_decision(self, recognize):
        recognize.return_value = {
            "fields": {
                "name": "OCR 姓名",
                "id_number": "B123456789",
                "id_number_valid": True,
            },
            "warnings": [],
        }
        order = self.make_order()
        SalesOrder.objects.filter(pk=order.pk).update(
            is_trade_in_subsidy=True,
            old_owner_same_as_owner=False,
            old_owner_name="人工姓名",
            old_owner_id_number="A123456789",
        )
        self.client.force_login(self.user)
        upload_url = reverse("subsidy_document_upload", args=[order.pk])
        for document_type, filename in (
            (SubsidyDocument.DocumentType.OLD_OWNER_ID_FRONT, "front.jpg"),
            (SubsidyDocument.DocumentType.OLD_OWNER_ID_BACK, "back.jpg"),
        ):
            self.client.post(
                upload_url,
                {
                    "document_type": document_type,
                    "file": uploaded_test_jpeg(filename),
                },
            )

        order.refresh_from_db()
        self.assertEqual(order.old_owner_name, "人工姓名")
        self.assertEqual(order.old_owner_ocr_name, "OCR 姓名")
        detail = self.client.get(reverse("order_detail", args=[order.pk]))
        self.assertContains(detail, "採用辨識結果")

        response = self.client.post(
            reverse("subsidy_ocr_decision", args=[order.pk]),
            {"decision": "apply"},
        )
        self.assertRedirects(
            response,
            f"{reverse('order_detail', args=[order.pk])}?tab=subsidy",
        )
        order.refresh_from_db()
        self.assertEqual(order.old_owner_name, "OCR 姓名")
        self.assertEqual(order.old_owner_id_number, "B123456789")
        self.assertEqual(order.old_owner_ocr_name, "")

    def test_non_subsidy_order_rejects_subsidy_upload(self):
        order = self.make_order()
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("subsidy_document_upload", args=[order.pk]),
            {
                "document_type": SubsidyDocument.DocumentType.RECYCLING_RECEIPT,
                "file": uploaded_test_pdf("receipt.pdf"),
            },
        )

        self.assertRedirects(response, reverse("order_detail", args=[order.pk]))
        self.assertFalse(order.subsidy_documents.exists())

    def test_registration_requires_data_and_all_fixed_documents(self):
        order = self.make_order()
        order.allocate(self.vehicle)
        self.client.force_login(self.user)

        missing = order.missing_registration_requirements()
        self.assertIn("新行照照片", missing)
        self.assertIn("新車領牌登記書", missing)
        self.assertIn("發票", missing)
        self.assertNotIn("監理站單據", missing)
        self.assertNotIn("強制險單", missing)

        response = self.client.post(
            reverse("registration_complete", args=[order.pk])
        )

        self.assertRedirects(response, reverse("order_detail", args=[order.pk]))
        order.refresh_from_db()
        self.assertFalse(order.is_registration_complete)
        self.assertEqual(order.status, SalesOrder.Status.ALLOCATED)

    def test_retired_registration_documents_are_hidden_and_rejected(self):
        order = self.make_order()
        order.allocate(self.vehicle)
        self.client.force_login(self.user)

        detail = self.client.get(reverse("order_detail", args=[order.pk]))
        self.assertNotContains(detail, "監理站單據")
        self.assertNotContains(detail, "強制險單")
        self.assertContains(detail, "data-document-upload")
        self.assertContains(detail, "data-auto-upload")
        self.assertContains(detail, "document-upload-progress.js")
        self.assertNotContains(detail, "this.form.submit()")
        progress_script = Path("static/js/document-upload-progress.js").read_text(
            encoding="utf-8"
        )
        self.assertIn('request.upload.addEventListener("progress"', progress_script)
        self.assertIn('request.upload.addEventListener("load"', progress_script)
        self.assertIn("new FormData(form)", progress_script)
        self.assertIn('fileInput.value = ""', progress_script)

        response = self.client.post(
            reverse("registration_document_upload", args=[order.pk]),
            {
                "document_type": (
                    RegistrationDocument.DocumentType.MOTOR_VEHICLE_RECEIPT
                ),
                "name": "",
                "file": uploaded_test_pdf("retired.pdf"),
            },
            HTTP_ACCEPT="application/json",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(response.json()["ok"])
        self.assertIn("已取消", response.json()["message"])
        self.assertFalse(order.registration_documents.exists())

    def test_legacy_retired_document_does_not_lock_reallocation(self):
        order = self.make_order()
        order.allocate(self.vehicle)
        RegistrationDocument.objects.create(
            order=order,
            document_type=RegistrationDocument.DocumentType.COMPULSORY_INSURANCE,
            file=SimpleUploadedFile(
                "legacy.pdf", b"legacy", content_type="application/pdf"
            ),
        )

        self.assertFalse(order.has_registration_started)

        RegistrationDocument.objects.create(
            order=order,
            document_type=RegistrationDocument.DocumentType.INVOICE,
            file=SimpleUploadedFile(
                "invoice.pdf", b"invoice", content_type="application/pdf"
            ),
        )
        self.assertTrue(order.has_registration_started)

    def test_registration_document_async_upload_returns_progress_destination(self):
        order = self.make_order()
        order.allocate(self.vehicle)
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("registration_document_upload", args=[order.pk]),
            {
                "document_type": RegistrationDocument.DocumentType.INVOICE,
                "name": "",
                "file": uploaded_test_pdf("invoice.pdf"),
            },
            HTTP_ACCEPT="application/json",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["ok"])
        self.assertEqual(
            response.json()["redirect_url"],
            f"{reverse('order_detail', args=[order.pk])}?tab=registration",
        )
        self.assertTrue(
            order.registration_documents.filter(
                document_type=RegistrationDocument.DocumentType.INVOICE
            ).exists()
        )

    def test_registration_can_be_completed_after_required_uploads(self):
        order = self.make_order()
        order.allocate(self.vehicle)
        incentive_rule = VehicleIncentiveRule.objects.create(
            vehicle_model=self.model,
            sales_bonus=Decimal("1500"),
            promotion_subsidy=Decimal("2000"),
            installment_interest_subsidy=Decimal("800"),
            effective_from=date(2026, 7, 1),
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("registration_save", args=[order.pk]),
            {
                "registration_date": "2026-07-29",
                "registration_county": "新北市",
                "final_plate_number": "abc-1234",
            },
        )
        self.assertRedirects(response, reverse("order_detail", args=[order.pk]))

        required_types = [
            RegistrationDocument.DocumentType.NEW_LICENSE,
            RegistrationDocument.DocumentType.REGISTRATION_APPLICATION,
            RegistrationDocument.DocumentType.INVOICE,
        ]
        for index, document_type in enumerate(required_types):
            response = self.client.post(
                reverse("registration_document_upload", args=[order.pk]),
                {
                    "document_type": document_type,
                    "name": "",
                    "file": uploaded_test_pdf(f"document-{index}.pdf"),
                },
            )
            self.assertRedirects(
                response, reverse("order_detail", args=[order.pk])
            )

        response = self.client.post(
            reverse("registration_complete", args=[order.pk])
        )
        self.assertRedirects(response, reverse("order_detail", args=[order.pk]))
        order.refresh_from_db()
        self.assertTrue(order.is_registration_complete)
        self.assertEqual(order.registration_date, date(2026, 7, 29))
        self.assertEqual(order.final_plate_number, "ABC-1234")
        self.assertEqual(order.registration_completed_by, "tester")
        self.assertEqual(order.status, SalesOrder.Status.DELIVERY_PENDING)
        profile = order.operations
        self.assertEqual(profile.vehicle_cost, Decimal("60000"))
        self.assertEqual(profile.vehicle_cost_rule, self.settlement_rule)
        self.assertEqual(profile.vehicle_cost_county, "新北市")
        self.assertIsNotNone(profile.vehicle_cost_locked_at)
        self.assertEqual(profile.incentive_rule, incentive_rule)
        self.assertEqual(profile.sales_bonus, Decimal("1500"))
        self.assertEqual(profile.promotion_subsidy, Decimal("2000"))
        self.assertEqual(profile.installment_interest_subsidy, Decimal("800"))
        self.assertIsNotNone(profile.incentive_locked_at)

    def test_registration_completion_requires_matching_settlement_cost(self):
        order = self.make_order()
        order.allocate(self.vehicle)
        self.settlement_rule.delete()
        SalesOrder.objects.filter(pk=order.pk).update(
            registration_date=date(2026, 7, 29),
            registration_county="新北市",
            final_plate_number="ABC-1234",
        )
        for document_type in [
            RegistrationDocument.DocumentType.NEW_LICENSE,
            RegistrationDocument.DocumentType.REGISTRATION_APPLICATION,
            RegistrationDocument.DocumentType.INVOICE,
        ]:
            RegistrationDocument.objects.create(
                order=order,
                document_type=document_type,
                file=SimpleUploadedFile(
                    f"{document_type}.pdf",
                    b"registration document",
                    content_type="application/pdf",
                ),
            )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("registration_complete", args=[order.pk])
        )

        self.assertRedirects(response, reverse("order_detail", args=[order.pk]))
        order.refresh_from_db()
        self.assertFalse(order.is_registration_complete)
        messages = [str(message) for message in response.wsgi_request._messages]
        self.assertTrue(any("代銷結算成本規則" in message for message in messages))

    def test_plate_selection_document_is_conditionally_required(self):
        order = self.make_order()
        order.allocate(self.vehicle)
        SalesOrder.objects.filter(pk=order.pk).update(
            registration_date=date(2026, 7, 29),
            registration_county="新北市",
            final_plate_number="ABC-1234",
            plate_choice=SalesOrder.PlateChoice.PREFERENCE,
        )
        order.refresh_from_db()
        for document_type in [
            RegistrationDocument.DocumentType.NEW_LICENSE,
            RegistrationDocument.DocumentType.REGISTRATION_APPLICATION,
            RegistrationDocument.DocumentType.INVOICE,
        ]:
            RegistrationDocument.objects.create(
                order=order,
                document_type=document_type,
                file=SimpleUploadedFile(
                    f"{document_type}.pdf", b"document", content_type="application/pdf"
                ),
            )

        self.assertIn("選號單", order.missing_registration_requirements())

    def test_other_insurance_accepts_multiple_named_documents(self):
        order = self.make_order()
        order.allocate(self.vehicle)
        self.client.force_login(self.user)

        for name in ("第三人責任險", "車體險"):
            response = self.client.post(
                reverse("registration_document_upload", args=[order.pk]),
                {
                    "document_type": RegistrationDocument.DocumentType.OTHER_INSURANCE,
                    "name": name,
                    "file": uploaded_test_pdf(f"{name}.pdf"),
                },
            )
            self.assertRedirects(
                response, reverse("order_detail", args=[order.pk])
            )

        self.assertEqual(
            order.registration_documents.filter(
                document_type=RegistrationDocument.DocumentType.OTHER_INSURANCE
            ).count(),
            2,
        )

    def test_fixed_registration_document_can_be_replaced_and_downloaded(self):
        order = self.make_order()
        order.allocate(self.vehicle)
        self.client.force_login(self.user)
        upload_url = reverse("registration_document_upload", args=[order.pk])

        response = self.client.post(
            upload_url,
            {
                "document_type": RegistrationDocument.DocumentType.INVOICE,
                "name": "",
                "file": uploaded_test_pdf("first.pdf"),
            },
        )
        self.assertRedirects(response, reverse("order_detail", args=[order.pk]))
        first_document = order.registration_documents.get(
            document_type=RegistrationDocument.DocumentType.INVOICE
        )
        first_path = Path(first_document.file.path)
        self.assertTrue(first_path.exists())

        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(
                upload_url,
                {
                    "document_type": RegistrationDocument.DocumentType.INVOICE,
                    "name": "",
                    "file": uploaded_test_pdf("replacement.pdf"),
                },
            )
        self.assertRedirects(response, reverse("order_detail", args=[order.pk]))
        self.assertFalse(first_path.exists())

        documents = order.registration_documents.filter(
            document_type=RegistrationDocument.DocumentType.INVOICE
        )
        self.assertEqual(documents.count(), 1)
        document = documents.get()
        response = self.client.get(
            reverse("registration_document_file", args=[document.pk])
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("private", response["Cache-Control"])
        self.assertIn("no-store", response["Cache-Control"])
        self.assertEqual(response["Pragma"], "no-cache")
        self.assertEqual(response["X-Content-Type-Options"], "nosniff")

    def test_dealer_order_can_deliver_before_registration(self):
        order = self.make_order()
        self.assertFalse(order.can_deliver)

        SalesOrder.objects.filter(pk=order.pk).update(
            source_type=SalesOrder.SourceType.DEALER
        )
        order.refresh_from_db()

        self.assertTrue(order.can_deliver)

    def test_same_vehicle_cannot_allocate_twice(self):
        first = self.make_order(signed=True)
        second = self.make_order(signed=True)
        first.allocate(self.vehicle)

        with self.assertRaisesMessage(ValidationError, "目前不可配車"):
            second.allocate(self.vehicle)

    def test_gas_vehicle_requires_engine_number(self):
        vehicle = VehicleInventory(
            vehicle_model=self.model,
            color=self.color,
            frame_number="FRAME-WRONG",
            ownership_store=self.store_a,
            location_store=self.store_a,
        )

        with self.assertRaises(ValidationError) as context:
            vehicle.full_clean()

        self.assertIn("engine_number", context.exception.message_dict)

    def test_dashboard_global_searches_phone_suffix(self):
        order = self.make_order()
        self.client.force_login(self.user)

        response = self.client.get(reverse("dashboard"), {"q": "5678"})

        self.assertContains(response, order.owner_name)
        self.assertContains(response, order.number)

    def test_dashboard_shows_clickable_in_progress_metric_and_empty_section(self):
        self.client.force_login(self.user)

        empty_response = self.client.get(reverse("dashboard"))
        self.assertContains(empty_response, "訂單進行中")
        self.assertContains(empty_response, "?status=in_progress")

        order = self.make_order()
        order.status = SalesOrder.Status.ALLOCATED
        order.save(update_fields=["status"])

        response = self.client.get(reverse("dashboard"))
        self.assertEqual(response.context["counts"]["in_progress"], 1)
        self.assertEqual(response.context["dashboard"]["workload"]["in_progress"], 1)
        self.assertContains(response, order.number)

        filtered = self.client.get(reverse("order_list"), {"status": "in_progress"})
        self.assertContains(filtered, order.number)

    def test_dynamic_accessory_and_fee_rows_can_be_deleted(self):
        order = self.make_order()
        accessory = AccessoryLine.objects.create(
            order=order,
            name="手機架",
            quantity=1,
            line_type=AccessoryLine.LineType.PURCHASE,
            amount=Decimal("850"),
        )
        fee = OtherFeeLine.objects.create(
            order=order,
            name="代辦費",
            amount=Decimal("300"),
        )

        accessory_formset = AccessoryFormSet(
            {
                "accessories-TOTAL_FORMS": "1",
                "accessories-INITIAL_FORMS": "1",
                "accessories-MIN_NUM_FORMS": "0",
                "accessories-MAX_NUM_FORMS": "1000",
                "accessories-0-id": str(accessory.pk),
                "accessories-0-DELETE": "on",
            },
            instance=order,
        )
        fee_formset = OtherFeeFormSet(
            {
                "other_fees-TOTAL_FORMS": "1",
                "other_fees-INITIAL_FORMS": "1",
                "other_fees-MIN_NUM_FORMS": "0",
                "other_fees-MAX_NUM_FORMS": "1000",
                "other_fees-0-id": str(fee.pk),
                "other_fees-0-DELETE": "on",
            },
            instance=order,
            prefix="other_fees",
        )

        self.assertTrue(accessory_formset.is_valid(), accessory_formset.errors)
        self.assertTrue(fee_formset.is_valid(), fee_formset.errors)
        accessory_formset.save()
        fee_formset.save()
        self.assertFalse(order.accessories.exists())
        self.assertFalse(order.other_fees.exists())

    def test_server_error_page_explains_what_user_can_do(self):
        from sales.views import server_error

        request = RequestFactory().get("/")
        request.user = AnonymousUser()
        response = server_error(request)

        self.assertEqual(response.status_code, 500)
        self.assertIn("系統暫時無法完成這個動作", response.content.decode())

    def test_contract_print_returns_two_page_dynamic_pdf(self):
        order = self.make_order()
        self.model.model_number = "INTERNAL-125"
        self.model.model_year = 2026
        self.model.vehicle_type = "abs_double_disc"
        self.model.save()
        self.client.force_login(self.user)

        response = self.client.get(reverse("contract_print", args=[order.pk]))
        content = b"".join(response.streaming_content)
        reader = PdfReader(BytesIO(content))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertIn("inline;", response["Content-Disposition"])
        self.assertEqual(len(reader.pages), 2)
        extracted = "\n".join(page.extract_text() for page in reader.pages)
        self.assertIn(order.number, extracted)
        self.assertIn(order.owner_name, extracted)
        self.assertIn("店家留存聯", extracted)
        self.assertIn("客戶留存聯", extracted)
        self.assertIn("財產登記制", extracted)
        self.assertIn("領牌後無法辦理退換貨", extracted)
        self.assertIn(
            "本人（或本公司）了解並確認本訂購單所載車型",
            extracted,
        )
        self.assertIn(
            "本人（或本公司）了解機車屬於財產登記制",
            extracted,
        )
        self.assertIn("應收", extracted)
        self.assertNotIn("現場應收", extracted)
        self.assertNotIn("分期總額", extracted)
        self.assertNotIn("收款說明", extracted)
        self.assertIn("領牌＋強制險，依單據收款", extracted)
        self.assertIn("稅金", extracted)
        self.assertIn("測試廠牌 通勤 125／白", extracted)

    def test_installment_contract_keeps_installment_total(self):
        order = self.make_order()
        order.payment_type = SalesOrder.PaymentType.INSTALLMENT
        order.installment_amount = Decimal("75000")
        order.installment_company = "和潤"
        order.installment_periods = 24
        order.installment_opening_fee = Decimal("2500")
        order.installment_monthly = Decimal("3125")
        order.plate_selection_fee = Decimal("300")
        AccessoryLine.objects.create(
            order=order,
            name="手機架",
            quantity=1,
            amount=Decimal("850"),
        )
        OtherFeeLine.objects.create(order=order, name="代辦費", amount=Decimal("500"))
        order.actual_balance = order.calculate_balance()
        order.save()
        self.client.force_login(self.user)

        response = self.client.get(reverse("contract_print", args=[order.pk]))
        content = b"".join(response.streaming_content)
        extracted = "\n".join(
            page.extract_text() for page in PdfReader(BytesIO(content)).pages
        )

        self.assertIn("分期總額", extracted)
        self.assertIn("應收", extracted)
        self.assertNotIn("收款說明", extracted)
        self.assertIn("領牌＋強制險，依單據收款", extracted)
        self.assertIn("選號費", extracted)
        self.assertIn("須以現金或匯款支付", extracted)
        self.assertLess(extracted.index("手機架"), extracted.index("分期開辦費"))
        self.assertLess(extracted.index("代辦費"), extracted.index("分期開辦費"))

    def test_privacy_consent_uses_latest_order_name_date_and_plate(self):
        order = self.make_order()
        order.owner_name = "馮華微"
        order.order_date = date(2026, 7, 29)
        order.final_plate_number = "ABC-1234"
        order.save()
        self.client.force_login(self.user)

        response = self.client.get(reverse("privacy_consent_print", args=[order.pk]))
        content = b"".join(response.streaming_content)
        reader = PdfReader(BytesIO(content))
        extracted = "\n".join(page.extract_text() for page in reader.pages)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(reader.pages), 1)
        self.assertIn("馮華微", extracted)
        self.assertIn("ABC-1234", extracted)
        self.assertIn("中華民國 115 年 07 月 29 日", extracted)
        self.assertIn("馭盛國際有限公司", extracted)

        order.owner_name = "馮華薇"
        order.order_date = date(2026, 8, 1)
        order.save()
        updated_response = self.client.get(
            reverse("privacy_consent_print", args=[order.pk])
        )
        updated_content = b"".join(updated_response.streaming_content)
        updated_text = "\n".join(
            page.extract_text()
            for page in PdfReader(BytesIO(updated_content)).pages
        )
        self.assertIn("馮華薇", updated_text)
        self.assertIn("中華民國 115 年 08 月 01 日", updated_text)
        self.assertNotIn("馮華微", updated_text)

    def test_combined_signing_documents_have_three_pages(self):
        order = self.make_order()
        self.client.force_login(self.user)

        response = self.client.get(reverse("order_documents_print", args=[order.pk]))
        content = b"".join(response.streaming_content)
        reader = PdfReader(BytesIO(content))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(reader.pages), 3)

        detail = self.client.get(
            reverse("order_detail", args=[order.pk]),
            {"created": "1"},
        )
        self.assertContains(detail, "只印個資同意書")
        self.assertContains(detail, "全部一起列印")
        self.assertContains(detail, "一鍵列印全部文件")
        self.assertContains(detail, "列印空白合約")
        self.assertContains(detail, "列印空白同意書")
        self.assertContains(detail, "簽署文件留存")
        self.assertContains(detail, "個資同意書附件")
        self.assertContains(detail, reverse("privacy_consent_upload", args=[order.pk]))

    def test_identity_document_print_adds_selected_watermark_and_audit_event(self):
        order = self.make_order()
        order.id_front = uploaded_test_jpeg("front.jpg")
        order.id_back = uploaded_test_jpeg("back.jpg")
        order.save(update_fields=["id_front", "id_back", "updated_at"])
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("identity_documents_print", args=[order.pk]),
            {
                "purpose": "registration",
                "sides": ["id_front", "id_front", "id_back", "id_back"],
            },
        )
        content = b"".join(response.streaming_content)
        reader = PdfReader(BytesIO(content))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(reader.pages), 2)
        self.assertIn("private", response["Cache-Control"])
        self.assertIn("no-store", response["Cache-Control"])
        self.assertEqual(response["X-Content-Type-Options"], "nosniff")
        self.assertTrue(
            OrderEvent.objects.filter(
                order=order,
                event_type="identity_document_printed",
                description__contains="限領牌使用",
                actor_name=self.user.username,
            ).exists()
        )

        detail = self.client.get(reverse("order_detail", args=[order.pk]))
        self.assertContains(detail, "列印／下載浮水印證件")
        self.assertContains(detail, "限申請補助使用")

    def test_privacy_consent_can_be_uploaded_separately(self):
        order = self.make_order()
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("privacy_consent_upload", args=[order.pk]),
            {
                "privacy_consent": uploaded_test_pdf("privacy.pdf")
            },
        )

        signed_documents_url = (
            f"{reverse('order_detail', args=[order.pk])}?tab=order#signed-documents"
        )
        self.assertRedirects(response, signed_documents_url)
        order.refresh_from_db()
        self.assertTrue(order.has_privacy_consent)
        self.assertIsNotNone(order.privacy_consent_uploaded_at)
        self.assertTrue(
            order.events.filter(description__contains="個資同意書").exists()
        )

        first_path = Path(order.privacy_consent.path)
        self.assertTrue(first_path.exists())
        with self.captureOnCommitCallbacks(execute=True):
            replacement = self.client.post(
                reverse("privacy_consent_upload", args=[order.pk]),
                {"privacy_consent": uploaded_test_pdf("privacy-replacement.pdf")},
            )
        self.assertRedirects(replacement, signed_documents_url)
        self.assertFalse(first_path.exists())

    def test_signed_contract_upload_returns_to_order_attachment_section(self):
        order = self.make_order()
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("contract_upload", args=[order.pk]),
            {"signed_contract": uploaded_test_pdf("signed-contract.pdf")},
        )

        signed_documents_url = (
            f"{reverse('order_detail', args=[order.pk])}?tab=order#signed-documents"
        )
        self.assertRedirects(response, signed_documents_url)
        order.refresh_from_db()
        self.assertTrue(order.has_signed_contract)
        self.assertIsNotNone(order.signed_contract_uploaded_at)
        self.assertTrue(
            order.events.filter(description__contains="訂購合約附件").exists()
        )

    def test_signed_document_uploads_support_progress_json_response(self):
        order = self.make_order()
        self.client.force_login(self.user)
        signed_documents_url = (
            f"{reverse('order_detail', args=[order.pk])}?tab=order#signed-documents"
        )

        response = self.client.post(
            reverse("contract_upload", args=[order.pk]),
            {"signed_contract": uploaded_test_pdf("signed-contract.pdf")},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
            HTTP_ACCEPT="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.json(),
            {
                "ok": True,
                "message": "訂購合約附件已上傳。",
                "redirect_url": signed_documents_url,
            },
        )

        invalid = self.client.post(
            reverse("privacy_consent_upload", args=[order.pk]),
            {},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
            HTTP_ACCEPT="application/json",
        )
        self.assertEqual(invalid.status_code, 400)
        self.assertFalse(invalid.json()["ok"])
        self.assertEqual(invalid.json()["redirect_url"], signed_documents_url)

    def test_completed_order_can_still_upload_signed_documents(self):
        order = self.make_order()
        SalesOrder.objects.filter(pk=order.pk).update(
            status=SalesOrder.Status.COMPLETED,
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("privacy_consent_upload", args=[order.pk]),
            {"privacy_consent": uploaded_test_pdf("completed-privacy.pdf")},
        )

        self.assertRedirects(
            response,
            f"{reverse('order_detail', args=[order.pk])}"
            "?tab=order#signed-documents",
        )
        order.refresh_from_db()
        self.assertTrue(order.has_privacy_consent)

    def test_mobile_order_and_inventory_pages_render(self):
        self.client.force_login(self.user)

        order_response = self.client.get(reverse("order_create"))
        inventory_response = self.client.get(reverse("inventory_create"))

        self.assertEqual(order_response.status_code, 200)
        self.assertContains(order_response, "建立新訂單")
        self.assertContains(order_response, "訂單建檔")
        self.assertNotContains(order_response, "手機下單")
        self.assertContains(order_response, "editing-presence")
        self.assertContains(order_response, "refreshDraftPresence")
        self.assertContains(order_response, "自動辨識姓名")
        self.assertContains(order_response, "移除正面照片")
        self.assertContains(order_response, "移除反面照片")
        self.assertContains(order_response, "ocrRequestVersion")
        self.assertContains(order_response, "identity-compare")
        self.assertContains(order_response, "form-validation-summary")
        self.assertContains(order_response, "requestPhotoVersion")
        self.assertContains(
            order_response,
            "領牌日期於後續編輯時填寫，牌險依單據收款",
        )
        self.assertContains(
            order_response,
            'name="registration_date"',
        )
        self.assertNotContains(
            order_response,
            '<input type="date" name="registration_date"',
        )
        order_form_template = (
            __import__("pathlib").Path("templates/sales/order_form.html").read_text(
                encoding="utf-8"
            )
        )
        self.assertIn(
            "{{ form.registration_date.as_hidden }}",
            order_form_template,
        )
        self.assertContains(order_response, "showSavedPhoto")
        self.assertContains(order_response, "正在辨識證件資料")
        self.assertContains(order_response, "請上傳身分證正反面")
        self.assertContains(
            order_response,
            "法人不使用證件自動辨識，請自行填寫公司名稱、統一編號、公司地址與聯絡資料。",
        )
        self.assertContains(
            order_response,
            'showConditional("corporate", ownerType.value === "corporate")',
        )
        self.assertContains(order_response, "資料欄位會在辨識完成後自動顯示")
        self.assertContains(order_response, "data-ocr-confirmation-fields")
        self.assertContains(order_response, "is-floating-processing")
        self.assertNotContains(order_response, "data-ocr-photo-progress")
        self.assertContains(order_response, 'role="status"')
        self.assertContains(order_response, 'setOcrState("processing")')
        self.assertContains(order_response, "applyOcrResult();")
        self.assertContains(order_response, 'setOcrState("manual"')
        self.assertContains(order_response, 'aria-current="step"')
        self.assertContains(order_response, "updateActiveSection")
        self.assertNotContains(order_response, "data-save-draft")
        self.assertNotContains(order_response, ">暫存</button>")
        content = order_response.content.decode()
        self.assertEqual(content.count('id="id_owner_name"'), 1)
        self.assertEqual(content.count('id="id_owner_id_number"'), 1)
        self.assertEqual(content.count('id="id_owner_birth_date"'), 1)
        self.assertEqual(content.count('id="id_owner_address"'), 1)
        self.assertGreater(
            content.index('id="add-accessory"'),
            content.index('id="accessory-forms"'),
        )
        self.assertGreater(
            content.index('id="add-other-fee"'),
            content.index('id="other-fee-forms"'),
        )
        self.assertContains(order_response, 'class="accessory-row dynamic-row')
        self.assertContains(order_response, "或從手機相簿選擇圖片")
        self.assertContains(order_response, "＋ 新增費用")
        self.assertContains(order_response, "data-remove-row")
        self.assertContains(order_response, "刪除此筆配件")
        self.assertContains(order_response, "刪除此筆費用")
        self.assertContains(order_response, "現金")
        self.assertContains(order_response, "分期")
        self.assertContains(order_response, "刷卡")
        self.assertContains(order_response, "分期公司")
        self.assertContains(order_response, "分期期數")
        self.assertContains(order_response, "分期開辦費")
        self.assertContains(order_response, "每期金額")
        self.assertContains(order_response, "送達地點／託運目的地")
        self.assertTrue(order_response.context["form"]["plate_choice"].field.required)
        self.assertTrue(order_response.context["form"]["delivery_method"].field.required)
        self.assertNotIn(
            "old_vehicle_valuation", order_response.context["form"].fields
        )
        self.assertIn(
            "old_owner_same_as_owner", order_response.context["form"].fields
        )
        self.assertNotIn("old_vehicle_tax", order_response.context["form"].fields)
        self.assertNotContains(order_response, "分期申請金額")
        self.assertNotContains(order_response, "分期申請日期")
        self.assertNotContains(order_response, "核准／拒絕日期")
        self.assertLess(content.index("分期資料"), content.index("其他費用"))
        self.assertNotContains(order_response, "折扣／抵扣")
        self.assertEqual(
            order_response.context["form"]["deposit_date"].value(),
            timezone.localdate(),
        )
        for field_name in (
            "vehicle_price",
            "plate_insurance_fee",
            "deposit_amount",
            "installment_periods",
            "installment_opening_fee",
            "installment_monthly",
        ):
            self.assertIsNone(order_response.context["form"][field_name].value())
        self.assertIsNone(
            order_response.context["formset"].forms[0]["amount"].value()
        )
        self.assertEqual(
            order_response.context["formset"].forms[0]["quantity"].value(), 1
        )
        self.assertContains(
            order_response,
            f'value="{timezone.localdate():%Y-%m-%d}"',
        )
        self.assertLess(content.index("id_id_front"), content.index("id_owner_name"))
        self.assertNotIn('capture="environment"', content)
        self.assertContains(order_response, 'type="date"')
        self.assertEqual(inventory_response.status_code, 200)
        self.assertContains(inventory_response, "新增進車")

    def test_delivery_destination_is_required_only_for_delivery_or_carrier(self):
        required_data = {
            "delivery_method": SalesOrder.DeliveryMethod.DIRECT_DELIVERY,
        }
        direct_delivery_form = SalesOrderForm(data=required_data)
        direct_delivery_form.is_valid()
        self.assertIn(
            "delivery_destination", direct_delivery_form.errors
        )

        carrier_form = SalesOrderForm(
            data={"delivery_method": SalesOrder.DeliveryMethod.CARRIER}
        )
        carrier_form.is_valid()
        self.assertIn("delivery_destination", carrier_form.errors)

        pickup_form = SalesOrderForm(
            data={"delivery_method": SalesOrder.DeliveryMethod.STORE_PICKUP}
        )
        pickup_form.is_valid()
        self.assertNotIn("delivery_destination", pickup_form.errors)

    def test_accessory_fields_are_required_only_after_product_is_selected(self):
        empty_form = AccessoryLineForm(
            data={
                "accessory_product": "",
                "quantity": "",
                "line_type": "",
                "amount": "",
                "labor_fee": "",
            }
        )
        self.assertTrue(empty_form.is_valid())
        self.assertFalse(empty_form.fields["accessory_product"].required)
        self.assertFalse(empty_form.fields["quantity"].required)
        self.assertFalse(empty_form.fields["line_type"].required)
        self.assertFalse(empty_form.fields["amount"].required)

        product = AccessoryProduct.objects.create(
            name="後箱",
            sale_price=Decimal("1200"),
            labor_fee=Decimal("300"),
        )
        selected_form = AccessoryLineForm(
            data={
                "accessory_product": product.pk,
                "quantity": "",
                "line_type": "",
                "amount": "99999",
                "labor_fee": "99999",
            }
        )
        self.assertFalse(selected_form.is_valid())
        self.assertIn("quantity", selected_form.errors)
        self.assertIn("line_type", selected_form.errors)
        self.assertNotIn("amount", selected_form.errors)

    def test_accessory_master_prices_are_snapshotted_and_cannot_be_overridden(self):
        product = AccessoryProduct.objects.create(
            name="行車記錄器",
            sale_price=Decimal("6500"),
            labor_fee=Decimal("800"),
        )
        form = AccessoryLineForm(
            data={
                "accessory_product": product.pk,
                "quantity": "2",
                "line_type": AccessoryLine.LineType.PURCHASE,
                "amount": "1",
                "labor_fee": "2",
                "note": "測試",
            }
        )

        self.assertTrue(form.is_valid(), form.errors)
        line = form.save(commit=False)
        self.assertEqual(line.name, "行車記錄器")
        self.assertEqual(line.amount, Decimal("6500"))
        self.assertEqual(line.labor_fee, Decimal("800"))
        self.assertEqual(line.line_total, Decimal("14600"))

        product.sale_price = Decimal("7000")
        product.labor_fee = Decimal("900")
        product.save()
        self.assertEqual(line.amount, Decimal("6500"))
        self.assertEqual(line.labor_fee, Decimal("800"))

    def test_accessory_maintenance_pages_are_available(self):
        self.client.force_login(self.user)
        product = AccessoryProduct.objects.create(
            name="手機架",
            sale_price=Decimal("850"),
            labor_fee=Decimal("200"),
        )

        listing = self.client.get(reverse("accessory_product_list"))
        editing = self.client.get(reverse("accessory_product_edit", args=[product.pk]))

        self.assertEqual(listing.status_code, 200)
        self.assertContains(listing, "手機架")
        self.assertContains(listing, "售價")
        self.assertContains(listing, "工資")
        self.assertEqual(editing.status_code, 200)
        self.assertContains(editing, "參考成本不會顯示在客戶訂單")

    def test_other_fee_is_optional_but_partial_row_requires_both_fields(self):
        empty_form = OtherFeeLineForm(data={"name": "", "amount": ""})
        self.assertTrue(empty_form.is_valid())

        missing_name = OtherFeeLineForm(data={"name": "", "amount": "300"})
        self.assertFalse(missing_name.is_valid())
        self.assertIn("name", missing_name.errors)

        missing_amount = OtherFeeLineForm(data={"name": "代辦費", "amount": ""})
        self.assertFalse(missing_amount.is_valid())
        self.assertIn("amount", missing_amount.errors)

    def test_cash_payment_does_not_require_hidden_installment_fields(self):
        form = SalesOrderForm()
        for field_name in (
            "installment_company",
            "installment_periods",
            "installment_opening_fee",
            "installment_monthly",
        ):
            self.assertFalse(form.fields[field_name].required)

    def test_app_version_endpoint_is_not_cached(self):
        response = self.client.get(reverse("app_version"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.json()["version"]), 12)
        self.assertIn("no-store", response["Cache-Control"])

    def test_authenticated_page_contains_update_controls(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse("dashboard"))

        self.assertContains(response, "檢查系統更新")
        self.assertContains(response, "檢查更新")
        self.assertContains(response, 'class="update-check-icon"')
        self.assertContains(response, "js/layout-audit")
        self.assertContains(response, "app-update-banner")
        self.assertContains(response, "js/app-update")

    def test_allocation_summary_keeps_nested_panel_inside_card_spacing(self):
        css = (
            __import__("pathlib").Path("static/css/app.css").read_text(
                encoding="utf-8"
            )
        )

        self.assertIn(
            ".allocation-summary > .data-list { padding: 18px 20px 0; }",
            css,
        )
        self.assertIn(
            ".allocation-summary > .reallocation-panel,",
            css,
        )
        self.assertIn(
            ".allocation-summary > .allocation-lock-note { margin: 18px 20px 20px; }",
            css,
        )
    def test_messages_have_timed_dismiss_and_manual_close(self):
        from pathlib import Path

        order = self.make_order()
        self.client.force_login(self.user)
        self.client.post(reverse("allocate_vehicle", args=[order.pk]), {})

        response = self.client.get(reverse("order_detail", args=[order.pk]))

        self.assertContains(response, "data-message-toast")
        self.assertContains(response, "data-message-close")
        self.assertContains(response, "js/message-toasts")
        script = Path("static/js/message-toasts.js").read_text(encoding="utf-8")
        self.assertIn("success: 4000", script)
        self.assertIn("info: 5000", script)
        self.assertIn("warning: 8000", script)
        self.assertIn('toast.classList.contains("error") ? null', script)
        self.assertIn('toast.addEventListener("mouseenter", pause)', script)
        self.assertIn('toast.addEventListener("touchstart", pause', script)


class OrderOperationsTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username="operations-user", password="test-pass-123"
        )
        self.store = Store.objects.create(name="總店", code="HQ")
        self.model = VehicleModel.objects.create(
            brand="測試廠牌",
            name="營運 125",
            energy_type=VehicleModel.EnergyType.GAS,
        )
        self.color = VehicleColor.objects.create(
            vehicle_model=self.model,
            name="黑",
        )
        self.order = SalesOrder.objects.create(
            owner_name="營運測試",
            owner_phone="0911222333",
            owner_address="新北市測試區",
            owner_id_number="A123456789",
            vehicle_model=self.model,
            color=self.color,
            vehicle_price=Decimal("80000"),
            deposit_amount=Decimal("5000"),
            actual_balance=Decimal("75000"),
            id_verified=True,
            status=SalesOrder.Status.ALLOCATION_PENDING,
        )
        self.client.force_login(self.user)

    def test_financial_totals_include_all_income_and_expense_fields(self):
        profile = self.order.operations
        profile.actual_disbursement = Decimal("80000")
        profile.vehicle_cost = Decimal("60000")
        profile.registration_tax_expense = Decimal("1000")
        profile.installment_interest_subsidy = Decimal("2000")
        profile.agency_fee_income = Decimal("500")
        profile.sales_bonus = Decimal("1500")
        profile.save()
        PaymentRecord.objects.create(
            order=self.order,
            item_name="訂金",
            received_amount=Decimal("5000"),
            confirmed=True,
        )
        PaymentRecord.objects.create(
            order=self.order,
            item_name="尾款",
            received_amount=Decimal("75000"),
            confirmed=False,
        )

        self.assertEqual(profile.total_income, Decimal("500"))
        self.assertEqual(profile.total_expense, Decimal("1000"))
        self.assertEqual(profile.net_profit, Decimal("23000"))
        self.assertEqual(profile.total_received, Decimal("5000"))

    def test_prefetched_payment_totals_do_not_issue_per_order_queries(self):
        PaymentRecord.objects.create(
            order=self.order,
            item_name="已確認收款",
            received_amount=Decimal("12000"),
            confirmed=True,
        )
        PaymentRecord.objects.create(
            order=self.order,
            item_name="待確認收款",
            received_amount=Decimal("3000"),
            confirmed=False,
        )
        profile = OrderOperationsProfile.objects.select_related("order").prefetch_related(
            "order__payment_records"
        ).get(order=self.order)

        with self.assertNumQueries(0):
            total = profile.total_received

        self.assertEqual(total, Decimal("12000"))

    def test_internal_discount_requires_approval_before_changing_receivable(self):
        original_balance = self.order.actual_balance

        request_response = self.client.post(
            reverse("order_discount_request", args=[self.order.pk]),
            {"amount": "500", "reason": "最終成交抹零"},
        )
        self.order.refresh_from_db()

        self.assertRedirects(request_response, reverse("order_operations", args=[self.order.pk]))
        self.assertEqual(self.order.discount_status, SalesOrder.DiscountStatus.PENDING)
        self.assertEqual(self.order.approved_discount_amount, Decimal("0"))
        self.assertEqual(self.order.actual_balance, original_balance)

        decision_response = self.client.post(
            reverse("order_discount_decide", args=[self.order.pk]),
            {"decision": "approve", "note": "主管確認"},
        )
        self.order.refresh_from_db()

        self.assertRedirects(decision_response, reverse("order_operations", args=[self.order.pk]))
        self.assertEqual(self.order.discount_status, SalesOrder.DiscountStatus.APPROVED)
        self.assertEqual(self.order.approved_discount_amount, Decimal("500"))
        self.assertEqual(self.order.calculated_balance, Decimal("74500"))
        self.assertEqual(self.order.actual_balance, Decimal("74500"))
        contract = self.client.get(reverse("contract_print", args=[self.order.pk]))
        from pypdf import PdfReader
        pdf_bytes = b"".join(contract.streaming_content)
        text = "".join(page.extract_text() or "" for page in PdfReader(BytesIO(pdf_bytes)).pages)
        self.assertIn("已核准優惠", text)

    def test_financial_totals_include_split_expenses_and_scrap_income(self):
        profile = self.order.operations
        profile.actual_disbursement = Decimal("80000")
        profile.vehicle_cost = Decimal("60000")
        profile.gift_expense = Decimal("100")
        profile.shipping_expense = Decimal("200")
        profile.card_fee_expense = Decimal("50")
        profile.scrap_agency_income = Decimal("300")
        profile.scrap_vehicle_income = Decimal("400")
        profile.sales_bonus = Decimal("1500")
        profile.installment_interest_subsidy = Decimal("2000")
        profile.save()

        self.assertEqual(profile.total_expense, Decimal("350"))
        self.assertEqual(profile.total_income, Decimal("700"))
        self.assertEqual(profile.net_profit, Decimal("23850"))

    def test_card_payment_details_require_card_payment_method(self):
        form = PaymentRecordForm(
            data={
                "item_name": "刷卡尾款",
                "expected_amount": "30000",
                "received_amount": "31500",
                "card_principal": "30000",
                "card_fee_charged": "1500",
                "bank_card_fee": "900",
                "payment_method": "現金",
            }
        )

        self.assertFalse(form.is_valid())
        self.assertIn("付款方式必須選擇", form.errors["payment_method"][0])

    def test_card_payment_fee_difference_is_explicit(self):
        payment = PaymentRecord(
            order=self.order,
            item_name="刷卡尾款",
            payment_method="刷卡",
            card_principal=Decimal("30000"),
            card_fee_charged=Decimal("1500"),
            bank_card_fee=Decimal("900"),
        )

        self.assertEqual(payment.card_fee_difference, Decimal("600"))

    def test_registration_financial_pairs_sync_until_manually_adjusted(self):
        self.order.registration_plate_fee = Decimal("300")
        self.order.registration_license_fee = Decimal("150")
        self.order.registration_inspection_fee = Decimal("200")
        self.order.road_maintenance_fee = Decimal("188")
        self.order.license_tax_fee = Decimal("0")
        self.order.compulsory_insurance_fee = Decimal("658")
        self.order.plate_selection_fee = Decimal("300")
        self.order.plate_insurance_fee = Decimal("1796")
        self.order.balance_adjustment_reason = "測試保留既有尾款"
        self.order.save()

        profile = self.order.operations
        profile.refresh_from_db()
        self.assertEqual(profile.registration_tax_expense, Decimal("838"))
        self.assertEqual(profile.compulsory_insurance_expense, Decimal("658"))
        self.assertEqual(profile.plate_selection_expense, Decimal("300"))
        self.assertEqual(profile.registration_tax_income, Decimal("838"))
        self.assertEqual(profile.compulsory_insurance_income, Decimal("658"))
        self.assertEqual(profile.plate_selection_income, Decimal("300"))

        profile.registration_tax_expense = Decimal("900")
        profile.compulsory_insurance_income = Decimal("700")
        profile.manual_financial_fields = [
            "registration_tax_expense",
            "compulsory_insurance_income",
        ]
        profile.save()

        self.order.registration_plate_fee = Decimal("400")
        self.order.compulsory_insurance_fee = Decimal("700")
        self.order.plate_selection_fee = Decimal("500")
        self.order.plate_insurance_fee = Decimal("2200")
        self.order.save()

        profile.refresh_from_db()
        self.assertEqual(profile.registration_tax_expense, Decimal("900"))
        self.assertEqual(profile.compulsory_insurance_income, Decimal("700"))
        self.assertEqual(profile.plate_selection_expense, Decimal("500"))
        self.assertEqual(profile.registration_tax_income, Decimal("1000"))
        self.assertEqual(profile.plate_selection_income, Decimal("500"))

    def test_registration_financial_pairs_are_editable_and_grouped(self):
        response = self.client.get(
            reverse("order_operations", args=[self.order.pk])
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "領牌相關收支")
        for field_name in (
            "registration_tax_expense",
            "compulsory_insurance_expense",
            "plate_selection_expense",
            "registration_tax_income",
            "compulsory_insurance_income",
            "plate_selection_income",
        ):
            self.assertFalse(response.context["form"].fields[field_name].disabled)
            self.assertContains(
                response,
                f'name="operations-{field_name}"',
            )

    def test_order_creation_automatically_builds_operations_and_receivables(self):
        profile = self.order.operations
        payments = {
            item.system_key: item
            for item in self.order.payment_records.exclude(system_key="")
        }

        self.assertEqual(set(payments), {"deposit", "balance"})
        self.assertEqual(payments["deposit"].expected_amount, Decimal("5000"))
        self.assertEqual(payments["deposit"].received_amount, Decimal("5000"))
        self.assertEqual(payments["balance"].expected_amount, Decimal("75000"))
        self.assertFalse(profile.payment_confirmed)
        self.assertEqual(profile.actual_disbursement, Decimal("80000"))

    def test_installment_change_rebuilds_receivables_without_duplicates(self):
        self.order.payment_type = SalesOrder.PaymentType.INSTALLMENT
        self.order.installment_company = "和潤"
        self.order.installment_amount = Decimal("70000")
        self.order.installment_periods = 24
        self.order.installment_monthly = Decimal("3200")
        self.order.actual_balance = Decimal("75000")
        self.order.save()
        self.order.save()

        payments = {
            item.system_key: item
            for item in self.order.payment_records.exclude(system_key="")
        }
        self.assertEqual(
            set(payments),
            {"deposit", "balance", "installment_disbursement"},
        )
        self.assertEqual(
            payments["installment_disbursement"].expected_amount,
            Decimal("70000"),
        )
        self.assertEqual(payments["balance"].expected_amount, Decimal("5000"))
        self.assertEqual(
            self.order.operations.installment_info,
            "和潤／24期／每期 3200 元",
        )

    def test_data_maintenance_hub_and_customer_list_reuse_existing_data(self):
        hub = self.client.get(reverse("data_maintenance"))
        customers = self.client.get(reverse("customer_list"), {"q": "營運測試"})

        self.assertEqual(hub.status_code, 200)
        self.assertContains(hub, "客戶資料")
        self.assertContains(hub, "車型資料")
        self.assertContains(hub, "庫存資料")
        self.assertEqual(customers.status_code, 200)
        self.assertContains(customers, self.order.owner_name)
        self.assertContains(customers, "1 張")
        self.assertNotContains(customers, self.order.owner_id_number)

    def test_customer_detail_shows_every_order_for_same_customer(self):
        second = SalesOrder.objects.create(
            owner_name=self.order.owner_name,
            owner_phone=self.order.owner_phone,
            owner_address=self.order.owner_address,
            owner_id_number=self.order.owner_id_number,
            vehicle_model=self.model,
            color=self.color,
            vehicle_price=Decimal("82000"),
            actual_balance=Decimal("82000"),
            id_verified=True,
        )

        response = self.client.get(
            reverse("customer_detail", args=[second.pk])
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.order.number)
        self.assertContains(response, second.number)
        self.assertContains(response, "2 張歷史訂單")

    def test_dashboard_uses_registration_date_for_monthly_performance(self):
        profile = self.order.operations
        profile.actual_disbursement = Decimal("80000")
        profile.vehicle_cost = Decimal("60000")
        profile.manual_financial_fields = ["actual_disbursement"]
        profile.save()
        self.order.status = SalesOrder.Status.COMPLETED
        self.order.registration_date = timezone.localdate()
        self.order.save(update_fields=["status", "registration_date", "updated_at"])

        response = self.client.get(reverse("dashboard"))

        performance = response.context["dashboard"]["performance"]
        self.assertEqual(performance["count"], 1)
        self.assertEqual(performance["sales_total"], Decimal("80000"))
        self.assertEqual(performance["profit_total"], Decimal("20000"))
        self.assertContains(response, "營運戰情看板")

    def test_settlement_cost_uses_county_then_locks_historical_snapshot(self):
        VehicleSettlementCostRule.objects.create(
            vehicle_model=self.model,
            registration_county="",
            amount=Decimal("62000"),
            effective_from=date(2026, 7, 1),
        )
        county_rule = VehicleSettlementCostRule.objects.create(
            vehicle_model=self.model,
            registration_county="新北市",
            amount=Decimal("61000"),
            effective_from=date(2026, 8, 1),
        )
        self.order.registration_date = date(2026, 8, 5)
        self.order.registration_county = "新北市"
        self.order.save(
            update_fields=["registration_date", "registration_county", "updated_at"]
        )

        from sales.services.settlement_cost import apply_order_settlement_cost

        profile = apply_order_settlement_cost(self.order, "tester", lock=True)
        self.assertEqual(profile.vehicle_cost, Decimal("61000"))
        self.assertEqual(profile.vehicle_cost_rule, county_rule)
        self.assertIsNotNone(profile.vehicle_cost_locked_at)

        county_rule.amount = Decimal("63000")
        county_rule.save()
        profile = apply_order_settlement_cost(self.order, "tester")
        self.assertEqual(profile.vehicle_cost, Decimal("61000"))

    def test_future_cost_does_not_apply_before_effective_date(self):
        current = VehicleSettlementCostRule.objects.create(
            vehicle_model=self.model,
            amount=Decimal("60000"),
            effective_from=date(2026, 7, 1),
        )
        VehicleSettlementCostRule.objects.create(
            vehicle_model=self.model,
            amount=Decimal("65000"),
            announced_on=date(2026, 7, 25),
            effective_from=date(2026, 8, 1),
        )
        from sales.services.settlement_cost import resolve_settlement_cost

        self.assertEqual(
            resolve_settlement_cost(self.model.pk, "臺北市", date(2026, 7, 31)),
            current,
        )
        self.assertEqual(
            resolve_settlement_cost(
                self.model.pk,
                "臺北市",
                date(2026, 8, 1),
            ).amount,
            Decimal("65000"),
        )

    def test_incentive_rule_uses_registration_date_and_locks_snapshot(self):
        current = VehicleIncentiveRule.objects.create(
            vehicle_model=self.model,
            sales_bonus=Decimal("1500"),
            promotion_subsidy=Decimal("2000"),
            installment_interest_subsidy=Decimal("800"),
            effective_from=date(2026, 7, 1),
        )
        future = VehicleIncentiveRule.objects.create(
            vehicle_model=self.model,
            sales_bonus=Decimal("1800"),
            promotion_subsidy=Decimal("2500"),
            installment_interest_subsidy=Decimal("900"),
            effective_from=date(2026, 8, 1),
        )
        current_rate = VehicleIncentiveInstallmentRate.objects.create(
            incentive_rule=current,
            periods=24,
            rate=Decimal("92.50"),
        )
        future_rate = VehicleIncentiveInstallmentRate.objects.create(
            incentive_rule=future,
            periods=24,
            rate=Decimal("95.00"),
        )
        self.order.payment_type = SalesOrder.PaymentType.INSTALLMENT
        self.order.installment_periods = 24
        self.order.registration_date = date(2026, 7, 31)
        self.order.save(
            update_fields=[
                "payment_type",
                "installment_periods",
                "registration_date",
                "updated_at",
            ]
        )

        from sales.services.incentive_rule import apply_order_incentive_rule

        profile = apply_order_incentive_rule(self.order, "tester")
        self.assertEqual(profile.incentive_rule, current)
        self.assertEqual(profile.sales_bonus, Decimal("1500"))
        self.assertEqual(profile.promotion_subsidy, Decimal("2000"))
        self.assertEqual(profile.installment_interest_subsidy, Decimal("800"))
        self.assertEqual(profile.actual_disbursement, Decimal("74000"))
        self.assertEqual(profile.incentive_installment_rate_rule, current_rate)
        self.assertEqual(profile.incentive_installment_periods, 24)
        self.assertEqual(profile.incentive_installment_rate, Decimal("92.50"))

        self.order.registration_date = date(2026, 8, 1)
        self.order.save(update_fields=["registration_date", "updated_at"])
        profile = apply_order_incentive_rule(self.order, "tester", lock=True)
        self.assertEqual(profile.incentive_rule, future)
        self.assertEqual(profile.sales_bonus, Decimal("1800"))
        self.assertEqual(profile.actual_disbursement, Decimal("76000"))
        self.assertEqual(profile.incentive_installment_rate_rule, future_rate)
        self.assertIsNotNone(profile.incentive_locked_at)

        future.sales_bonus = Decimal("3000")
        future.save()
        profile = apply_order_incentive_rule(self.order, "tester")
        self.assertEqual(profile.sales_bonus, Decimal("1800"))

    def test_manual_incentive_adjustment_is_not_overwritten_by_new_version(self):
        VehicleIncentiveRule.objects.create(
            vehicle_model=self.model,
            sales_bonus=Decimal("1000"),
            promotion_subsidy=Decimal("2000"),
            installment_interest_subsidy=Decimal("700"),
            effective_from=date(2026, 7, 1),
        )
        VehicleIncentiveRule.objects.create(
            vehicle_model=self.model,
            sales_bonus=Decimal("1500"),
            promotion_subsidy=Decimal("2500"),
            installment_interest_subsidy=Decimal("900"),
            effective_from=date(2026, 8, 1),
        )
        self.order.registration_date = date(2026, 7, 20)
        self.order.save(update_fields=["registration_date", "updated_at"])

        from sales.services.incentive_rule import apply_order_incentive_rule

        profile = apply_order_incentive_rule(self.order)
        profile.sales_bonus = Decimal("1200")
        profile.manual_financial_fields = ["sales_bonus"]
        profile.save()

        self.order.registration_date = date(2026, 8, 5)
        self.order.save(update_fields=["registration_date", "updated_at"])
        profile = apply_order_incentive_rule(self.order)
        self.assertEqual(profile.sales_bonus, Decimal("1200"))
        self.assertEqual(profile.promotion_subsidy, Decimal("2500"))
        self.assertEqual(profile.installment_interest_subsidy, Decimal("900"))

    def test_installment_disbursement_uses_matching_period_and_missing_period_is_not_guessed(self):
        rule = VehicleIncentiveRule.objects.create(
            vehicle_model=self.model,
            effective_from=date(2026, 7, 1),
        )
        VehicleIncentiveInstallmentRate.objects.create(
            incentive_rule=rule,
            periods=12,
            rate=Decimal("96"),
        )
        VehicleIncentiveInstallmentRate.objects.create(
            incentive_rule=rule,
            periods=24,
            rate=Decimal("92.5"),
        )
        self.order.payment_type = SalesOrder.PaymentType.INSTALLMENT
        self.order.installment_periods = 12
        self.order.registration_date = date(2026, 7, 20)
        self.order.save(
            update_fields=[
                "payment_type",
                "installment_periods",
                "registration_date",
                "updated_at",
            ]
        )

        from sales.services.incentive_rule import apply_order_incentive_rule

        profile = apply_order_incentive_rule(self.order)
        self.assertEqual(profile.actual_disbursement, Decimal("76800"))
        self.assertEqual(profile.incentive_installment_rate, Decimal("96"))

        self.order.installment_periods = 24
        self.order.save(update_fields=["installment_periods", "updated_at"])
        profile = apply_order_incentive_rule(self.order)
        self.assertEqual(profile.actual_disbursement, Decimal("74000"))
        self.assertEqual(profile.incentive_installment_rate, Decimal("92.5"))

        self.order.installment_periods = 36
        self.order.save(update_fields=["installment_periods", "updated_at"])
        profile = apply_order_incentive_rule(self.order)
        self.assertEqual(profile.actual_disbursement, Decimal("0"))
        self.assertEqual(profile.incentive_installment_periods, 36)
        self.assertIsNone(profile.incentive_installment_rate)

    def test_platform_order_keeps_manually_entered_disbursement(self):
        platform = SalesSource.objects.create(
            source_type=SalesSource.SourceType.PLATFORM,
            name="測試平台",
        )
        rule = VehicleIncentiveRule.objects.create(
            vehicle_model=self.model,
            effective_from=date(2026, 7, 1),
        )
        self.order.source_type = SalesOrder.SourceType.PLATFORM
        self.order.source = platform
        self.order.payment_type = SalesOrder.PaymentType.INSTALLMENT
        self.order.registration_date = date(2026, 7, 20)
        self.order.save(
            update_fields=[
                "source_type",
                "source",
                "payment_type",
                "registration_date",
                "updated_at",
            ]
        )
        profile = self.order.operations
        profile.actual_disbursement = Decimal("73000")
        profile.save(update_fields=["actual_disbursement", "updated_at"])

        from sales.services.incentive_rule import apply_order_incentive_rule

        profile = apply_order_incentive_rule(self.order)
        self.assertEqual(profile.incentive_rule, rule)
        self.assertEqual(profile.actual_disbursement, Decimal("73000"))

    def test_incentive_rule_maintenance_pages_are_available(self):
        rule = VehicleIncentiveRule.objects.create(
            vehicle_model=self.model,
            sales_bonus=Decimal("1500"),
            promotion_subsidy=Decimal("2000"),
            installment_interest_subsidy=Decimal("800"),
            effective_from=date(2026, 8, 1),
        )

        listing = self.client.get(reverse("incentive_rule_list"))
        editing = self.client.get(reverse("incentive_rule_edit", args=[rule.pk]))

        self.assertEqual(listing.status_code, 200)
        self.assertContains(listing, "車型獎勵與補助")
        self.assertContains(listing, "1500")
        self.assertContains(listing, "持續有效")
        self.assertEqual(editing.status_code, 200)
        self.assertContains(editing, "實銷獎勵金")
        self.assertContains(editing, "結束日期")

    def test_vehicle_model_edit_embeds_incentive_history_and_editor(self):
        rule = VehicleIncentiveRule.objects.create(
            vehicle_model=self.model,
            sales_bonus=Decimal("1500"),
            effective_from=date(2026, 8, 1),
        )

        response = self.client.get(
            reverse("vehicle_model_edit", args=[self.model.pk])
        )
        edit_response = self.client.get(
            reverse("vehicle_model_edit", args=[self.model.pk]),
            {"edit_incentive": rule.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "獎勵、補助與分期撥款")
        self.assertContains(response, "新增獎勵補助版本")
        self.assertContains(response, "1500")
        self.assertNotContains(
            response,
            reverse("incentive_rule_create"),
        )
        self.assertContains(edit_response, "編輯 2026/08/01 版本")

    def test_vehicle_model_edit_can_create_and_update_incentive_inline(self):
        edit_url = reverse("vehicle_model_edit", args=[self.model.pk])
        create_response = self.client.post(
            edit_url,
            {
                "action": "save_incentive",
                "incentive-sales_bonus": "1800",
                "incentive-promotion_subsidy": "2200",
                "incentive-installment_interest_subsidy": "900",
                "incentive-announced_on": "2026-07-20",
                "incentive-effective_from": "2026-08-01",
                "incentive-effective_to": "",
                "incentive-note": "八月版本",
                "incentive-active": "on",
                "rates-TOTAL_FORMS": "2",
                "rates-INITIAL_FORMS": "0",
                "rates-MIN_NUM_FORMS": "0",
                "rates-MAX_NUM_FORMS": "1000",
                "rates-0-periods": "12",
                "rates-0-rate": "96",
                "rates-1-periods": "24",
                "rates-1-rate": "92.5",
            },
        )

        rule = VehicleIncentiveRule.objects.get(vehicle_model=self.model)
        self.assertRedirects(
            create_response,
            f"{edit_url}#incentive-rules",
        )
        self.assertEqual(rule.sales_bonus, Decimal("1800"))
        self.assertEqual(
            list(rule.installment_rates.values_list("periods", "rate")),
            [(12, Decimal("96")), (24, Decimal("92.5"))],
        )

        rate_12 = rule.installment_rates.get(periods=12)
        rate_24 = rule.installment_rates.get(periods=24)
        update_response = self.client.post(
            edit_url,
            {
                "action": "save_incentive",
                "rule_id": rule.pk,
                "incentive-sales_bonus": "2000",
                "incentive-promotion_subsidy": "2200",
                "incentive-installment_interest_subsidy": "900",
                "incentive-announced_on": "2026-07-20",
                "incentive-effective_from": "2026-08-01",
                "incentive-effective_to": "",
                "incentive-note": "更新版本",
                "incentive-active": "on",
                "rates-TOTAL_FORMS": "2",
                "rates-INITIAL_FORMS": "2",
                "rates-MIN_NUM_FORMS": "0",
                "rates-MAX_NUM_FORMS": "1000",
                "rates-0-id": rate_12.pk,
                "rates-0-periods": "12",
                "rates-0-rate": "97",
                "rates-1-id": rate_24.pk,
                "rates-1-periods": "24",
                "rates-1-rate": "93",
            },
        )

        self.assertRedirects(
            update_response,
            f"{edit_url}#incentive-rules",
        )
        rule.refresh_from_db()
        self.assertEqual(rule.sales_bonus, Decimal("2000"))
        self.assertEqual(rule.installment_rates.get(periods=12).rate, Decimal("97"))
        self.assertEqual(rule.installment_rates.get(periods=24).rate, Decimal("93"))

    def test_settlement_cost_maintenance_pages_are_available(self):
        rule = VehicleSettlementCostRule.objects.create(
            vehicle_model=self.model,
            registration_county="新北市",
            amount=Decimal("61000"),
            effective_from=date(2026, 8, 1),
        )

        listing = self.client.get(reverse("settlement_cost_rule_list"))
        editing = self.client.get(
            reverse("settlement_cost_rule_edit", args=[rule.pk])
        )

        self.assertEqual(listing.status_code, 200)
        self.assertContains(listing, "代銷結算成本")
        self.assertContains(listing, "新北市")
        self.assertContains(listing, "61000")
        self.assertEqual(editing.status_code, 200)
        self.assertContains(editing, "領牌縣市")

    def test_payment_page_starts_with_one_manual_row_and_collapses_system_items(self):
        response = self.client.get(
            reverse("order_operations", args=[self.order.pk])
        )
        html = response.content.decode()

        self.assertContains(response, "系統應收摘要")
        self.assertContains(response, "需要確認收款時再展開")
        # 一筆預設人工列，加上一筆供動態新增使用的 template。
        self.assertEqual(
            html.count(
                '<button type="button" class="button danger-outline small" '
                "data-delete-payment>"
            ),
            2,
        )

    def test_installment_disbursement_can_be_confirmed_from_reconciliation_list(self):
        self.order.payment_type = SalesOrder.PaymentType.INSTALLMENT
        self.order.installment_company = "和潤"
        self.order.installment_amount = Decimal("70000")
        self.order.installment_periods = 24
        self.order.installment_monthly = Decimal("3200")
        self.order.save()
        record = self.order.payment_records.get(
            system_key="installment_disbursement"
        )

        listing = self.client.get(reverse("reconciliation_list"))
        self.assertEqual(listing.status_code, 200)
        self.assertContains(listing, "和潤")
        self.assertContains(listing, "70000")

        response = self.client.post(
            reverse("reconciliation_update", args=[record.pk]),
            {
                "received_amount": "69800",
                "received_on": "2026-08-05",
                "receiving_account": "公司帳戶",
                "confirmed": "on",
                "note": "人工核對帳本",
            },
        )

        self.assertRedirects(response, reverse("reconciliation_list"))
        record.refresh_from_db()
        self.order.operations.refresh_from_db()
        self.assertEqual(record.received_amount, Decimal("69800"))
        self.assertTrue(record.confirmed)
        self.assertEqual(record.confirmed_by, "operations-user")
        self.assertTrue(self.order.operations.installment_transfer_confirmed)
        self.assertEqual(
            self.order.operations.actual_disbursement,
            Decimal("69800"),
        )
        self.assertIn(
            "actual_disbursement",
            self.order.operations.manual_financial_fields,
        )
        self.assertTrue(
            OrderChange.objects.filter(
                order=self.order,
                reason__startswith="統一對帳更新",
            ).exists()
        )

    def test_reconciliation_update_accepts_same_site_next_url(self):
        self.order.payment_type = SalesOrder.PaymentType.INSTALLMENT
        self.order.installment_amount = Decimal("70000")
        self.order.save()
        record = self.order.payment_records.get(
            system_key="installment_disbursement"
        )
        target = reverse("reconciliation_list") + "?status=pending"

        response = self.client.post(
            reverse("reconciliation_update", args=[record.pk]),
            {
                "received_amount": "5000",
                "received_on": "2026-08-05",
                "receiving_account": "公司帳戶",
                "note": "同源返回測試",
                "next": target,
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], target)

    def test_reconciliation_update_rejects_protocol_relative_next_url(self):
        self.order.payment_type = SalesOrder.PaymentType.INSTALLMENT
        self.order.installment_amount = Decimal("70000")
        self.order.save()
        record = self.order.payment_records.get(
            system_key="installment_disbursement"
        )

        response = self.client.post(
            reverse("reconciliation_update", args=[record.pk]),
            {
                "received_amount": "5000",
                "received_on": "2026-08-05",
                "receiving_account": "公司帳戶",
                "note": "外部返回測試",
                "next": "//evil.example/steal-session",
            },
        )

        self.assertRedirects(response, reverse("reconciliation_list"))

    def test_existing_order_without_profile_can_open_operations_page(self):
        self.order.operations.delete()
        response = self.client.get(
            reverse("order_operations", args=[self.order.pk])
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "收款與分期對帳")
        self.assertTrue(
            OrderOperationsProfile.objects.filter(order=self.order).exists()
        )

    def test_secret_is_encrypted_and_reveal_is_audited(self):
        from sales.services.secret_fields import encrypt_secret

        encrypted = encrypt_secret("safe-password")
        profile = self.order.operations
        profile.vehicle_control_password_encrypted = encrypted
        profile.save()

        self.assertNotIn("safe-password", profile.vehicle_control_password_encrypted)
        self.assertEqual(decrypt_secret(encrypted), "safe-password")
        response = self.client.post(
            reverse("order_secret_reveal", args=[self.order.pk]),
            {"field": "vehicle_control_password"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["value"], "safe-password")
        self.assertTrue(
            OrderEvent.objects.filter(
                order=self.order,
                event_type="secret_viewed",
            ).exists()
        )

    def test_operations_report_and_excel_export(self):
        profile = self.order.operations
        profile.vehicle_cost = Decimal("60000")
        profile.save()
        list_response = self.client.get(reverse("operations_report"))
        export_response = self.client.get(reverse("operations_report_export"))

        self.assertEqual(list_response.status_code, 200)
        self.assertContains(list_response, self.order.number)
        self.assertContains(list_response, "車款成交額")
        self.assertContains(list_response, "實際領牌日期")
        self.assertContains(list_response, "依車型查看營運表現")
        self.assertContains(list_response, 'class="analysis-table"')
        self.assertContains(list_response, "$ 80,000")
        self.assertEqual(list_response.context["analysis_summary"]["count"], 1)
        self.assertEqual(
            list_response.context["analysis_summary"]["vehicle_sales"],
            Decimal("80000"),
        )
        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(
            export_response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(export_response.content))
        sheet = workbook["營運總表"]
        headers = [cell.value for cell in sheet[1]]
        self.assertIn("單筆淨利", headers)
        self.assertIn("電池合約方案", headers)
        self.assertEqual(sheet.cell(row=2, column=1).value, self.order.number)

    def test_operations_excel_export_escapes_formula_like_user_text(self):
        self.order.owner_name = "=HYPERLINK(\"https://evil.example\",\"點我\")"
        self.order.owner_address = "+CMD|' /C calc'!A0"
        self.order.delivery_destination = "-2+3"
        self.order.old_owner_name = "@SUM(1+1)"
        self.order.save(
            update_fields=[
                "owner_name",
                "owner_address",
                "delivery_destination",
                "old_owner_name",
                "updated_at",
            ]
        )

        response = self.client.get(reverse("operations_report_export"))

        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(response.content), data_only=False)
        sheet = workbook["營運總表"]
        columns = {cell.value: cell.column for cell in sheet[1]}
        expected = {
            "車主名稱": "'=HYPERLINK(\"https://evil.example\",\"點我\")",
            "戶籍地址": "'+CMD|' /C calc'!A0",
            "自送托運地點": "'-2+3",
            "舊車車主": "'@SUM(1+1)",
        }
        for header, value in expected.items():
            cell = sheet.cell(row=2, column=columns[header])
            self.assertEqual(cell.value, value)
            self.assertEqual(cell.data_type, "s")

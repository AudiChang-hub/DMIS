"""同步「車行、網路平台.xlsx」的車行聯絡簿資料。

此服務把 Excel 車行頁視為聯絡資料的唯一來源，但不刪除系統中未出現在
Excel 的車行。同步可重跑，且台鈴欄位依現行營運語意拆成油車／電車合作。
"""

from dataclasses import dataclass
from datetime import date
from pathlib import Path
import re

from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from openpyxl import load_workbook

from sales.models import (
    SalesSource,
    SalesSourceBrandPolicy,
    SalesSourceCategory,
    SalesSourceCooperationProfile,
)


DEALER_NAME_ALIASES = {
    "東湖上慶": "上慶",
    "湖州": "湖洲",
}


@dataclass(frozen=True)
class DealerWorkbookRow:
    row_number: int
    holiday_gift: bool
    has_line_group: bool
    name: str
    responsible_person: str
    phone: str
    phone_secondary: str
    mobile: str
    other_contact: str
    address: str
    sym_value: str
    suzuki_value: str
    sym_capacity: int | None
    suzuki_capacity: int | None
    note: str

    @property
    def canonical_name(self):
        return DEALER_NAME_ALIASES.get(self.name, self.name)

    @property
    def sym_cooperates(self):
        return self.sym_value.upper() in {"V", "專銷"}

    @property
    def sym_relationship(self):
        if self.sym_value == "專銷":
            return SalesSourceCooperationProfile.RelationshipType.EXCLUSIVE
        return SalesSourceCooperationProfile.RelationshipType.GENERAL

    @property
    def suzuki_gas_cooperates(self):
        # 現行聯絡簿：V 代表台鈴油車、電車都有合作。
        return self.suzuki_value.upper() == "V"

    @property
    def suzuki_electric_cooperates(self):
        # 「電動車」只代表台鈴電車；V 則代表油電都有。
        value = self.suzuki_value.replace(" ", "").upper()
        return value == "V" or value == "電動車"


@dataclass
class DealerWorkbookSyncResult:
    source_rows: int = 0
    created: int = 0
    updated: int = 0
    unchanged: int = 0
    aliases_used: int = 0
    database_only: int = 0


def _text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _capacity(value):
    text = _text(value)
    if not text:
        return None
    try:
        parsed = int(float(text))
    except (TypeError, ValueError):
        return None
    return parsed if parsed >= 0 else None


def read_dealer_workbook(path):
    """讀取現行聯絡簿的車行頁，第三列起為資料。"""
    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        if "車行" not in workbook.sheetnames:
            raise ValueError("找不到「車行」工作表。")
        sheet = workbook["車行"]
        headers = [_text(cell) for cell in next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))]
        if len(headers) < 14 or headers[2] != "店名" or headers[9] != "三陽" or headers[10] != "台鈴":
            raise ValueError("車行工作表欄位格式不符，請確認第二列仍為聯絡簿欄位列。")
        rows = []
        for row_number, values in enumerate(sheet.iter_rows(min_row=3, values_only=True), 3):
            values = tuple(values) + (None,) * max(0, 14 - len(values))
            name = _text(values[2])
            if not name:
                continue
            rows.append(
                DealerWorkbookRow(
                    row_number=row_number,
                    holiday_gift=bool(_text(values[0])),
                    has_line_group=bool(_text(values[1])),
                    name=name,
                    responsible_person=_text(values[3]),
                    phone=_text(values[4]),
                    phone_secondary=_text(values[5]),
                    mobile=_text(values[6]),
                    other_contact=_text(values[7]),
                    address=_text(values[8]),
                    sym_value=_text(values[9]),
                    suzuki_value=_text(values[10]),
                    sym_capacity=_capacity(values[11]),
                    suzuki_capacity=_capacity(values[12]),
                    note=_text(values[13]),
                )
            )
        return rows
    finally:
        workbook.close()


def _dealer_category():
    category, _ = SalesSourceCategory.objects.get_or_create(
        name="合作車行",
        defaults={
            "system_behavior": SalesSource.SourceType.DEALER,
            "active": True,
        },
    )
    return category


def _next_dealer_code(relationship_type, used_codes, today=None):
    today = today or date.today()
    prefix = "S" if relationship_type == SalesSourceCooperationProfile.RelationshipType.EXCLUSIVE else "N"
    stem = f"{prefix}{today:%y%m%d}"
    suffixes = []
    for code in used_codes:
        match = re.fullmatch(rf"{re.escape(stem)}(\d{{2}})", code or "")
        if match:
            suffixes.append(int(match.group(1)))
    for suffix in range(max(suffixes, default=0) + 1, 100):
        candidate = f"{stem}{suffix:02d}"
        if candidate not in used_codes:
            used_codes.add(candidate)
            return candidate
    raise ValueError(f"{stem} 當日車行代碼已用盡。")


def _profile_values(row):
    return {
        SalesSourceBrandPolicy.CooperationScope.SYM: (
            row.sym_cooperates,
            row.sym_relationship,
        ),
        SalesSourceBrandPolicy.CooperationScope.SUZUKI_GAS: (
            row.suzuki_gas_cooperates,
            SalesSourceCooperationProfile.RelationshipType.GENERAL,
        ),
        SalesSourceBrandPolicy.CooperationScope.SUZUKI_ELECTRIC: (
            row.suzuki_electric_cooperates,
            SalesSourceCooperationProfile.RelationshipType.GENERAL,
        ),
    }


def _sync_current_policy(source, scope, cooperates):
    """同步目前生效的商務規則，並保留既有傭金調整與備註。"""
    today = timezone.localdate()
    current = (
        source.brand_policies.filter(
            cooperation_scope=scope,
            effective_from__lte=today,
        )
        .filter(Q(effective_to__isnull=True) | Q(effective_to__gte=today))
        .order_by("-effective_from", "-pk")
        .first()
    )
    if current and current.cooperates == cooperates:
        return False
    SalesSourceBrandPolicy.objects.update_or_create(
        source=source,
        cooperation_scope=scope,
        effective_from=today,
        defaults={
            "cooperates": cooperates,
            "commission_adjustment": current.commission_adjustment if current else 0,
            "effective_to": None,
            "note": current.note if current else "",
        },
    )
    return True


@transaction.atomic
def sync_dealer_workbook(path, *, apply=False):
    rows = read_dealer_workbook(path)
    result = DealerWorkbookSyncResult(source_rows=len(rows))
    category = _dealer_category()
    dealers = list(SalesSource.objects.filter(source_type=SalesSource.SourceType.DEALER))
    by_name = {dealer.name.casefold(): dealer for dealer in dealers}
    used_codes = set(SalesSource.objects.exclude(code="").values_list("code", flat=True))
    workbook_names = set()

    for row in rows:
        workbook_names.add(row.canonical_name.casefold())
        source = by_name.get(row.canonical_name.casefold())
        created = source is None
        if created:
            source = SalesSource(
                source_type=SalesSource.SourceType.DEALER,
                name=row.canonical_name,
                category=category,
                code=_next_dealer_code(row.sym_relationship, used_codes),
            )
            result.created += 1
            by_name[source.name.casefold()] = source
        else:
            if row.name != row.canonical_name:
                result.aliases_used += 1

        values = {
            "category": category,
            "responsible_person": row.responsible_person,
            "phone": row.phone,
            "phone_secondary": row.phone_secondary,
            "mobile": row.mobile,
            "other_contact": row.other_contact,
            "address": row.address,
            "sym_vehicle_capacity": row.sym_capacity,
            "suzuki_vehicle_capacity": row.suzuki_capacity,
            "holiday_gift": row.holiday_gift,
            "has_line_group": row.has_line_group,
            # 備註以 Excel 為準；空白也會清除舊匯入產生的額外文字。
            "note": row.note,
        }
        changed = created
        for field_name, value in values.items():
            if getattr(source, field_name) != value:
                setattr(source, field_name, value)
                changed = True
        source.save()

        for scope, (cooperates, relationship_type) in _profile_values(row).items():
            profile, profile_created = SalesSourceCooperationProfile.objects.get_or_create(
                source=source,
                cooperation_scope=scope,
            )
            profile_values = {
                "cooperates": cooperates,
                "relationship_type": relationship_type,
                "note": "",
            }
            profile_changed = profile_created
            for field_name, value in profile_values.items():
                if getattr(profile, field_name) != value:
                    setattr(profile, field_name, value)
                    profile_changed = True
            if profile_changed:
                profile.save()
                changed = True
            if _sync_current_policy(source, scope, cooperates):
                changed = True

        if not created:
            if changed:
                result.updated += 1
            else:
                result.unchanged += 1

    result.database_only = sum(
        dealer.name.casefold() not in workbook_names for dealer in dealers
    )
    if not apply:
        transaction.set_rollback(True)
    return result

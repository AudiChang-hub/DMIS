import hashlib
import json
import re
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation

from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from sales.models import (
    LegacyImportBatch,
    LegacyImportCorrection,
    LegacyImportMasterMapping,
    LegacyImportRow,
    LegacySalesSnapshot,
    OrderOperationsProfile,
    PaymentRecord,
    SalesOrder,
    SalesSource,
    SalesSourceBrandPolicy,
    SalesSourceCategory,
    SalesSourceCooperationProfile,
    Store,
    VehicleColor,
    VehicleInventory,
    VehicleModel,
    normalize_legacy_master_value,
    normalize_vehicle_model_master_value,
    normalize_vehicle_identifier,
)


DUPLICATE_IDENTIFIER_MESSAGE = "同一工作表存在重複的標準化車輛識別號碼"
MULTIPLE_NEW_SALES_MESSAGE = "同一識別號碼存在多筆新車銷售；請確認後續交易是否為中古車"
DUPLICATE_SALES_TRANSACTION_MESSAGE = "同一筆銷售交易在工作表重複出現"
MISSING_IDENTIFIER_MESSAGE = "缺少引擎／車身號碼"
EMPTY_SALES_PLACEHOLDER_MESSAGE = "Excel 空白公式列，系統自動略過"
NON_VEHICLE_SALES_NOISE_MESSAGE = "缺少有效車輛序號且無交易資料，系統自動略過"
INVALID_EMAIL_MESSAGE = "Email 格式不正確，請修正或清空後再匯入"
PREVIEW_SCHEMA_VERSION = 6
SYSTEM_VALIDATION_MESSAGES = {
    DUPLICATE_IDENTIFIER_MESSAGE,
    MULTIPLE_NEW_SALES_MESSAGE,
    DUPLICATE_SALES_TRANSACTION_MESSAGE,
    MISSING_IDENTIFIER_MESSAGE,
    EMPTY_SALES_PLACEHOLDER_MESSAGE,
    NON_VEHICLE_SALES_NOISE_MESSAGE,
    INVALID_EMAIL_MESSAGE,
}


def _has_invalid_email(value):
    if not value:
        return False
    try:
        validate_email(value)
    except ValidationError:
        return True
    return False


def friendly_import_message(message):
    text = str(message)
    if "owner_email" in text and ("電子郵件" in text or "email" in text.lower()):
        return INVALID_EMAIL_MESSAGE
    return text


def _friendly_import_exception(exc):
    if isinstance(exc, ValidationError):
        if hasattr(exc, "message_dict"):
            parts = []
            field_labels = {"owner_email": "Email"}
            for field, field_messages in exc.message_dict.items():
                label = field_labels.get(field, field)
                parts.extend(f"{label}：{message}" for message in field_messages)
            return friendly_import_message("；".join(parts))
        return "；".join(exc.messages)
    return friendly_import_message(exc)


def _normalize_master_mapping_value(mapping_type, source_value):
    if mapping_type == LegacyImportMasterMapping.MappingType.VEHICLE_MODEL:
        return normalize_vehicle_model_master_value(source_value)
    return normalize_legacy_master_value(source_value)


def _legacy_master_mapping(mapping_type, source_value):
    normalized = _normalize_master_mapping_value(mapping_type, source_value)
    if not normalized:
        return None
    return (
        LegacyImportMasterMapping.objects.select_related(
            "vehicle_model", "sales_source"
        )
        .filter(
            mapping_type=mapping_type,
            normalized_source_value=normalized,
        )
        .first()
    )


@transaction.atomic
def save_import_master_mapping(
    *,
    mapping_type,
    source_value,
    actor_name,
    vehicle_model=None,
    sales_source=None,
    ignored=False,
    note="",
):
    normalized = _normalize_master_mapping_value(mapping_type, source_value)
    if not normalized:
        raise ValueError("缺少要處理的 Excel 原始名稱。")
    mapping = (
        LegacyImportMasterMapping.objects.select_for_update()
        .filter(
            mapping_type=mapping_type,
            normalized_source_value=normalized,
        )
        .first()
        or LegacyImportMasterMapping(
            mapping_type=mapping_type,
            normalized_source_value=normalized,
        )
    )
    mapping.source_value = str(source_value).strip()
    mapping.vehicle_model = vehicle_model
    mapping.sales_source = sales_source
    mapping.ignored = ignored
    mapping.note = note.strip()
    mapping.updated_by = actor_name
    mapping.save()
    return mapping


def build_import_master_workspace(batch):
    """建立同頁補主檔清單，不改寫 Excel 原始列。"""
    validation = (batch.preview_summary or {}).get("validation", {})
    unmapped_models = validation.get("unmapped_models", [])
    unmapped_sources = validation.get("unmapped_sources", [])
    wanted_model_keys = {
        normalize_vehicle_model_master_value(value): value for value in unmapped_models
    }
    wanted_source_keys = {
        normalize_legacy_master_value(value): value for value in unmapped_sources
    }
    model_stats = {
        key: {
            "source_value": value,
            "row_count": 0,
            "colors": set(),
            "sales_examples": [],
            "inventory_examples": [],
        }
        for key, value in wanted_model_keys.items()
    }
    source_stats = {
        key: {"source_value": value, "row_count": 0}
        for key, value in wanted_source_keys.items()
    }
    rows = batch.rows.filter(excluded=False).values_list(
        "sheet_name", "source_row", "mapped_data"
    )
    for sheet_name, source_row, data in rows.iterator(chunk_size=500):
        model_key = normalize_vehicle_model_master_value(data.get("model_number"))
        if model_key in model_stats:
            model_stat = model_stats[model_key]
            model_stat["row_count"] += 1
            color = str(data.get("color") or "").strip()
            if color:
                model_stat["colors"].add(color)
            example_bucket = (
                model_stat["sales_examples"]
                if sheet_name == "銷貨"
                else model_stat["inventory_examples"]
            )
            if len(example_bucket) < 5:
                example_bucket.append(
                    {
                        "sheet_name": sheet_name,
                        "source_row": source_row,
                        "context_label": (
                            "中古車銷售"
                            if data.get("vehicle_category")
                            == SalesOrder.VehicleCategory.USED
                            else "新車銷售"
                            if sheet_name == "銷貨"
                            else "庫存進貨"
                        ),
                        "color": color or "未記錄",
                        "identifier": str(
                            data.get("identifier_raw")
                            or data.get("identifier")
                            or ""
                        ).strip(),
                        "plate_number": str(data.get("plate_number") or "").strip(),
                        "owner_name": str(data.get("owner_name") or "").strip(),
                        "registration_date": data.get("registration_date") or "",
                        "received_on": data.get("received_on") or "",
                        "manufactured_year_month": data.get("manufactured_year_month")
                        or "",
                    }
                )
        if sheet_name == "銷貨":
            source_key = normalize_legacy_master_value(data.get("dealer_name"))
            if source_key in source_stats:
                source_stats[source_key]["row_count"] += 1
    model_items = []
    for value in model_stats.values():
        value["colors"] = sorted(value["colors"])
        value["colors_text"] = "、".join(value["colors"])
        value["examples"] = (
            value.pop("sales_examples") + value.pop("inventory_examples")
        )[:5]
        model_items.append(value)
    model_items = sorted(
        model_items, key=lambda item: item["source_value"].casefold()
    )
    source_items = sorted(
        source_stats.values(), key=lambda item: item["source_value"].casefold()
    )
    return {
        "models": model_items,
        "sources": source_items,
        "total": len(model_items) + len(source_items),
    }


def file_sha256(uploaded_file):
    digest = hashlib.sha256()
    for chunk in uploaded_file.chunks():
        digest.update(chunk)
    uploaded_file.seek(0)
    return digest.hexdigest()


def _json_value(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def _row_dict(sheet, row_number, header_row):
    values = {}
    for column in range(1, sheet.max_column + 1):
        header = sheet.cell(header_row, column).value
        value = sheet.cell(row_number, column).value
        if header not in (None, "") or value not in (None, ""):
            label = str(header).strip() if header not in (None, "") else get_column_letter(column)
            if label in values:
                label = f"{label}（{get_column_letter(column)}欄）"
            values[label] = _json_value(value)
    return values


def _row_dict_values(headers, row_values):
    values = {}
    for column, (header, value) in enumerate(zip(headers, row_values), 1):
        if header not in (None, "") or value not in (None, ""):
            label = str(header).strip() if header not in (None, "") else get_column_letter(column)
            if label in values:
                label = f"{label}（{get_column_letter(column)}欄）"
            values[label] = _json_value(value)
    return values


def _value(row_values, column_letter):
    index = column_index_from_string(column_letter) - 1
    return row_values[index] if index < len(row_values) else None


def _cell(sheet, row_number, column_letter):
    return sheet[f"{column_letter}{row_number}"].value


def _text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _decimal(value):
    if value in (None, "", "-"):
        return Decimal("0")
    try:
        return Decimal(str(value).replace(",", "").replace("$", "").strip())
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value).replace(".", "/").replace("-", "/")
    if not text:
        return None
    for fmt in ("%Y/%m/%d", "%Y/%m", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _infer_sales_vehicle_category(raw, dealer_name):
    """只使用明確的中古車交易文字，避免把新車的舊車估價誤判為中古車。"""
    dealer = _text(dealer_name).replace(" ", "")
    note = _text(raw.get("備註"))
    subsidy = _text(raw.get("補助方案"))
    if dealer == "中古車":
        return SalesOrder.VehicleCategory.USED, "車行欄位為中古車"
    if "中古車買賣" in note or note.strip() == "中古車":
        return SalesOrder.VehicleCategory.USED, "備註標示中古車交易"
    if "中古車過戶" in subsidy:
        return SalesOrder.VehicleCategory.USED, "補助方案標示中古車過戶"
    return SalesOrder.VehicleCategory.NEW, "未發現中古車交易標記"


TRANSACTION_TYPE_MARKERS = (
    ("試乘", SalesOrder.TransactionType.TEST_RIDE),
    ("中獎", SalesOrder.TransactionType.PRIZE),
    ("領牌車", SalesOrder.TransactionType.REGISTERED),
    ("一般新車", SalesOrder.TransactionType.REGULAR_NEW),
)


def _infer_sales_transaction_type(raw, dealer_name, vehicle_category, sales_category=""):
    """交易類型與來源分開判斷，避免把「試乘車」建立成通路。"""
    if vehicle_category == SalesOrder.VehicleCategory.USED:
        return SalesOrder.TransactionType.USED, "車輛類別為中古車"
    text = " ".join(
        filter(
            None,
            (
                _text(dealer_name),
                _text(sales_category),
                _text(raw.get("備註")),
                _text(raw.get("銷售方案分類")),
            ),
        )
    ).replace(" ", "")
    for marker, transaction_type in TRANSACTION_TYPE_MARKERS:
        if marker in text:
            return transaction_type, f"來源資料含有「{marker}」"
    return SalesOrder.TransactionType.REGULAR_NEW, "未發現特殊交易類型標記"


def _clean_sales_source_name(value):
    """移除車行欄位內的訂單註記；純註記文字不視為通路。"""
    text = _text(value).strip()
    if not text:
        return ""
    compact = text.replace(" ", "")
    pure_markers = {
        "試乘車",
        "試乘",
        "中獎車",
        "中獎",
        "領牌車",
        "一般新車",
        "中古車",
        "代申請補助",
    }
    if compact in pure_markers:
        return ""
    marker_pattern = r"(?:試乘車?|中獎車?|領牌車|一般新車|代申請補助)"
    cleaned = re.sub(
        rf"[\(（\[【]\s*{marker_pattern}\s*[\)）\]】]",
        "",
        text,
        flags=re.IGNORECASE,
    )
    cleaned = re.sub(
        rf"\s*[-－—／/]\s*{marker_pattern}\s*$",
        "",
        cleaned,
        flags=re.IGNORECASE,
    )
    return cleaned.strip(" -－—／/")


def _join_unique_note_lines(*values):
    """合併歷史備註，避免匯入或重新驗證時重複加入相同標記。"""
    lines = []
    for value in values:
        for line in _text(value).splitlines():
            normalized = line.strip()
            if normalized and normalized not in lines:
                lines.append(normalized)
    return "\n".join(lines)


def _sales_order_note(raw, dealer_name, transaction_type, current_note=""):
    """把車行名稱中的交易註記移到訂單備註，不污染通路主檔名稱。"""
    note = _join_unique_note_lines(current_note, raw.get("備註"))
    if (
        transaction_type == SalesOrder.TransactionType.TEST_RIDE
        and "試乘" in _text(dealer_name).replace(" ", "")
    ):
        note = _join_unique_note_lines(note, "試乘車")
    if "代申請補助" in _text(dealer_name).replace(" ", ""):
        note = _join_unique_note_lines(note, "代申請補助")
    return note


def _sales_transaction_key(data):
    """建立可重跑的銷售交易鍵；同車後續中古車轉售應是另一筆交易。"""
    identifier = data.get("identifier") or ""
    vehicle_fallback = ":".join(
        filter(None, (data.get("model_number", ""), data.get("plate_number", "")))
    )
    vehicle_key = identifier or vehicle_fallback or "unknown-vehicle"
    category = data.get("vehicle_category") or SalesOrder.VehicleCategory.NEW
    transaction_date = (
        data.get("registration_date")
        or data.get("invoice_date")
        or data.get("order_date")
        or "no-date"
    )
    owner_identity = data.get("owner_id_number") or data.get("owner_name") or "unknown-owner"
    owner_digest = hashlib.sha256(str(owner_identity).strip().upper().encode("utf-8")).hexdigest()[:12]
    return f"sales:{vehicle_key}:{category}:{transaction_date}:{owner_digest}"


def _is_empty_sales_placeholder(data):
    """辨識 Excel 尾端只有公式 0、沒有交易內容的空白列。"""
    owner = _text(data.get("owner_name"))
    return owner in {"", "0"} and not any(
        _text(data.get(key))
        for key in (
            "model_number",
            "identifier",
            "plate_number",
            "registration_date",
            "invoice_date",
            "order_date",
            "owner_id_number",
            "owner_phone",
        )
    )


def _is_plausible_vehicle_identifier(value):
    """排除電話、Email 等 Excel 欄位錯位值，只接受合理的引擎／車身序號。"""
    normalized = normalize_vehicle_identifier(_text(value)) or ""
    return bool(
        6 <= len(normalized) <= 30
        and re.fullmatch(r"[A-Z0-9]+", normalized)
        and re.search(r"[A-Z]", normalized)
        and re.search(r"[0-9]", normalized)
    )


def _is_non_vehicle_sales_noise(data):
    """辨識誤落在銷貨欄位中的聯絡資料，不把它建立成車輛或歷史訂單。"""
    if _is_plausible_vehicle_identifier(
        data.get("identifier_raw") or data.get("identifier")
    ):
        return False
    return not any(
        _text(data.get(key))
        for key in (
            "owner_name",
            "plate_number",
            "registration_date",
            "invoice_date",
            "order_date",
            "owner_id_number",
            "owner_phone",
            "owner_email",
            "dealer_name",
        )
    ) and not any(
        _decimal(data.get(key))
        for key in (
            "historical_received_price",
            "cash_received",
            "card_received",
        )
    )


def _year_month(value):
    parsed = _date(value)
    if parsed:
        return parsed.strftime("%Y/%m")
    digits = "".join(character for character in _text(value) if character.isdigit())
    if len(digits) == 6 and 1 <= int(digits[4:]) <= 12:
        return f"{digits[:4]}/{digits[4:]}"
    return ""


def _fingerprint(payload):
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str).encode()
    return hashlib.sha256(encoded).hexdigest()


def _operations_sales_rows(batch, workbook):
    if "銷貨" not in workbook.sheetnames:
        return [], {"銷貨": "找不到銷貨工作表"}
    sheet = workbook["銷貨"]
    headers = next(sheet.iter_rows(min_row=3, max_row=3, values_only=True))
    rows = []
    for row_number, row_values in enumerate(sheet.iter_rows(min_row=4, values_only=True), 4):
        model_number = _text(_value(row_values, "C"))
        identifier_raw = _text(_value(row_values, "D"))
        owner = _text(_value(row_values, "AT")) or _text(_value(row_values, "E"))
        if not any((model_number, identifier_raw, owner)) or (
            not model_number and not identifier_raw and owner == "0"
        ):
            continue
        identifier = normalize_vehicle_identifier(identifier_raw) or ""
        raw = _row_dict_values(headers, row_values)
        dealer_name_raw = _text(_value(row_values, "AN"))
        vehicle_category, vehicle_category_reason = _infer_sales_vehicle_category(
            raw,
            dealer_name_raw,
        )
        sales_category = _text(_value(row_values, "CL"))
        transaction_type, transaction_type_reason = _infer_sales_transaction_type(
            raw,
            dealer_name_raw,
            vehicle_category,
            sales_category,
        )
        dealer_name = _clean_sales_source_name(dealer_name_raw)
        order_note = _sales_order_note(raw, dealer_name_raw, transaction_type)
        mapped = {
            "vehicle_category": vehicle_category,
            "vehicle_category_reason": vehicle_category_reason,
            "transaction_type": transaction_type,
            "transaction_type_reason": transaction_type_reason,
            "model_number": model_number,
            "identifier_raw": identifier_raw,
            "identifier": identifier,
            "owner_name": owner,
            "owner_name_primary": _text(_value(row_values, "E")),
            "owner_name_detail": _text(_value(row_values, "AT")),
            "color": _text(_value(row_values, "F")) or "未記錄",
            "registration_date": _json_value(_date(_value(row_values, "B")) or _date(_value(row_values, "AR"))),
            "order_date": _json_value(_date(_value(row_values, "CH"))),
            "plate_number": _text(_value(row_values, "AS")),
            "historical_received_price": str(_decimal(_value(row_values, "G"))),
            "cash_received": str(_decimal(_value(row_values, "J"))),
            "card_received": str(_decimal(_value(row_values, "K"))),
            "payment_confirmed": _text(_value(row_values, "AO")).upper() == "V",
            "dealer_name": dealer_name,
            "dealer_name_raw": dealer_name_raw,
            "installment_company": _text(_value(row_values, "AP")),
            "installment_periods": int(_decimal(_value(row_values, "AQ"))),
            "owner_birth_date": _json_value(_date(_value(row_values, "AU"))),
            "owner_id_number": _text(_value(row_values, "AW")),
            "owner_address": _text(_value(row_values, "AX")),
            "owner_phone": _text(_value(row_values, "AY")),
            "owner_email": _text(_value(row_values, "AZ")),
            "invoice_date": _json_value(_date(_value(row_values, "BB"))),
            "balance_invoice_number": _text(_value(row_values, "BC")),
            "subsidy_type": _text(_value(row_values, "BD")),
            "subsidy_amount": str(_decimal(_value(row_values, "BE"))),
            "remittance_account": _text(_value(row_values, "BF")),
            "bank_name": _text(_value(row_values, "BG")),
            "trade_in_plate": _text(_value(row_values, "BM")),
            "old_owner_name": _text(_value(row_values, "BK")),
            "old_owner_id_number": _text(_value(row_values, "BL")),
            "old_vehicle_engine_number": _text(_value(row_values, "BN")),
            "old_vehicle_brand": _text(_value(row_values, "BO")),
            "old_vehicle_manufactured_year_month": _year_month(_value(row_values, "BR")),
            "vehicle_control_account": _text(_value(row_values, "BU")),
            "battery_plan": _text(_value(row_values, "BW")),
            "battery_account": _text(_value(row_values, "BY")),
            "standard_gift": _text(_value(row_values, "CC")),
            "company_gift": _text(_value(row_values, "CI")),
            "sales_category": sales_category,
            "note": order_note,
        }
        natural_key = _sales_transaction_key(mapped)
        name_mismatch = bool(mapped["owner_name_primary"] and mapped["owner_name_detail"] and mapped["owner_name_primary"] != mapped["owner_name_detail"])
        messages = ["銷貨與車主資料區姓名不同，採車主資料區"] if name_mismatch else []
        rows.append(
            LegacyImportRow(
                batch=batch,
                sheet_name="銷貨",
                source_row=row_number,
                fingerprint=_fingerprint(raw),
                natural_key=natural_key,
                action=LegacyImportRow.Action.CREATE,
                raw_data=raw,
                mapped_data=mapped,
                messages=messages,
            )
        )
    return rows, {}


def _operations_inventory_rows(batch, workbook):
    if "進貨" not in workbook.sheetnames:
        return [], {"進貨": "找不到進貨工作表"}
    sheet = workbook["進貨"]
    headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    rows = []
    seen = set()
    duplicates = set()
    for row_number, row_values in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        model_number = _text(_value(row_values, "B"))
        identifier_raw = _text(_value(row_values, "C"))
        if not model_number and not identifier_raw:
            continue
        identifier = normalize_vehicle_identifier(identifier_raw) or ""
        if identifier and identifier in seen:
            duplicates.add(identifier)
        seen.add(identifier)
        raw = _row_dict_values(headers, row_values)
        mapped = {
            "received_on": _json_value(_date(_value(row_values, "A"))),
            "model_number": model_number,
            "identifier_raw": identifier_raw,
            "identifier": identifier,
            "color": _text(_value(row_values, "D")) or "未記錄",
            "quantity": int(_decimal(_value(row_values, "F"))),
            "manufactured_year_month": _year_month(_value(row_values, "J")),
        }
        existing = VehicleInventory.objects.filter(
            normalized_engine_number=identifier
        ).exists() or VehicleInventory.objects.filter(normalized_frame_number=identifier).exists()
        rows.append(
            LegacyImportRow(
                batch=batch, sheet_name="進貨", source_row=row_number,
                fingerprint=_fingerprint(raw), natural_key=identifier or f"inventory:{row_number}",
                action=LegacyImportRow.Action.SKIP if existing else LegacyImportRow.Action.CREATE,
                raw_data=raw, mapped_data=mapped, messages=[],
            )
        )
    for row in rows:
        if row.mapped_data["identifier"] in duplicates:
            row.action = LegacyImportRow.Action.CONFLICT
            row.messages.append(DUPLICATE_IDENTIFIER_MESSAGE)
        elif not row.mapped_data["identifier"]:
            row.action = LegacyImportRow.Action.ERROR
            row.messages.append(MISSING_IDENTIFIER_MESSAGE)
    return rows, {}


def _channel_rows(batch, workbook):
    rows = []
    errors = {}
    for sheet_name, source_type, header_row in (
        ("車行", SalesSource.SourceType.DEALER, 2),
        ("網路平台", SalesSource.SourceType.PLATFORM, 1),
    ):
        if sheet_name not in workbook.sheetnames:
            errors[sheet_name] = f"找不到{sheet_name}工作表"
            continue
        sheet = workbook[sheet_name]
        headers = next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
        for row_number, row_values in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
            name = _text(_value(row_values, "A"))
            if not name:
                continue
            raw = _row_dict_values(headers, row_values)
            if source_type == SalesSource.SourceType.DEALER:
                mapped = {
                    "name": name, "source_type": source_type,
                    "contact_name": _text(_value(row_values, "B")),
                    "phone": _text(_value(row_values, "C")),
                    "phone_2": _text(_value(row_values, "D")),
                    "mobile": _text(_value(row_values, "E")),
                    "fax": _text(_value(row_values, "F")),
                    "address": _text(_value(row_values, "G")),
                    "brands": [brand for brand, column in (("SYM", "H"), ("SUZUKI", "I")) if _text(_value(row_values, column))],
                    "vehicle_capacity": int(_decimal(_value(row_values, "J"))) or None,
                    "note": _text(_value(row_values, "K")),
                }
            else:
                mapped = {
                    "name": name, "source_type": source_type,
                    "contact_name": _text(_value(row_values, "B")),
                    "phone": _text(_value(row_values, "C")),
                    "extension": _text(_value(row_values, "D")),
                    "mobile": _text(_value(row_values, "E")),
                    "email": _text(_value(row_values, "F")),
                }
            existing = SalesSource.objects.filter(source_type=source_type, name=name).exists()
            rows.append(LegacyImportRow(
                batch=batch, sheet_name=sheet_name, source_row=row_number,
                fingerprint=_fingerprint(raw), natural_key=f"{source_type}:{name}",
                action=LegacyImportRow.Action.UPDATE if existing else LegacyImportRow.Action.CREATE,
                raw_data=raw, mapped_data=mapped, messages=[],
            ))
    return rows, errors


@transaction.atomic
def build_import_preview(batch):
    batch.rows.all().delete()
    batch.source_file.open("rb")
    workbook = load_workbook(batch.source_file, read_only=True, data_only=True)
    try:
        if batch.import_type == LegacyImportBatch.ImportType.CHANNELS:
            rows, errors = _channel_rows(batch, workbook)
        else:
            inventory_rows, inventory_errors = _operations_inventory_rows(batch, workbook)
            sales_rows, sales_errors = _operations_sales_rows(batch, workbook)
            rows = inventory_rows + sales_rows
            errors = {**inventory_errors, **sales_errors}
        LegacyImportRow.objects.bulk_create(rows, batch_size=500)
        batch.source_sheets = list(workbook.sheetnames)
        batch.preview_summary = {"errors": errors, "sheets": list(workbook.sheetnames)}
        batch.status = LegacyImportBatch.Status.PREVIEW
        batch.save(update_fields=["source_sheets", "preview_summary", "status", "updated_at"])
        return revalidate_import_batch(batch)
    finally:
        workbook.close()
        batch.source_file.close()


def _base_messages(row):
    return [message for message in row.messages if message not in SYSTEM_VALIDATION_MESSAGES]


def _json_clean_value(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


@transaction.atomic
def revalidate_import_batch(batch):
    """依目前暫存修正重新判斷動作與摘要，不變更原始 Excel 資料。"""
    if batch.status != LegacyImportBatch.Status.PREVIEW:
        raise ValueError("只有待確認批次可以重新驗證。")
    rows = list(batch.rows.order_by("sheet_name", "source_row"))
    active_rows = [row for row in rows if not row.excluded]
    # 舊版預覽批次沒有車輛類別與交易類型；重新驗證時補上推論結果。
    # 已經人工指定的內容一律保留。
    for row in active_rows:
        if row.sheet_name != "銷貨":
            continue
        mapped_data = dict(row.mapped_data)
        if not mapped_data.get("vehicle_category"):
            category, reason = _infer_sales_vehicle_category(
                row.raw_data,
                mapped_data.get("dealer_name_raw") or mapped_data.get("dealer_name", ""),
            )
            mapped_data["vehicle_category"] = category
            mapped_data["vehicle_category_reason"] = reason
        if not row.manually_corrected:
            mapped_data["dealer_name_raw"] = (
                mapped_data.get("dealer_name_raw")
                or mapped_data.get("dealer_name", "")
            )
            mapped_data["dealer_name"] = _clean_sales_source_name(
                mapped_data["dealer_name_raw"]
            )
            transaction_type, reason = _infer_sales_transaction_type(
                row.raw_data,
                mapped_data.get("dealer_name_raw", ""),
                mapped_data["vehicle_category"],
                mapped_data.get("sales_category", ""),
            )
            mapped_data["transaction_type"] = transaction_type
            mapped_data["transaction_type_reason"] = reason
        elif not mapped_data.get("transaction_type"):
            transaction_type, reason = _infer_sales_transaction_type(
                row.raw_data,
                mapped_data.get("dealer_name_raw", ""),
                mapped_data["vehicle_category"],
                mapped_data.get("sales_category", ""),
            )
            mapped_data["transaction_type"] = transaction_type
            mapped_data["transaction_type_reason"] = reason
        mapped_data["note"] = _sales_order_note(
            row.raw_data,
            mapped_data.get("dealer_name_raw", ""),
            mapped_data.get("transaction_type"),
            mapped_data.get("note", ""),
        )
        row.mapped_data = mapped_data
    inventory_identifier_counts = {}
    sales_new_identifier_counts = {}
    sales_transaction_counts = {}
    for row in active_rows:
        identifier = row.mapped_data.get("identifier", "")
        if row.sheet_name == "進貨" and identifier:
            inventory_identifier_counts[identifier] = (
                inventory_identifier_counts.get(identifier, 0) + 1
            )
        elif row.sheet_name == "銷貨":
            if _is_empty_sales_placeholder(
                row.mapped_data
            ) or _is_non_vehicle_sales_noise(row.mapped_data):
                continue
            transaction_key = _sales_transaction_key(row.mapped_data)
            sales_transaction_counts[transaction_key] = (
                sales_transaction_counts.get(transaction_key, 0) + 1
            )
            if (
                identifier
                and row.mapped_data.get("vehicle_category", SalesOrder.VehicleCategory.NEW)
                == SalesOrder.VehicleCategory.NEW
            ):
                sales_new_identifier_counts[identifier] = (
                    sales_new_identifier_counts.get(identifier, 0) + 1
                )

    inventory_identifiers = set(
        VehicleInventory.objects.exclude(normalized_engine_number__isnull=True)
        .values_list("normalized_engine_number", flat=True)
    ) | set(
        VehicleInventory.objects.exclude(normalized_frame_number__isnull=True)
        .values_list("normalized_frame_number", flat=True)
    )
    inventory_identifiers.discard("")
    completed_sales_keys = {
        _sales_transaction_key(mapped_data)
        for mapped_data in LegacyImportRow.objects.filter(
            sheet_name="銷貨",
            batch__status=LegacyImportBatch.Status.COMPLETED,
            action__in=[
                LegacyImportRow.Action.CREATE,
                LegacyImportRow.Action.UPDATE,
                LegacyImportRow.Action.SKIP,
            ],
        ).values_list("mapped_data", flat=True)
    }
    existing_sources = set(
        SalesSource.objects.values_list("source_type", "name")
    )
    now = timezone.now()
    for row in rows:
        messages = _base_messages(row)
        if row.excluded:
            row.action = LegacyImportRow.Action.EXCLUDE
        elif row.sheet_name == "進貨":
            identifier = row.mapped_data.get("identifier", "")
            row.natural_key = identifier or f"inventory:{row.source_row}"
            if identifier and inventory_identifier_counts.get(identifier, 0) > 1:
                row.action = LegacyImportRow.Action.CONFLICT
                messages.append(DUPLICATE_IDENTIFIER_MESSAGE)
            elif not identifier:
                row.action = LegacyImportRow.Action.ERROR
                messages.append(MISSING_IDENTIFIER_MESSAGE)
            elif identifier in inventory_identifiers:
                row.action = LegacyImportRow.Action.SKIP
            else:
                row.action = LegacyImportRow.Action.CREATE
        elif row.sheet_name == "銷貨":
            identifier = row.mapped_data.get("identifier", "")
            category = row.mapped_data.get(
                "vehicle_category",
                SalesOrder.VehicleCategory.NEW,
            )
            is_placeholder = _is_empty_sales_placeholder(row.mapped_data)
            is_non_vehicle_noise = _is_non_vehicle_sales_noise(row.mapped_data)
            if is_placeholder or is_non_vehicle_noise:
                row.natural_key = f"sales-placeholder:{row.source_row}"
                row.action = LegacyImportRow.Action.SKIP
                messages.append(
                    EMPTY_SALES_PLACEHOLDER_MESSAGE
                    if is_placeholder
                    else NON_VEHICLE_SALES_NOISE_MESSAGE
                )
            else:
                row.natural_key = _sales_transaction_key(row.mapped_data)
                if sales_transaction_counts.get(row.natural_key, 0) > 1:
                    row.action = LegacyImportRow.Action.CONFLICT
                    messages.append(DUPLICATE_SALES_TRANSACTION_MESSAGE)
                elif (
                    category == SalesOrder.VehicleCategory.NEW
                    and identifier
                    and sales_new_identifier_counts.get(identifier, 0) > 1
                ):
                    row.action = LegacyImportRow.Action.CONFLICT
                    messages.append(MULTIPLE_NEW_SALES_MESSAGE)
                elif row.natural_key in completed_sales_keys:
                    row.action = LegacyImportRow.Action.SKIP
                elif _has_invalid_email(row.mapped_data.get("owner_email", "")):
                    row.action = LegacyImportRow.Action.ERROR
                    messages.append(INVALID_EMAIL_MESSAGE)
                else:
                    row.action = LegacyImportRow.Action.CREATE
        else:
            source_type = row.mapped_data.get("source_type", "")
            name = row.mapped_data.get("name", "")
            row.natural_key = f"{source_type}:{name}"
            row.action = (
                LegacyImportRow.Action.UPDATE
                if (source_type, name) in existing_sources
                else LegacyImportRow.Action.CREATE
            )
        row.messages = messages
        row.updated_at = now
    if rows:
        LegacyImportRow.objects.bulk_update(
            rows,
            ["mapped_data", "action", "messages", "natural_key", "updated_at"],
            batch_size=500,
        )

    counts = {choice: 0 for choice, _ in LegacyImportRow.Action.choices}
    for row in rows:
        counts[row.action] += 1
    validation = {}
    if batch.import_type == LegacyImportBatch.ImportType.OPERATIONS:
        sales_rows = [row for row in active_rows if row.sheet_name == "銷貨"]
        inventory_rows = [row for row in active_rows if row.sheet_name == "進貨"]
        known_models = {
            normalize_vehicle_model_master_value(value)
            for pair in VehicleModel.objects.values_list("model_number", "name")
            for value in pair
            if value
        }
        known_models.update(
            LegacyImportMasterMapping.objects.filter(
                mapping_type=LegacyImportMasterMapping.MappingType.VEHICLE_MODEL
            ).values_list("normalized_source_value", flat=True)
        )
        known_sources = {
            normalize_legacy_master_value(value)
            for value in SalesSource.objects.values_list("name", flat=True)
        }
        known_sources.update(
            LegacyImportMasterMapping.objects.filter(
                mapping_type=LegacyImportMasterMapping.MappingType.SALES_SOURCE
            ).values_list("normalized_source_value", flat=True)
        )
        source_available = {
            row.mapped_data.get("identifier", "")
            for row in inventory_rows
            if row.mapped_data.get("quantity") == 1 and row.mapped_data.get("identifier")
        }
        system_available = set(
            VehicleInventory.objects.filter(status=VehicleInventory.Status.AVAILABLE)
            .values_list("normalized_engine_number", flat=True)
        ) | set(
            VehicleInventory.objects.filter(status=VehicleInventory.Status.AVAILABLE)
            .values_list("normalized_frame_number", flat=True)
        )
        system_available.discard(None)
        system_available.discard("")
        validation = {
            "owner_name_differences": sum(
                bool(row.mapped_data.get("owner_name_primary") and row.mapped_data.get("owner_name_detail") and row.mapped_data.get("owner_name_primary") != row.mapped_data.get("owner_name_detail"))
                for row in sales_rows
            ),
            "unmapped_models": sorted({
                row.mapped_data.get("model_number", "")
                for row in active_rows
                if row.mapped_data.get("model_number")
                and not _is_empty_sales_placeholder(row.mapped_data)
                and not _is_non_vehicle_sales_noise(row.mapped_data)
                and normalize_vehicle_model_master_value(
                    row.mapped_data["model_number"]
                )
                not in known_models
            }),
            "unmapped_sources": sorted({
                row.mapped_data.get("dealer_name", "")
                for row in sales_rows
                if row.mapped_data.get("dealer_name")
                and normalize_legacy_master_value(row.mapped_data["dealer_name"])
                not in known_sources
            }),
            "used_vehicle_sales": sum(
                row.mapped_data.get("vehicle_category") == SalesOrder.VehicleCategory.USED
                for row in sales_rows
            ),
            "combined_fee_income_rows": sum(
                bool(_decimal(row.raw_data.get("刷卡、分期手續費收入")))
                for row in sales_rows
            ),
            "source_available_count": len(source_available),
            "system_available_count": len(system_available),
            "source_only_inventory": sorted(source_available - system_available)[:100],
            "system_only_inventory": sorted(system_available - source_available)[:100],
        }
    else:
        current_keys = {
            f"{row.mapped_data.get('source_type', '')}:{row.mapped_data.get('name', '')}"
            for row in active_rows
        }
        prior_keys = set(
            LegacyImportRow.objects.filter(
                batch__import_type=LegacyImportBatch.ImportType.CHANNELS,
                batch__status=LegacyImportBatch.Status.COMPLETED,
                sheet_name__in=["車行", "網路平台"],
                action__in=[
                    LegacyImportRow.Action.CREATE,
                    LegacyImportRow.Action.UPDATE,
                    LegacyImportRow.Action.SKIP,
                ],
            ).values_list("natural_key", flat=True)
        )
        validation = {"removed_source_keys": sorted(prior_keys - current_keys)}
    previous = batch.preview_summary or {}
    summary = {
        "parser_schema_version": PREVIEW_SCHEMA_VERSION,
        "source_rows": len(rows),
        "counts": counts,
        "errors": previous.get("errors", {}),
        "sheets": batch.source_sheets,
        "validation": validation,
    }
    batch.preview_summary = summary
    batch.save(update_fields=["preview_summary", "updated_at"])
    return summary


@transaction.atomic
def apply_import_row_decision(row, mapping, decision, reason, actor_name):
    if row.batch.status != LegacyImportBatch.Status.PREVIEW:
        raise ValueError("只有待確認批次可以修正。")
    before = dict(row.mapped_data)
    was_excluded = row.excluded
    if decision == LegacyImportCorrection.Decision.EXCLUDE:
        row.excluded = True
        correction_decision = LegacyImportCorrection.Decision.EXCLUDE
    else:
        updated = dict(row.mapped_data)
        updated.update({key: _json_clean_value(value) for key, value in mapping.items()})
        if "identifier_raw" in mapping:
            updated["identifier"] = normalize_vehicle_identifier(mapping.get("identifier_raw")) or ""
        if (
            row.sheet_name == "銷貨"
            and mapping.get("vehicle_category")
            and mapping.get("vehicle_category") != before.get("vehicle_category")
        ):
            updated["vehicle_category_reason"] = "人工調整"
        if (
            row.sheet_name == "銷貨"
            and mapping.get("transaction_type")
            and mapping.get("transaction_type") != before.get("transaction_type")
        ):
            updated["transaction_type_reason"] = "人工調整"
        row.mapped_data = updated
        row.excluded = False
        correction_decision = (
            LegacyImportCorrection.Decision.RESTORE
            if was_excluded
            else LegacyImportCorrection.Decision.CORRECT
        )
    row.manually_corrected = True
    row.corrected_by = actor_name
    row.corrected_at = timezone.now()
    row.save(
        update_fields=[
            "mapped_data",
            "excluded",
            "manually_corrected",
            "corrected_by",
            "corrected_at",
            "updated_at",
        ]
    )
    LegacyImportCorrection.objects.create(
        row=row,
        decision=correction_decision,
        before_data=before,
        after_data=dict(row.mapped_data),
        reason=reason,
        corrected_by=actor_name,
    )
    return revalidate_import_batch(row.batch)


def _placeholder_model(model_number):
    from sales.services.vehicle_brands import canonical_vehicle_brand_name

    brand = canonical_vehicle_brand_name("歷史資料", create_missing=True)
    model, _ = VehicleModel.objects.get_or_create(
        brand=brand, name=model_number or "未辨識車型", model_number=model_number or "UNKNOWN",
        model_year=None, model_code="",
        defaults={"energy_type": VehicleModel.EnergyType.GAS, "displacement_cc": 125, "active": False},
    )
    return model


def _model_for_number(model_number):
    mapping = _legacy_master_mapping(
        LegacyImportMasterMapping.MappingType.VEHICLE_MODEL,
        model_number,
    )
    if mapping and mapping.vehicle_model_id:
        return mapping.vehicle_model
    return (
        VehicleModel.objects.filter(model_number__iexact=model_number).first()
        or VehicleModel.objects.filter(
            factory_model_codes__normalized_code=normalize_vehicle_model_master_value(
                model_number
            )
        ).first()
        or VehicleModel.objects.filter(name__iexact=model_number).first()
        or _placeholder_model(model_number)
    )


def _source_for_name(source_name):
    mapping = _legacy_master_mapping(
        LegacyImportMasterMapping.MappingType.SALES_SOURCE,
        source_name,
    )
    if mapping:
        return mapping.sales_source if not mapping.ignored else None
    return SalesSource.objects.filter(name__iexact=source_name).first()


def _color_for_model(model, name):
    color, _ = VehicleColor.objects.get_or_create(vehicle_model=model, name=name or "未記錄", defaults={"active": False})
    return color


def _default_store():
    return Store.objects.order_by("id").first() or Store.objects.create(name="總店", code="MAIN")


def _default_source_category(source_type):
    default_names = {
        SalesSource.SourceType.STORE: "其他來源",
        SalesSource.SourceType.DEALER: "合作車行",
        SalesSource.SourceType.PLATFORM: "網路平台",
    }
    category, _ = SalesSourceCategory.objects.get_or_create(
        name=default_names[source_type],
        defaults={"system_behavior": source_type, "active": True},
    )
    return category


def _merge_note_lines(base_note, lines):
    """Keep imported contact details readable without restoring a contact sub-table."""
    paragraphs = [part.strip() for part in (base_note or "").split("\n") if part.strip()]
    for line in lines:
        cleaned = (line or "").strip()
        if cleaned and cleaned not in paragraphs:
            paragraphs.append(cleaned)
    return "\n".join(paragraphs)


def _channel_contact_note(data):
    details = []
    if data.get("phone_2"):
        details.append(f"電話二：{data['phone_2']}")
    if data.get("extension"):
        details.append(f"分機：{data['extension']}")
    if data.get("mobile"):
        details.append(f"手機：{data['mobile']}")
    if data.get("email"):
        details.append(f"Email：{data['email']}")
    contact_name = (data.get("contact_name") or "").strip()
    if not contact_name and not details:
        return ""
    label = f"歷史聯絡資料：{contact_name}" if contact_name else "歷史其他聯絡資料"
    return f"{label}｜{'／'.join(details)}" if details else label


def _commit_channel_row(row):
    data = row.mapped_data
    source, _ = SalesSource.objects.get_or_create(
        source_type=data["source_type"],
        name=data["name"],
    )
    source.category = _default_source_category(data["source_type"])
    source.responsible_person = data.get("contact_name", "")
    source.phone = data.get("phone", "")
    source.phone_secondary = data.get("phone_2", "")
    source.mobile = data.get("mobile", "")
    source.fax = data.get("fax", "")
    other_contacts = []
    if data.get("fax"):
        other_contacts.append(f"傳真：{data['fax']}")
    if data.get("extension"):
        other_contacts.append(f"分機：{data['extension']}")
    if data.get("email"):
        other_contacts.append(f"Email：{data['email']}")
    source.other_contact = "／".join(other_contacts)
    source.address = data.get("address", "")
    source.vehicle_capacity = data.get("vehicle_capacity")
    source.note = _merge_note_lines(
        source.note,
        [data.get("note", ""), _channel_contact_note(data)],
    )
    source.save()
    if source.source_type == SalesSource.SourceType.DEALER:
        imported_scopes = set()
        if "SYM" in data.get("brands", []):
            imported_scopes.add(SalesSourceBrandPolicy.CooperationScope.SYM)
        if "SUZUKI" in data.get("brands", []):
            imported_scopes.update(
                {
                    SalesSourceBrandPolicy.CooperationScope.SUZUKI_GAS,
                    SalesSourceBrandPolicy.CooperationScope.SUZUKI_ELECTRIC,
                }
            )
        today = timezone.localdate()
        for scope, _label in SalesSourceBrandPolicy.CooperationScope.choices:
            cooperates = scope in imported_scopes
            SalesSourceCooperationProfile.objects.update_or_create(
                source=source,
                cooperation_scope=scope,
                defaults={
                    "cooperates": cooperates,
                    "relationship_type": SalesSourceCooperationProfile.RelationshipType.GENERAL,
                    "vehicle_capacity": data.get("vehicle_capacity") if cooperates else None,
                },
            )
            SalesSourceBrandPolicy.objects.update_or_create(
                source=source,
                cooperation_scope=scope,
                effective_from=today,
                defaults={
                    "cooperates": cooperates,
                    "commission_adjustment": 0,
                    "effective_to": None,
                    "note": "",
                },
            )
    row.committed_model = "SalesSource"
    row.committed_pk = str(source.pk)


def _commit_inventory_row(row):
    data = row.mapped_data
    identifier = data["identifier"]
    existing = VehicleInventory.objects.filter(normalized_engine_number=identifier).first() or VehicleInventory.objects.filter(normalized_frame_number=identifier).first()
    if existing:
        row.action = LegacyImportRow.Action.SKIP
        row.committed_model, row.committed_pk = "VehicleInventory", str(existing.pk)
        return
    model = _model_for_number(data["model_number"])
    color = _color_for_model(model, data["color"])
    store = _default_store()
    status = VehicleInventory.Status.AVAILABLE if data["quantity"] == 1 else VehicleInventory.Status.INACTIVE
    vehicle = VehicleInventory(
        vehicle_model=model, color=color, ownership_store=store, location_store=store,
        received_on=_date(data["received_on"]) or timezone.localdate(), manufactured_year_month=data["manufactured_year_month"],
        status=status,
    )
    if model.energy_type == VehicleModel.EnergyType.GAS:
        vehicle.engine_number = data["identifier_raw"]
    else:
        vehicle.frame_number = data["identifier_raw"]
    vehicle.save()
    row.committed_model, row.committed_pk = "VehicleInventory", str(vehicle.pk)


def _commit_sales_row(row, actor_name):
    data = row.mapped_data
    model = _model_for_number(data["model_number"])
    color = _color_for_model(model, data["color"])
    vehicle_category = data.get("vehicle_category") or SalesOrder.VehicleCategory.NEW
    vehicle = None
    if vehicle_category == SalesOrder.VehicleCategory.NEW and data["identifier"]:
        vehicle = VehicleInventory.objects.filter(normalized_engine_number=data["identifier"]).first() or VehicleInventory.objects.filter(normalized_frame_number=data["identifier"]).first()
    order_date = _date(data["order_date"]) or _date(data["registration_date"]) or timezone.localdate()
    owner_id = data["owner_id_number"] or f"HIST-{str(row.batch_id)[:8]}-{row.source_row}"
    source = _source_for_name(data.get("dealer_name", ""))
    source_type = SalesOrder.SourceType.STORE
    if source:
        source_type = source.source_type
    order = SalesOrder.objects.create(
        order_date=order_date, status=SalesOrder.Status.ALLOCATION_PENDING,
        source=source, source_type=source_type,
        owner_name=data["owner_name"] or "歷史資料未填", owner_phone=data["owner_phone"] or "未提供",
        owner_email=data["owner_email"], owner_birth_date=_date(data["owner_birth_date"]),
        owner_address=data["owner_address"] or "未提供", owner_id_number=owner_id,
        vehicle_model=model, color=color, vehicle_price=0, vehicle_category=vehicle_category,
        transaction_type=data.get("transaction_type") or SalesOrder.TransactionType.REGULAR_NEW,
        registration_date=_date(data["registration_date"]), final_plate_number=data["plate_number"],
        payment_type=SalesOrder.PaymentType.INSTALLMENT if data["installment_periods"] else (SalesOrder.PaymentType.CARD if _decimal(data["card_received"]) else SalesOrder.PaymentType.CASH),
        installment_company=data["installment_company"], installment_periods=data["installment_periods"],
        trade_in_plate=data["trade_in_plate"], old_owner_name=data["old_owner_name"], old_owner_id_number=data["old_owner_id_number"],
        subsidy_type=data["subsidy_type"],
        note=data.get("note", ""),
        is_trade_in_subsidy=(
            vehicle_category == SalesOrder.VehicleCategory.NEW
            and bool(data["subsidy_type"] or data["trade_in_plate"])
        ),
        allocated_vehicle=vehicle,
    )
    delivered_at = timezone.make_aware(datetime.combine(_date(data["registration_date"]) or order_date, time(hour=12)))
    SalesOrder.objects.filter(pk=order.pk).update(
        status=SalesOrder.Status.COMPLETED, delivered_at=delivered_at, delivered_by="歷史資料匯入",
        registration_completed_at=delivered_at if data["registration_date"] else None,
        registration_completed_by="歷史資料匯入" if data["registration_date"] else "",
    )
    if vehicle:
        VehicleInventory.objects.filter(pk=vehicle.pk).update(status=VehicleInventory.Status.SOLD)
    profile, _ = OrderOperationsProfile.objects.get_or_create(order=order)
    profile.dealer_name = data.get("dealer_name", "")
    profile.vehicle_cost = _decimal(row.raw_data.get("成本"))
    profile.registration_tax_expense = _decimal(row.raw_data.get("領牌稅金支出"))
    profile.compulsory_insurance_expense = _decimal(row.raw_data.get("強制險支出"))
    profile.plate_selection_expense = _decimal(row.raw_data.get("選號支出"))
    profile.dealer_commission_expense = _decimal(row.raw_data.get("車行傭金支出"))
    profile.registration_tax_income = _decimal(row.raw_data.get("領牌稅金收入"))
    profile.compulsory_insurance_income = _decimal(row.raw_data.get("強制險收入"))
    profile.agency_fee_income = _decimal(row.raw_data.get("代辦費收入"))
    profile.plate_selection_income = _decimal(row.raw_data.get("選號收入"))
    profile.sales_bonus = _decimal(row.raw_data.get("實銷獎勵金"))
    profile.promotion_subsidy = _decimal(row.raw_data.get("促銷補助金"))
    profile.installment_interest_subsidy = _decimal(row.raw_data.get("分期補貼息"))
    profile.insurance_commission = _decimal(row.raw_data.get("強制險傭金"))
    profile.credit_card_commission = _decimal(row.raw_data.get("信用卡傭金"))
    profile.payment_confirmed = data["payment_confirmed"]
    profile.invoice_date = _date(data["invoice_date"])
    profile.balance_invoice_number = data["balance_invoice_number"]
    profile.subsidy_amount = _decimal(data["subsidy_amount"])
    profile.bank_name = data["bank_name"]
    profile.remittance_account = data["remittance_account"]
    profile.old_vehicle_engine_number = data["old_vehicle_engine_number"]
    profile.old_vehicle_brand = data["old_vehicle_brand"]
    profile.vehicle_control_account = data["vehicle_control_account"]
    profile.battery_plan = data["battery_plan"]
    profile.battery_account = data["battery_account"]
    profile.helmet = data["standard_gift"]
    profile.company_gift_or_remittance = data["company_gift"]
    profile.updated_by = actor_name
    profile.save()
    for key, amount, method in (("legacy_cash", data["cash_received"], "cash"), ("legacy_card", data["card_received"], "card")):
        if _decimal(amount):
            PaymentRecord.objects.create(order=order, system_key=key, item_name="歷史收款", expected_amount=_decimal(amount), received_amount=_decimal(amount), received_on=order_date, payment_method=method, confirmed=data["payment_confirmed"])
    LegacySalesSnapshot.objects.create(
        order=order, import_row=row,
        historical_received_price=_decimal(data["historical_received_price"]),
        cash_received=_decimal(data["cash_received"]), card_received=_decimal(data["card_received"]),
        vehicle_identifier=data.get("identifier_raw") or data.get("identifier", ""),
        sales_category=data["sales_category"], raw_financials=row.raw_data,
    )
    from sales.services.order_search import rebuild_order_search_index
    rebuild_order_search_index(order.pk)
    row.committed_model, row.committed_pk = "SalesOrder", str(order.pk)


def _save_import_progress(batch, completed, total, result):
    LegacyImportBatch.objects.filter(pk=batch.pk).update(
        processing_completed=completed,
        processing_total=total,
        processing_heartbeat_at=timezone.now(),
        result_summary=result,
        updated_at=timezone.now(),
    )


def confirm_import(batch, actor_name):
    if batch.status not in {
        LegacyImportBatch.Status.PREVIEW,
        LegacyImportBatch.Status.PROCESSING,
        LegacyImportBatch.Status.FAILED,
    }:
        raise ValueError("此批次已確認或已失敗，不能重複匯入。")
    unresolved = batch.rows.filter(
        action__in=[LegacyImportRow.Action.CONFLICT, LegacyImportRow.Action.ERROR]
    ).count()
    if unresolved:
        raise ValueError(f"尚有 {unresolved} 筆衝突或錯誤資料，請先修正或排除後再匯入。")
    rows = list(batch.rows.order_by("sheet_name", "source_row"))
    # 進貨必須先建立，銷貨才能連結實體車輛。
    rows.sort(key=lambda row: {"進貨": 0, "車行": 1, "網路平台": 1, "銷貨": 2}.get(row.sheet_name, 9))
    total = len(rows)
    result = {
        "created": 0,
        "updated": 0,
        "skipped": 0,
        "excluded": 0,
        "conflicts": 0,
        "errors": 0,
    }
    now = timezone.now()
    if batch.status != LegacyImportBatch.Status.PROCESSING:
        batch.status = LegacyImportBatch.Status.PROCESSING
        batch.processing_started_at = batch.processing_started_at or now
        batch.processing_finished_at = None
        batch.processing_error = ""
        batch.processing_total = total
        batch.processing_completed = 0
        batch.save(
            update_fields=[
                "status",
                "processing_started_at",
                "processing_finished_at",
                "processing_error",
                "processing_total",
                "processing_completed",
                "updated_at",
            ]
        )
    for completed, row in enumerate(rows, 1):
        if row.committed_model and row.committed_pk:
            result["updated" if row.action == LegacyImportRow.Action.UPDATE else "created"] += 1
            if completed % 25 == 0 or completed == total:
                _save_import_progress(batch, completed, total, result)
            continue
        if row.action == LegacyImportRow.Action.CONFLICT:
            result["conflicts"] += 1
        elif row.action == LegacyImportRow.Action.ERROR:
            result["errors"] += 1
        elif row.action == LegacyImportRow.Action.SKIP:
            result["skipped"] += 1
        elif row.action == LegacyImportRow.Action.EXCLUDE:
            result["excluded"] += 1
        else:
            try:
                with transaction.atomic():
                    if row.sheet_name in {"車行", "網路平台"}:
                        _commit_channel_row(row)
                    elif row.sheet_name == "進貨":
                        _commit_inventory_row(row)
                    elif row.sheet_name == "銷貨":
                        _commit_sales_row(row, actor_name)
                    row.save(update_fields=["action", "committed_model", "committed_pk", "updated_at"])
                result["updated" if row.action == LegacyImportRow.Action.UPDATE else "created"] += 1
            except Exception as exc:
                row.action = LegacyImportRow.Action.ERROR
                row.messages = [*row.messages, _friendly_import_exception(exc)]
                row.save(update_fields=["action", "messages", "updated_at"])
                result["errors"] += 1
        if completed % 25 == 0 or completed == total:
            _save_import_progress(batch, completed, total, result)
    batch.status = LegacyImportBatch.Status.COMPLETED
    batch.result_summary = result
    batch.confirmed_by = actor_name
    batch.confirmed_at = timezone.now()
    batch.processing_completed = total
    batch.processing_total = total
    batch.processing_finished_at = timezone.now()
    batch.processing_heartbeat_at = timezone.now()
    batch.processing_error = ""
    batch.save(
        update_fields=[
            "status",
            "result_summary",
            "confirmed_by",
            "confirmed_at",
            "processing_completed",
            "processing_total",
            "processing_finished_at",
            "processing_heartbeat_at",
            "processing_error",
            "updated_at",
        ]
    )
    return result


@transaction.atomic
def retry_completed_import_row(row, mapping, decision, reason, actor_name):
    row = LegacyImportRow.objects.select_for_update().select_related("batch").get(pk=row.pk)
    batch = LegacyImportBatch.objects.select_for_update().get(pk=row.batch_id)
    if batch.status != LegacyImportBatch.Status.COMPLETED:
        raise ValueError("只有已完成批次中的失敗資料可以補匯。")
    if row.action != LegacyImportRow.Action.ERROR or row.committed_model:
        raise ValueError("這筆資料已經處理完成，不需要再次補匯。")

    before = dict(row.mapped_data)
    updated = dict(row.mapped_data)
    updated.update({key: _json_clean_value(value) for key, value in mapping.items()})
    if "identifier_raw" in mapping:
        updated["identifier"] = normalize_vehicle_identifier(mapping.get("identifier_raw")) or ""
    row.mapped_data = updated
    row.excluded = decision == LegacyImportCorrection.Decision.EXCLUDE
    row.manually_corrected = True
    row.corrected_by = actor_name
    row.corrected_at = timezone.now()

    succeeded = False
    error_text = ""
    if row.excluded:
        row.action = LegacyImportRow.Action.EXCLUDE
        row.messages = []
        succeeded = True
    else:
        row.messages = []
        try:
            with transaction.atomic():
                if row.sheet_name in {"車行", "網路平台"}:
                    _commit_channel_row(row)
                elif row.sheet_name == "進貨":
                    _commit_inventory_row(row)
                elif row.sheet_name == "銷貨":
                    _commit_sales_row(row, actor_name)
                else:
                    raise ValueError("不支援的匯入工作表。")
        except Exception as exc:
            error_text = _friendly_import_exception(exc)
            row.action = LegacyImportRow.Action.ERROR
            row.messages = [error_text]
        else:
            row.action = LegacyImportRow.Action.CREATE
            succeeded = True

    row.save(
        update_fields=[
            "mapped_data",
            "excluded",
            "manually_corrected",
            "corrected_by",
            "corrected_at",
            "action",
            "messages",
            "committed_model",
            "committed_pk",
            "updated_at",
        ]
    )
    LegacyImportCorrection.objects.create(
        row=row,
        decision=(
            LegacyImportCorrection.Decision.EXCLUDE
            if row.excluded
            else LegacyImportCorrection.Decision.CORRECT
        ),
        before_data=before,
        after_data=row.mapped_data,
        reason=reason,
        corrected_by=actor_name,
    )
    if succeeded:
        result = dict(batch.result_summary or {})
        result["errors"] = max(int(result.get("errors", 0)) - 1, 0)
        result["excluded" if row.excluded else "created"] = int(
            result.get("excluded" if row.excluded else "created", 0)
        ) + 1
        batch.result_summary = result
        batch.save(update_fields=["result_summary", "updated_at"])
    return {"ok": succeeded, "error": error_text, "excluded": row.excluded}

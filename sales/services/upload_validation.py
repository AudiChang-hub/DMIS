"""Server-side validation for user supplied documents.

Browser ``accept`` attributes are only a convenience.  Every upload that is
persisted by the sales application must also pass extension, declared MIME and
file-signature/parser checks here.  The stream position is restored before the
form/model saves the file.
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
import warnings
from zipfile import BadZipFile, ZipFile

from django import forms
from django.core.files.uploadedfile import UploadedFile
from openpyxl import load_workbook
from PIL import Image
from pypdf import PdfReader


MIB = 1024 * 1024
DOCUMENT_MAX_BYTES = 20 * MIB
TEMPLATE_MAX_BYTES = 20 * MIB
SUBSIDY_MAX_BYTES = 20 * MIB
IMAGE_MAX_BYTES = 12 * MIB
ZIP_MAX_UNCOMPRESSED_BYTES = 100 * MIB
ZIP_MAX_ENTRIES = 10_000

IMAGE_MIMES = {
    ".jpg": {"image/jpeg", "image/jpg"},
    ".jpeg": {"image/jpeg", "image/jpg"},
    ".png": {"image/png"},
    ".webp": {"image/webp"},
}
IMAGE_FORMATS = {
    ".jpg": "JPEG",
    ".jpeg": "JPEG",
    ".png": "PNG",
    ".webp": "WEBP",
}
PDF_MIMES = {"application/pdf", "application/x-pdf"}
GENERIC_MIMES = {"", "application/octet-stream"}

SUBSIDY_MIMES = {
    **IMAGE_MIMES,
    ".heic": {"image/heic", "image/heif", "image/heic-sequence"},
    ".heif": {"image/heif", "image/heic", "image/heif-sequence"},
    ".pdf": PDF_MIMES,
    ".doc": {"application/msword", "application/vnd.ms-office"},
    ".docx": {
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    },
    ".odt": {"application/vnd.oasis.opendocument.text"},
    ".xls": {"application/vnd.ms-excel", "application/vnd.ms-office"},
    ".xlsx": {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    },
    ".ods": {"application/vnd.oasis.opendocument.spreadsheet"},
    ".txt": {"text/plain"},
    ".csv": {"text/csv", "application/csv", "text/plain"},
}


def _read_upload(upload: UploadedFile, max_bytes: int) -> bytes:
    if upload.size > max_bytes:
        raise forms.ValidationError(
            f"單一檔案不可超過 {max_bytes // MIB} MB。"
        )
    try:
        upload.seek(0)
        data = upload.read(max_bytes + 1)
    except (OSError, ValueError) as exc:
        raise forms.ValidationError("無法讀取這個檔案，請重新選擇。") from exc
    finally:
        try:
            upload.seek(0)
        except (OSError, ValueError):
            pass
    if len(data) > max_bytes:
        raise forms.ValidationError(
            f"單一檔案不可超過 {max_bytes // MIB} MB。"
        )
    if not data:
        raise forms.ValidationError("檔案是空的，請重新選擇。")
    return data


def _extension(upload: UploadedFile, allowed: set[str]) -> str:
    extension = Path(upload.name or "").suffix.lower()
    if extension not in allowed:
        readable = "、".join(sorted(value.removeprefix(".").upper() for value in allowed))
        raise forms.ValidationError(f"僅支援 {readable} 檔案。")
    return extension


def _validate_declared_mime(upload: UploadedFile, extension: str, allowed_mimes: dict[str, set[str]]):
    declared = (getattr(upload, "content_type", "") or "").split(";", 1)[0].strip().lower()
    if declared not in allowed_mimes[extension] | GENERIC_MIMES:
        raise forms.ValidationError(
            "檔案類型與副檔名不符，請確認檔案沒有被改名或損壞。"
        )


def _validate_image(data: bytes, extension: str):
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("error", Image.DecompressionBombWarning)
            with Image.open(BytesIO(data)) as image:
                image.verify()
                actual_format = (image.format or "").upper()
    except (Image.DecompressionBombError, Image.DecompressionBombWarning, OSError, SyntaxError, ValueError) as exc:
        raise forms.ValidationError("圖片已損壞、過大或不是真實的圖片檔。") from exc
    if actual_format != IMAGE_FORMATS[extension]:
        raise forms.ValidationError(
            "圖片內容與副檔名不符，請改用正確格式後再上傳。"
        )


def _resolved_pdf_object(value):
    try:
        return value.get_object()
    except (AttributeError, KeyError, TypeError, ValueError):
        return value


def _unsafe_pdf_action(value) -> bool:
    action = _resolved_pdf_object(value)
    if not hasattr(action, "get"):
        return False
    return str(action.get("/S", "")) in {
        "/JavaScript",
        "/Launch",
        "/SubmitForm",
        "/ImportData",
        "/GoToR",
    } or bool(action.get("/JS"))


def _pdf_has_active_content(reader: PdfReader) -> bool:
    root = _resolved_pdf_object(reader.trailer.get("/Root"))
    if not hasattr(root, "get"):
        return True
    if root.get("/OpenAction") or root.get("/AA"):
        return True
    names = _resolved_pdf_object(root.get("/Names"))
    if hasattr(names, "get") and (
        names.get("/JavaScript") or names.get("/EmbeddedFiles")
    ):
        return True
    acro_form = _resolved_pdf_object(root.get("/AcroForm"))
    if hasattr(acro_form, "get") and acro_form.get("/AA"):
        return True
    for page in reader.pages:
        if page.get("/AA"):
            return True
        annotations = _resolved_pdf_object(page.get("/Annots")) or []
        for annotation_ref in list(annotations)[:1000]:
            annotation = _resolved_pdf_object(annotation_ref)
            if not hasattr(annotation, "get"):
                continue
            if str(annotation.get("/Subtype", "")) == "/FileAttachment":
                return True
            if annotation.get("/AA") or _unsafe_pdf_action(annotation.get("/A")):
                return True
    return False


def _validate_pdf(data: bytes):
    if not data[:1024].lstrip().startswith(b"%PDF-"):
        raise forms.ValidationError("這不是有效的 PDF 檔案。")
    lowered = data.lower()
    dangerous_markers = (b"/javascript", b"/embeddedfiles", b"/openaction", b"/launch")
    if any(marker in lowered for marker in dangerous_markers):
        raise forms.ValidationError("PDF 含有系統不允許的腳本或嵌入檔案。")
    try:
        reader = PdfReader(BytesIO(data), strict=False)
        if reader.is_encrypted or len(reader.pages) < 1:
            raise ValueError("encrypted or empty PDF")
        # Accessing the first page forces pypdf to resolve the page tree.
        reader.pages[0].mediabox
        if _pdf_has_active_content(reader):
            raise forms.ValidationError(
                "PDF 含有系統不允許的腳本、動作或嵌入檔案。"
            )
    except forms.ValidationError:
        raise
    except Exception as exc:  # pypdf raises several malformed-object exception types
        raise forms.ValidationError("無法解析這個 PDF，檔案可能已損壞或加密。") from exc


def _safe_zip(data: bytes) -> ZipFile:
    try:
        archive = ZipFile(BytesIO(data))
        entries = archive.infolist()
    except BadZipFile as exc:
        raise forms.ValidationError("這不是有效的 Office 檔案。") from exc
    if len(entries) > ZIP_MAX_ENTRIES:
        archive.close()
        raise forms.ValidationError("Office 檔案內容異常，包含過多項目。")
    total = 0
    for entry in entries:
        normalized = entry.filename.replace("\\", "/")
        if normalized.startswith("/") or "../" in f"/{normalized}":
            archive.close()
            raise forms.ValidationError("Office 檔案含有不安全的路徑。")
        total += entry.file_size
        if total > ZIP_MAX_UNCOMPRESSED_BYTES:
            archive.close()
            raise forms.ValidationError("Office 檔案解壓後過大，無法安全處理。")
        if entry.compress_size == 0 and entry.file_size > MIB:
            archive.close()
            raise forms.ValidationError("Office 檔案壓縮結構異常。")
        if entry.compress_size and entry.file_size / entry.compress_size > 500:
            archive.close()
            raise forms.ValidationError("Office 檔案壓縮比異常，無法安全處理。")
    return archive


def _require_zip_members(archive: ZipFile, required: set[str]):
    missing = required - set(archive.namelist())
    if missing:
        raise forms.ValidationError("Office 檔案結構不完整或已損壞。")


def _validate_ooxml(data: bytes, extension: str):
    with _safe_zip(data) as archive:
        common = {"[Content_Types].xml", "_rels/.rels"}
        if extension in {".xlsx", ".xlsm"}:
            _require_zip_members(archive, common | {"xl/workbook.xml"})
        elif extension == ".docx":
            _require_zip_members(archive, common | {"word/document.xml"})


def _validate_excel(data: bytes, extension: str):
    _validate_ooxml(data, extension)
    try:
        workbook = load_workbook(
            BytesIO(data),
            read_only=True,
            data_only=False,
            keep_vba=extension == ".xlsm",
            keep_links=False,
        )
        if not workbook.sheetnames:
            raise ValueError("workbook has no worksheets")
        workbook.close()
    except Exception as exc:
        raise forms.ValidationError("Excel 檔案已損壞或結構不完整。") from exc


def _validate_open_document(data: bytes, expected_mimetype: bytes):
    with _safe_zip(data) as archive:
        _require_zip_members(archive, {"mimetype", "content.xml"})
        if archive.read("mimetype") != expected_mimetype:
            raise forms.ValidationError("OpenDocument 檔案類型與副檔名不符。")


def _validate_heif(data: bytes):
    if len(data) < 12 or data[4:8] != b"ftyp":
        raise forms.ValidationError("這不是有效的 HEIC/HEIF 圖片。")
    brands = {b"heic", b"heix", b"hevc", b"hevx", b"mif1", b"msf1"}
    if data[8:12] not in brands and not any(brand in data[8:64] for brand in brands):
        raise forms.ValidationError("這不是有效的 HEIC/HEIF 圖片。")


def _validate_text(data: bytes):
    if b"\x00" in data:
        raise forms.ValidationError("文字檔包含無法讀取的二進位內容。")
    prefix = data[:1024].lstrip().lower()
    if prefix.startswith((b"<!doctype html", b"<html", b"<script", b"<?php", b"mz")):
        raise forms.ValidationError("文字檔內容不安全或與檔案類型不符。")


def validate_document_upload(upload, *, max_bytes: int = DOCUMENT_MAX_BYTES):
    """Validate image/PDF uploads used for contracts and operational proof."""
    if not isinstance(upload, UploadedFile):
        return upload
    allowed = set(IMAGE_MIMES) | {".pdf"}
    extension = _extension(upload, allowed)
    mime_map = {**IMAGE_MIMES, ".pdf": PDF_MIMES}
    _validate_declared_mime(upload, extension, mime_map)
    data = _read_upload(upload, max_bytes)
    if extension == ".pdf":
        _validate_pdf(data)
    else:
        _validate_image(data, extension)
    return upload


def validate_image_upload(upload, *, max_bytes: int = IMAGE_MAX_BYTES):
    """Validate camera/gallery uploads by parsing their real image content."""
    if not isinstance(upload, UploadedFile):
        return upload
    extension = _extension(upload, set(IMAGE_MIMES))
    _validate_declared_mime(upload, extension, IMAGE_MIMES)
    data = _read_upload(upload, max_bytes)
    _validate_image(data, extension)
    return upload


def validate_excel_upload(upload, *, max_bytes: int = 30 * MIB):
    """Validate the .xlsx workbook used by the legacy import preview."""
    if not isinstance(upload, UploadedFile):
        return upload
    extension = _extension(upload, {".xlsx"})
    _validate_declared_mime(upload, extension, {".xlsx": SUBSIDY_MIMES[".xlsx"]})
    data = _read_upload(upload, max_bytes)
    _validate_excel(data, extension)
    return upload


def validate_template_background(upload):
    """Validate image/PDF/Excel backgrounds before template rendering."""
    if not isinstance(upload, UploadedFile):
        return upload
    allowed = set(IMAGE_MIMES) | {".pdf", ".xlsx", ".xlsm"}
    extension = _extension(upload, allowed)
    mime_map = {
        **IMAGE_MIMES,
        ".pdf": PDF_MIMES,
        ".xlsx": SUBSIDY_MIMES[".xlsx"],
        ".xlsm": {
            "application/vnd.ms-excel.sheet.macroenabled.12",
            "application/vnd.ms-excel",
        },
    }
    _validate_declared_mime(upload, extension, mime_map)
    data = _read_upload(upload, TEMPLATE_MAX_BYTES)
    if extension == ".pdf":
        _validate_pdf(data)
    elif extension in {".xlsx", ".xlsm"}:
        _validate_excel(data, extension)
    else:
        _validate_image(data, extension)
    return upload


def validate_subsidy_upload(upload):
    """Retain all existing subsidy formats while validating their structure."""
    if not isinstance(upload, UploadedFile):
        return upload
    extension = _extension(upload, set(SUBSIDY_MIMES))
    _validate_declared_mime(upload, extension, SUBSIDY_MIMES)
    data = _read_upload(upload, SUBSIDY_MAX_BYTES)
    if extension in IMAGE_MIMES:
        _validate_image(data, extension)
    elif extension in {".heic", ".heif"}:
        _validate_heif(data)
    elif extension == ".pdf":
        _validate_pdf(data)
    elif extension == ".docx":
        _validate_ooxml(data, extension)
    elif extension == ".xlsx":
        _validate_excel(data, extension)
    elif extension == ".odt":
        _validate_open_document(data, b"application/vnd.oasis.opendocument.text")
    elif extension == ".ods":
        _validate_open_document(data, b"application/vnd.oasis.opendocument.spreadsheet")
    elif extension in {".doc", ".xls"}:
        if not data.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
            raise forms.ValidationError("這不是有效的舊版 Office 檔案。")
    else:
        _validate_text(data)
    return upload

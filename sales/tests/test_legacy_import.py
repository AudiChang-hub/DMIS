from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from sales.models import (
    LegacyImportBatch,
    LegacyImportCorrection,
    LegacyImportMasterMapping,
    LegacyImportRow,
    LegacySalesSnapshot,
    SalesOrder,
    SalesSource,
    SalesSourceBrandPolicy,
    SalesSourceCooperationProfile,
    SalesSourceCategory,
    Store,
    VehicleColor,
    VehicleInventory,
    VehicleModel,
)
from sales.forms import LegacyImportRowCorrectionForm, LegacyImportUploadForm
from sales.jobs import run_legacy_import_job
from sales.services.legacy_import import (
    INVALID_EMAIL_MESSAGE,
    PREVIEW_SCHEMA_VERSION,
    _clean_sales_source_name,
    _infer_sales_transaction_type,
    _infer_sales_vehicle_category,
    _sales_order_note,
    apply_import_row_decision,
    build_import_master_workspace,
    build_import_preview,
    confirm_import,
    file_sha256,
    revalidate_import_batch,
    retry_completed_import_row,
    save_import_master_mapping,
)


def workbook_bytes(kind="operations"):
    workbook = Workbook()
    if kind == "operations":
        sales = workbook.active
        sales.title = "銷貨"
        sales["C3"] = "車種型號"
        sales["D3"] = "識別號碼"
        sales["E3"] = "車主名稱"
        sales["G3"] = "收款價"
        sales["H3"] = "成本"
        sales["J3"] = "現金"
        sales["B4"] = "2026/08/01"
        sales["C4"] = "TEST125"
        sales["D4"] = " ab-123 "
        sales["E4"] = "舊欄姓名"
        sales["AT4"] = "正式車主"
        sales["F4"] = "白"
        sales["G4"] = 70000
        sales["H4"] = 60000
        sales["J4"] = 70000
        sales["AS4"] = "ABC-1234"
        sales["AW4"] = "A123456789"
        sales["AX4"] = "新北市測試路1號"
        sales["AY4"] = "0912345678"
        sales["CH4"] = "2026/07/30"
        inventory = workbook.create_sheet("進貨")
        for column, label in enumerate(("進貨日期", "車種型號", "車身號碼", "顏色", "尺碼", "數量", "單價", "總額", "月份", "出廠日期"), 1):
            inventory.cell(1, column, label)
        inventory.append(["2026/07/01", "TEST125", " AB-123 ", "白", "", 0, "", "", "", "2026/06"])
    else:
        dealer = workbook.active
        dealer.title = "車行"
        dealer.append(["", "", "", "", "", "", "", "月餅", "價格表", "容量"])
        dealer.append(["店名", "負責人", "電話一", "電話二", "手機", "傳真", "地址", "三陽", "台鈴", "排車容量", "備註"])
        dealer.append(["測試車行", "王先生", "02-1234", "", "0912", "", "新北市", "", "V", 5, "合作中"])
        platform = workbook.create_sheet("網路平台")
        platform.append(["平台", "聯絡人", "電話", "分機", "手機", "信箱"])
        platform.append(["測試平台", "李小姐", "02-5678", "123", "", "test@example.com"])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def conflict_workbook_bytes():
    workbook = load_workbook(BytesIO(workbook_bytes()))
    inventory = workbook["進貨"]
    inventory.append(["2026/07/02", "TEST125", "AB 123", "黑", "", 1, "", "", "", "2026/07"])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def used_vehicle_resale_workbook_bytes(mark_as_used=True):
    workbook = load_workbook(BytesIO(workbook_bytes()))
    sales = workbook["銷貨"]
    sales["AN3"] = "車行"
    sales["AT3"] = "車主名稱2"
    sales["AW3"] = "身分證字號"
    sales["B5"] = "2026/08/05"
    sales["C5"] = "TEST125"
    sales["D5"] = "AB-123"
    sales["E5"] = "中古車主"
    sales["AT5"] = "中古車主"
    sales["F5"] = "白"
    sales["G5"] = 30000
    sales["J5"] = 30000
    sales["AN5"] = "中古車" if mark_as_used else ""
    sales["AS5"] = "ABC-1234"
    sales["AW5"] = "B223456789"
    sales["AX5"] = "新北市中古路2號"
    sales["AY5"] = "0987654321"
    # 舊 Excel 尾端常因公式留下只有 0 的空白列，不能產生假訂單。
    sales["AT6"] = 0
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


class LegacyImportTests(TestCase):
    def setUp(self):
        Store.objects.create(name="總店", code="MAIN")
        self.user = get_user_model().objects.create_user(username="importer", password="test-pass")
        self.client.force_login(self.user)
        self.tempdir = TemporaryDirectory()
        self.override = override_settings(MEDIA_ROOT=self.tempdir.name)
        self.override.enable()

    def tearDown(self):
        self.override.disable()
        self.tempdir.cleanup()

    def make_batch(self, kind):
        content = workbook_bytes(kind)
        upload = SimpleUploadedFile(f"{kind}.xlsx", content)
        digest = file_sha256(upload)
        return LegacyImportBatch.objects.create(
            import_type=kind,
            source_file=upload,
            original_filename=f"{kind}.xlsx",
            file_sha256=digest,
            file_size=len(content),
            uploaded_by="tester",
        )

    def make_conflict_batch(self):
        content = conflict_workbook_bytes()
        upload = SimpleUploadedFile("conflicts.xlsx", content)
        digest = file_sha256(upload)
        return LegacyImportBatch.objects.create(
            import_type=LegacyImportBatch.ImportType.OPERATIONS,
            source_file=upload,
            original_filename="conflicts.xlsx",
            file_sha256=digest,
            file_size=len(content),
            uploaded_by="tester",
        )

    def make_used_vehicle_batch(self, mark_as_used=True):
        content = used_vehicle_resale_workbook_bytes(mark_as_used=mark_as_used)
        upload = SimpleUploadedFile("used-vehicle.xlsx", content)
        digest = file_sha256(upload)
        return LegacyImportBatch.objects.create(
            import_type=LegacyImportBatch.ImportType.OPERATIONS,
            source_file=upload,
            original_filename="used-vehicle.xlsx",
            file_sha256=digest,
            file_size=len(content),
            uploaded_by="tester",
        )

    def test_upload_form_only_contains_type_and_file(self):
        self.assertEqual(list(LegacyImportUploadForm().fields), ["import_type", "source_file"])

    def test_invalid_row_can_be_excluded_without_filling_required_import_fields(self):
        batch = self.make_batch(LegacyImportBatch.ImportType.OPERATIONS)
        build_import_preview(batch)
        row = batch.rows.get(sheet_name="進貨")
        form = LegacyImportRowCorrectionForm(
            {"decision": "exclude", "reason": "來源列無法確認，先不匯入"},
            row=row,
        )
        self.assertTrue(form.is_valid(), form.errors)

    def test_operations_preview_and_confirm_preserves_historical_price(self):
        batch = self.make_batch(LegacyImportBatch.ImportType.OPERATIONS)
        summary = build_import_preview(batch)
        self.assertEqual(summary["source_rows"], 2)
        self.assertEqual(batch.rows.filter(sheet_name="銷貨").count(), 1)
        result = confirm_import(batch, "tester")
        self.assertEqual(result["created"], 2)
        order = SalesOrder.objects.get(owner_name="正式車主")
        self.assertEqual(order.vehicle_price, 0)
        self.assertEqual(order.allocated_vehicle.normalized_engine_number, "AB123")
        self.assertEqual(order.legacy_snapshot.historical_received_price, 70000)
        self.assertEqual(LegacySalesSnapshot.objects.count(), 1)
        self.assertEqual(VehicleInventory.objects.get().status, VehicleInventory.Status.SOLD)

    def test_trial_vehicle_source_suffix_becomes_order_note(self):
        SalesSource.objects.create(
            name="昌勝",
            source_type=SalesSource.SourceType.DEALER,
            active=True,
        )
        workbook = load_workbook(BytesIO(workbook_bytes()))
        workbook["銷貨"]["AN3"] = "車行"
        workbook["銷貨"]["AN4"] = "昌勝(試乘車)"
        stream = BytesIO()
        workbook.save(stream)
        upload = SimpleUploadedFile("trial-vehicle.xlsx", stream.getvalue())
        batch = LegacyImportBatch.objects.create(
            import_type=LegacyImportBatch.ImportType.OPERATIONS,
            source_file=upload,
            original_filename="trial-vehicle.xlsx",
            file_sha256=file_sha256(upload),
            file_size=len(stream.getvalue()),
            uploaded_by="tester",
        )

        build_import_preview(batch)
        sales_row = batch.rows.get(sheet_name="銷貨")
        self.assertEqual(sales_row.mapped_data["dealer_name"], "昌勝")
        self.assertEqual(sales_row.mapped_data["transaction_type"], "test_ride")
        self.assertEqual(sales_row.mapped_data["note"], "試乘車")

        confirm_import(batch, "tester")

        order = SalesOrder.objects.get(owner_name="正式車主")
        self.assertEqual(order.source.name, "昌勝")
        self.assertEqual(order.transaction_type, SalesOrder.TransactionType.TEST_RIDE)
        self.assertEqual(order.note, "試乘車")

    def test_subsidy_application_pseudo_source_becomes_order_note(self):
        workbook = load_workbook(BytesIO(workbook_bytes()))
        workbook["銷貨"]["AN3"] = "車行"
        workbook["銷貨"]["AN4"] = "代申請補助"
        stream = BytesIO()
        workbook.save(stream)
        upload = SimpleUploadedFile("subsidy-application.xlsx", stream.getvalue())
        batch = LegacyImportBatch.objects.create(
            import_type=LegacyImportBatch.ImportType.OPERATIONS,
            source_file=upload,
            original_filename="subsidy-application.xlsx",
            file_sha256=file_sha256(upload),
            file_size=len(stream.getvalue()),
            uploaded_by="tester",
        )

        build_import_preview(batch)
        sales_row = batch.rows.get(sheet_name="銷貨")
        self.assertEqual(sales_row.mapped_data["dealer_name"], "")
        self.assertEqual(sales_row.mapped_data["transaction_type"], "regular_new")
        self.assertEqual(sales_row.mapped_data["note"], "代申請補助")

        confirm_import(batch, "tester")

        order = SalesOrder.objects.get(owner_name="正式車主")
        self.assertIsNone(order.source)
        self.assertEqual(order.source_type, SalesOrder.SourceType.STORE)
        self.assertEqual(
            order.transaction_type, SalesOrder.TransactionType.REGULAR_NEW
        )
        self.assertEqual(order.note, "代申請補助")
        self.assertEqual(order.operations.dealer_name, "")

    def test_invalid_email_is_reported_during_preview(self):
        workbook = load_workbook(BytesIO(workbook_bytes()))
        workbook["銷貨"]["AZ4"] = "新北市測試路1號"
        stream = BytesIO()
        workbook.save(stream)
        upload = SimpleUploadedFile("invalid-email.xlsx", stream.getvalue())
        batch = LegacyImportBatch.objects.create(
            import_type=LegacyImportBatch.ImportType.OPERATIONS,
            source_file=upload,
            original_filename="invalid-email.xlsx",
            file_sha256=file_sha256(upload),
            file_size=len(stream.getvalue()),
            uploaded_by="tester",
        )

        summary = build_import_preview(batch)

        sales_row = batch.rows.get(sheet_name="銷貨")
        self.assertEqual(sales_row.action, LegacyImportRow.Action.ERROR)
        self.assertIn(INVALID_EMAIL_MESSAGE, sales_row.messages)
        self.assertEqual(summary["counts"]["error"], 1)

    @patch("sales.views.django_rq.get_queue")
    def test_confirm_starts_background_import_and_status_endpoint_reports_progress(self, get_queue):
        batch = self.make_batch(LegacyImportBatch.ImportType.OPERATIONS)
        build_import_preview(batch)
        get_queue.return_value.enqueue.return_value.id = "job-123"

        response = self.client.post(reverse("legacy_import_confirm", args=[batch.pk]))

        self.assertRedirects(response, reverse("legacy_import_detail", args=[batch.pk]))
        batch.refresh_from_db()
        self.assertEqual(batch.status, LegacyImportBatch.Status.PROCESSING)
        self.assertEqual(batch.processing_total, 2)
        self.assertEqual(batch.processing_job_id, "job-123")
        self.assertFalse(SalesOrder.objects.exists())
        get_queue.return_value.enqueue.assert_called_once()

        status_response = self.client.get(reverse("legacy_import_status", args=[batch.pk]))
        self.assertEqual(status_response.status_code, 200)
        self.assertEqual(status_response.json()["status"], "processing")
        self.assertEqual(status_response.json()["percent"], 0)

        run_legacy_import_job(str(batch.pk), "tester")
        batch.refresh_from_db()
        self.assertEqual(batch.status, LegacyImportBatch.Status.COMPLETED)
        self.assertEqual(batch.processing_completed, 2)
        self.assertEqual(SalesOrder.objects.count(), 1)

    @patch("sales.views.django_rq.get_queue")
    def test_duplicate_confirm_does_not_enqueue_second_job(self, get_queue):
        batch = self.make_batch(LegacyImportBatch.ImportType.OPERATIONS)
        build_import_preview(batch)
        get_queue.return_value.enqueue.return_value.id = "job-123"

        self.client.post(reverse("legacy_import_confirm", args=[batch.pk]))
        self.client.post(reverse("legacy_import_confirm", args=[batch.pk]))

        get_queue.return_value.enqueue.assert_called_once()

    @patch("sales.views.django_rq.get_queue")
    def test_enqueue_failure_returns_batch_to_preview(self, get_queue):
        batch = self.make_batch(LegacyImportBatch.ImportType.OPERATIONS)
        build_import_preview(batch)
        get_queue.return_value.enqueue.side_effect = RuntimeError("redis unavailable")

        self.client.post(reverse("legacy_import_confirm", args=[batch.pk]))

        batch.refresh_from_db()
        self.assertEqual(batch.status, LegacyImportBatch.Status.PREVIEW)
        self.assertIn("無法啟動", batch.processing_error)
        self.assertFalse(SalesOrder.objects.exists())

    def test_failed_background_import_can_resume_without_duplicate_records(self):
        batch = self.make_batch(LegacyImportBatch.ImportType.OPERATIONS)
        build_import_preview(batch)
        confirm_import(batch, "tester")
        batch.refresh_from_db()
        batch.status = LegacyImportBatch.Status.FAILED
        batch.processing_error = "模擬 worker 中斷"
        batch.save(update_fields=["status", "processing_error", "updated_at"])

        result = confirm_import(batch, "tester")

        self.assertEqual(result["created"], 2)
        self.assertEqual(SalesOrder.objects.count(), 1)
        self.assertEqual(VehicleInventory.objects.count(), 1)

    def test_completed_error_row_can_be_repaired_and_imported_alone(self):
        workbook = load_workbook(BytesIO(workbook_bytes()))
        workbook["銷貨"]["AZ4"] = "0912345678"
        stream = BytesIO()
        workbook.save(stream)
        upload = SimpleUploadedFile("repair-email.xlsx", stream.getvalue())
        batch = LegacyImportBatch.objects.create(
            import_type=LegacyImportBatch.ImportType.OPERATIONS,
            source_file=upload,
            original_filename="repair-email.xlsx",
            file_sha256=file_sha256(upload),
            file_size=len(stream.getvalue()),
            uploaded_by="tester",
        )
        build_import_preview(batch)
        sales_row = batch.rows.get(sheet_name="銷貨")
        apply_import_row_decision(
            sales_row,
            {},
            LegacyImportCorrection.Decision.EXCLUDE,
            "先完成其他資料",
            "tester",
        )
        confirm_import(batch, "tester")
        sales_row.refresh_from_db()
        sales_row.action = LegacyImportRow.Action.ERROR
        sales_row.excluded = False
        sales_row.messages = [INVALID_EMAIL_MESSAGE]
        sales_row.save(update_fields=["action", "excluded", "messages", "updated_at"])
        batch.refresh_from_db()
        batch.result_summary = {**batch.result_summary, "excluded": 0, "errors": 1}
        batch.save(update_fields=["result_summary", "updated_at"])

        result = retry_completed_import_row(
            sales_row,
            {"owner_email": ""},
            LegacyImportCorrection.Decision.CORRECT,
            "Email 欄誤放電話，清空後補匯",
            "tester",
        )

        self.assertTrue(result["ok"])
        sales_row.refresh_from_db()
        batch.refresh_from_db()
        self.assertEqual(sales_row.committed_model, "SalesOrder")
        self.assertEqual(SalesOrder.objects.count(), 1)
        self.assertEqual(batch.result_summary["errors"], 0)

    def test_explicit_used_vehicle_signals_do_not_match_trade_in_wording(self):
        self.assertEqual(
            _infer_sales_vehicle_category({}, "中古車")[0],
            SalesOrder.VehicleCategory.USED,
        )
        self.assertEqual(
            _infer_sales_vehicle_category({"備註": "中古車買賣"}, "")[0],
            SalesOrder.VehicleCategory.USED,
        )
        self.assertEqual(
            _infer_sales_vehicle_category({"補助方案": "FUN 中古車過戶"}, "")[0],
            SalesOrder.VehicleCategory.USED,
        )
        self.assertEqual(
            _infer_sales_vehicle_category({"備註": "新車成交，另有中古車估價"}, "")[0],
            SalesOrder.VehicleCategory.NEW,
        )

    def test_transaction_type_is_separated_from_sales_source_name(self):
        self.assertEqual(_clean_sales_source_name("昌勝(試乘車)"), "昌勝")
        self.assertEqual(_clean_sales_source_name("東永-試乘車"), "東永")
        self.assertEqual(_clean_sales_source_name("中獎車"), "")
        self.assertEqual(_clean_sales_source_name("代申請補助"), "")
        self.assertEqual(
            _clean_sales_source_name("昌勝(代申請補助)"), "昌勝"
        )
        self.assertEqual(
            _infer_sales_transaction_type(
                {}, "昌勝(試乘車)", SalesOrder.VehicleCategory.NEW
            )[0],
            SalesOrder.TransactionType.TEST_RIDE,
        )
        self.assertEqual(
            _infer_sales_transaction_type(
                {}, "中獎車", SalesOrder.VehicleCategory.NEW
            )[0],
            SalesOrder.TransactionType.PRIZE,
        )
        self.assertEqual(
            _infer_sales_transaction_type(
                {}, "中古車", SalesOrder.VehicleCategory.USED
            )[0],
            SalesOrder.TransactionType.USED,
        )
        self.assertEqual(
            _sales_order_note(
                {}, "昌勝(試乘車)", SalesOrder.TransactionType.TEST_RIDE
            ),
            "試乘車",
        )
        self.assertEqual(
            _sales_order_note(
                {"備註": "客戶指定"},
                "昌勝（試乘車）",
                SalesOrder.TransactionType.TEST_RIDE,
                "試乘車",
            ),
            "試乘車\n客戶指定",
        )
        self.assertEqual(
            _sales_order_note(
                {},
                "代申請補助",
                SalesOrder.TransactionType.REGULAR_NEW,
            ),
            "代申請補助",
        )

    def test_special_platform_source_name_is_moved_to_order_note(self):
        rules = (
            ("momo員購", "momo"),
            ("小樹購員購", "小樹購"),
            ("Yahoo+假展場", "Yahoo"),
        )
        for legacy_name, canonical_name in rules:
            with self.subTest(legacy_name=legacy_name):
                self.assertEqual(
                    _clean_sales_source_name(legacy_name), canonical_name
                )
                self.assertEqual(
                    _sales_order_note(
                        {},
                        legacy_name,
                        SalesOrder.TransactionType.REGULAR_NEW,
                    ),
                    legacy_name,
                )

    def test_employee_purchase_platform_name_is_moved_to_store_order_note(self):
        source_names = (
            "上海商銀員購",
            "台新銀員購",
            "台新銀行員購",
            "華新麗華員購",
        )
        for source_name in source_names:
            with self.subTest(source_name=source_name):
                self.assertEqual(_clean_sales_source_name(source_name), "")
                self.assertEqual(
                    _sales_order_note(
                        {},
                        source_name,
                        SalesOrder.TransactionType.REGULAR_NEW,
                    ),
                    source_name,
                )
                self.assertEqual(
                    _sales_order_note(
                        {"備註": source_name},
                        source_name,
                        SalesOrder.TransactionType.REGULAR_NEW,
                        source_name,
                    ),
                    source_name,
                )

        self.assertEqual(_clean_sales_source_name("博客來"), "博客來")
        self.assertEqual(
            _sales_order_note(
                {}, "博客來", SalesOrder.TransactionType.REGULAR_NEW
            ),
            "",
        )

    def test_used_vehicle_resale_can_share_identifier_without_reusing_inventory(self):
        batch = self.make_used_vehicle_batch()
        summary = build_import_preview(batch)
        self.assertEqual(summary["counts"]["conflict"], 0)
        self.assertEqual(summary["validation"]["used_vehicle_sales"], 1)
        sales_rows = list(batch.rows.filter(sheet_name="銷貨").order_by("source_row"))
        self.assertEqual(len(sales_rows), 2)
        self.assertEqual(sales_rows[0].mapped_data["vehicle_category"], "new")
        self.assertEqual(sales_rows[1].mapped_data["vehicle_category"], "used")
        self.assertNotEqual(sales_rows[0].natural_key, sales_rows[1].natural_key)

        confirm_import(batch, "tester")

        new_order = SalesOrder.objects.get(owner_name="正式車主")
        used_order = SalesOrder.objects.get(owner_name="中古車主")
        self.assertEqual(new_order.vehicle_category, SalesOrder.VehicleCategory.NEW)
        self.assertIsNotNone(new_order.allocated_vehicle)
        self.assertEqual(used_order.vehicle_category, SalesOrder.VehicleCategory.USED)
        self.assertIsNone(used_order.allocated_vehicle)
        self.assertEqual(used_order.legacy_snapshot.vehicle_identifier, "AB-123")
        self.assertIn("ab123", used_order.search_index.search_text)
        detail = self.client.get(reverse("order_detail", args=[used_order.pk]))
        self.assertContains(detail, "歷史中古車交易")
        self.assertContains(detail, "不需占用新車庫存")

    def test_revalidate_backfills_vehicle_category_for_existing_preview(self):
        batch = self.make_used_vehicle_batch()
        build_import_preview(batch)
        for row in batch.rows.filter(sheet_name="銷貨"):
            mapped_data = dict(row.mapped_data)
            mapped_data.pop("vehicle_category", None)
            mapped_data.pop("vehicle_category_reason", None)
            row.mapped_data = mapped_data
            row.save(update_fields=["mapped_data", "updated_at"])
        placeholder = LegacyImportRow.objects.create(
            batch=batch,
            sheet_name="銷貨",
            source_row=99,
            fingerprint="0" * 64,
            natural_key="legacy-placeholder",
            action=LegacyImportRow.Action.CREATE,
            raw_data={"車主名稱": 0},
            mapped_data={
                "owner_name": "0",
                "model_number": "",
                "identifier": "",
                "plate_number": "",
            },
            messages=[],
        )

        summary = revalidate_import_batch(batch)

        rows = list(batch.rows.filter(sheet_name="銷貨").order_by("source_row"))
        self.assertEqual(summary["counts"]["conflict"], 0)
        self.assertEqual(summary["counts"]["skip"], 1)
        self.assertEqual(summary["validation"]["used_vehicle_sales"], 1)
        self.assertEqual(rows[0].mapped_data["vehicle_category"], "new")
        self.assertEqual(rows[1].mapped_data["vehicle_category"], "used")
        placeholder.refresh_from_db()
        self.assertEqual(placeholder.action, LegacyImportRow.Action.SKIP)
        self.assertIn("Excel 空白公式列，系統自動略過", placeholder.messages)

    def test_opening_old_preview_automatically_applies_current_parser_rules(self):
        batch = self.make_batch(LegacyImportBatch.ImportType.OPERATIONS)
        build_import_preview(batch)
        row = batch.rows.get(sheet_name="銷貨")
        mapped_data = dict(row.mapped_data)
        mapped_data["dealer_name"] = "試乘車"
        mapped_data["dealer_name_raw"] = "試乘車"
        mapped_data["transaction_type"] = SalesOrder.TransactionType.REGULAR_NEW
        mapped_data["transaction_type_reason"] = "舊版預設值"
        row.mapped_data = mapped_data
        row.save(update_fields=["mapped_data", "updated_at"])
        batch.preview_summary = {
            **batch.preview_summary,
            "validation": {"unmapped_sources": ["試乘車"]},
        }
        batch.preview_summary.pop("parser_schema_version", None)
        batch.save(update_fields=["preview_summary", "updated_at"])

        response = self.client.get(reverse("legacy_import_detail", args=[batch.pk]))

        self.assertEqual(response.status_code, 200)
        row.refresh_from_db()
        batch.refresh_from_db()
        self.assertEqual(row.mapped_data["dealer_name_raw"], "試乘車")
        self.assertEqual(row.mapped_data["dealer_name"], "")
        self.assertEqual(
            row.mapped_data["transaction_type"],
            SalesOrder.TransactionType.TEST_RIDE,
        )
        self.assertEqual(
            batch.preview_summary["parser_schema_version"],
            PREVIEW_SCHEMA_VERSION,
        )
        self.assertNotIn(
            "試乘車", batch.preview_summary["validation"]["unmapped_sources"]
        )

    def test_ambiguous_second_new_sale_stays_conflict_until_marked_used(self):
        batch = self.make_used_vehicle_batch(mark_as_used=False)
        summary = build_import_preview(batch)
        self.assertEqual(summary["counts"]["conflict"], 2)
        later_row = batch.rows.filter(sheet_name="銷貨").order_by("source_row").last()
        mapping = dict(later_row.mapped_data)
        mapping["vehicle_category"] = SalesOrder.VehicleCategory.USED
        summary = apply_import_row_decision(
            later_row,
            mapping,
            LegacyImportCorrection.Decision.CORRECT,
            "確認為中古車交易",
            "tester",
        )
        later_row.refresh_from_db()
        self.assertEqual(summary["counts"]["conflict"], 0)
        self.assertEqual(later_row.mapped_data["vehicle_category_reason"], "人工調整")

    def test_channel_preview_groups_source_and_contact(self):
        batch = self.make_batch(LegacyImportBatch.ImportType.CHANNELS)
        summary = build_import_preview(batch)
        self.assertEqual(summary["source_rows"], 2)
        result = confirm_import(batch, "tester")
        self.assertEqual(result["created"], 2)
        self.assertEqual(batch.rows.filter(committed_model="SalesSource").count(), 2)
        dealer = SalesSource.objects.get(name="測試車行")
        platform = SalesSource.objects.get(name="測試平台")
        self.assertEqual(dealer.responsible_person, "王先生")
        self.assertEqual(dealer.phone, "02-1234")
        self.assertEqual(dealer.mobile, "0912")
        self.assertEqual(dealer.address, "新北市")
        self.assertEqual(dealer.note, "合作中")
        self.assertIn("歷史聯絡資料：李小姐", platform.note)
        self.assertIn("分機：123", platform.note)
        self.assertIn("Email：test@example.com", platform.note)
        profiles = {
            profile.cooperation_scope: profile
            for profile in dealer.cooperation_profiles.all()
        }
        self.assertEqual(len(profiles), 3)
        self.assertTrue(
            profiles[SalesSourceBrandPolicy.CooperationScope.SUZUKI_GAS].cooperates
        )
        self.assertTrue(
            profiles[SalesSourceBrandPolicy.CooperationScope.SUZUKI_ELECTRIC].cooperates
        )
        self.assertFalse(
            profiles[SalesSourceBrandPolicy.CooperationScope.SYM].cooperates
        )
        self.assertIsNone(dealer.sym_vehicle_capacity)
        self.assertEqual(dealer.suzuki_vehicle_capacity, 5)

    def test_current_dealer_workbook_maps_suzuki_electric_without_gas(self):
        workbook = Workbook()
        dealer = workbook.active
        dealer.title = "車行"
        dealer.append(["月餅", "LINE群組", "", "", "", "", "", "", "", "價格表", "", "排車容量", "", ""])
        dealer.append(["", "", "店名", "負責人", "電話一", "電話二", "手機", "手機/傳真", "地址", "三陽", "台鈴", "三陽", "台鈴", "備註"])
        dealer.append(["", "E", "電動車行", "王先生", "02-1234", "", "0912", "", "基隆市仁愛區", "", "電動車", "", 3, "Excel 備註"])
        dealer.append(["", "", "油電車行", "李小姐", "02-5678", "", "0922", "", "新北市汐止區", "", "V", "", 5, "油電備註"])
        workbook.create_sheet("網路平台").append(["平台", "聯絡人", "電話", "分機", "手機", "信箱"])
        stream = BytesIO()
        workbook.save(stream)
        upload = SimpleUploadedFile("channels-current.xlsx", stream.getvalue())
        batch = LegacyImportBatch.objects.create(
            import_type=LegacyImportBatch.ImportType.CHANNELS,
            source_file=upload,
            original_filename="channels-current.xlsx",
            file_sha256=file_sha256(upload),
            file_size=len(stream.getvalue()),
            uploaded_by="tester",
        )

        build_import_preview(batch)
        confirm_import(batch, "tester")

        source = SalesSource.objects.get(name="電動車行")
        profiles = {
            item.cooperation_scope: item.cooperates
            for item in source.cooperation_profiles.all()
        }
        self.assertFalse(profiles[SalesSourceBrandPolicy.CooperationScope.SUZUKI_GAS])
        self.assertTrue(profiles[SalesSourceBrandPolicy.CooperationScope.SUZUKI_ELECTRIC])
        self.assertEqual(source.note, "Excel 備註")
        self.assertTrue(source.has_line_group)

        oil_and_electric = SalesSource.objects.get(name="油電車行")
        oil_and_electric_profiles = {
            item.cooperation_scope: item.cooperates
            for item in oil_and_electric.cooperation_profiles.all()
        }
        self.assertTrue(
            oil_and_electric_profiles[SalesSourceBrandPolicy.CooperationScope.SUZUKI_GAS]
        )
        self.assertTrue(
            oil_and_electric_profiles[SalesSourceBrandPolicy.CooperationScope.SUZUKI_ELECTRIC]
        )

    def test_same_batch_cannot_be_confirmed_twice(self):
        batch = self.make_batch(LegacyImportBatch.ImportType.CHANNELS)
        build_import_preview(batch)
        confirm_import(batch, "tester")
        with self.assertRaises(ValueError):
            confirm_import(batch, "tester")

    def test_unresolved_conflict_blocks_confirm(self):
        batch = self.make_conflict_batch()
        summary = build_import_preview(batch)
        self.assertEqual(summary["counts"]["conflict"], 2)
        with self.assertRaisesMessage(ValueError, "尚有 2 筆衝突或錯誤資料"):
            confirm_import(batch, "tester")
        self.assertFalse(SalesOrder.objects.exists())

    def test_correcting_identifier_revalidates_entire_batch_and_keeps_audit(self):
        batch = self.make_conflict_batch()
        build_import_preview(batch)
        row = batch.rows.filter(sheet_name="進貨").order_by("source_row").last()
        summary = apply_import_row_decision(
            row,
            {"identifier_raw": "AB-124"},
            LegacyImportCorrection.Decision.CORRECT,
            "第二筆識別號碼輸入錯誤",
            "tester",
        )
        row.refresh_from_db()
        self.assertEqual(summary["counts"]["conflict"], 0)
        self.assertEqual(row.mapped_data["identifier"], "AB124")
        self.assertTrue(row.manually_corrected)
        self.assertEqual(row.corrections.get().reason, "第二筆識別號碼輸入錯誤")

    def test_excluding_duplicate_resolves_conflict_without_changing_source(self):
        batch = self.make_conflict_batch()
        build_import_preview(batch)
        row = batch.rows.filter(sheet_name="進貨").order_by("source_row").last()
        original_raw = dict(row.raw_data)
        summary = apply_import_row_decision(
            row,
            {},
            LegacyImportCorrection.Decision.EXCLUDE,
            "Excel 重複列",
            "tester",
        )
        row.refresh_from_db()
        self.assertEqual(summary["counts"]["conflict"], 0)
        self.assertEqual(summary["counts"]["exclude"], 1)
        self.assertEqual(row.raw_data, original_raw)
        result = confirm_import(batch, "tester")
        self.assertEqual(result["excluded"], 1)

    def test_conflict_workspace_renders_and_row_can_be_corrected_in_browser_flow(self):
        batch = self.make_conflict_batch()
        build_import_preview(batch)
        conflict_url = reverse("legacy_import_detail", args=[batch.pk]) + "?action=conflict"
        response = self.client.get(conflict_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "重複資料比較")
        self.assertContains(response, "2 筆需要確認")
        row = batch.rows.filter(sheet_name="進貨").order_by("source_row").last()
        edit_response = self.client.get(conflict_url + f"&edit={row.pk}")
        self.assertEqual(edit_response.status_code, 200)
        self.assertContains(edit_response, "修正或排除此列")
        post_response = self.client.post(
            reverse("legacy_import_row_decide", args=[batch.pk, row.pk]),
            {
                "decision": "correct",
                "received_on": "2026-07-02",
                "model_number": "TEST125",
                "identifier_raw": "AB-124",
                "color": "黑",
                "quantity": "1",
                "manufactured_year_month": "2026/07",
                "reason": "識別號碼更正",
            },
        )
        self.assertEqual(post_response.status_code, 302)
        batch.refresh_from_db()
        self.assertEqual(batch.preview_summary["counts"]["conflict"], 0)

    def test_unmapped_vehicle_can_link_existing_master_without_changing_excel(self):
        batch = self.make_batch(LegacyImportBatch.ImportType.OPERATIONS)
        build_import_preview(batch)
        original_raw = dict(batch.rows.get(sheet_name="進貨").raw_data)
        target = VehicleModel.objects.create(
            brand="SUZUKI",
            name="SUI 125",
            model_number="UQ125DA",
            energy_type=VehicleModel.EnergyType.GAS,
            model_year=2026,
            model_code=VehicleModel.ModelType.FRONT_DISC_REAR_DRUM,
            displacement_cc=125,
        )
        response = self.client.post(
            reverse(
                "legacy_import_master_resolve",
                args=[batch.pk, LegacyImportMasterMapping.MappingType.VEHICLE_MODEL],
            ),
            {
                "source_value": "TEST125",
                "resolution_action": "link",
                "model-link-vehicle_model": target.pk,
            },
        )
        self.assertEqual(response.status_code, 302)
        batch.refresh_from_db()
        self.assertEqual(batch.preview_summary["validation"]["unmapped_models"], [])
        mapping = LegacyImportMasterMapping.objects.get()
        self.assertEqual(mapping.vehicle_model, target)
        self.assertEqual(batch.rows.get(sheet_name="進貨").raw_data, original_raw)

        confirm_import(batch, "tester")
        self.assertEqual(VehicleInventory.objects.get().vehicle_model, target)
        self.assertEqual(SalesOrder.objects.get().vehicle_model, target)

    def test_quick_create_source_and_mapping_are_reused_by_next_batch(self):
        content = workbook_bytes()
        workbook = load_workbook(BytesIO(content))
        workbook["銷貨"]["AN4"] = "新合作車行"
        stream = BytesIO()
        workbook.save(stream)
        payload = stream.getvalue()
        batch = LegacyImportBatch.objects.create(
            import_type=LegacyImportBatch.ImportType.OPERATIONS,
            source_file=SimpleUploadedFile("source.xlsx", payload),
            original_filename="source.xlsx",
            file_sha256="1" * 64,
            file_size=len(payload),
            uploaded_by="tester",
        )
        build_import_preview(batch)
        self.assertEqual(
            batch.preview_summary["validation"]["unmapped_sources"],
            ["新合作車行"],
        )
        response = self.client.post(
            reverse(
                "legacy_import_master_resolve",
                args=[batch.pk, LegacyImportMasterMapping.MappingType.SALES_SOURCE],
            ),
            {
                "source_value": "新合作車行",
                "resolution_action": "create",
                "source-create-source_category": SalesSourceCategory.objects.get(
                    name="合作車行"
                ).pk,
                "source-create-name": "新合作車行",
                "source-create-address": "新北市測試路",
            },
        )
        self.assertEqual(response.status_code, 302)
        source = SalesSource.objects.get(name="新合作車行")
        self.assertEqual(source.category.name, "合作車行")
        self.assertEqual(source.source_type, SalesSource.SourceType.DEALER)
        batch.refresh_from_db()
        self.assertEqual(batch.preview_summary["validation"]["unmapped_sources"], [])

        second = LegacyImportBatch.objects.create(
            import_type=LegacyImportBatch.ImportType.OPERATIONS,
            source_file=SimpleUploadedFile("source-2.xlsx", payload),
            original_filename="source-2.xlsx",
            file_sha256="2" * 64,
            file_size=len(payload),
            uploaded_by="tester",
        )
        summary = build_import_preview(second)
        self.assertEqual(summary["validation"]["unmapped_sources"], [])
        confirm_import(second, "tester")
        order = SalesOrder.objects.get()
        self.assertEqual(order.source, source)
        self.assertEqual(order.transaction_type, SalesOrder.TransactionType.REGULAR_NEW)

    def test_quick_create_can_add_staff_category_inline(self):
        content = workbook_bytes()
        workbook = load_workbook(BytesIO(content))
        workbook["銷貨"]["AN4"] = "文傑"
        stream = BytesIO()
        workbook.save(stream)
        payload = stream.getvalue()
        batch = LegacyImportBatch.objects.create(
            import_type=LegacyImportBatch.ImportType.OPERATIONS,
            source_file=SimpleUploadedFile("staff.xlsx", payload),
            original_filename="staff.xlsx",
            file_sha256="4" * 64,
            file_size=len(payload),
            uploaded_by="tester",
        )
        build_import_preview(batch)

        response = self.client.post(
            reverse(
                "legacy_import_master_resolve",
                args=[batch.pk, LegacyImportMasterMapping.MappingType.SALES_SOURCE],
            ),
            {
                "source_value": "文傑",
                "resolution_action": "create",
                "source-create-source_category": "",
                "source-create-new_category_name": "本店員工",
                "source-create-new_category_behavior": SalesSourceCategory.SystemBehavior.STORE,
                "source-create-name": "文傑",
                "source-create-address": "",
            },
        )

        self.assertEqual(response.status_code, 302)
        source = SalesSource.objects.get(name="文傑")
        self.assertEqual(source.category.name, "本店員工")
        self.assertEqual(source.source_type, SalesSource.SourceType.STORE)

    def test_keep_historical_source_text_removes_warning_without_polluting_master(self):
        content = workbook_bytes()
        workbook = load_workbook(BytesIO(content))
        workbook["銷貨"]["AN4"] = "朋友推薦"
        stream = BytesIO()
        workbook.save(stream)
        payload = stream.getvalue()
        batch = LegacyImportBatch.objects.create(
            import_type=LegacyImportBatch.ImportType.OPERATIONS,
            source_file=SimpleUploadedFile("referral.xlsx", payload),
            original_filename="referral.xlsx",
            file_sha256="3" * 64,
            file_size=len(payload),
            uploaded_by="tester",
        )
        build_import_preview(batch)
        response = self.client.post(
            reverse(
                "legacy_import_master_resolve",
                args=[batch.pk, LegacyImportMasterMapping.MappingType.SALES_SOURCE],
            ),
            {
                "source_value": "朋友推薦",
                "resolution_action": "ignore",
                "note": "歷史分類文字，不是通路",
            },
        )
        self.assertEqual(response.status_code, 302)
        batch.refresh_from_db()
        self.assertEqual(batch.preview_summary["validation"]["unmapped_sources"], [])
        self.assertFalse(SalesSource.objects.filter(name="朋友推薦").exists())
        mapping = LegacyImportMasterMapping.objects.get()
        self.assertTrue(mapping.ignored)

    def test_master_workspace_shows_occurrence_count_and_observed_colors(self):
        batch = self.make_batch(LegacyImportBatch.ImportType.OPERATIONS)
        build_import_preview(batch)
        workspace = build_import_master_workspace(batch)
        self.assertEqual(workspace["total"], 1)
        self.assertEqual(workspace["models"][0]["source_value"], "TEST125")
        self.assertEqual(workspace["models"][0]["row_count"], 2)
        self.assertEqual(workspace["models"][0]["colors"], ["白"])
        self.assertEqual(len(workspace["models"][0]["examples"]), 2)
        sales_example = workspace["models"][0]["examples"][0]
        self.assertEqual(sales_example["context_label"], "新車銷售")
        self.assertEqual(sales_example["identifier"], "ab-123")
        self.assertEqual(sales_example["plate_number"], "ABC-1234")
        self.assertEqual(sales_example["owner_name"], "正式車主")
        response = self.client.get(reverse("legacy_import_detail", args=[batch.pk]))
        self.assertContains(response, "待補主檔工作台")
        self.assertContains(response, "Excel 出現 2 筆")
        self.assertContains(response, "查看車牌、車色與識別號碼")
        self.assertContains(response, "引擎／車身號碼")
        self.assertContains(response, "ABC-1234")

    def test_dr_z4sm_import_aliases_share_one_master_mapping(self):
        target = VehicleModel.objects.create(
            brand="SUZUKI",
            name="DR-Z4SM (滑胎版)",
            model_number="DR-Z4SM",
            energy_type=VehicleModel.EnergyType.GAS,
            model_year=2026,
            model_code=VehicleModel.ModelType.ABS_DUAL_DISC,
            displacement_cc=398,
        )

        save_import_master_mapping(
            mapping_type=LegacyImportMasterMapping.MappingType.VEHICLE_MODEL,
            source_value="DRZ-4SM",
            vehicle_model=target,
            actor_name="tester",
        )
        save_import_master_mapping(
            mapping_type=LegacyImportMasterMapping.MappingType.VEHICLE_MODEL,
            source_value="DR-Z4SM (滑胎版)",
            vehicle_model=target,
            actor_name="tester",
        )

        mapping = LegacyImportMasterMapping.objects.get()
        self.assertEqual(mapping.normalized_source_value, "dr-z4sm")
        self.assertEqual(mapping.vehicle_model_id, target.pk)

    def test_dr_z4s_import_aliases_share_one_master_mapping(self):
        target = VehicleModel.objects.create(
            brand="SUZUKI",
            name="DR-Z4S (越野版)",
            model_number="DR-Z4S",
            energy_type=VehicleModel.EnergyType.GAS,
            model_year=2026,
            model_code=VehicleModel.ModelType.ABS_DUAL_DISC,
            displacement_cc=398,
        )

        save_import_master_mapping(
            mapping_type=LegacyImportMasterMapping.MappingType.VEHICLE_MODEL,
            source_value="DRZ-4S",
            vehicle_model=target,
            actor_name="tester",
        )
        save_import_master_mapping(
            mapping_type=LegacyImportMasterMapping.MappingType.VEHICLE_MODEL,
            source_value="DR-Z4S",
            vehicle_model=target,
            actor_name="tester",
        )

        mapping = LegacyImportMasterMapping.objects.get()
        self.assertEqual(mapping.normalized_source_value, "dr-z4s")
        self.assertEqual(mapping.vehicle_model_id, target.pk)

    def test_preview_rows_search_by_linked_machine_model_number_and_identifier(self):
        batch = self.make_batch(LegacyImportBatch.ImportType.OPERATIONS)
        build_import_preview(batch)
        target = VehicleModel.objects.create(
            brand="SUZUKI",
            name="SUI 125",
            model_number="UQ125DA",
            energy_type=VehicleModel.EnergyType.GAS,
            model_year=2026,
            model_code=VehicleModel.ModelType.FRONT_DISC_REAR_DRUM,
        )
        save_import_master_mapping(
            mapping_type=LegacyImportMasterMapping.MappingType.VEHICLE_MODEL,
            source_value="TEST125",
            vehicle_model=target,
            actor_name="tester",
        )

        for query in ("SUI 125", "UQ125DA", "ab123"):
            with self.subTest(query=query):
                response = self.client.get(
                    reverse("legacy_import_detail", args=[batch.pk]),
                    {"q": query},
                )
                self.assertEqual(response.status_code, 200)
                self.assertEqual(response.context["page_obj"].paginator.count, 2)
                self.assertContains(response, "TEST125")

        empty = self.client.get(
            reverse("legacy_import_detail", args=[batch.pk]),
            {"q": "完全不存在"},
        )
        self.assertEqual(empty.context["page_obj"].paginator.count, 0)

    def test_shifted_contact_values_are_not_treated_as_vehicle_models(self):
        content = workbook_bytes()
        workbook = load_workbook(BytesIO(content))
        sales = workbook["銷貨"]
        sales["C5"] = "02"
        sales["D5"] = "26951112"
        sales["C6"] = "7000021"
        sales["D6"] = "usmartmotor@gmail.com"
        sales["C7"] = "RARE125"
        sales["D7"] = "CGA2-750464"
        stream = BytesIO()
        workbook.save(stream)
        payload = stream.getvalue()
        batch = LegacyImportBatch.objects.create(
            import_type=LegacyImportBatch.ImportType.OPERATIONS,
            source_file=SimpleUploadedFile("shifted-contact.xlsx", payload),
            original_filename="shifted-contact.xlsx",
            file_sha256="4" * 64,
            file_size=len(payload),
            uploaded_by="tester",
        )

        summary = build_import_preview(batch)

        self.assertEqual(summary["counts"]["skip"], 2)
        self.assertEqual(summary["validation"]["unmapped_models"], ["RARE125", "TEST125"])
        ignored_rows = batch.rows.filter(sheet_name="銷貨", source_row__in=[5, 6])
        self.assertEqual(ignored_rows.filter(action="skip").count(), 2)
        self.assertTrue(
            all(
                "缺少有效車輛序號且無交易資料，系統自動略過" in row.messages
                for row in ignored_rows
            )
        )
        workspace = build_import_master_workspace(batch)
        self.assertEqual(
            [item["source_value"] for item in workspace["models"]],
            ["RARE125", "TEST125"],
        )

    def test_preview_batch_can_be_deleted_with_uploaded_file(self):
        batch = self.make_batch(LegacyImportBatch.ImportType.CHANNELS)
        build_import_preview(batch)
        stored_path = Path(batch.source_file.path)
        self.assertTrue(stored_path.exists())
        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(reverse("legacy_import_delete", args=[batch.pk]))
        self.assertRedirects(response, reverse("legacy_import_list"))
        self.assertFalse(LegacyImportBatch.objects.filter(pk=batch.pk).exists())
        self.assertFalse(stored_path.exists())

    def test_completed_batch_cannot_be_deleted_but_can_be_archived_and_restored(self):
        batch = self.make_batch(LegacyImportBatch.ImportType.CHANNELS)
        build_import_preview(batch)
        confirm_import(batch, "tester")
        delete_response = self.client.post(reverse("legacy_import_delete", args=[batch.pk]))
        self.assertEqual(delete_response.status_code, 302)
        self.assertTrue(LegacyImportBatch.objects.filter(pk=batch.pk).exists())
        self.client.post(reverse("legacy_import_archive", args=[batch.pk]))
        batch.refresh_from_db()
        self.assertIsNotNone(batch.archived_at)
        self.client.post(reverse("legacy_import_restore", args=[batch.pk]))
        batch.refresh_from_db()
        self.assertIsNone(batch.archived_at)

    def test_excluded_sales_row_is_not_treated_as_previously_imported_next_time(self):
        first = self.make_batch(LegacyImportBatch.ImportType.OPERATIONS)
        build_import_preview(first)
        sales_row = first.rows.get(sheet_name="銷貨")
        apply_import_row_decision(
            sales_row,
            {},
            LegacyImportCorrection.Decision.EXCLUDE,
            "本次先不匯入銷貨",
            "tester",
        )
        confirm_import(first, "tester")
        second = self.make_batch(LegacyImportBatch.ImportType.OPERATIONS)
        build_import_preview(second)
        self.assertEqual(
            second.rows.get(sheet_name="銷貨").action,
            "create",
        )

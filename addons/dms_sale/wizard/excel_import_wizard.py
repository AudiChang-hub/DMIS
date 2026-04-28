import base64
import io
import logging
from datetime import date, datetime

from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

# Excel 欄位索引（0-based）
COL = {
    'excel_sync_id':        0,   # A 序號
    'registration_date':    1,   # B 領牌日期
    'product_sku':          2,   # C 車種型號
    'vin_or_engine':        3,   # D 油：引擎號碼、電：車身號碼
    'customer_name':        4,   # E 車主名稱
    'color_name':           5,   # F 顏色
    'received_amount':      6,   # G 收款價
    'cost':                 7,   # H 成本
    # I 月份 → 不匯入
    'cash':                 9,   # J 現金
    'credit_card':         10,   # K 信用卡
    'out_credit_card_fee': 11,   # L 信用卡手續費支出
    'out_installment_fee': 12,   # M 分期手續費支出
    'out_plate_tax':       13,   # N 領牌稅金支出
    'out_compulsory_ins':  14,   # O 強制險支出
    'out_plate_select':    15,   # P 選號支出
    'out_used_car':        16,   # Q 中古車支出
    'out_gift_shipping':   17,   # R 贈品、運費支出
    'out_dealer_commission': 18, # S 車行傭金支出
    'out_friendly_dealer_bonus': 19,  # T 友善車行獎金支出
    'out_first_sale_bonus': 20,  # U 首賣獎金支出
    'out_unit_bonus':      21,   # V 台數獎金支出
    'in_plate_tax':        22,   # W 領牌稅金收入
    'in_compulsory_ins':   23,   # X 強制險收入
    'in_agency_fee':       24,   # Y 代辦費收入
    'in_scrap_agency':     25,   # Z 報廢代辦收入
    'in_plate_select':     26,   # AA 選號收入
    'in_used_car':         27,   # AB 中古車收入
    'in_scrap_car':        28,   # AC 報廢車收入
    'in_card_installment_fee': 29, # AD 刷卡、分期手續費收入
    'in_yamaha_bonus':     30,   # AE 山葉獎金收入
    'in_friendly_dealer_bonus': 31, # AF 友善車行獎金收入
    'in_other':            32,   # AG 其他收入
    'in_actual_sales_bonus': 33, # AH 實銷獎勵金
    'in_promo_subsidy':    34,   # AI 促銷補助金
    'in_installment_subsidy': 35, # AJ 分期補貼息
    'in_compulsory_ins_commission': 36, # AK 強制險傭金
    'in_credit_card_commission': 37,    # AL 信用卡傭金
    # AM 單筆淨利 → 計算欄位，不匯入
    'dealer_name':         39,   # AN 車行
    'dealer_receipt_flag': 40,   # AO 車行收款（V=已結清）
    'finance_company':     41,   # AP 分期公司
    'installment_periods': 42,   # AQ 期數
    # AR 領牌日期2 → 不匯入（與B重複）
    'plate_number':        44,   # AS 車牌號碼
    # AT 車主名稱2 → 不匯入（與E重複）
    'birthday_ad':         46,   # AU 生日
    # AV 民國生日 → 計算欄位，不匯入
    'id_number':           48,   # AW 身分證字號
    'address_registered':  49,   # AX 戶籍地址
    'customer_phone':      50,   # AY 手機
    'customer_email':      51,   # AZ Email
    'moea_invoice_no':     52,   # BA 工業局發票號碼
    'moea_invoice_date':   53,   # BB 發票日期
    'balance_invoice_no':  54,   # BC 尾款發票號碼（實為Balance Invoice）
    'subsidy_plan':        55,   # BD 補助方案（注意：BD=subsidy_plan? 確認欄56=subsidy_amount?）
    'subsidy_amount':      56,   # BE 補助金額
    'remittance_account':  57,   # BF 銀行欄（名稱寫反→實為帳號）
    'remittance_bank':     58,   # BG 匯款帳戶欄（名稱寫反→實為銀行）
    'subsidy_moenv_date':  59,   # BH 申請日 → 環境部申請日
    'subsidy_boie_status': 60,   # BI 工業局申請狀態（V 或文字）
    'subsidy_moenv_status': 61,  # BJ 環境部申請狀態（V 或文字）
    'subsidy_city_status': 62,   # BK 縣市政府申請狀態（V 或文字）
    'used_car_owner':      63,   # BL 舊車車主
    'used_car_owner_id_no': 64,  # BM 舊車車主身分證
    'used_car_plate':      65,   # BN 舊車牌照號碼
    'used_car_engine_no':  66,   # BO 舊車引擎號碼
    'used_car_brand':      67,   # BP 舊車廠牌
    'used_car_displacement': 68, # BQ 排氣量
    'used_car_manufacture_date': 69, # BR 出廠日期
    'used_car_scrap_date': 70,   # BS 報廢日期
    'used_car_recycle_date': 71, # BT 回收日期
    'used_car_owner_address': 72, # BU 舊車戶籍
    'used_car_owner_phone': 73,  # BV 舊車車主電話
    'ev_control_account':  74,   # BW 車控帳號
    'ev_control_password': 75,   # BX 車控密碼
    'ev_battery_plan':     76,   # BY 電池合約方案
    'ev_battery_start_date': 77, # BZ 電池合約啟用日期
    'ev_battery_account':  78,   # CA 電池合約帳號
    'ev_battery_password': 79,   # CB 電池合約密碼
    'extra_helmet':        80,   # CC 安全帽
    'extra_gift_voucher':  81,   # CD 公司禮卷、匯款
    'extra_other':         82,   # CE 其他
    'extra_platform_gift': 83,   # CF 平台贈品
    # CG 欄1（空） → 不匯入
    'note':                85,   # CG→實為CH 備註
    'order_date':          86,   # CH→CI 訂單日期
    'extra_company_gift':  87,   # CI→CJ 公司贈品
    'extra_customer_service_phone': 88, # CJ→CK 客服電話
    'install_info':        89,   # CK→CL 分期資訊（append到備註）
    'extra_special_plan':  90,   # CL→CM 特殊方案
    # CM 領牌年月 → 不匯入
}

# 工業局/環境部/縣市政府補助金額欄（能轉數字才填）
SUBSIDY_AMT_COL = {
    'subsidy_moea':  60,  # BI 工業局
    'subsidy_moenv': 61,  # BJ 環境部
    'subsidy_local': 62,  # BK 縣市政府
}

# 特自字行名稱 → 店面
STORE_DEALER_NAMES = {'馭盛', ''}


def _to_float(val):
    """嘗試轉成 float，失敗回 0.0"""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _to_date(val):
    """datetime/date/str → date，失敗回 None"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        val = val.strip()
        for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%d/%m/%Y'):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                continue
    return None


def _to_str(val):
    """清理成字串，None/空/nan → None"""
    if val is None:
        return None
    s = str(val).strip()
    if s == '' or s.lower() == 'none' or s.lower() == 'nan':
        return None
    return s


def _cell(row, idx):
    """安全取 row[idx]"""
    try:
        return row[idx]
    except IndexError:
        return None


class ExcelImportWizard(models.TransientModel):
    _name = 'dms.excel.import.wizard'
    _description = 'Excel 銷貨資料匯入'

    file_data = fields.Binary(string='Excel 檔案', required=True)
    file_name = fields.Char(string='檔名')

    # 預覽結果
    preview_insert = fields.Integer(string='將新增', readonly=True)
    preview_update = fields.Integer(string='將更新', readonly=True)
    preview_skip = fields.Integer(string='略過（無序號）', readonly=True)
    preview_done = fields.Boolean(default=False)

    # 錯誤摘要
    error_summary = fields.Text(string='警告/錯誤', readonly=True)

    def _parse_excel(self):
        """讀取 Excel，回傳 list of row tuple（第4列起）"""
        try:
            import openpyxl
        except ImportError:
            raise UserError('伺服器缺少 openpyxl 套件，請聯絡管理員安裝。')

        raw = base64.b64decode(self.file_data)
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)

        if '銷貨' not in wb.sheetnames:
            raise UserError('Excel 中找不到「銷貨」頁籤。')

        ws = wb['銷貨']
        rows = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            rows.append(row)
        return rows

    def _build_vals(self, row, errors):
        """將一列資料轉換成 dms.sale.order vals dict"""
        vals = {}

        # ── 來源序號（主鍵）
        sync_id = _to_str(_cell(row, COL['excel_sync_id']))
        if not sync_id:
            return None  # 無序號 → 跳過
        vals['excel_sync_id'] = sync_id
        vals['sale_origin'] = 'excel'

        # ── 訂單日期
        order_date = _to_date(_cell(row, COL['order_date']))
        vals['order_date'] = order_date or date.today()

        # ── 狀態（歷史資料全部已成立）
        vals['state'] = 'confirmed'

        # ── 客戶資訊
        vals['customer_name'] = _to_str(_cell(row, COL['customer_name'])) or '（未填）'
        vals['customer_phone'] = _to_str(_cell(row, COL['customer_phone']))
        vals['customer_email'] = _to_str(_cell(row, COL['customer_email']))
        vals['id_number'] = _to_str(_cell(row, COL['id_number']))
        vals['birthday_ad'] = _to_date(_cell(row, COL['birthday_ad']))
        vals['address_registered'] = _to_str(_cell(row, COL['address_registered']))

        # ── 車款（SKU 查找，以 model 欄位比對）
        sku = _to_str(_cell(row, COL['product_sku']))
        if sku:
            product = self.env['dms.product'].search(
                [('model', '=', sku)], limit=1)
            if product:
                vals['product_id'] = product.id
                # 顏色：永遠寫 source_color_name，有找到再連結 color_id
                color_name = _to_str(_cell(row, COL['color_name']))
                if color_name:
                    vals['source_color_name'] = color_name
                    color = self.env['dms.product.color'].search(
                        [('product_id', '=', product.id),
                         ('name', '=', color_name)], limit=1)
                    if color:
                        vals['color_id'] = color.id
                # 引擎/車身號碼
                vin = _to_str(_cell(row, COL['vin_or_engine']))
                if vin:
                    if product.energy_type == 'electric':
                        vals['frame_number'] = vin
                    else:
                        vals['engine_number'] = vin
            else:
                vals['source_product_name'] = sku
                # 車款找不到時，顏色仍備存原始字串
                color_name = _to_str(_cell(row, COL['color_name']))
                if color_name:
                    vals['source_color_name'] = color_name
                errors.append(f"序號 {sync_id}：SKU「{sku}」找不到對應車款，車款留空")
        else:
            # 無SKU但有引擎/車身號碼
            vin = _to_str(_cell(row, COL['vin_or_engine']))
            if vin:
                vals['engine_number'] = vin

        # ── 車輛資訊
        vals['registration_date'] = _to_date(_cell(row, COL['registration_date']))
        vals['plate_number'] = _to_str(_cell(row, COL['plate_number']))

        # ── 車行與交易類型
        dealer_name = _to_str(_cell(row, COL['dealer_name']))
        if not dealer_name or dealer_name in STORE_DEALER_NAMES:
            vals['sale_type'] = 'store'
        else:
            # 車行名稱採「包含」比對（ilike），不要求完全一致：
            # 1) Excel 端與主檔可能存在簡稱 / 全名差異（如「ABC」vs「ABC 機車行」）。
            # 2) 英文名稱大小寫常不一致；ilike 為大小寫不敏感，可一併涵蓋。
            # 比對策略：以 Excel 名稱當 needle，搜尋主檔 name 包含該字串的候選，
            # 取最短 name 的候選（最接近 Excel 字串），避免被過長 / 過泛的名稱誤對。
            candidates = self.env['dms.dealer'].search(
                [('name', 'ilike', dealer_name)])
            dealer = self.env['dms.dealer'].browse()
            if candidates:
                # 優先挑選 strip().upper() 完全相等者；否則取 name 最短的候選。
                target_upper = dealer_name.strip().upper()
                exact = candidates.filtered(
                    lambda c: (c.name or '').strip().upper() == target_upper)
                if exact:
                    dealer = exact[0]
                else:
                    dealer = sorted(
                        candidates, key=lambda c: len((c.name or '').strip()))[0]
            if dealer:
                vals['sale_type'] = 'dealer'
                vals['dealer_id'] = dealer.id
            else:
                vals['sale_type'] = 'dealer'
                vals['source_dealer_name'] = dealer_name
                errors.append(f"序號 {sync_id}：車行「{dealer_name}」找不到，已記錄原始名稱")

        # ── 付款方式
        periods = int(_to_float(_cell(row, COL['installment_periods'])))
        credit = _to_float(_cell(row, COL['credit_card']))
        cash = _to_float(_cell(row, COL['cash']))
        fin_co = _to_str(_cell(row, COL['finance_company']))
        credit_card_fee = _to_float(_cell(row, COL['out_credit_card_fee']))
        if periods > 0 or fin_co:
            vals['payment_method'] = 'installment'
        elif credit > 0 or credit_card_fee > 0:
            vals['payment_method'] = 'credit'
        else:
            vals['payment_method'] = 'cash'
        vals['installment_periods'] = periods

        # 分期公司
        if fin_co:
            known = {'和潤', '遠信', '仲信'}
            if fin_co in known:
                vals['finance_company'] = fin_co
            else:
                vals['finance_company'] = 'other'
                vals['finance_company_other'] = fin_co

        # ── 收款價 / 總成交價
        recv = _to_float(_cell(row, COL['received_amount']))
        vals['received_amount'] = recv
        vals['amount_total'] = recv

        # ── 成本
        vals['cost'] = _to_float(_cell(row, COL['cost']))

        # ── 已結清
        dealer_receipt = _to_str(_cell(row, COL['dealer_receipt_flag']))
        if dealer_receipt and dealer_receipt.strip().upper() == 'V':
            vals['is_settled'] = True
            vals['settle_date'] = vals.get('registration_date')

        # ── 收益統計：支出
        for field, col in [
            ('out_credit_card_fee', COL['out_credit_card_fee']),
            ('out_installment_fee', COL['out_installment_fee']),
            ('out_plate_tax', COL['out_plate_tax']),
            ('out_compulsory_ins', COL['out_compulsory_ins']),
            ('out_plate_select', COL['out_plate_select']),
            ('out_used_car', COL['out_used_car']),
            ('out_gift_shipping', COL['out_gift_shipping']),
            ('out_dealer_commission', COL['out_dealer_commission']),
            ('out_friendly_dealer_bonus', COL['out_friendly_dealer_bonus']),
            ('out_first_sale_bonus', COL['out_first_sale_bonus']),
            ('out_unit_bonus', COL['out_unit_bonus']),
        ]:
            vals[field] = _to_float(_cell(row, col))

        # ── 收益統計：收入
        for field, col in [
            ('in_plate_tax', COL['in_plate_tax']),
            ('in_compulsory_ins', COL['in_compulsory_ins']),
            ('in_agency_fee', COL['in_agency_fee']),
            ('in_scrap_agency', COL['in_scrap_agency']),
            ('in_plate_select', COL['in_plate_select']),
            ('in_used_car', COL['in_used_car']),
            ('in_scrap_car', COL['in_scrap_car']),
            ('in_card_installment_fee', COL['in_card_installment_fee']),
            ('in_yamaha_bonus', COL['in_yamaha_bonus']),
            ('in_friendly_dealer_bonus', COL['in_friendly_dealer_bonus']),
            ('in_other', COL['in_other']),
            ('in_actual_sales_bonus', COL['in_actual_sales_bonus']),
            ('in_promo_subsidy', COL['in_promo_subsidy']),
            ('in_installment_subsidy', COL['in_installment_subsidy']),
            ('in_compulsory_ins_commission', COL['in_compulsory_ins_commission']),
            ('in_credit_card_commission', COL['in_credit_card_commission']),
        ]:
            vals[field] = _to_float(_cell(row, col))

        # ── 發票與補助申辦
        vals['moea_invoice_no'] = _to_str(_cell(row, COL['moea_invoice_no']))
        vals['moea_invoice_date'] = _to_date(_cell(row, COL['moea_invoice_date']))
        vals['balance_invoice_no'] = _to_str(_cell(row, COL['balance_invoice_no']))
        vals['subsidy_plan'] = _to_str(_cell(row, COL['subsidy_plan']))
        vals['subsidy_amount'] = _to_float(_cell(row, COL['subsidy_amount']))
        vals['remittance_account'] = _to_str(_cell(row, COL['remittance_account']))
        vals['remittance_bank'] = _to_str(_cell(row, COL['remittance_bank']))
        vals['subsidy_moenv_date'] = _to_date(_cell(row, COL['subsidy_moenv_date']))

        # 補助金額：能轉數字才填
        for field, col in SUBSIDY_AMT_COL.items():
            raw_val = _cell(row, col)
            try:
                vals[field] = float(raw_val) if raw_val is not None else 0.0
            except (ValueError, TypeError):
                vals[field] = 0.0

        # 補助申請狀態（文字）
        vals['subsidy_boie_status'] = _to_str(_cell(row, COL['subsidy_boie_status']))
        vals['subsidy_moenv_status'] = _to_str(_cell(row, COL['subsidy_moenv_status']))
        vals['subsidy_city_status'] = _to_str(_cell(row, COL['subsidy_city_status']))

        # ── 舊車資訊
        used_car_fields = {
            'used_car_owner': COL['used_car_owner'],
            'used_car_owner_id_no': COL['used_car_owner_id_no'],
            'used_car_plate': COL['used_car_plate'],
            'used_car_engine_no': COL['used_car_engine_no'],
            'used_car_brand': COL['used_car_brand'],
            'used_car_displacement': COL['used_car_displacement'],
            'used_car_owner_address': COL['used_car_owner_address'],
            'used_car_owner_phone': COL['used_car_owner_phone'],
        }
        has_used_car = False
        for field, col in used_car_fields.items():
            v = _to_str(_cell(row, col))
            vals[field] = v
            if v:
                has_used_car = True

        vals['used_car_manufacture_date'] = _to_date(_cell(row, COL['used_car_manufacture_date']))
        vals['used_car_scrap_date'] = _to_date(_cell(row, COL['used_car_scrap_date']))
        vals['used_car_recycle_date'] = _to_date(_cell(row, COL['used_car_recycle_date']))
        if vals['used_car_manufacture_date'] or vals['used_car_scrap_date'] or vals['used_car_recycle_date']:
            has_used_car = True
        vals['is_trade_in'] = has_used_car

        # ── 電車資訊
        vals['ev_control_account'] = _to_str(_cell(row, COL['ev_control_account']))
        vals['ev_control_password'] = _to_str(_cell(row, COL['ev_control_password']))
        vals['ev_battery_plan'] = _to_str(_cell(row, COL['ev_battery_plan']))
        vals['ev_battery_start_date'] = _to_date(_cell(row, COL['ev_battery_start_date']))
        vals['ev_battery_account'] = _to_str(_cell(row, COL['ev_battery_account']))
        vals['ev_battery_password'] = _to_str(_cell(row, COL['ev_battery_password']))

        # ── 其他資訊
        vals['extra_helmet'] = _to_str(_cell(row, COL['extra_helmet']))
        vals['extra_gift_voucher'] = _to_str(_cell(row, COL['extra_gift_voucher']))
        vals['extra_other'] = _to_str(_cell(row, COL['extra_other']))
        vals['extra_platform_gift'] = _to_str(_cell(row, COL['extra_platform_gift']))
        vals['extra_company_gift'] = _to_str(_cell(row, COL['extra_company_gift']))
        vals['extra_customer_service_phone'] = _to_str(_cell(row, COL['extra_customer_service_phone']))
        vals['extra_special_plan'] = _to_str(_cell(row, COL['extra_special_plan']))

        # ── 備註（主備註 + append 分期資訊）
        note = _to_str(_cell(row, COL['note']))
        install_info = _to_str(_cell(row, COL['install_info']))
        if install_info:
            note = f"{note}\n【分期資訊】{install_info}" if note else f"【分期資訊】{install_info}"
        vals['note'] = note

        return vals

    def action_preview(self):
        """解析 Excel，計算新增/更新筆數"""
        rows = self._parse_excel()
        errors = []
        insert_count = 0
        update_count = 0
        skip_count = 0

        existing_ids = set(
            self.env['dms.sale.order'].search([
                ('excel_sync_id', '!=', False)
            ]).mapped('excel_sync_id')
        )

        for row in rows:
            sync_id = _to_str(_cell(row, COL['excel_sync_id']))
            if not sync_id:
                skip_count += 1
                continue
            if sync_id in existing_ids:
                update_count += 1
            else:
                insert_count += 1

        self.write({
            'preview_insert': insert_count,
            'preview_update': update_count,
            'preview_skip': skip_count,
            'preview_done': True,
            'error_summary': '\n'.join(errors) if errors else False,
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'dms.excel.import.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_import(self):
        """執行匯入"""
        rows = self._parse_excel()
        errors = []
        inserted = 0
        updated = 0
        skipped = 0

        SaleOrder = self.env['dms.sale.order']

        existing = {
            r.excel_sync_id: r
            for r in SaleOrder.search([('excel_sync_id', '!=', False)])
        }

        for i, row in enumerate(rows, start=4):
            try:
                vals = self._build_vals(row, errors)
            except Exception as e:
                errors.append(f"第 {i} 列解析失敗：{e}")
                continue

            if vals is None:
                skipped += 1
                continue

            sync_id = vals['excel_sync_id']

            if sync_id in existing:
                try:
                    existing[sync_id].write(vals)
                    updated += 1
                except Exception as e:
                    errors.append(f"序號 {sync_id} 更新失敗：{e}")
            else:
                try:
                    SaleOrder.create(vals)
                    inserted += 1
                except Exception as e:
                    errors.append(f"序號 {sync_id} 新增失敗：{e}")

        summary = f"✅ 匯入完成：新增 {inserted} 筆，更新 {updated} 筆，略過 {skipped} 筆。"
        if errors:
            summary += f"\n\n⚠️ 警告（共 {len(errors)} 筆）：\n" + '\n'.join(errors[:50])
            if len(errors) > 50:
                summary += f"\n... 共 {len(errors)} 筆警告（僅顯示前50筆）"

        self.write({
            'preview_insert': inserted,
            'preview_update': updated,
            'preview_skip': skipped,
            'error_summary': summary,
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'dms.excel.import.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader, PdfWriter
from reportlab.lib.colors import Color
from reportlab.lib.pagesizes import A4, A5, landscape
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas

from sales.services.order_contract_pdf import FONT_REGULAR


PRINT_FIELD_CHOICES = (
    ("owner_name", "車主姓名／公司名稱"),
    ("owner_id_number", "身分證字號／統一編號"),
    ("owner_birthday", "西元生日"),
    ("owner_roc_birthday", "民國生日"),
    ("owner_phone", "手機"),
    ("owner_email", "Email"),
    ("owner_address", "戶籍／公司地址"),
    ("order_number", "訂單編號"),
    ("order_date", "訂單日期"),
    ("source_name", "車行／平台／本店"),
    ("vehicle_brand_model", "品牌車型"),
    ("vehicle_model_number", "純型號"),
    ("vehicle_year", "車輛年份"),
    ("vehicle_color", "顏色"),
    ("vehicle_identifier", "引擎／車身號碼"),
    ("manufactured_year_month", "出廠年月"),
    ("plate_number", "車牌號碼"),
    ("registration_date", "實際領牌日期"),
    ("registration_county", "領牌縣市"),
    ("vehicle_price", "車款售價"),
    ("plate_insurance_fee", "牌險費"),
    ("final_receivable", "訂單應收總額"),
    ("subsidy_total", "補助總額"),
    ("trade_in_plate", "舊車牌照號碼"),
    ("old_owner_name", "舊車車主"),
    ("old_owner_id_number", "舊車主身分證"),
    ("delivery_destination", "自送／託運地點"),
    ("invoice_date", "發票日期"),
    ("invoice_number", "尾款發票號碼"),
    ("today", "列印日期"),
)


def template_page_size(template):
    if template.paper_size == template.PaperSize.A5:
        size = A5
    elif template.paper_size == template.PaperSize.CUSTOM:
        size = (float(template.width_mm) * mm, float(template.height_mm) * mm)
    else:
        size = A4
    return landscape(size) if template.orientation == template.Orientation.LANDSCAPE else size


def _date(value):
    return value.strftime("%Y/%m/%d") if value else ""


def _roc_date(value):
    return f"民國{value.year - 1911}年{value.month}月{value.day}日" if value else ""


def _money(value):
    return f"{value:,.0f}" if value is not None else ""


def order_print_values(order, today=None):
    from django.utils import timezone

    today = today or timezone.localdate()
    vehicle = order.allocated_vehicle
    model = order.vehicle_model
    source_name = order.source.name if order.source_id else "馭盛"
    profile = getattr(order, "operations", None)
    return {
        "owner_name": order.owner_name,
        "owner_id_number": order.owner_id_number,
        "owner_birthday": _date(order.owner_birth_date),
        "owner_roc_birthday": _roc_date(order.owner_birth_date),
        "owner_phone": order.owner_phone,
        "owner_email": order.owner_email,
        "owner_address": order.owner_address,
        "order_number": order.number,
        "order_date": _date(order.order_date),
        "source_name": source_name,
        "vehicle_brand_model": f"{model.brand} {model.name}".strip(),
        "vehicle_model_number": model.model_number,
        "vehicle_year": str(model.model_year or ""),
        "vehicle_color": order.color.name,
        "vehicle_identifier": vehicle.identifier if vehicle else "",
        "manufactured_year_month": vehicle.manufactured_year_month if vehicle else "",
        "plate_number": order.final_plate_number,
        "registration_date": _date(order.registration_date),
        "registration_county": order.registration_county,
        "vehicle_price": _money(order.vehicle_price),
        "plate_insurance_fee": _money(order.plate_insurance_fee),
        "final_receivable": _money(order.actual_balance + order.deposit_amount),
        "subsidy_total": _money(order.subsidy_total),
        "trade_in_plate": order.trade_in_plate,
        "old_owner_name": order.old_owner_name,
        "old_owner_id_number": order.old_owner_id_number,
        "delivery_destination": order.delivery_destination,
        "invoice_date": _date(getattr(profile, "invoice_date", None)),
        "invoice_number": getattr(profile, "balance_invoice_number", ""),
        "today": _date(today),
    }


def _color(value, default=None):
    if not isinstance(value, str) or not value or value in {"00000000", "000000"}:
        return default
    value = value[-6:]
    try:
        return Color(int(value[0:2], 16) / 255, int(value[2:4], 16) / 255, int(value[4:6], 16) / 255)
    except ValueError:
        return default


def _excel_background(path, page_size):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    min_row, min_col, max_row, max_col = 1, 1, ws.max_row, ws.max_column
    if ws.print_area:
        try:
            area = list(ws.print_area.ranges)[0]
            min_col, min_row, max_col, max_row = area.bounds
        except (IndexError, TypeError, ValueError):
            pass
    col_widths = []
    for col in range(min_col, max_col + 1):
        letter = ws.cell(min_row, col).column_letter
        width = ws.column_dimensions[letter].width or 8.43
        col_widths.append(max(float(width) * 5.25, 8))
    row_heights = [float(ws.row_dimensions[row].height or 15) for row in range(min_row, max_row + 1)]
    page_w, page_h = page_size
    margin = 8 * mm
    scale = min((page_w - 2 * margin) / sum(col_widths), (page_h - 2 * margin) / sum(row_heights))
    x_positions = [margin]
    for width in col_widths:
        x_positions.append(x_positions[-1] + width * scale)
    y_positions = [page_h - margin]
    for height in row_heights:
        y_positions.append(y_positions[-1] - height * scale)
    merged = {}
    covered = set()
    for area in ws.merged_cells.ranges:
        if area.max_row < min_row or area.min_row > max_row or area.max_col < min_col or area.min_col > max_col:
            continue
        merged[(area.min_row, area.min_col)] = area
        for row in range(area.min_row, area.max_row + 1):
            for col in range(area.min_col, area.max_col + 1):
                if (row, col) != (area.min_row, area.min_col):
                    covered.add((row, col))
    output = BytesIO()
    c = canvas.Canvas(output, pagesize=page_size)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            if (row, col) in covered:
                continue
            cell = ws.cell(row, col)
            area = merged.get((row, col))
            end_row = area.max_row if area else row
            end_col = area.max_col if area else col
            x = x_positions[col - min_col]
            right = x_positions[end_col - min_col + 1]
            top = y_positions[row - min_row]
            bottom = y_positions[end_row - min_row + 1]
            fill = _color(getattr(cell.fill.fgColor, "rgb", None))
            if fill:
                c.setFillColor(fill)
                c.rect(x, bottom, right - x, top - bottom, stroke=0, fill=1)
            c.setStrokeColor(Color(0.72, 0.75, 0.73))
            c.setLineWidth(0.35)
            c.rect(x, bottom, right - x, top - bottom, stroke=1, fill=0)
            value = "" if cell.value is None else str(cell.value)
            if value:
                font_size = max(5.5, min(float(cell.font.sz or 10) * scale, 14))
                c.setFont("MSJH", font_size)
                c.setFillColor(_color(getattr(cell.font.color, "rgb", None), Color(0.08, 0.1, 0.09)))
                text_y = bottom + max(1.5, (top - bottom - font_size) / 2)
                align = cell.alignment.horizontal or "left"
                if align == "center":
                    c.drawCentredString((x + right) / 2, text_y, value[:120])
                elif align == "right":
                    c.drawRightString(right - 2, text_y, value[:120])
                else:
                    c.drawString(x + 2, text_y, value[:120])
    c.save()
    output.seek(0)
    return output


def _background_pdf(template, page_size):
    if not template.background_file:
        output = BytesIO()
        c = canvas.Canvas(output, pagesize=page_size)
        c.showPage()
        c.save()
        output.seek(0)
        return output
    path = Path(template.background_file.path)
    extension = path.suffix.lower()
    if extension in {".xlsx", ".xlsm"}:
        return _excel_background(path, page_size)
    if extension == ".pdf":
        return open(path, "rb")
    if extension in {".jpg", ".jpeg", ".png", ".webp"}:
        output = BytesIO()
        c = canvas.Canvas(output, pagesize=page_size)
        c.drawImage(str(path), 0, 0, width=page_size[0], height=page_size[1], preserveAspectRatio=True, anchor="c")
        c.showPage()
        c.save()
        output.seek(0)
        return output
    raise ValueError("背景檔案僅支援 Excel、PDF、JPG、PNG 或 WebP。")


def build_positioned_template_pdf(template, order):
    page_size = template_page_size(template)
    page_w, page_h = page_size
    values = order_print_values(order)
    overlay = BytesIO()
    c = canvas.Canvas(overlay, pagesize=page_size)
    offset_x = float(template.printer_offset_x_mm) * mm
    offset_y = float(template.printer_offset_y_mm) * mm
    for field in template.fields.filter(active=True):
        value = f"{field.prefix}{values.get(field.field_key, '')}{field.suffix}"
        if not value:
            continue
        x = float(field.x_mm) * mm + offset_x
        y = page_h - float(field.y_mm) * mm - float(field.font_size) + offset_y
        width = float(field.width_mm) * mm
        font_size = float(field.font_size)
        c.setFont("MSJH", font_size)
        c.setFillColor(Color(0.05, 0.08, 0.07))
        if field.alignment == field.Alignment.CENTER:
            c.drawCentredString(x + width / 2, y, value)
        elif field.alignment == field.Alignment.RIGHT:
            c.drawRightString(x + width, y, value)
        else:
            c.drawString(x, y, value)
    c.showPage()
    c.save()
    overlay.seek(0)

    background_stream = _background_pdf(template, page_size)
    background_reader = PdfReader(background_stream)
    page = background_reader.pages[0]
    page.scale_to(page_w, page_h)
    page.merge_page(PdfReader(overlay).pages[0])
    writer = PdfWriter()
    writer.add_page(page)
    output = BytesIO()
    writer.write(output)
    output.seek(0)
    if hasattr(background_stream, "close"):
        background_stream.close()
    return output

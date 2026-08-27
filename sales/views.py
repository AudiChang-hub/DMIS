import json
import logging
import shutil
import uuid
from functools import wraps
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
from pathlib import Path

from django.contrib import messages
from django.contrib.auth import get_user_model, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.sessions.models import Session
from django.conf import settings
from django.core.files.base import ContentFile
from django.core.exceptions import PermissionDenied, ValidationError
from django.db import IntegrityError, connection, transaction
from django.db.models import Count, DecimalField, Exists, OuterRef, Prefetch, Q, Subquery, Sum
from django.core.paginator import Paginator
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils.http import url_has_allowed_host_and_scheme
from django.utils import timezone
from django.views.decorators.http import require_http_methods
import django_rq
from rq import Retry, Worker

from .services.order_contract_pdf import build_order_contract_pdf
from .services.privacy_consent_pdf import build_privacy_consent_pdf
from .services.excel_export import sanitize_excel_row
from .forms import (
    AccessoryProductForm,
    AdminPasswordResetForm,
    AdminUserCreateForm,
    AdminUserEditForm,
    AccessoryFormSet,
    AllocationForm,
    BusinessHolidayForm,
    BrandRegistrationFeeRuleForm,
    CancellationRequestForm,
    DealerVolumeBonusAdjustmentForm,
    DealerVolumeBonusRuleForm,
    DealerVolumeBonusSettlementForm,
    DealerVolumeBonusTierFormSet,
    DiscountDecisionForm,
    DiscountRequestForm,
    DeliveryCompletionForm,
    DeliveryPaymentForm,
    InstallmentCompanyForm,
    InstallmentPlanOptionFormSet,
    InstallmentPlanVersionForm,
    LegacyImportUploadForm,
    LegacyImportRowCorrectionForm,
    LegacySalesSourceLinkForm,
    LegacySalesSourceQuickCreateForm,
    LegacyVehicleModelLinkForm,
    LegacyVehicleModelQuickCreateForm,
    OtherFeeFormSet,
    OrderOperationsForm,
    OrderEditForm,
    PrivacyConsentForm,
    PaymentRecordFormSet,
    PositionedPrintFieldFormSet,
    PositionedPrintTemplateForm,
    ReconciliationRecordForm,
    RefundCompletionForm,
    QuickInventoryEntryFormSet,
    ReallocationForm,
    RegistrationDocumentUploadForm,
    RegistrationStageForm,
    RequiredPasswordChangeForm,
    SalesOrderForm,
    SalesSourceBrandPolicyFormSet,
    SalesSourceCategoryForm,
    SalesSourceCooperationForm,
    SalesSourceCooperationProfileForm,
    SalesSourceForm,
    SignedContractForm,
    SubsidyDataForm,
    SubsidyItemFormSet,
    SubsidyDocumentUploadForm,
    VehicleInventoryForm,
    VehicleBrandForm,
    VehicleColorMasterFormSet,
    VehicleIncentiveRuleForm,
    VehicleModelCommissionForm,
    VehicleModelFamilyMoveForm,
    VehicleModelMasterForm,
    VehicleModelVersionMergeForm,
    VehicleModelYearCorrectionForm,
    VehiclePriceVersionForm,
    VehicleSettlementCostRuleForm,
)
from .models import (
    AccessoryProduct,
    BusinessHoliday,
    BrandRegistrationFeeRule,
    DealerVolumeBonusRule,
    DealerVolumeBonusSettlement,
    DeliveryRecord,
    InstallmentCompany,
    InstallmentPlanVersion,
    LegacyImportBatch,
    LegacyImportMasterMapping,
    LegacyImportRow,
    OrderEvent,
    OrderOperationsProfile,
    PaymentRecord,
    PositionedPrintTemplate,
    OrderChange,
    OrderDraft,
    IdOcrJob,
    RegistrationDocument,
    SalesOrder,
    SalesOrderSearchIndex,
    SalesSource,
    SalesSourceBrandPolicy,
    SalesSourceCategory,
    SalesSourceCooperationProfile,
    Store,
    SubsidyDocument,
    VehicleColor,
    VehicleBrand,
    VehicleInventory,
    VehicleInventoryHistory,
    VehicleIncentiveRule,
    VehicleModel,
    VehicleModelFamily,
    VehiclePriceVersion,
    VehicleSettlementCostRule,
    UserAccountAuditLog,
    UserAppearancePreference,
    UserSecurityProfile,
    normalize_legacy_master_value,
    normalize_vehicle_model_master_value,
    normalize_vehicle_identifier,
)
from .themes import DEFAULT_THEME, THEME_VALUES
from .services.vehicle_brands import (
    rename_vehicle_brand_references,
    vehicle_brand_search_q,
    vehicle_brand_is_used,
)
from .services.vehicle_model_family import (
    correct_vehicle_model_year,
    delete_unused_vehicle_model,
    merge_vehicle_model_versions,
    move_vehicle_model_to_family,
    rename_vehicle_model_family,
    vehicle_model_delete_blockers,
    vehicle_model_relation_summary,
)
from .jobs import delete_id_ocr_job_files, run_id_ocr_job, run_legacy_import_job
from .services.id_ocr import recognize_id_card
from .services.order_change_display import build_order_change_cards
from .services.order_next_actions import build_order_next_actions
from .services.order_search import (
    build_order_match_summary,
    build_order_search_query,
)
from .services.operations_sync import sync_order_operations
from .services.operations_sync import refresh_payment_confirmation
from .services.price_version import (
    apply_order_price_snapshot,
    recommended_price_from_snapshot,
    recommended_vehicle_price,
    resolve_vehicle_price_version,
)
from .services.settlement_cost import (
    apply_order_settlement_cost,
    resolve_settlement_cost,
)
from .services.incentive_rule import (
    apply_order_incentive_rule,
)
from .services.installment_plan import (
    apply_order_installment_snapshot,
    installment_option_payload,
    resolve_installment_plan_version,
)
from .services.dealer_commission import apply_order_dealer_commission
from .services.dealer_commission import (
    create_volume_bonus_settlement,
    preview_volume_bonus,
    revise_volume_bonus_settlement,
)
from .services.dashboard_metrics import build_dashboard_metrics
from .services.secret_fields import decrypt_secret, encrypt_secret
from .services.legacy_import import (
    PREVIEW_SCHEMA_VERSION,
    apply_import_row_decision,
    build_import_master_workspace,
    build_import_preview,
    file_sha256,
    friendly_import_message,
    revalidate_import_batch,
    retry_completed_import_row,
    save_import_master_mapping,
)
from .services.positioned_template_pdf import build_positioned_template_pdf
from .services.identity_document_pdf import (
    PURPOSE_LABELS as IDENTITY_DOCUMENT_PURPOSES,
    build_identity_document_pdf,
)
from .services.upload_validation import validate_image_upload


def _form_error_text(form):
    """將表單錯誤整理成可直接給一般使用者閱讀的一句話。"""
    return " ".join(
        str(error)
        for errors in form.errors.values()
        for error in errors
    )


def _protect_private_response(response):
    """避免含個資的檔案被瀏覽器或中介快取保存。"""
    response["Cache-Control"] = "private, no-store, max-age=0"
    response["Pragma"] = "no-cache"
    response["X-Content-Type-Options"] = "nosniff"
    return response


def system_health(request):
    """Minimal readiness check for container and deployment smoke tests."""
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT 1")
            cursor.fetchone()
    except Exception:
        logger.exception("health_check_failed request_id=%s", getattr(request, "request_id", ""))
        return JsonResponse({"ok": False}, status=503)
    return JsonResponse({"ok": True})


@login_required
@require_http_methods(["POST"])
def appearance_theme_update(request):
    theme = request.POST.get("theme", "")
    return_to = request.POST.get("next", "")
    if not url_has_allowed_host_and_scheme(
        return_to,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return_to = reverse("dashboard")

    if theme not in THEME_VALUES:
        messages.error(request, "無法套用這個配色，請重新選擇。")
        return redirect(return_to)

    UserAppearancePreference.objects.update_or_create(
        user=request.user,
        defaults={"theme": theme},
    )
    if theme == DEFAULT_THEME:
        messages.success(request, "已恢復系統預設配色。")
    else:
        messages.success(request, "外觀配色已儲存，其他裝置登入後也會自動套用。")
    return redirect(return_to)


def superuser_required(view_func):
    """只允許目前仍啟用的系統管理者進入帳號管理功能。"""

    @wraps(view_func)
    @login_required
    def wrapped(request, *args, **kwargs):
        if not request.user.is_active or not request.user.is_superuser:
            raise PermissionDenied("只有系統管理者可以管理帳號。")
        return view_func(request, *args, **kwargs)

    return wrapped


def _account_audit(*, actor, target, action, description, metadata=None):
    return UserAccountAuditLog.objects.create(
        actor=actor,
        target=target,
        target_username=target.get_username(),
        action=action,
        description=description,
        metadata=metadata or {},
    )


def _invalidate_user_sessions(user):
    """重設密碼後立即登出該帳號在其他裝置上的既有登入。"""

    session_keys = []
    for session in Session.objects.filter(expire_date__gte=timezone.now()):
        try:
            if str(session.get_decoded().get("_auth_user_id")) == str(user.pk):
                session_keys.append(session.session_key)
        except Exception:
            continue
    if session_keys:
        Session.objects.filter(session_key__in=session_keys).delete()


def _active_superuser_count():
    return get_user_model().objects.filter(is_active=True, is_superuser=True).count()


@login_required
def system_diagnostics(request):
    """提供內部人員可理解、且不洩漏連線資訊的服務狀態。"""
    checks = []

    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT 1")
            cursor.fetchone()
        checks.append(
            {"key": "database", "label": "訂單資料庫", "tone": "success", "summary": "連線正常"}
        )
    except Exception:
        logger.exception("diagnostics_database_failed request_id=%s", getattr(request, "request_id", ""))
        checks.append(
            {"key": "database", "label": "訂單資料庫", "tone": "danger", "summary": "目前無法連線"}
        )

    queue_details = []
    if settings.REDIS_URL:
        try:
            for queue_name, label in (
                ("ocr", "證件辨識"),
                ("search", "搜尋索引"),
                ("imports", "歷史資料匯入"),
            ):
                queue = django_rq.get_queue(queue_name)
                queue.connection.ping()
                queue_details.append(
                    {
                        "label": label,
                        "waiting": queue.count,
                        "workers": Worker.count(connection=queue.connection, queue=queue),
                    }
                )
            waiting_total = sum(item["waiting"] for item in queue_details)
            worker_total = sum(item["workers"] for item in queue_details)
            tone = "success" if worker_total >= 3 else "warning"
            summary = f"{worker_total} 個背景工作執行中，{waiting_total} 件等待處理"
            checks.append(
                {"key": "workers", "label": "背景工作", "tone": tone, "summary": summary}
            )
        except Exception:
            logger.exception("diagnostics_redis_failed request_id=%s", getattr(request, "request_id", ""))
            checks.append(
                {"key": "workers", "label": "背景工作", "tone": "danger", "summary": "Redis 或背景工作無法連線"}
            )
    else:
        checks.append(
            {"key": "workers", "label": "背景工作", "tone": "neutral", "summary": "本機模式：工作會立即執行"}
        )

    total_orders = SalesOrder.objects.count()
    indexed_orders = SalesOrderSearchIndex.objects.count()
    missing_indexes = max(total_orders - indexed_orders, 0)
    checks.append(
        {
            "key": "search",
            "label": "全欄位搜尋",
            "tone": "success" if missing_indexes == 0 else "warning",
            "summary": (
                f"{indexed_orders:,} 張訂單已建立索引"
                if missing_indexes == 0
                else f"尚有 {missing_indexes:,} 張訂單等待建立搜尋索引"
            ),
        }
    )

    media_path = Path(settings.MEDIA_ROOT)
    usage_path = media_path if media_path.exists() else Path(settings.BASE_DIR)
    disk = shutil.disk_usage(usage_path)
    free_percent = (disk.free / disk.total * 100) if disk.total else 0
    free_gb = disk.free / (1024 ** 3)
    disk_tone = "success" if free_percent >= 20 else "warning" if free_percent >= 10 else "danger"
    checks.append(
        {
            "key": "storage",
            "label": "照片與文件空間",
            "tone": disk_tone,
            "summary": f"剩餘 {free_gb:,.1f} GB（{free_percent:.0f}%）",
        }
    )

    if any(item["tone"] == "danger" for item in checks):
        overall_tone = "danger"
    elif any(item["tone"] == "warning" for item in checks):
        overall_tone = "warning"
    else:
        overall_tone = "success"

    return render(
        request,
        "sales/system_diagnostics.html",
        {
            "checks": checks,
            "queue_details": queue_details,
            "checked_at": timezone.localtime(),
            "overall_tone": overall_tone,
        },
    )


logger = logging.getLogger(__name__)
DRAFT_PRESENCE_TIMEOUT = timedelta(seconds=90)
ORDER_PRESENCE_TIMEOUT = timedelta(seconds=90)


def _editing_name(user):
    return user.get_full_name() or user.get_username()


def _document_upload_response(
    request,
    *,
    order_pk,
    tab,
    ok,
    message,
    status=200,
    section="",
):
    """同時支援一般表單與可顯示上傳進度的非同步表單。"""
    detail_url = reverse("order_detail", args=[order_pk])
    target_url = f"{detail_url}?tab={tab}"
    if section:
        target_url = f"{target_url}#{section}"
    accepts_json = (
        request.headers.get("X-Requested-With") == "XMLHttpRequest"
        or "application/json" in request.headers.get("Accept", "")
    )
    if accepts_json:
        return JsonResponse(
            {
                "ok": ok,
                "message": message,
                "redirect_url": target_url,
            },
            status=status,
        )
    if ok:
        messages.success(request, message)
    else:
        messages.error(request, message)
    return redirect(target_url if section else detail_url)


def _order_detail_section_url(order_pk, section):
    return f"{reverse('order_detail', args=[order_pk])}?tab=order#{section}"


@login_required
def data_maintenance(request):
    context = {
        "customer_count": SalesOrder.objects.values("owner_id_number").distinct().count(),
        "vehicle_model_count": VehicleModelFamily.objects.count(),
        "vehicle_model_version_count": VehicleModel.objects.count(),
        "vehicle_brand_count": VehicleBrand.objects.count(),
        "accessory_product_count": AccessoryProduct.objects.count(),
        "inventory_count": VehicleInventory.objects.count(),
        "sales_source_count": SalesSource.objects.count(),
        "installment_company_count": InstallmentCompany.objects.count(),
        "settlement_cost_rule_count": VehicleSettlementCostRule.objects.count(),
        "incentive_rule_count": VehicleIncentiveRule.objects.count(),
        "dealer_bonus_rule_count": DealerVolumeBonusRule.objects.count(),
        "holiday_count": BusinessHoliday.objects.filter(active=True).count(),
        "registration_fee_rule_count": BrandRegistrationFeeRule.objects.filter(active=True).count(),
    }
    if request.user.is_superuser:
        context["system_user_count"] = get_user_model().objects.count()
    return render(
        request,
        "sales/data_maintenance.html",
        context,
    )


@superuser_required
def user_management(request):
    query = request.GET.get("q", "").strip()
    status = request.GET.get("status", "all")
    users = get_user_model().objects.select_related("security_profile")
    if query:
        users = users.filter(
            Q(username__icontains=query)
            | Q(first_name__icontains=query)
            | Q(last_name__icontains=query)
        )
    if status == "active":
        users = users.filter(is_active=True)
    elif status == "inactive":
        users = users.filter(is_active=False)
    elif status == "admins":
        users = users.filter(is_superuser=True)
    elif status == "never_login":
        users = users.filter(last_login__isnull=True)
    else:
        status = "all"

    users = list(users.order_by("-is_active", "-is_superuser", "username"))
    for account in users:
        profile = getattr(account, "security_profile", None)
        account.must_change_password = bool(profile and profile.must_change_password)

    all_users = get_user_model().objects.all()
    return render(
        request,
        "sales/user_management.html",
        {
            "accounts": users,
            "query": query,
            "status": status,
            "counts": {
                "all": all_users.count(),
                "active": all_users.filter(is_active=True).count(),
                "inactive": all_users.filter(is_active=False).count(),
                "admins": all_users.filter(is_superuser=True).count(),
            },
            "audit_logs": UserAccountAuditLog.objects.select_related("actor", "target")[:20],
        },
    )


@superuser_required
@require_http_methods(["GET", "POST"])
def user_account_create(request):
    form = AdminUserCreateForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        with transaction.atomic():
            user = get_user_model().objects.create_user(
                username=form.cleaned_data["username"],
                password=form.cleaned_data["password1"],
                first_name=form.cleaned_data["display_name"].strip(),
                is_active=form.cleaned_data["is_active"],
                is_staff=form.cleaned_data["is_superuser"],
                is_superuser=form.cleaned_data["is_superuser"],
            )
            UserSecurityProfile.objects.create(
                user=user,
                must_change_password=form.cleaned_data["must_change_password"],
            )
            _account_audit(
                actor=request.user,
                target=user,
                action=UserAccountAuditLog.Action.CREATE,
                description=f"建立帳號 {user.get_username()}。",
                metadata={
                    "is_active": user.is_active,
                    "is_superuser": user.is_superuser,
                    "must_change_password": form.cleaned_data["must_change_password"],
                },
            )
        messages.success(request, f"已建立 {user.get_username()}；臨時密碼不會再次顯示，請立即交付本人。")
        return redirect("user_management")
    return render(
        request,
        "sales/user_account_form.html",
        {"form": form, "mode": "create", "account": None},
    )


@superuser_required
@require_http_methods(["GET", "POST"])
def user_account_edit(request, pk):
    account = get_object_or_404(get_user_model(), pk=pk)
    form = AdminUserEditForm(request.POST or None, instance=account)
    if request.method == "POST" and form.is_valid():
        new_active = form.cleaned_data["is_active"]
        new_superuser = form.cleaned_data["is_superuser"]
        if account.pk == request.user.pk and not new_active:
            form.add_error("is_active", "不能停用自己目前正在使用的帳號。")
        if account.pk == request.user.pk and not new_superuser:
            form.add_error("is_superuser", "不能移除自己目前的管理者權限。")
        if (
            account.is_active
            and account.is_superuser
            and (not new_active or not new_superuser)
            and _active_superuser_count() <= 1
        ):
            form.add_error(None, "系統至少要保留一個啟用中的管理者帳號。")

        if not form.errors:
            before = {
                "username": account.username,
                "display_name": account.get_full_name(),
                "is_active": account.is_active,
                "is_superuser": account.is_superuser,
            }
            with transaction.atomic():
                account.username = form.cleaned_data["username"]
                account.first_name = form.cleaned_data["display_name"].strip()
                account.last_name = ""
                account.is_active = new_active
                account.is_superuser = new_superuser
                account.is_staff = new_superuser
                account.save(update_fields=["username", "first_name", "last_name", "is_active", "is_superuser", "is_staff"])
                after = {
                    "username": account.username,
                    "display_name": account.get_full_name(),
                    "is_active": account.is_active,
                    "is_superuser": account.is_superuser,
                }
                _account_audit(
                    actor=request.user,
                    target=account,
                    action=UserAccountAuditLog.Action.UPDATE,
                    description=f"修改帳號 {account.get_username()} 的基本資料或權限。",
                    metadata={"before": before, "after": after},
                )
            messages.success(request, f"已儲存 {account.get_username()} 的帳號設定。")
            return redirect("user_management")
    return render(
        request,
        "sales/user_account_form.html",
        {"form": form, "mode": "edit", "account": account},
    )


@superuser_required
@require_http_methods(["POST"])
def user_account_status(request, pk):
    account = get_object_or_404(get_user_model(), pk=pk)
    requested_action = request.POST.get("action")
    if requested_action not in {"activate", "deactivate"}:
        messages.error(request, "無法辨識帳號狀態操作，請重新操作。")
        return redirect("user_management")
    activate = requested_action == "activate"
    if account.pk == request.user.pk and not activate:
        messages.error(request, "不能停用自己目前正在使用的帳號。")
        return redirect("user_management")
    if account.is_active and account.is_superuser and not activate and _active_superuser_count() <= 1:
        messages.error(request, "系統至少要保留一個啟用中的管理者帳號。")
        return redirect("user_management")
    account.is_active = activate
    account.save(update_fields=["is_active"])
    action = UserAccountAuditLog.Action.ACTIVATE if activate else UserAccountAuditLog.Action.DEACTIVATE
    _account_audit(
        actor=request.user,
        target=account,
        action=action,
        description=f"{'啟用' if activate else '停用'}帳號 {account.get_username()}。",
    )
    messages.success(request, f"已{'啟用' if activate else '停用'} {account.get_username()}。")
    return redirect("user_management")


@superuser_required
@require_http_methods(["GET", "POST"])
def user_account_reset_password(request, pk):
    account = get_object_or_404(get_user_model(), pk=pk)
    if account.pk == request.user.pk:
        messages.info(request, "自己的密碼請使用『變更我的密碼』，避免中斷目前登入。")
        return redirect("password_change_required")
    form = AdminPasswordResetForm(request.POST or None, user=account)
    if request.method == "POST" and form.is_valid():
        with transaction.atomic():
            account.set_password(form.cleaned_data["password1"])
            account.save(update_fields=["password"])
            profile, _ = UserSecurityProfile.objects.get_or_create(user=account)
            profile.must_change_password = form.cleaned_data["must_change_password"]
            profile.password_changed_at = (
                None if profile.must_change_password else timezone.now()
            )
            profile.save(update_fields=["must_change_password", "password_changed_at", "updated_at"])
            _invalidate_user_sessions(account)
            _account_audit(
                actor=request.user,
                target=account,
                action=UserAccountAuditLog.Action.RESET_PASSWORD,
                description=f"重設帳號 {account.get_username()} 的密碼並登出既有裝置。",
                metadata={"must_change_password": profile.must_change_password},
            )
        messages.success(request, f"已重設 {account.get_username()} 的密碼，既有裝置已登出。")
        return redirect("user_management")
    return render(request, "sales/user_password_reset.html", {"form": form, "account": account})


@login_required
@require_http_methods(["GET", "POST"])
def password_change_required(request):
    form = RequiredPasswordChangeForm(request.user, request.POST or None)
    if request.method == "POST" and form.is_valid():
        user = form.save()
        update_session_auth_hash(request, user)
        profile, _ = UserSecurityProfile.objects.get_or_create(user=user)
        profile.must_change_password = False
        profile.password_changed_at = timezone.now()
        profile.save(update_fields=["must_change_password", "password_changed_at", "updated_at"])
        _account_audit(
            actor=user,
            target=user,
            action=UserAccountAuditLog.Action.CHANGE_PASSWORD,
            description=f"帳號 {user.get_username()} 已自行變更密碼。",
        )
        messages.success(request, "密碼已更新，可以繼續使用系統。")
        return redirect("dashboard")
    profile = getattr(request.user, "security_profile", None)
    return render(
        request,
        "registration/password_change_required.html",
        {"form": form, "is_required": bool(profile and profile.must_change_password)},
    )


@login_required
def business_holiday_list(request):
    editing = None
    edit_pk = request.GET.get("edit")
    if edit_pk:
        editing = get_object_or_404(BusinessHoliday, pk=edit_pk)
    form = BusinessHolidayForm(request.POST or None, instance=editing)
    if request.method == "POST" and form.is_valid():
        holiday = form.save()
        messages.success(request, f"已儲存工作日排除日期：{holiday.date} {holiday.name}。")
        return redirect("business_holiday_list")
    holidays = list(BusinessHoliday.objects.all())
    official_holidays = [
        holiday
        for holiday in holidays
        if holiday.source == BusinessHoliday.Source.DGPA
    ]
    latest_official_sync = max(
        (holiday.updated_at for holiday in official_holidays),
        default=None,
    )
    return render(
        request,
        "sales/business_holiday_list.html",
        {
            "holidays": holidays,
            "form": form,
            "editing": editing,
            "latest_official_sync": latest_official_sync,
            "official_years": sorted(
                {holiday.date.year for holiday in official_holidays}
            ),
        },
    )


@login_required
def business_holiday_delete(request, pk):
    holiday = get_object_or_404(BusinessHoliday, pk=pk)
    if request.method == "POST":
        label = str(holiday)
        holiday.delete()
        messages.success(request, f"已刪除 {label}。")
    return redirect("business_holiday_list")


@login_required
def brand_registration_fee_rule_list(request):
    editing = None
    edit_pk = request.GET.get("edit")
    if edit_pk:
        editing = get_object_or_404(BrandRegistrationFeeRule, pk=edit_pk)
    form = BrandRegistrationFeeRuleForm(request.POST or None, instance=editing)
    if request.method == "POST" and form.is_valid():
        rule = form.save()
        messages.success(request, f"已儲存 {rule.brand} 的牌險計算規則。")
        return redirect("brand_registration_fee_rule_list")
    return render(
        request,
        "sales/brand_registration_fee_rule_list.html",
        {
            "rules": BrandRegistrationFeeRule.objects.all(),
            "form": form,
            "editing": editing,
        },
    )


@login_required
def brand_registration_fee_rule_delete(request, pk):
    rule = get_object_or_404(BrandRegistrationFeeRule, pk=pk)
    if request.method == "POST":
        label = f"{rule.brand}／{rule.get_calculation_type_display()}"
        rule.delete()
        messages.success(request, f"已刪除 {label}。")
    return redirect("brand_registration_fee_rule_list")


@login_required
def positioned_template_list(request):
    return render(
        request,
        "sales/positioned_template_list.html",
        {
            "templates": PositionedPrintTemplate.objects.prefetch_related("fields"),
        },
    )


@login_required
@transaction.atomic
def positioned_template_form(request, pk=None):
    template = get_object_or_404(PositionedPrintTemplate, pk=pk) if pk else PositionedPrintTemplate()
    form = PositionedPrintTemplateForm(request.POST or None, request.FILES or None, instance=template)
    field_formset = PositionedPrintFieldFormSet(
        request.POST or None,
        instance=template,
        prefix="print_fields",
    )
    if request.method == "POST" and form.is_valid() and field_formset.is_valid():
        template = form.save()
        field_formset.instance = template
        field_formset.save()
        messages.success(request, f"已儲存列印範本：{template.name}。")
        return redirect("positioned_template_edit", pk=template.pk)
    sample_order = SalesOrder.objects.select_related(
        "source", "vehicle_model", "color", "allocated_vehicle", "operations"
    ).order_by("-created_at").first()
    return render(
        request,
        "sales/positioned_template_form.html",
        {
            "form": form,
            "field_formset": field_formset,
            "template_object": template,
            "sample_order": sample_order,
        },
    )


@login_required
def positioned_template_preview(request, pk, order_pk=None):
    template = get_object_or_404(PositionedPrintTemplate.objects.prefetch_related("fields"), pk=pk)
    orders = SalesOrder.objects.select_related(
        "source", "vehicle_model", "color", "allocated_vehicle", "operations"
    )
    order = get_object_or_404(orders, pk=order_pk) if order_pk else orders.order_by("-created_at").first()
    if not order:
        messages.error(request, "目前沒有可供預覽的訂單，請先建立一張訂單。")
        return redirect("positioned_template_edit", pk=pk)
    try:
        output = build_positioned_template_pdf(template, order)
    except (ValueError, OSError) as exc:
        messages.error(request, f"無法產生套表：{exc}")
        return redirect("positioned_template_edit", pk=pk)
    response = FileResponse(
        output,
        content_type="application/pdf",
        filename=f"{template.document_type}-{order.number}.pdf",
    )
    response["Cache-Control"] = "private, no-store"
    return _protect_private_response(response)


@login_required
@transaction.atomic
def positioned_template_delete(request, pk):
    template = get_object_or_404(PositionedPrintTemplate, pk=pk)
    if request.method == "POST":
        background = template.background_file
        label = template.name
        template.delete()
        if background:
            background.delete(save=False)
        messages.success(request, f"已刪除列印範本：{label}。")
    return redirect("positioned_template_list")


@login_required
def sales_source_list(request):
    keyword = request.GET.get("q", "").strip()
    source_type = request.GET.get("type", "")
    category_id = request.GET.get("category", "")
    cooperation_scope = request.GET.get(
        "cooperation_scope", request.GET.get("brand", "")
    ).strip()
    holiday_gift = request.GET.get("holiday_gift", "")
    line_group = request.GET.get("line_group", "").strip()
    relationship_type = request.GET.get("relationship_type", "").strip()
    sources = SalesSource.objects.select_related("category").order_by(
        "source_type", "category__name", "name", "id"
    )
    if keyword:
        keyword_filter = (
            Q(name__icontains=keyword)
            | Q(code__icontains=keyword)
            | Q(address__icontains=keyword)
            | Q(responsible_person__icontains=keyword)
            | Q(phone__icontains=keyword)
            | Q(phone_secondary__icontains=keyword)
            | Q(mobile__icontains=keyword)
            | Q(other_contact__icontains=keyword)
            | Q(note__icontains=keyword)
            | Q(category__name__icontains=keyword)
            | Q(cooperation_profiles__note__icontains=keyword)
        )
        for relationship_value, relationship_label in (
            SalesSourceCooperationProfile.RelationshipType.choices
        ):
            if keyword in relationship_label:
                keyword_filter |= Q(
                    cooperation_profiles__relationship_type=relationship_value
                )
        if any(keyword in label for label in ("年節送禮", "送禮", "月餅")):
            keyword_filter |= Q(holiday_gift=True)
        if "line" in keyword.casefold() or "群組" in keyword:
            keyword_filter |= Q(has_line_group=True)
        sources = sources.filter(keyword_filter).distinct()
    if source_type in {value for value, _ in SalesSource.SourceType.choices}:
        sources = sources.filter(source_type=source_type)
    if category_id.isdigit():
        sources = sources.filter(category_id=category_id)
    valid_scopes = {
        value for value, _ in SalesSourceBrandPolicy.CooperationScope.choices
    }
    if cooperation_scope in valid_scopes:
        today = timezone.localdate()
        selected_profile = SalesSourceCooperationProfile.objects.filter(
            source_id=OuterRef("pk"), cooperation_scope=cooperation_scope
        )
        latest_scope_state = (
            SalesSourceBrandPolicy.objects.filter(
                source_id=OuterRef("pk"),
                cooperation_scope=cooperation_scope,
                effective_from__lte=today,
            )
            .filter(Q(effective_to__isnull=True) | Q(effective_to__gte=today))
            .order_by("-effective_from", "-pk")
            .values("cooperates")[:1]
        )
        sources = sources.annotate(
            selected_scope_profile_exists=Exists(selected_profile),
            selected_scope_profile_cooperates=Exists(
                selected_profile.filter(cooperates=True)
            ),
            selected_scope_policy_cooperates=Subquery(latest_scope_state),
        ).filter(
            Q(selected_scope_profile_cooperates=True)
            | Q(
                selected_scope_profile_exists=False,
                selected_scope_policy_cooperates=True,
            )
        )
    valid_relationship_types = {
        value for value, _ in SalesSourceCooperationProfile.RelationshipType.choices
    }
    if relationship_type in valid_relationship_types:
        sources = sources.filter(
            cooperation_profiles__cooperates=True,
            cooperation_profiles__relationship_type=relationship_type,
        )
    if holiday_gift in {"yes", "no"}:
        sources = sources.filter(
            source_type=SalesSource.SourceType.DEALER,
            holiday_gift=holiday_gift == "yes",
        )
    if line_group == "yes":
        sources = sources.filter(
            source_type=SalesSource.SourceType.DEALER,
            has_line_group=True,
        )
    elif line_group == "no":
        sources = sources.filter(
            source_type=SalesSource.SourceType.DEALER,
            has_line_group=False,
        )
    today = timezone.localdate()
    current_policy_queryset = SalesSourceBrandPolicy.objects.filter(
        cooperation_scope__in=valid_scopes,
        effective_from__lte=today,
    ).filter(
        Q(effective_to__isnull=True) | Q(effective_to__gte=today)
    ).order_by("cooperation_scope", "-effective_from", "-pk")
    page = Paginator(sources.distinct().prefetch_related(
        Prefetch(
            "brand_policies",
            queryset=current_policy_queryset,
            to_attr="current_cooperation_policies",
        ),
        Prefetch(
            "cooperation_profiles",
            queryset=SalesSourceCooperationProfile.objects.order_by(
                "cooperation_scope"
            ),
            to_attr="current_cooperation_profiles",
        ),
    ), 100).get_page(
        request.GET.get("page")
    )
    for item in page.object_list:
        item.cooperation_overview = _sales_source_brand_overview(
            item,
            policies=item.current_cooperation_policies,
            profiles=item.current_cooperation_profiles,
        )
        item.list_type_label = _sales_source_list_type_label(
            item,
            item.cooperation_overview,
        )
        item.line_group_marker = _sales_source_line_group_marker(
            item,
            item.cooperation_overview,
        )
    return render(
        request,
        "sales/sales_source_list.html",
        {
            "page_obj": page,
            "sources": page.object_list,
            "source_types": SalesSource.SourceType.choices,
            "source_categories": SalesSourceCategory.objects.filter(active=True).order_by(
                "system_behavior", "name"
            ),
            "cooperation_scopes": SalesSourceBrandPolicy.CooperationScope.choices,
            "relationship_types": SalesSourceCooperationProfile.RelationshipType.choices,
            "holiday_gift_count": SalesSource.objects.filter(
                source_type=SalesSource.SourceType.DEALER,
                holiday_gift=True,
            ).count(),
            "selected": {
                "q": keyword,
                "type": source_type,
                "category": category_id,
                "cooperation_scope": cooperation_scope,
                "holiday_gift": holiday_gift,
                "line_group": line_group,
                "relationship_type": relationship_type,
            },
        },
    )


@login_required
@transaction.atomic
def vehicle_brand_list(request):
    editing = None
    edit_pk = request.GET.get("edit")
    if edit_pk:
        editing = get_object_or_404(VehicleBrand, pk=edit_pk)
    original_name = editing.name if editing else ""
    form = VehicleBrandForm(request.POST or None, instance=editing)
    if request.method == "POST" and form.is_valid():
        brand = form.save(commit=False)
        try:
            with transaction.atomic():
                if editing and original_name != brand.name:
                    rename_vehicle_brand_references(original_name, brand.name)
                brand.save()
        except IntegrityError:
            form.add_error(
                "name",
                "無法改名：既有車型或規則會與另一筆資料重複，請先檢查相關資料。",
            )
        else:
            messages.success(request, f"已儲存品牌：{brand.name}。")
            return redirect("vehicle_brand_list")
    brands = list(VehicleBrand.objects.select_related("parent").all())
    brands.sort(
        key=lambda brand: (
            brand.parent.display_order if brand.parent_id else brand.display_order,
            1 if brand.parent_id else 0,
            brand.display_order,
            brand.name.casefold(),
        )
    )
    for brand in brands:
        brand.is_used = vehicle_brand_is_used(brand.name)
    return render(
        request,
        "sales/vehicle_brand_list.html",
        {"brands": brands, "form": form, "editing": editing},
    )

@login_required
@require_http_methods(["GET", "POST"])
@transaction.atomic
def sales_source_holiday_gift_manage(request):
    dealers = SalesSource.objects.filter(
        source_type=SalesSource.SourceType.DEALER
    ).order_by("-holiday_gift", "name", "id")

    if request.method == "POST":
        raw_ids = request.POST.getlist("source_ids")
        try:
            selected_ids = {int(raw_id) for raw_id in raw_ids}
        except (TypeError, ValueError):
            messages.error(request, "名單未更新：送出的車行資料格式不正確，請重新操作。")
            return redirect("sales_source_holiday_gift_manage")

        valid_ids = set(
            SalesSource.objects.filter(
                source_type=SalesSource.SourceType.DEALER,
                pk__in=selected_ids,
            ).values_list("pk", flat=True)
        )
        if valid_ids != selected_ids:
            messages.error(request, "名單未更新：其中包含不存在或不適用的車行。")
            return redirect("sales_source_holiday_gift_manage")

        changed_at = timezone.now()
        SalesSource.objects.filter(
            source_type=SalesSource.SourceType.DEALER,
            holiday_gift=True,
        ).exclude(pk__in=selected_ids).update(
            holiday_gift=False,
            updated_at=changed_at,
        )
        SalesSource.objects.filter(
            source_type=SalesSource.SourceType.DEALER,
            holiday_gift=False,
            pk__in=selected_ids,
        ).update(
            holiday_gift=True,
            updated_at=changed_at,
        )
        messages.success(request, f"年節送禮名單已更新，目前共 {len(selected_ids)} 家。")
        return redirect(f"{reverse('sales_source_list')}?holiday_gift=yes")

    dealer_rows = list(dealers)
    return render(
        request,
        "sales/sales_source_holiday_gift_manage.html",
        {
            "dealers": dealer_rows,
            "dealer_count": len(dealer_rows),
            "holiday_gift_count": sum(dealer.holiday_gift for dealer in dealer_rows),
        },
    )


@login_required
@transaction.atomic
def sales_source_form(request, pk=None):
    source = get_object_or_404(SalesSource, pk=pk) if pk else SalesSource()
    post_data = request.POST or None
    brand_overview = _sales_source_brand_overview(source)
    form = SalesSourceForm(post_data, instance=source)
    cooperation_sections = _sales_source_cooperation_sections(
        source, post_data=post_data, brand_overview=brand_overview
    )
    scoped_policies = (
        source.brand_policies.filter(
            cooperation_scope__in=SalesSourceCooperationForm.FIELD_BY_SCOPE
        )
        if source.pk
        else SalesSourceBrandPolicy.objects.none()
    )
    policy_formset = SalesSourceBrandPolicyFormSet(
        post_data,
        instance=source,
        prefix="policies",
        queryset=scoped_policies,
    )
    form_valid = form.is_valid() if request.method == "POST" else False
    is_dealer = bool(
        form_valid
        and form.cleaned_data["category"].system_behavior
        == SalesSource.SourceType.DEALER
    )
    cooperation_valid = all(
        section["form"].is_valid() for section in cooperation_sections
    ) if request.method == "POST" and is_dealer else True
    policy_valid = policy_formset.is_valid() if request.method == "POST" else False
    if request.method == "POST" and form_valid and cooperation_valid and policy_valid:
        source = form.save()
        policy_formset.instance = source
        policy_formset.save()
        desired_states = {}
        if source.source_type == SalesSource.SourceType.DEALER:
            for section in cooperation_sections:
                profile = section["form"].save(commit=False)
                profile.source = source
                profile.cooperation_scope = section["scope"]
                profile.save()
                desired_states[section["scope"]] = profile.cooperates
            _sync_sales_source_cooperation_scopes(source, desired_states)
        messages.success(
            request,
            f"已儲存{source.category.name if source.category_id else source.get_source_type_display()}：{source.name}。",
        )
        return redirect("sales_source_list")
    return render(
        request,
        "sales/sales_source_form.html",
        {
            "form": form,
            "cooperation_sections": cooperation_sections,
            "policy_formset": policy_formset,
            "source": source if source.pk else None,
            "brand_overview": brand_overview,
            "category_behaviors": {
                str(category.pk): category.system_behavior
                for category in form.fields["category"].queryset
            },
        },
    )


def _sales_source_cooperation_sections(source, post_data=None, brand_overview=None):
    profiles = {
        profile.cooperation_scope: profile
        for profile in (
            source.cooperation_profiles.all() if source.pk else []
        )
    }
    brand_overview = brand_overview or _sales_source_brand_overview(source)
    sections = []
    for scope, label in SalesSourceBrandPolicy.CooperationScope.choices:
        profile = profiles.get(scope)
        initial = None
        if profile is None:
            profile = SalesSourceCooperationProfile(cooperation_scope=scope)
            initial = {
                "cooperates": brand_overview["states"].get(scope, False),
                "relationship_type": SalesSourceCooperationProfile.RelationshipType.GENERAL,
                "vehicle_capacity": (
                    source.vehicle_capacity
                    if brand_overview["states"].get(scope, False)
                    else None
                ),
            }
        sections.append(
            {
                "scope": scope,
                "label": label,
                "form": SalesSourceCooperationProfileForm(
                    post_data,
                    instance=profile,
                    initial=initial,
                    prefix=f"cooperation-{scope}",
                ),
            }
        )
    return sections


def _sync_sales_source_cooperation_scopes(source, desired_states):
    """Persist the three user-facing cooperation switches as effective rules."""
    if source.source_type != SalesSource.SourceType.DEALER:
        return

    today = timezone.localdate()
    for scope, cooperates in desired_states.items():
        current = (
            source.brand_policies.filter(
                cooperation_scope=scope,
                effective_from__lte=today,
            )
            .filter(Q(effective_to__isnull=True) | Q(effective_to__gte=today))
            .order_by("-effective_from", "-pk")
            .first()
        )
        if current and current.cooperates == cooperates:
            continue
        defaults = {
            "cooperates": cooperates,
            "commission_adjustment": (
                current.commission_adjustment if current else 0
            ),
            "effective_to": None,
            "note": current.note if current else "",
        }
        policy, _ = SalesSourceBrandPolicy.objects.update_or_create(
            source=source,
            cooperation_scope=scope,
            effective_from=today,
            defaults=defaults,
        )
        policy.save()


def _sales_source_brand_overview(source, policies=None, profiles=None):
    """Return the currently effective cooperation state for the editor summary."""
    today = timezone.localdate()
    current_by_scope = {}
    if policies is None and source.pk:
        policies = source.brand_policies.filter(
            cooperation_scope__in=SalesSourceCooperationForm.FIELD_BY_SCOPE,
            effective_from__lte=today,
        ).filter(
            Q(effective_to__isnull=True) | Q(effective_to__gte=today)
        ).order_by("cooperation_scope", "-effective_from", "-pk")
    for policy in policies or []:
        if policy.cooperation_scope:
            current_by_scope.setdefault(policy.cooperation_scope, policy)
    if profiles is None and source.pk:
        profiles = source.cooperation_profiles.all()
    profile_by_scope = {
        profile.cooperation_scope: profile for profile in (profiles or [])
    }

    priority = []
    states = {}
    cooperating = []
    for cooperation_scope, label in SalesSourceBrandPolicy.CooperationScope.choices:
        policy = current_by_scope.get(cooperation_scope)
        profile = profile_by_scope.get(cooperation_scope)
        cooperates = profile.cooperates if profile is not None else bool(
            policy and policy.cooperates
        )
        if profile is None and policy is None:
            status = "unset"
            status_label = "尚未設定"
        elif cooperates:
            status = "cooperates"
            status_label = "有配合"
        else:
            status = "not_cooperating"
            status_label = "未配合"
        states[cooperation_scope] = cooperates
        if cooperates:
            cooperating.append(
                {
                    "scope": cooperation_scope,
                    "label": label,
                    "price_list_label": (
                        policy.price_list_label if policy else f"{label}價格表"
                    ),
                    "commission_adjustment": (
                        policy.commission_adjustment if policy else 0
                    ),
                    "relationship_type": (
                        profile.get_relationship_type_display() if profile else "一般"
                    ),
                    "relationship_type_code": (
                        profile.relationship_type
                        if profile
                        else SalesSourceCooperationProfile.RelationshipType.GENERAL
                    ),
                    "vehicle_capacity": profile.vehicle_capacity if profile else None,
                    "note": profile.note if profile else "",
                }
            )
        priority.append(
            {
                "cooperation_scope": cooperation_scope,
                "label": label,
                "status": status,
                "status_label": status_label,
                "price_list_label": f"{label}價格表",
                "relationship_type": (
                    profile.get_relationship_type_display() if profile else "一般"
                ),
                "vehicle_capacity": profile.vehicle_capacity if profile else None,
            }
        )
    return {
        "priority": priority,
        "cooperating": cooperating,
        "cooperating_count": len(cooperating),
        "states": states,
    }


def _sales_source_list_type_label(source, cooperation_overview):
    """Return a concise, user-facing classification for the source list."""
    if source.source_type != SalesSource.SourceType.DEALER:
        return (
            source.category.name
            if source.category_id
            else source.get_source_type_display()
        )

    relationship_types = {
        item["relationship_type"]
        for item in cooperation_overview.get("cooperating", [])
    }
    if "股東" in relationship_types:
        return "股東車行"
    if "專銷" in relationship_types:
        return "專銷車行"
    return "一般車行"


def _sales_source_line_group_marker(source, cooperation_overview):
    """Return the familiar LINE-group colour used for each cooperation scope."""
    if not source.has_line_group:
        return {
            "tone": "none",
            "description": "無 LINE 群組",
        }

    cooperation_by_scope = {
        item["scope"]: item
        for item in cooperation_overview.get("cooperating", [])
    }
    has_sym = SalesSourceBrandPolicy.CooperationScope.SYM in cooperation_by_scope
    has_suzuki = any(
        scope in cooperation_by_scope
        for scope in (
            SalesSourceBrandPolicy.CooperationScope.SUZUKI_GAS,
            SalesSourceBrandPolicy.CooperationScope.SUZUKI_ELECTRIC,
        )
    )
    sym_profile = cooperation_by_scope.get(
        SalesSourceBrandPolicy.CooperationScope.SYM
    )
    sym_is_exclusive = bool(
        sym_profile
        and sym_profile["relationship_type_code"]
        == SalesSourceCooperationProfile.RelationshipType.EXCLUSIVE
    )

    if has_sym and has_suzuki and sym_is_exclusive:
        return {
            "tone": "sym-exclusive",
            "description": "有 LINE 群組：三陽專銷，並與台鈴合作",
        }
    if has_sym and has_suzuki:
        return {
            "tone": "mixed",
            "description": "有 LINE 群組：三陽與台鈴一般合作",
        }
    if has_suzuki:
        return {
            "tone": "suzuki",
            "description": "有 LINE 群組：僅台鈴合作",
        }
    if has_sym:
        return {
            "tone": "sym",
            "description": "有 LINE 群組：僅三陽合作",
        }
    return {
        "tone": "unclassified",
        "description": "有 LINE 群組：合作範圍尚未設定",
    }


@login_required
def installment_company_list(request):
    editing = None
    edit_pk = request.GET.get("edit")
    if edit_pk:
        editing = get_object_or_404(InstallmentCompany, pk=edit_pk)
    form = InstallmentCompanyForm(request.POST or None, instance=editing)
    show_editor = bool(
        editing or request.GET.get("new") == "1" or request.method == "POST"
    )
    if request.method == "POST" and form.is_valid():
        company = form.save()
        messages.success(request, f"已儲存分期公司：{company.name}。")
        return redirect("installment_company_list")
    return render(
        request,
        "sales/installment_company_list.html",
        {
            "companies": InstallmentCompany.objects.annotate(
                option_count=Count("plan_options", distinct=True),
                order_count=Count("orders", distinct=True),
            ),
            "form": form,
            "editing": editing,
            "show_editor": show_editor,
        },
    )


@login_required
@require_http_methods(["POST"])
def installment_company_quick_create(request):
    name = request.POST.get("name", "").strip()
    existing = (
        InstallmentCompany.objects.filter(name__iexact=name).first()
        if name
        else None
    )
    if existing:
        if not existing.active:
            return JsonResponse(
                {
                    "ok": False,
                    "errors": {
                        "name": [
                            "這家公司已存在但目前停用，請到「分期公司」維護頁重新啟用。"
                        ]
                    },
                },
                status=409,
            )
        return JsonResponse(
            {
                "ok": True,
                "created": False,
                "company": {"id": existing.pk, "name": existing.name},
            }
        )

    form = InstallmentCompanyForm(
        {
            "name": name,
            "customer_service_phone": request.POST.get(
                "customer_service_phone", ""
            ).strip(),
            "active": True,
            "note": "",
        }
    )
    if not form.is_valid():
        return JsonResponse(
            {"ok": False, "errors": form.errors.get_json_data()}, status=400
        )
    try:
        company = form.save()
    except IntegrityError:
        company = InstallmentCompany.objects.filter(name__iexact=name).first()
        if company and company.active:
            return JsonResponse(
                {
                    "ok": True,
                    "created": False,
                    "company": {"id": company.pk, "name": company.name},
                }
            )
        return JsonResponse(
            {"ok": False, "errors": {"name": [{"message": "公司名稱已被使用。"}]}},
            status=409,
        )
    return JsonResponse(
        {
            "ok": True,
            "created": True,
            "company": {"id": company.pk, "name": company.name},
        },
        status=201,
    )


@login_required
@transaction.atomic
def vehicle_installment_plan_list(request, model_pk):
    vehicle_model = get_object_or_404(VehicleModel, pk=model_pk)
    today = timezone.localdate()
    editing = None
    edit_pk = request.POST.get("plan_id") or request.GET.get("edit")
    if edit_pk:
        editing = get_object_or_404(
            InstallmentPlanVersion, pk=edit_pk, vehicle_model=vehicle_model
        )
    plan = editing or InstallmentPlanVersion(vehicle_model=vehicle_model)
    post_data = request.POST or None
    form = InstallmentPlanVersionForm(post_data, instance=plan, prefix="plan")
    option_formset = InstallmentPlanOptionFormSet(
        post_data, instance=plan, prefix="options"
    )
    if request.method == "GET" and not (plan.pk and plan.options.exists()):
        # 新版本或尚未設定期數的版本，提供第一列方便開始填寫；
        # 已有期數時只呈現實際資料，避免讓使用者誤以為尚有一筆未完成。
        option_formset.extra = 1
    if request.method == "POST" and form.is_valid() and option_formset.is_valid():
        plan = form.save(commit=False)
        plan.vehicle_model = vehicle_model
        plan.save()
        option_formset.instance = plan
        option_formset.save()
        messages.success(request, "分期方案版本已儲存。")
        return redirect(
            f"{reverse('vehicle_installment_plan_list', args=[vehicle_model.pk])}"
            f"?edit={plan.pk}&saved=1"
        )
    plans = list(
        vehicle_model.installment_plan_versions.prefetch_related(
            "options__company"
        )
    )
    current_plan = resolve_installment_plan_version(vehicle_model.pk, today)
    editing_options = list(editing.options.select_related("company")) if editing else []
    editing_lifecycle = None
    if editing:
        if not editing.active:
            editing_lifecycle = {
                "key": "inactive",
                "label": "已儲存，目前停用",
                "detail": "這個版本不會自動套用到新訂單。",
            }
        elif current_plan and current_plan.pk == editing.pk:
            editing_lifecycle = {
                "key": "current",
                "label": "已儲存，目前生效中",
                "detail": "符合有效日期的新訂單會自動套用這個版本。",
            }
        elif editing.effective_from > today:
            editing_lifecycle = {
                "key": "scheduled",
                "label": "已儲存，等待生效",
                "detail": f"將於 {editing.effective_from:%Y/%m/%d} 起套用到符合日期的新訂單。",
            }
        elif editing.effective_to and editing.effective_to < today:
            editing_lifecycle = {
                "key": "expired",
                "label": "已儲存，效期已結束",
                "detail": "這個版本保留供歷史訂單查閱，不會再套用到新訂單。",
            }
        else:
            editing_lifecycle = {
                "key": "superseded",
                "label": "已儲存，已有較新版本生效",
                "detail": "這個版本仍會依原有效日期保留，不會覆蓋目前版本。",
            }
    return render(
        request,
        "sales/installment_plan_list.html",
        {
            "vehicle_model": vehicle_model,
            "plans": plans,
            "form": form,
            "option_formset": option_formset,
            "editing": editing,
            "editing_options": editing_options,
            "editing_lifecycle": editing_lifecycle,
            "saved_confirmation": bool(editing and request.GET.get("saved") == "1"),
            "current_plan": current_plan,
            "today": today,
        },
    )


@login_required
def dealer_volume_bonus_list(request):
    show_all = request.GET.get("show") == "all"
    rules = (
        DealerVolumeBonusRule.objects.filter(
            dealer__source_type=SalesSource.SourceType.DEALER
        )
        .select_related("dealer")
        .prefetch_related("tiers", "settlement__allocations")
    )
    rows = []
    hidden_count = 0
    for rule in rules:
        preview = preview_volume_bonus(rule)
        has_settlement = hasattr(rule, "settlement")
        if not show_all and preview["quantity"] == 0 and not has_settlement:
            hidden_count += 1
            continue
        rows.append({"rule": rule, "preview": preview})
    return render(
        request,
        "sales/dealer_volume_bonus_list.html",
        {
            "rows": rows,
            "show_all": show_all,
            "hidden_count": hidden_count,
        },
    )


@login_required
@transaction.atomic
def dealer_volume_bonus_form(request, pk=None):
    rule = get_object_or_404(DealerVolumeBonusRule, pk=pk) if pk else DealerVolumeBonusRule()
    if rule.pk and hasattr(rule, "settlement"):
        messages.error(request, "此規則已完成結算，為保留明細不可再修改。")
        return redirect("dealer_volume_bonus_list")
    post_data = request.POST or None
    form = DealerVolumeBonusRuleForm(post_data, instance=rule)
    tier_formset = DealerVolumeBonusTierFormSet(
        post_data, instance=rule, prefix="tiers"
    )
    if request.method == "POST" and form.is_valid() and tier_formset.is_valid():
        rule = form.save()
        tier_formset.instance = rule
        tier_formset.save()
        messages.success(request, "車行台數獎金規則已儲存。")
        return redirect("dealer_volume_bonus_list")
    return render(
        request,
        "sales/dealer_volume_bonus_form.html",
        {"form": form, "tier_formset": tier_formset, "rule": rule if rule.pk else None},
    )


@login_required
@transaction.atomic
def dealer_volume_bonus_settle(request, pk):
    rule = get_object_or_404(DealerVolumeBonusRule.objects.select_related("dealer"), pk=pk)
    if hasattr(rule, "settlement"):
        messages.error(request, "此規則已完成結算，不可重複結算。")
        return redirect("dealer_volume_bonus_list")
    preview = preview_volume_bonus(rule)
    settlement = DealerVolumeBonusSettlement(
        rule=rule,
        expected_amount=preview["expected_amount"],
        actual_amount=preview["expected_amount"],
    )
    form = DealerVolumeBonusSettlementForm(request.POST or None, instance=settlement)
    if request.method == "POST" and form.is_valid():
        try:
            create_volume_bonus_settlement(
                rule,
                _editing_name(request.user),
                form.cleaned_data["actual_amount"],
                form.cleaned_data["adjustment_reason"],
            )
        except (ValueError, ValidationError) as exc:
            messages.error(request, str(exc))
        else:
            messages.success(request, "台數獎金已結算並保存逐單明細。")
            return redirect("dealer_volume_bonus_list")
    return render(
        request,
        "sales/dealer_volume_bonus_settle.html",
        {"rule": rule, "preview": preview, "form": form},
    )


@login_required
@transaction.atomic
def dealer_volume_bonus_revise(request, pk):
    settlement = get_object_or_404(
        DealerVolumeBonusSettlement.objects.select_related("rule__dealer").prefetch_related(
            "allocations__order__vehicle_model", "adjustments"
        ),
        pk=pk,
    )
    form = DealerVolumeBonusAdjustmentForm(
        request.POST or None,
        initial={"actual_amount": settlement.actual_amount},
    )
    if request.method == "POST" and form.is_valid():
        try:
            revise_volume_bonus_settlement(
                settlement,
                _editing_name(request.user),
                form.cleaned_data["actual_amount"],
                form.cleaned_data["reason"],
            )
        except (ValueError, ValidationError) as exc:
            messages.error(request, str(exc))
        else:
            messages.success(request, "實際入帳金額已重新分攤，並保存調整紀錄。")
            return redirect("dealer_volume_bonus_list")
    return render(
        request,
        "sales/dealer_volume_bonus_revise.html",
        {"settlement": settlement, "form": form},
    )


@login_required
def legacy_import_list(request):
    form = LegacyImportUploadForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        uploaded = form.cleaned_data["source_file"]
        batch = form.save(commit=False)
        batch.original_filename = uploaded.name
        batch.file_size = uploaded.size
        batch.file_sha256 = file_sha256(uploaded)
        batch.uploaded_by = _editing_name(request.user)
        batch.save()
        try:
            build_import_preview(batch)
        except Exception as exc:
            logger.exception("建立歷史資料匯入預覽失敗")
            batch.status = LegacyImportBatch.Status.FAILED
            batch.result_summary = {"error": str(exc)}
            batch.save(update_fields=["status", "result_summary", "updated_at"])
            messages.error(request, f"檔案解析失敗：{exc}")
        else:
            messages.success(request, "檔案已解析，請先檢查預覽與衝突報告再確認匯入。")
        return redirect("legacy_import_detail", pk=batch.pk)
    show_archived = request.GET.get("archived") == "1"
    batches = LegacyImportBatch.objects.all()
    batches = batches.filter(archived_at__isnull=not show_archived)
    return render(
        request,
        "sales/legacy_import_list.html",
        {"form": form, "batches": batches[:50], "show_archived": show_archived},
    )


@login_required
@require_http_methods(["GET", "POST"])
@transaction.atomic
def sales_source_category_list(request):
    editing = None
    edit_pk = request.GET.get("edit")
    if edit_pk:
        editing = get_object_or_404(SalesSourceCategory, pk=edit_pk)
    if request.method == "POST" and request.POST.get("action") == "delete":
        category = get_object_or_404(SalesSourceCategory, pk=request.POST.get("category_id"))
        if category.sources.exists():
            messages.error(request, f"無法刪除「{category.name}」：仍有通路正在使用。可改為停用。")
        else:
            label = category.name
            category.delete()
            messages.success(request, f"已刪除未使用的通路分類：{label}。")
        return redirect("sales_source_category_list")
    form = SalesSourceCategoryForm(request.POST or None, instance=editing)
    if request.method == "POST" and form.is_valid():
        category = form.save()
        messages.success(request, f"已儲存通路分類：{category.name}。")
        return redirect("sales_source_category_list")
    categories = SalesSourceCategory.objects.annotate(
        source_count=Count("sources")
    ).order_by("system_behavior", "name", "id")
    return render(
        request,
        "sales/sales_source_category_list.html",
        {"form": form, "categories": categories, "editing": editing},
    )


@login_required
def legacy_import_detail(request, pk):
    batch = get_object_or_404(LegacyImportBatch, pk=pk)
    if (
        batch.status == LegacyImportBatch.Status.PREVIEW
        and (batch.preview_summary or {}).get("parser_schema_version", 0)
        < PREVIEW_SCHEMA_VERSION
    ):
        revalidate_import_batch(batch)
        batch.refresh_from_db()
    rows = batch.rows.all()
    action = request.GET.get("action", "")
    if action in {value for value, _ in LegacyImportRow.Action.choices}:
        rows = rows.filter(action=action)
    search_query = request.GET.get("q", "").strip()
    if search_query:
        normalized_query = normalize_vehicle_model_master_value(search_query)
        normalized_identifier = normalize_vehicle_identifier(search_query) or ""
        matching_models = VehicleModel.objects.filter(
            Q(brand__icontains=search_query)
            | Q(name__icontains=search_query)
            | Q(model_number__icontains=search_query)
            | Q(factory_model_codes__code__icontains=search_query)
        ).distinct().values_list("pk", "name", "model_number")
        matching_model_ids = []
        matching_model_aliases = set()
        for model_id, name, model_number in matching_models:
            matching_model_ids.append(model_id)
            matching_model_aliases.update(
                normalize_vehicle_model_master_value(value)
                for value in (name, model_number)
                if value
            )
        matching_model_aliases.update(
            normalize_vehicle_model_master_value(value)
            for value in VehicleModel.objects.filter(pk__in=matching_model_ids)
            .values_list("factory_model_codes__code", flat=True)
            .distinct()
            if value
        )
        mapped_source_values = set(
            LegacyImportMasterMapping.objects.filter(
                mapping_type=LegacyImportMasterMapping.MappingType.VEHICLE_MODEL,
                vehicle_model_id__in=matching_model_ids,
            ).values_list("source_value", flat=True)
        )
        source_model_values = set(
            batch.rows.exclude(mapped_data__model_number="").values_list(
                "mapped_data__model_number", flat=True
            )
        )
        matching_source_models = {
            value
            for value in source_model_values
            if value
            and (
                normalized_query in normalize_vehicle_model_master_value(value)
                or normalize_vehicle_model_master_value(value)
                in matching_model_aliases
            )
        }
        matching_source_models.update(mapped_source_values)
        search_filter = (
            Q(mapped_data__model_number__in=matching_source_models)
            | Q(mapped_data__owner_name__icontains=search_query)
            | Q(mapped_data__plate_number__icontains=search_query)
            | Q(mapped_data__color__icontains=search_query)
            | Q(mapped_data__dealer_name__icontains=search_query)
            | Q(mapped_data__dealer_name_raw__icontains=search_query)
            | Q(mapped_data__name__icontains=search_query)
            | Q(mapped_data__contact_name__icontains=search_query)
        )
        transaction_aliases = {
            "一般新車": SalesOrder.TransactionType.REGULAR_NEW,
            "領牌車": SalesOrder.TransactionType.REGISTERED,
            "試乘車": SalesOrder.TransactionType.TEST_RIDE,
            "中獎車": SalesOrder.TransactionType.PRIZE,
            "中古車交易": SalesOrder.TransactionType.USED,
        }
        for label, value in transaction_aliases.items():
            if search_query in label:
                search_filter |= Q(mapped_data__transaction_type=value)
        if normalized_identifier:
            search_filter |= Q(mapped_data__identifier__icontains=normalized_identifier)
        rows = rows.filter(search_filter)
    filtered_rows = rows
    page = Paginator(rows, 100).get_page(request.GET.get("page"))
    for preview_row in page.object_list:
        preview_row.display_messages = [
            friendly_import_message(message) for message in preview_row.messages
        ]
    editing_row = None
    correction_form = None
    editing_id = request.GET.get("edit", "")
    if editing_id:
        candidate = get_object_or_404(batch.rows.prefetch_related("corrections"), pk=editing_id)
        can_edit_candidate = batch.status == LegacyImportBatch.Status.PREVIEW or (
            batch.status == LegacyImportBatch.Status.COMPLETED
            and candidate.action == LegacyImportRow.Action.ERROR
            and not candidate.committed_model
        )
        if can_edit_candidate:
            editing_row = candidate
            correction_form = LegacyImportRowCorrectionForm(row=editing_row)
    conflict_groups = []
    if action == LegacyImportRow.Action.CONFLICT:
        grouped = {}
        for conflict_row in filtered_rows:
            comparison_key = (
                conflict_row.mapped_data.get("identifier")
                if conflict_row.sheet_name == "銷貨"
                else conflict_row.natural_key
            )
            key = (conflict_row.sheet_name, comparison_key)
            grouped.setdefault(key, []).append(conflict_row)
        conflict_groups = [
            {"sheet_name": key[0], "comparison_key": key[1], "rows": group_rows}
            for key, group_rows in grouped.items()
        ]
    counts = {value: 0 for value, _label in LegacyImportRow.Action.choices}
    for item in batch.rows.values("action").annotate(total=Count("id")):
        counts[item["action"]] = item["total"]
    unresolved_count = counts.get("conflict", 0) + counts.get("error", 0)
    action_labels = dict(LegacyImportRow.Action.choices)
    master_workspace = {"models": [], "sources": [], "total": 0}
    if (
        batch.status == LegacyImportBatch.Status.PREVIEW
        and batch.import_type == LegacyImportBatch.ImportType.OPERATIONS
    ):
        master_workspace = build_import_master_workspace(batch)
    return render(
        request,
        "sales/legacy_import_detail.html",
        {
            "batch": batch,
            "rows": page.object_list,
            "page_obj": page,
            "selected_action": action,
            "selected_action_label": action_labels.get(action, "全部資料"),
            "search_query": search_query,
            "counts": counts,
            "unresolved_count": unresolved_count,
            "can_confirm": batch.status == LegacyImportBatch.Status.PREVIEW and unresolved_count == 0,
            "can_resume": (
                batch.status == LegacyImportBatch.Status.FAILED
                and batch.processing_started_at is not None
                and unresolved_count == 0
            ),
            "processing_percent": (
                min(100, int(batch.processing_completed * 100 / batch.processing_total))
                if batch.processing_total
                else 0
            ),
            "conflict_groups": conflict_groups,
            "editing_row": editing_row,
            "correction_form": correction_form,
            "master_workspace": master_workspace,
            "vehicle_model_link_form": LegacyVehicleModelLinkForm(prefix="model-link"),
            "sales_source_link_form": LegacySalesSourceLinkForm(prefix="source-link"),
            "vehicle_model_quick_form": LegacyVehicleModelQuickCreateForm(prefix="model-create"),
            "sales_source_quick_form": LegacySalesSourceQuickCreateForm(prefix="source-create"),
        },
    )


@login_required
@require_http_methods(["POST"])
@transaction.atomic
def legacy_import_master_resolve(request, pk, mapping_type):
    batch = get_object_or_404(
        LegacyImportBatch.objects.select_for_update(),
        pk=pk,
        status=LegacyImportBatch.Status.PREVIEW,
        import_type=LegacyImportBatch.ImportType.OPERATIONS,
    )
    type_config = {
        LegacyImportMasterMapping.MappingType.VEHICLE_MODEL: (
            "unmapped_models",
            "車型",
        ),
        LegacyImportMasterMapping.MappingType.SALES_SOURCE: (
            "unmapped_sources",
            "通路",
        ),
    }
    if mapping_type not in type_config:
        raise Http404("不支援的主檔類型")
    validation_key, type_label = type_config[mapping_type]
    source_value = request.POST.get("source_value", "").strip()
    unresolved_values = (
        (batch.preview_summary or {}).get("validation", {}).get(validation_key, [])
    )
    normalize_unresolved = (
        normalize_vehicle_model_master_value
        if mapping_type == LegacyImportMasterMapping.MappingType.VEHICLE_MODEL
        else normalize_legacy_master_value
    )
    unresolved_keys = {normalize_unresolved(value) for value in unresolved_values}
    if normalize_unresolved(source_value) not in unresolved_keys:
        messages.info(request, f"「{source_value or type_label}」已由其他操作處理，清單已更新。")
        return redirect(
            f"{reverse('legacy_import_detail', args=[batch.pk])}#master-data-workspace"
        )

    action = request.POST.get("resolution_action", "")
    actor_name = _editing_name(request.user)
    try:
        if action == "link":
            if mapping_type == LegacyImportMasterMapping.MappingType.VEHICLE_MODEL:
                form = LegacyVehicleModelLinkForm(request.POST, prefix="model-link")
                if not form.is_valid():
                    raise ValueError(_form_error_text(form))
                target = form.cleaned_data["vehicle_model"]
                save_import_master_mapping(
                    mapping_type=mapping_type,
                    source_value=source_value,
                    actor_name=actor_name,
                    vehicle_model=target,
                )
            else:
                form = LegacySalesSourceLinkForm(request.POST, prefix="source-link")
                if not form.is_valid():
                    raise ValueError(_form_error_text(form))
                target = form.cleaned_data["sales_source"]
                save_import_master_mapping(
                    mapping_type=mapping_type,
                    source_value=source_value,
                    actor_name=actor_name,
                    sales_source=target,
                )
            success_text = f"已將「{source_value}」對應至「{target}」。"
        elif action == "create":
            if mapping_type == LegacyImportMasterMapping.MappingType.VEHICLE_MODEL:
                form = LegacyVehicleModelQuickCreateForm(request.POST, prefix="model-create")
                if not form.is_valid():
                    raise ValueError(_form_error_text(form))
                target = form.save()
                save_import_master_mapping(
                    mapping_type=mapping_type,
                    source_value=source_value,
                    actor_name=actor_name,
                    vehicle_model=target,
                    note="由歷史匯入補齊工作台快速建立",
                )
            else:
                form = LegacySalesSourceQuickCreateForm(request.POST, prefix="source-create")
                if not form.is_valid():
                    raise ValueError(_form_error_text(form))
                target = form.save()
                save_import_master_mapping(
                    mapping_type=mapping_type,
                    source_value=source_value,
                    actor_name=actor_name,
                    sales_source=target,
                    note="由歷史匯入補齊工作台快速建立",
                )
            success_text = f"已建立「{target}」，並完成「{source_value}」的對應。"
        elif action == "ignore":
            save_import_master_mapping(
                mapping_type=mapping_type,
                source_value=source_value,
                actor_name=actor_name,
                ignored=True,
                note=request.POST.get("note", ""),
            )
            success_text = f"已將「{source_value}」標記為保留歷史文字，不建立{type_label}主檔。"
        else:
            raise ValueError("請選擇對應、快速新增或保留歷史文字。")
        summary = revalidate_import_batch(batch)
    except (ValueError, ValidationError, IntegrityError) as exc:
        messages.error(request, f"無法處理「{source_value}」：{exc}")
    else:
        remaining = len(summary["validation"].get(validation_key, []))
        messages.success(request, f"{success_text} 尚有 {remaining} 個未對應{type_label}。")
    return redirect(
        f"{reverse('legacy_import_detail', args=[batch.pk])}#master-data-workspace"
    )


@login_required
@require_http_methods(["POST"])
@transaction.atomic
def legacy_import_row_decide(request, pk, row_pk):
    batch = get_object_or_404(LegacyImportBatch.objects.select_for_update(), pk=pk)
    row = get_object_or_404(batch.rows.select_for_update(), pk=row_pk)
    form = LegacyImportRowCorrectionForm(request.POST, row=row)
    if not form.is_valid():
        rows = batch.rows.filter(action=row.action)
        page = Paginator(rows, 100).get_page(1)
        for preview_row in page.object_list:
            preview_row.display_messages = [
                friendly_import_message(message) for message in preview_row.messages
            ]
        counts = {value: 0 for value, _label in LegacyImportRow.Action.choices}
        for item in batch.rows.values("action").annotate(total=Count("id")):
            counts[item["action"]] = item["total"]
        return render(
            request,
            "sales/legacy_import_detail.html",
            {
                "batch": batch,
                "rows": page.object_list,
                "page_obj": page,
                "selected_action": row.action,
                "selected_action_label": row.get_action_display(),
                "search_query": "",
                "counts": counts,
                "unresolved_count": counts.get("conflict", 0) + counts.get("error", 0),
                "can_confirm": False,
                "conflict_groups": [],
                "editing_row": row,
                "correction_form": form,
                "master_workspace": {"models": [], "sources": [], "total": 0},
                "vehicle_model_link_form": LegacyVehicleModelLinkForm(prefix="model-link"),
                "sales_source_link_form": LegacySalesSourceLinkForm(prefix="source-link"),
                "vehicle_model_quick_form": LegacyVehicleModelQuickCreateForm(prefix="model-create"),
                "sales_source_quick_form": LegacySalesSourceQuickCreateForm(prefix="source-create"),
            },
            status=400,
        )
    try:
        if batch.status == LegacyImportBatch.Status.COMPLETED:
            retry_result = retry_completed_import_row(
                row,
                form.cleaned_mapping(),
                form.cleaned_data["decision"],
                form.cleaned_data["reason"],
                _editing_name(request.user),
            )
            if not retry_result["ok"]:
                messages.error(request, f"仍無法匯入 Excel 第 {row.source_row} 列：{retry_result['error']}")
                return redirect(
                    f"{reverse('legacy_import_detail', args=[batch.pk])}?action=error&edit={row.pk}#row-editor"
                )
            if retry_result["excluded"]:
                messages.success(request, f"已排除 Excel 第 {row.source_row} 列，不影響其他已匯入資料。")
            else:
                messages.success(request, f"已修正並補匯 Excel 第 {row.source_row} 列。")
            if batch.rows.filter(action=LegacyImportRow.Action.ERROR).exists():
                return redirect(f"{reverse('legacy_import_detail', args=[batch.pk])}?action=error#row-preview")
            return redirect("legacy_import_detail", pk=batch.pk)
        summary = apply_import_row_decision(
            row,
            form.cleaned_mapping(),
            form.cleaned_data["decision"],
            form.cleaned_data["reason"],
            _editing_name(request.user),
        )
    except ValueError as exc:
        messages.error(request, str(exc))
    else:
        remaining = summary["counts"].get("conflict", 0) + summary["counts"].get("error", 0)
        if form.cleaned_data["decision"] == "exclude":
            messages.success(request, f"已排除 {row.sheet_name}第 {row.source_row} 列；尚有 {remaining} 筆問題待處理。")
        else:
            messages.success(request, f"已儲存修正並重新驗證；尚有 {remaining} 筆問題待處理。")
    remaining_action = "conflict" if batch.rows.filter(action="conflict").exists() else "error"
    if not batch.rows.filter(action__in=["conflict", "error"]).exists():
        return redirect("legacy_import_detail", pk=batch.pk)
    return redirect(f"{reverse('legacy_import_detail', args=[batch.pk])}?action={remaining_action}#issue-review")


@login_required
@require_http_methods(["POST"])
@transaction.atomic
def legacy_import_delete(request, pk):
    batch = get_object_or_404(LegacyImportBatch.objects.select_for_update(), pk=pk)
    if batch.status in {LegacyImportBatch.Status.PROCESSING, LegacyImportBatch.Status.COMPLETED}:
        if batch.status == LegacyImportBatch.Status.PROCESSING:
            messages.error(request, "背景匯入正在執行，為避免資料不完整，目前不能刪除此批次。")
            return redirect("legacy_import_detail", pk=batch.pk)
        messages.error(request, "已完成匯入的批次不能直接刪除，避免破壞正式訂單與庫存。")
        return redirect("legacy_import_detail", pk=batch.pk)
    filename = batch.original_filename
    stored_name = batch.source_file.name
    storage = batch.source_file.storage
    batch.delete()
    if stored_name:
        transaction.on_commit(lambda: storage.delete(stored_name))
    messages.success(request, f"已刪除匯入批次：{filename}。正式資料未受影響。")
    return redirect("legacy_import_list")


@login_required
@require_http_methods(["POST"])
def legacy_import_archive(request, pk):
    batch = get_object_or_404(LegacyImportBatch, pk=pk)
    if batch.status != LegacyImportBatch.Status.COMPLETED:
        messages.error(request, "只有已完成匯入的批次需要封存。")
        return redirect("legacy_import_detail", pk=batch.pk)
    batch.archived_at = timezone.now()
    batch.archived_by = _editing_name(request.user)
    batch.save(update_fields=["archived_at", "archived_by", "updated_at"])
    messages.success(request, "匯入紀錄已封存；正式資料與稽核紀錄仍完整保留。")
    return redirect("legacy_import_list")


@login_required
@require_http_methods(["POST"])
def legacy_import_restore(request, pk):
    batch = get_object_or_404(LegacyImportBatch, pk=pk)
    batch.archived_at = None
    batch.archived_by = ""
    batch.save(update_fields=["archived_at", "archived_by", "updated_at"])
    messages.success(request, "匯入紀錄已恢復顯示。")
    return redirect("legacy_import_detail", pk=batch.pk)


@login_required
@require_http_methods(["POST"])
def legacy_import_confirm(request, pk):
    actor_name = _editing_name(request.user)
    previous_status = None
    with transaction.atomic():
        batch = get_object_or_404(LegacyImportBatch.objects.select_for_update(), pk=pk)
        if batch.status == LegacyImportBatch.Status.PROCESSING:
            messages.info(request, "這個批次已在背景匯入，不需要重複送出。")
            return redirect("legacy_import_detail", pk=pk)
        if batch.status not in {LegacyImportBatch.Status.PREVIEW, LegacyImportBatch.Status.FAILED}:
            messages.error(request, "這個批次已完成，不能重複匯入。")
            return redirect("legacy_import_detail", pk=pk)
        if batch.status == LegacyImportBatch.Status.FAILED and not batch.processing_started_at:
            messages.error(request, "這是檔案解析錯誤，請修正 Excel 後重新上傳。")
            return redirect("legacy_import_detail", pk=pk)
        unresolved = batch.rows.filter(
            action__in=[LegacyImportRow.Action.CONFLICT, LegacyImportRow.Action.ERROR]
        ).count()
        if unresolved:
            messages.error(request, f"尚有 {unresolved} 筆衝突或錯誤資料，請先修正或排除。")
            return redirect("legacy_import_detail", pk=pk)
        previous_status = batch.status
        total = batch.rows.count()
        batch.status = LegacyImportBatch.Status.PROCESSING
        batch.processing_started_at = timezone.now()
        batch.processing_finished_at = None
        batch.processing_heartbeat_at = timezone.now()
        batch.processing_error = ""
        batch.processing_total = total
        batch.processing_completed = 0
        batch.processing_job_id = ""
        batch.save(
            update_fields=[
                "status",
                "processing_started_at",
                "processing_finished_at",
                "processing_heartbeat_at",
                "processing_error",
                "processing_total",
                "processing_completed",
                "processing_job_id",
                "updated_at",
            ]
        )
    job_id = f"legacy-import-{batch.pk}-{uuid.uuid4().hex[:12]}"
    try:
        queue = django_rq.get_queue("imports")
        job = queue.enqueue(
            run_legacy_import_job,
            str(batch.pk),
            actor_name,
            job_timeout=3600,
            job_id=job_id,
        )
    except Exception:
        logger.exception("無法排入歷史資料背景匯入", extra={"batch_id": str(batch.pk)})
        LegacyImportBatch.objects.filter(
            pk=batch.pk,
            status=LegacyImportBatch.Status.PROCESSING,
            processing_job_id="",
        ).update(
            status=previous_status,
            processing_error="無法啟動背景匯入，請稍後再試。",
            processing_finished_at=timezone.now(),
            updated_at=timezone.now(),
        )
        messages.error(request, "無法啟動背景匯入，資料尚未寫入，請稍後再試。")
    else:
        LegacyImportBatch.objects.filter(pk=batch.pk).update(
            processing_job_id=getattr(job, "id", job_id) or job_id,
            updated_at=timezone.now(),
        )
        messages.success(request, "已開始背景匯入。你可以留在此頁查看進度，也可以先處理其他工作。")
    return redirect("legacy_import_detail", pk=pk)


@login_required
def legacy_import_status(request, pk):
    batch = get_object_or_404(LegacyImportBatch, pk=pk)
    total = batch.processing_total
    completed = min(batch.processing_completed, total) if total else batch.processing_completed
    percent = min(100, int(completed * 100 / total)) if total else 0
    return JsonResponse(
        {
            "status": batch.status,
            "status_label": batch.get_status_display(),
            "total": total,
            "completed": completed,
            "percent": percent,
            "result": batch.result_summary or {},
            "error": batch.processing_error,
            "finished": batch.status in {LegacyImportBatch.Status.COMPLETED, LegacyImportBatch.Status.FAILED},
        }
    )


@login_required
def customer_list(request):
    latest_id = (
        SalesOrder.objects.filter(owner_id_number=OuterRef("owner_id_number"))
        .order_by("-order_date", "-id")
        .values("id")[:1]
    )
    order_count = (
        SalesOrder.objects.filter(owner_id_number=OuterRef("owner_id_number"))
        .values("owner_id_number")
        .annotate(total=Count("id"))
        .values("total")[:1]
    )
    customers = SalesOrder.objects.filter(id=Subquery(latest_id)).annotate(
        order_count=Subquery(order_count)
    )
    keyword = request.GET.get("q", "").strip()
    if keyword:
        customers = customers.filter(
            Q(owner_name__icontains=keyword)
            | Q(owner_phone__icontains=keyword)
            | Q(owner_email__icontains=keyword)
            | Q(owner_id_number__icontains=keyword)
            | Q(owner_address__icontains=keyword)
        )
    customers = customers.order_by("owner_name", "-order_date")
    page = Paginator(customers, 100).get_page(request.GET.get("page"))
    return render(
        request,
        "sales/customer_list.html",
        {
            "customers": page.object_list,
            "page_obj": page,
            "keyword": keyword,
        },
    )


@login_required
def customer_detail(request, pk):
    customer = get_object_or_404(SalesOrder, pk=pk)
    orders = (
        SalesOrder.objects.filter(owner_id_number=customer.owner_id_number)
        .select_related("vehicle_model", "color", "source")
        .order_by("-order_date", "-id")
    )
    page = Paginator(orders, 100).get_page(request.GET.get("page"))
    return render(
        request,
        "sales/customer_detail.html",
        {
            "customer": customer,
            "orders": page.object_list,
            "page_obj": page,
        },
    )


INVENTORY_HISTORY_FIELDS = {
    "vehicle_model": "車型",
    "color": "車色",
    "engine_number": "引擎號碼",
    "frame_number": "車身號碼",
    "current_dealer": "實際位置",
    "received_on": "進車日期",
    "condition_note": "車況說明",
    "condition_photo": "車況照片",
    "condition_resolution": "處理結果",
}


def _inventory_values(vehicle):
    return {
        "vehicle_model": (vehicle.vehicle_model_id, str(vehicle.vehicle_model)),
        "color": (vehicle.color_id, vehicle.color.name),
        "engine_number": (vehicle.engine_number or "", vehicle.engine_number or "未填寫"),
        "frame_number": (vehicle.frame_number or "", vehicle.frame_number or "未填寫"),
        "current_dealer": (
            vehicle.current_dealer_id or "store",
            vehicle.actual_location_label,
        ),
        "received_on": (str(vehicle.received_on), str(vehicle.received_on)),
        "condition_note": (vehicle.condition_note, vehicle.condition_note or "未填寫"),
        "condition_photo": (
            vehicle.condition_photo.name if vehicle.condition_photo else "",
            "有照片" if vehicle.condition_photo else "無照片",
        ),
        "condition_resolution": (
            vehicle.condition_resolution,
            vehicle.condition_resolution or "未填寫",
        ),
    }


def _create_inventory_history(
    vehicle,
    *,
    actor_name,
    event_type,
    reason="",
    changes=None,
    from_location_label="",
    to_location_label="",
):
    history = VehicleInventoryHistory(
        vehicle=vehicle,
        event_type=event_type,
        actor_name=actor_name,
        reason=reason,
        changes=changes or {},
        status_snapshot=vehicle.status,
        location_store_snapshot_id=vehicle.location_store_id,
        location_label_snapshot=vehicle.actual_location_label,
        condition_note_snapshot=vehicle.condition_note,
        condition_resolution_snapshot=vehicle.condition_resolution,
        from_location_label=from_location_label,
        to_location_label=to_location_label,
    )
    if vehicle.condition_photo:
        vehicle.condition_photo.open("rb")
        try:
            content = ContentFile(vehicle.condition_photo.read())
        finally:
            vehicle.condition_photo.close()
        suffix = Path(vehicle.condition_photo.name).suffix or ".jpg"
        history.condition_photo_snapshot.save(
            f"{uuid.uuid4().hex}{suffix}", content, save=False
        )
    history.save()
    return history


def _session_key(request):
    if not request.session.session_key:
        request.session.create()
    return request.session.session_key


def _claim_edit_lock(instance, request, timeout):
    session_key = _session_key(request)
    now = timezone.now()
    active_elsewhere = bool(
        instance.editing_session
        and instance.editing_session != session_key
        and instance.editing_at
        and instance.editing_at >= now - timeout
    )
    if active_elsewhere:
        return False
    instance.editing_session = session_key
    instance.editing_by = _editing_name(request.user)
    instance.editing_at = now
    instance.save(
        update_fields=["editing_session", "editing_by", "editing_at"]
    )
    return True


def _vehicle_rate_data():
    return {
        str(model.pk): {
            "energy_type": model.energy_type,
            "displacement_cc": model.displacement_cc,
            "electric_registration_class": model.electric_registration_class,
        }
        for model in VehicleModel.objects.filter(active=True).only(
            "id", "energy_type", "displacement_cc"
        )
    }


def _accessory_product_data():
    return {
        str(product.pk): {
            "name": product.name,
            "sale_price": str(product.sale_price),
            "labor_fee": str(product.labor_fee),
        }
        for product in AccessoryProduct.objects.filter(active=True).only(
            "id", "name", "sale_price", "labor_fee"
        )
    }


def app_version(request):
    from config.app_version import get_app_version

    response = JsonResponse({"version": get_app_version()})
    response["Cache-Control"] = "no-store, no-cache, must-revalidate"
    response["Pragma"] = "no-cache"
    return _protect_private_response(response)


@login_required
def dashboard(request):
    query = request.GET.get("q", "").strip()
    orders = SalesOrder.objects.select_related(
        "source",
        "vehicle_model",
        "color",
        "allocated_vehicle",
        "allocated_vehicle__ownership_store",
        "allocated_vehicle__location_store",
        "allocated_vehicle__current_dealer",
        "search_index",
    ).prefetch_related(
        "accessories",
        "other_fees",
        "subsidy_documents",
        "registration_documents",
        "events",
        "changes",
    )
    search_results = None
    search_result_count = 0
    if query:
        matched_orders = orders.filter(build_order_search_query(query)).distinct()
        paginator = Paginator(matched_orders, 50)
        search_results = paginator.get_page(request.GET.get("page"))
        search_result_count = paginator.count
        for order in search_results:
            order.search_matches = build_order_match_summary(order, query)

    urgent_statuses = [
        SalesOrder.Status.CANCEL_REFUND_PENDING,
        SalesOrder.Status.DELIVERED_DOCS_PENDING,
    ]
    metrics = build_dashboard_metrics() if not query else None
    urgent_orders = (
        SalesOrder.objects.select_related(
            "source", "vehicle_model", "color", "allocated_vehicle"
        )
        .filter(status__in=urgent_statuses)
        .order_by("-updated_at")[:5]
        if metrics
        else []
    )
    context = {
        "query": query,
        "search_results": search_results,
        "search_result_count": search_result_count,
        "dashboard": metrics,
        "urgent_orders": urgent_orders,
        "dealer_reminders": metrics["dealer_reminders"] if metrics else [],
        "counts": {
            "urgent": metrics["workload"]["urgent"] if metrics else 0,
            "allocation": metrics["workload"]["allocation"] if metrics else 0,
            "in_progress": metrics["workload"]["in_progress"] if metrics else 0,
            "inventory": metrics["inventory"]["available"] if metrics else 0,
        },
        "drafts": OrderDraft.objects.all()[:5],
    }
    return render(request, "sales/dashboard.html", context)


@login_required
def user_guide(request):
    """提供不含技術術語、可搜尋與列印的 End User 使用說明。"""
    return render(request, "help/user_guide.html")


@login_required
@transaction.atomic
def registration_fee_variance_confirm(request, pk):
    order = get_object_or_404(SalesOrder.objects.select_for_update(), pk=pk)
    if request.method == "POST":
        if order.registration_calculated_total == order.plate_insurance_fee:
            messages.info(request, "系統試算與實際牌險已相同，無須另外確認。")
        else:
            order.registration_fee_variance_confirmed_at = timezone.now()
            order.registration_fee_variance_confirmed_by = _editing_name(request.user)
            order.registration_fee_variance_confirmed_calculated_total = (
                order.registration_calculated_total
            )
            order.registration_fee_variance_confirmed_actual_total = (
                order.plate_insurance_fee
            )
            order.save(
                update_fields=[
                    "registration_fee_variance_confirmed_at",
                    "registration_fee_variance_confirmed_by",
                    "registration_fee_variance_confirmed_calculated_total",
                    "registration_fee_variance_confirmed_actual_total",
                    "updated_at",
                ]
            )
            OrderEvent.objects.create(
                order=order,
                event_type="registration_fee_variance_confirmed",
                description=(
                    f"確認牌險差額：系統 {order.registration_calculated_total:,.0f} 元／"
                    f"實際 {order.plate_insurance_fee:,.0f} 元"
                ),
                actor_name=_editing_name(request.user),
            )
            messages.success(request, "牌險差額已確認並保存紀錄。")
    return redirect("dashboard")


@login_required
def order_list(request):
    orders = SalesOrder.objects.select_related(
        "source", "vehicle_model", "color", "search_index"
    )
    query = request.GET.get("q", "").strip()
    if query:
        orders = orders.filter(build_order_search_query(query)).distinct()
    status = request.GET.get("status")
    if status == "registration_pending":
        orders = orders.filter(
            status__in=[
                SalesOrder.Status.ALLOCATED,
                SalesOrder.Status.TRANSFER_PENDING,
                SalesOrder.Status.IN_TRANSFER,
            ]
        )
    elif status == "in_progress":
        orders = orders.exclude(
            status__in=[
                SalesOrder.Status.ALLOCATION_PENDING,
                SalesOrder.Status.CANCEL_REFUND_PENDING,
                SalesOrder.Status.DELIVERED_DOCS_PENDING,
                SalesOrder.Status.COMPLETED,
                SalesOrder.Status.CANCELLED,
            ]
        )
    elif status == "urgent":
        orders = orders.filter(
            status__in=[
                SalesOrder.Status.CANCEL_REFUND_PENDING,
                SalesOrder.Status.DELIVERED_DOCS_PENDING,
            ]
        )
    elif status:
        orders = orders.filter(status=status)
    orders = orders.order_by("-order_date", "-created_at", "-pk")
    page = Paginator(orders, 50).get_page(request.GET.get("page"))
    if query:
        for order in page.object_list:
            order.search_matches = build_order_match_summary(order, query)
    query_params = request.GET.copy()
    query_params.pop("page", None)
    return render(
        request,
        "sales/order_list.html",
        {
            "orders": page.object_list,
            "page_obj": page,
            "query": query,
            "query_params": query_params.urlencode(),
            "statuses": SalesOrder.Status.choices,
        },
    )


def _operations_report_queryset(request):
    rows = SalesOrder.objects.select_related(
        "vehicle_model", "color", "allocated_vehicle", "operations", "source"
    ).prefetch_related("payment_records")
    keyword = request.GET.get("q", "").strip()
    if keyword:
        rows = rows.filter(build_order_search_query(keyword)).distinct()
    energy_type = request.GET.get("energy_type", "")
    if energy_type in {value for value, _ in VehicleModel.EnergyType.choices}:
        rows = rows.filter(vehicle_model__energy_type=energy_type)
    payment_status = request.GET.get("payment_status", "")
    if payment_status == "confirmed":
        rows = rows.filter(operations__payment_confirmed=True)
    elif payment_status == "pending":
        rows = rows.exclude(operations__payment_confirmed=True)
    date_basis = request.GET.get("date_basis", "registration")
    date_field = {
        "order": "order_date",
        "delivery": "delivered_at__date",
        "registration": "registration_date",
    }.get(date_basis, "registration_date")
    sort_field = "delivered_at" if date_basis == "delivery" else date_field
    if request.GET.get("date_from"):
        rows = rows.filter(**{f"{date_field}__gte": request.GET["date_from"]})
    if request.GET.get("date_to"):
        rows = rows.filter(**{f"{date_field}__lte": request.GET["date_to"]})
    if request.GET.get("include_cancelled") != "1":
        rows = rows.exclude(status=SalesOrder.Status.CANCELLED)
    return rows.order_by(f"-{sort_field}", "-id")


def _operations_analysis(rows):
    summary = {
        "count": 0,
        "vehicle_sales": Decimal("0"),
        "actual_received": Decimal("0"),
        "net_profit": Decimal("0"),
        "profit_ready": 0,
    }
    models = {}
    for order in rows:
        profile = getattr(order, "operations", None)
        received = profile.total_received if profile else Decimal("0")
        profit = profile.net_profit if profile and profile.vehicle_cost else Decimal("0")
        summary["count"] += 1
        summary["vehicle_sales"] += order.vehicle_price
        summary["actual_received"] += received
        summary["net_profit"] += profit
        if profile and profile.vehicle_cost:
            summary["profit_ready"] += 1
        key = order.vehicle_model_id
        bucket = models.setdefault(
            key,
            {
                "label": f"{order.vehicle_model.brand} {order.vehicle_model.name}".strip(),
                "count": 0,
                "vehicle_sales": Decimal("0"),
                "actual_received": Decimal("0"),
                "net_profit": Decimal("0"),
            },
        )
        bucket["count"] += 1
        bucket["vehicle_sales"] += order.vehicle_price
        bucket["actual_received"] += received
        bucket["net_profit"] += profit
    summary["average_price"] = summary["vehicle_sales"] / summary["count"] if summary["count"] else Decimal("0")
    return summary, sorted(models.values(), key=lambda item: (-item["count"], item["label"]))


@login_required
def operations_report(request):
    rows = _operations_report_queryset(request)
    analysis_summary, model_breakdown = _operations_analysis(rows)
    paginator = Paginator(rows, 100)
    page = paginator.get_page(request.GET.get("page"))
    for order in page.object_list:
        order.operation_data = getattr(order, "operations", None)
    return render(
        request,
        "sales/operations_report.html",
        {
            "orders": page.object_list,
            "page_obj": page,
            "energy_types": VehicleModel.EnergyType.choices,
            "selected": request.GET,
            "analysis_summary": analysis_summary,
            "model_breakdown": model_breakdown,
            "date_basis_label": {
                "order": "訂單日期",
                "delivery": "實際交付日期",
                "registration": "實際領牌日期",
            }.get(request.GET.get("date_basis", "registration"), "實際領牌日期"),
        },
    )


def _reconciliation_queryset(request):
    records = PaymentRecord.objects.select_related(
        "order",
        "order__source",
        "order__vehicle_model",
        "order__color",
    ).filter(
        Q(system_key="installment_disbursement")
        | Q(
            system_key="balance",
            order__source_type__in=[
                SalesOrder.SourceType.PLATFORM,
                SalesOrder.SourceType.DEALER,
            ],
        )
    )
    keyword = request.GET.get("q", "").strip()
    if keyword:
        records = records.filter(
            Q(order__number__icontains=keyword)
            | Q(order__owner_name__icontains=keyword)
            | Q(order__installment_company__icontains=keyword)
            | Q(order__source__name__icontains=keyword)
            | Q(order__final_plate_number__icontains=keyword)
        )
    channel = request.GET.get("channel", "")
    if channel == "installment":
        records = records.filter(system_key="installment_disbursement")
    elif channel == "platform":
        records = records.filter(
            system_key="balance",
            order__source_type=SalesOrder.SourceType.PLATFORM,
        )
    elif channel == "dealer":
        records = records.filter(
            system_key="balance",
            order__source_type=SalesOrder.SourceType.DEALER,
        )
    status = request.GET.get("status", "")
    if status == "confirmed":
        records = records.filter(confirmed=True)
    elif status == "pending":
        records = records.filter(confirmed=False)
    if request.GET.get("date_from"):
        records = records.filter(order__order_date__gte=request.GET["date_from"])
    if request.GET.get("date_to"):
        records = records.filter(order__order_date__lte=request.GET["date_to"])
    return records.order_by("confirmed", "-order__order_date", "-order_id")


def _decorate_reconciliation_record(record):
    if record.system_key == "installment_disbursement":
        record.reconciliation_channel = "installment"
        record.reconciliation_channel_label = "分期公司"
        record.reconciliation_party = record.order.installment_company or "未填分期公司"
    elif record.order.source_type == SalesOrder.SourceType.PLATFORM:
        record.reconciliation_channel = "platform"
        record.reconciliation_channel_label = "網路平台"
        record.reconciliation_party = (
            record.order.source.name if record.order.source_id else "未填平台"
        )
    else:
        record.reconciliation_channel = "dealer"
        record.reconciliation_channel_label = "合作車行"
        record.reconciliation_party = (
            record.order.source.name if record.order.source_id else "未填車行"
        )
    record.reconciliation_difference = (
        record.received_amount - record.expected_amount
    )
    return record


@login_required
def reconciliation_list(request):
    page = Paginator(_reconciliation_queryset(request), 100).get_page(
        request.GET.get("page")
    )
    records = [_decorate_reconciliation_record(record) for record in page.object_list]
    return render(
        request,
        "sales/reconciliation_list.html",
        {
            "records": records,
            "page_obj": page,
            "selected": request.GET,
        },
    )


@login_required
@transaction.atomic
def reconciliation_update(request, pk):
    record = get_object_or_404(
        PaymentRecord.objects.select_for_update().select_related("order"),
        pk=pk,
    )
    eligible = record.system_key == "installment_disbursement" or (
        record.system_key == "balance"
        and record.order.source_type
        in {SalesOrder.SourceType.PLATFORM, SalesOrder.SourceType.DEALER}
    )
    if request.method != "POST" or not eligible:
        messages.error(request, "此筆資料不屬於統一對帳範圍。")
        return redirect("reconciliation_list")
    before = {
        "預計金額": str(record.expected_amount),
        "預計金額調整原因": record.expected_amount_override_reason,
        "實際金額": str(record.received_amount),
        "入帳日期": str(record.received_on or ""),
        "收款帳戶": record.receiving_account,
        "確認狀態": "已確認" if record.confirmed else "待確認",
    }
    form = ReconciliationRecordForm(request.POST, instance=record)
    if form.is_valid():
        record = form.save(commit=False)
        if record.confirmed:
            record.confirmed_by = _editing_name(request.user)
            record.confirmed_at = timezone.now()
        else:
            record.confirmed_by = ""
            record.confirmed_at = None
        record.save()
        if record.confirmed and (
            record.system_key == "installment_disbursement"
            or record.order.source_type == SalesOrder.SourceType.PLATFORM
        ):
            profile, _created = OrderOperationsProfile.objects.get_or_create(
                order=record.order
            )
            profile.actual_disbursement = record.received_amount
            protected = set(profile.manual_financial_fields or [])
            protected.add("actual_disbursement")
            profile.manual_financial_fields = sorted(protected)
            profile.updated_by = _editing_name(request.user)
            profile.save(
                update_fields=[
                    "actual_disbursement",
                    "manual_financial_fields",
                    "updated_by",
                    "updated_at",
                ]
            )
        refresh_payment_confirmation(record.order_id)
        after = {
            "預計金額": str(record.expected_amount),
            "預計金額調整原因": record.expected_amount_override_reason,
            "實際金額": str(record.received_amount),
            "入帳日期": str(record.received_on or ""),
            "收款帳戶": record.receiving_account,
            "確認狀態": "已確認" if record.confirmed else "待確認",
        }
        changes = {
            key: {"before": before[key], "after": value}
            for key, value in after.items()
            if before[key] != value
        }
        OrderChange.objects.create(
            order=record.order,
            reason=f"統一對帳更新：{record.item_name}",
            changes=changes,
            actor_name=_editing_name(request.user),
        )
        OrderEvent.objects.create(
            order=record.order,
            event_type="reconciliation_updated",
            description=(
                f"更新{record.item_name}：預計 {record.expected_amount:.0f} 元、"
                f"實際 {record.received_amount:.0f} 元。"
            ),
            actor_name=_editing_name(request.user),
        )
        messages.success(request, f"已更新 {record.order.number} 的對帳資料。")
    else:
        messages.error(
            request,
            "對帳更新失敗：" + " ".join(
                error for errors in form.errors.values() for error in errors
            ),
        )
    next_url = request.POST.get("next")
    if not url_has_allowed_host_and_scheme(
        url=next_url or "",
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        next_url = None
    return redirect(next_url or "reconciliation_list")


@login_required
def operations_report_export(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "營運總表"
    headers = [
        "訂單編號", "訂單日期", "車種", "型號", "型式", "顏色", "引擎／車身號碼",
        "車主名稱", "車牌號碼", "車款售價", "實際撥款", "成本",
        "總收款金額", "確認收款", "分期公司", "期數", "每期金額",
        "分期公司確認匯款", "身分證字號", "西元生日", "民國生日",
        "戶籍地址", "手機", "Email", "自送托運地點", "發票日期",
        "尾款發票號碼", "補助方案", "補助金額", "銀行", "匯款帳戶",
        "申請日", "工業局", "環境部", "縣市政府", "舊車車主",
        "舊車車主身分證", "舊車牌照號碼", "舊車引擎號碼", "舊車廠牌",
        "排氣量", "出廠日期", "報廢日期", "回收日期",
        "領牌稅金支出", "強制險支出", "選號支出", "贈品支出", "運費支出",
        "車行傭金支出", "銀行刷卡手續費支出", "領牌稅金收入", "強制險收入",
        "代辦費收入", "選號收入", "分期手續費收入", "刷卡手續費收入",
        "其他收入", "報廢代辦收入", "報廢車收入", "實銷獎勵金", "促銷補助金", "分期補貼息", "強制險傭金",
        "信用卡傭金", "車控帳號", "電池合約方案", "電池合約啟用日期",
        "電池合約帳號", "安全帽", "公司禮券、匯款", "其他",
        "平台贈品", "客服電話", "分期資訊", "車行",
        "總收入", "總支出", "單筆淨利",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="174C3C")
    for order in _operations_report_queryset(request):
        profile = getattr(order, "operations", None)
        roc_birth = ""
        if order.owner_birth_date:
            roc_birth = (
                f"{order.owner_birth_date.year - 1911}/"
                f"{order.owner_birth_date.month:02d}/{order.owner_birth_date.day:02d}"
            )
        def op(name, default=""):
            return getattr(profile, name, default) if profile else default
        sheet.append(sanitize_excel_row([
            order.number, order.order_date, order.vehicle_model.name,
            order.vehicle_model.model_number,
            order.vehicle_model.get_model_code_display()
            if order.vehicle_model.model_code else "",
            order.color.name,
            order.allocated_vehicle.identifier if order.allocated_vehicle else "",
            order.owner_name, order.final_plate_number, order.vehicle_price,
            op("actual_disbursement", 0), op("vehicle_cost", 0),
            profile.total_received if profile else 0,
            "是" if op("payment_confirmed", False) else "否",
            order.installment_company, order.installment_periods,
            order.installment_monthly,
            "是" if op("installment_transfer_confirmed", False) else "否",
            order.owner_id_number, order.owner_birth_date, roc_birth,
            order.owner_address, order.owner_phone, order.owner_email,
            order.delivery_destination, op("invoice_date"),
            op("balance_invoice_number"), order.subsidy_type,
            op("subsidy_amount", 0), op("bank_name"), op("remittance_account"),
            op("subsidy_applied_on"),
            profile.get_industry_bureau_status_display() if profile else "",
            profile.get_environment_ministry_status_display() if profile else "",
            profile.get_local_government_status_display() if profile else "",
            order.old_owner_name, order.old_owner_id_number, order.trade_in_plate,
            op("old_vehicle_engine_number"), op("old_vehicle_brand"),
            op("old_vehicle_displacement_cc"), op("old_vehicle_manufactured_on"),
            op("scrapped_on"), op("recycled_on"),
            op("registration_tax_expense", 0),
            op("compulsory_insurance_expense", 0),
            op("plate_selection_expense", 0),
            op("gift_expense", 0),
            op("shipping_expense", 0),
            op("dealer_commission_expense", 0),
            op("card_fee_expense", 0),
            op("registration_tax_income", 0),
            op("compulsory_insurance_income", 0),
            op("agency_fee_income", 0),
            op("plate_selection_income", 0),
            op("installment_fee_income", 0),
            op("card_fee_income", 0),
            op("other_income", 0), op("scrap_agency_income", 0),
            op("scrap_vehicle_income", 0), op("sales_bonus", 0),
            op("promotion_subsidy", 0),
            op("installment_interest_subsidy", 0),
            op("insurance_commission", 0),
            op("credit_card_commission", 0), op("vehicle_control_account"),
            op("battery_plan"), op("battery_activated_on"),
            op("battery_account"), op("helmet"),
            op("company_gift_or_remittance"), op("other_fulfillment"),
            op("platform_gift"), op("customer_service_phone"),
            op("installment_info"), op("dealer_name"),
            profile.total_income if profile else 0,
            profile.total_expense if profile else 0,
            profile.net_profit if profile else 0,
        ]))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = min(
            max(len(str(cell.value or "")) for cell in column) + 2, 28
        )
    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="operations-report.xlsx"'
    return response


@login_required
@transaction.atomic
def order_create(request):
    draft_id = request.POST.get("_draft_id") or request.GET.get("draft")
    draft = (
        get_object_or_404(OrderDraft.objects.select_for_update(), pk=draft_id)
        if draft_id
        else None
    )
    if draft and not _claim_edit_lock(draft, request, DRAFT_PRESENCE_TIMEOUT):
        messages.error(
            request,
            f"此草稿目前由 {draft.editing_by or '其他人員'} 編輯，暫時無法進入。",
        )
        return redirect("dashboard")
    existing_documents = {
        "id_front": bool(draft and draft.id_front),
        "id_back": bool(draft and draft.id_back),
    }
    if request.method == "POST":
        post_data = request.POST
        form = SalesOrderForm(
            post_data,
            request.FILES,
            existing_documents=existing_documents,
        )
        formset = AccessoryFormSet(post_data)
        fee_formset = OtherFeeFormSet(post_data, prefix="other_fees")
        if form.is_valid() and formset.is_valid() and fee_formset.is_valid():
            order = form.save(commit=False)
            if draft:
                if not form.cleaned_data.get("id_front") and draft.id_front:
                    order.id_front = draft.id_front.name
                    draft.id_front = ""
                if not form.cleaned_data.get("id_back") and draft.id_back:
                    order.id_back = draft.id_back.name
                    draft.id_back = ""
            order.status = SalesOrder.Status.ALLOCATION_PENDING
            order.save()
            apply_order_price_snapshot(order)
            apply_order_installment_snapshot(order)
            formset.instance = order
            formset.save()
            fee_formset.instance = order
            fee_formset.save()
            order.calculated_balance = order.calculate_balance()
            order.actual_balance = order.calculated_balance
            order.save(
                update_fields=["calculated_balance", "actual_balance", "updated_at"]
            )
            OrderEvent.objects.create(
                order=order,
                event_type="created",
                description="建立訂單，進入待配車。",
                actor_name=request.user.get_username(),
            )
            if draft:
                draft.delete_with_files()
            messages.success(request, "訂單已建立並進入待配車。")
            return redirect(f"{reverse('order_detail', kwargs={'pk': order.pk})}?created=1")
    else:
        initial = _draft_form_initial(draft.data) if draft else None
        form = SalesOrderForm(initial=initial)
        formset = AccessoryFormSet(
            initial=_draft_lines(
                draft.data,
                "accessories",
                (
                    "accessory_product",
                    "quantity",
                    "line_type",
                    "amount",
                    "labor_fee",
                    "note",
                ),
            )
            if draft
            else None
        )
        fee_formset = OtherFeeFormSet(
            initial=_draft_lines(draft.data, "other_fees", ("name", "amount"))
            if draft
            else None,
            prefix="other_fees",
        )
    return render(
        request,
        "sales/order_form.html",
        {
            "form": form,
            "formset": formset,
            "fee_formset": fee_formset,
            "draft": draft,
            "document_source": draft,
            "document_model": "draft",
            "vehicle_rate_data": _vehicle_rate_data(),
            "accessory_product_data": _accessory_product_data(),
        },
    )


def _draft_form_initial(data):
    return {
        key: (value[-1] if isinstance(value, list) and value else value)
        for key, value in data.items()
        if not key.startswith("accessories-")
    }


def _field_versions(post_data):
    try:
        versions = json.loads(post_data.get("_field_versions", "{}"))
    except (TypeError, ValueError, json.JSONDecodeError):
        return {}
    if not isinstance(versions, dict):
        return {}
    result = {}
    for key, version in versions.items():
        try:
            result[str(key)] = max(0, int(version))
        except (TypeError, ValueError):
            continue
    return result


def _reconcile_collaborative_post(draft, post_data):
    if not draft:
        return post_data
    reconciled = post_data.copy()
    client_versions = _field_versions(post_data)
    for state in draft.field_states.all():
        if client_versions.get(state.field_key, 0) < state.version:
            reconciled.setlist(
                state.field_key,
                state.value if isinstance(state.value, list) else [str(state.value)],
            )
    return reconciled


def _draft_lines(data, prefix, fields):
    try:
        total = int(data.get(f"{prefix}-TOTAL_FORMS", 0))
    except (TypeError, ValueError):
        total = 0
    rows = []
    for index in range(min(total, 50)):
        if data.get(f"{prefix}-{index}-DELETE"):
            continue
        row = {
            field: data.get(f"{prefix}-{index}-{field}", "")
            for field in fields
        }
        if any(str(value).strip() for value in row.values()):
            rows.append(row)
    return rows


def _validate_draft_image(upload):
    return validate_image_upload(upload, max_bytes=10 * 1024 * 1024)


@login_required
@transaction.atomic
def draft_save(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "僅接受 POST。"}, status=405)
    draft_id = request.POST.get("_draft_id")
    is_new = not draft_id
    if draft_id:
        draft = get_object_or_404(OrderDraft.objects.select_for_update(), pk=draft_id)
        if not _claim_edit_lock(draft, request, DRAFT_PRESENCE_TIMEOUT):
            return JsonResponse(
                {
                    "ok": False,
                    "locked": True,
                    "error": f"此草稿目前由 {draft.editing_by or '其他人員'} 編輯。",
                },
                status=423,
            )
        try:
            client_revision = int(request.POST.get("_draft_revision", 0))
        except (TypeError, ValueError):
            client_revision = 0
        if client_revision != draft.revision:
            return JsonResponse(
                {
                    "ok": False,
                    "conflict": True,
                    "error": "此草稿已被其他人更新，請重新載入後再編輯。",
                    "revision": draft.revision,
                },
                status=409,
            )
    else:
        draft = OrderDraft(
            created_by=request.user.get_username(),
            editing_session=_session_key(request),
            editing_by=_editing_name(request.user),
            editing_at=timezone.now(),
        )

    excluded = {
        "csrfmiddlewaretoken",
        "_draft_id",
        "_draft_revision",
        "_remove_id_front",
        "_remove_id_back",
        "_field_versions",
    }
    draft.data = {
        key: values if len(values) > 1 else values[0]
        for key, values in request.POST.lists()
        if key not in excluded
    }
    for field_name in ("id_front", "id_back"):
        if request.POST.get(f"_remove_{field_name}") == "1":
            current = getattr(draft, field_name)
            if current:
                current.delete(save=False)
            setattr(draft, field_name, "")
        upload = request.FILES.get(field_name)
        if upload:
            try:
                _validate_draft_image(upload)
            except ValidationError as exc:
                return JsonResponse(
                    {"ok": False, "error": " ".join(exc.messages)}, status=400
                )
            current = getattr(draft, field_name)
            if current:
                current.delete(save=False)
            setattr(draft, field_name, upload)
    if not is_new:
        draft.revision += 1
    draft.updated_by = request.user.get_username()
    draft.save()
    photo_urls = {
        field_name: (
            reverse(
                "protected_media",
                args=["draft", str(draft.pk), field_name],
            )
            if getattr(draft, field_name)
            else ""
        )
        for field_name in ("id_front", "id_back")
    }
    return JsonResponse(
        {
            "ok": True,
            "id": str(draft.pk),
            "revision": draft.revision,
            "updated_at": timezone.localtime(draft.updated_at).strftime("%H:%M"),
            "edit_url": f"{reverse('order_create')}?draft={draft.pk}",
            "photos": photo_urls,
        }
    )


@login_required
@transaction.atomic
def draft_delete(request, pk):
    if request.method != "POST":
        if request.headers.get("Accept") == "application/json":
            return JsonResponse(
                {"ok": False, "error": "刪除草稿僅接受 POST。"},
                status=405,
            )
        return redirect(f"{reverse('order_create')}?draft={pk}")
    draft = get_object_or_404(OrderDraft.objects.select_for_update(), pk=pk)
    if not _claim_edit_lock(draft, request, DRAFT_PRESENCE_TIMEOUT):
        error = f"此草稿目前由 {draft.editing_by or '其他人員'} 編輯，無法刪除。"
        if request.headers.get("Accept") == "application/json":
            return JsonResponse({"ok": False, "error": error}, status=423)
        messages.error(
            request,
            error,
        )
        return redirect("dashboard")
    draft.delete_with_files()
    if request.headers.get("Accept") == "application/json":
        messages.success(request, "草稿與暫存證件照片已刪除。")
        return JsonResponse({"ok": True, "redirect_url": reverse("dashboard")})
    messages.success(request, "草稿與暫存證件照片已刪除。")
    return redirect("dashboard")


@login_required
@transaction.atomic
def draft_presence(request, pk):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "僅接受 POST。"}, status=405)

    if not request.session.session_key:
        request.session.create()
    session_key = request.session.session_key
    draft = get_object_or_404(OrderDraft.objects.select_for_update(), pk=pk)
    now = timezone.now()
    is_active = bool(
        draft.editing_session
        and draft.editing_at
        and draft.editing_at >= now - DRAFT_PRESENCE_TIMEOUT
    )
    is_current_session = draft.editing_session == session_key

    if request.POST.get("action") == "release":
        if is_current_session:
            draft.editing_session = ""
            draft.editing_by = ""
            draft.editing_at = None
            draft.save(
                update_fields=[
                    "editing_session",
                    "editing_by",
                    "editing_at",
                ]
            )
        return JsonResponse({"ok": True, "active": False})

    if is_active and not is_current_session:
        return JsonResponse(
            {
                "ok": True,
                "active": True,
                "mine": False,
                "editing_by": draft.editing_by or "其他人員",
            }
        )

    draft.editing_session = session_key
    draft.editing_by = _editing_name(request.user)
    draft.editing_at = now
    draft.save(
        update_fields=[
            "editing_session",
            "editing_by",
            "editing_at",
        ]
    )
    return JsonResponse(
        {
            "ok": True,
            "active": True,
            "mine": True,
            "editing_by": draft.editing_by,
        }
    )


@login_required
def order_detail(request, pk):
    order = get_object_or_404(
        SalesOrder.objects.select_related(
            "source",
            "vehicle_model",
            "color",
            "allocated_vehicle",
            "allocated_vehicle__location_store",
            "allocated_vehicle__current_dealer",
            "delivery_record",
            "operations",
        ).prefetch_related(
            "accessories",
            "other_fees",
            "events",
            "changes",
            "registration_documents",
            "subsidy_documents",
            "subsidy_items",
            "payment_records",
        ),
        pk=pk,
    )
    if request.GET.get("tab") == "documents":
        legacy_query = request.GET.copy()
        legacy_query["tab"] = "order"
        return redirect(
            f"{reverse('order_detail', args=[pk])}?{legacy_query.urlencode()}"
            "#signed-documents"
        )
    registration_documents = {
        document.document_type: document
        for document in order.registration_documents.all()
        if document.document_type
        != RegistrationDocument.DocumentType.OTHER_INSURANCE
    }
    required_types = order.required_registration_document_types()
    registration_document_rows = [
        {
            "type": document_type,
            "label": label,
            "required": document_type in required_types,
            "document": registration_documents.get(document_type),
        }
        for document_type, label in RegistrationDocument.DocumentType.choices
        if document_type in RegistrationDocument.active_fixed_document_types()
    ]
    subsidy_documents = {
        document.document_type: document
        for document in order.subsidy_documents.all()
    }
    subsidy_required_types = order.required_subsidy_document_types()
    subsidy_required_types_when_enabled = (
        order.subsidy_document_types_when_enabled()
    )
    subsidy_document_rows = [
        {
            "type": document_type,
            "label": label,
            "required": document_type in subsidy_required_types,
            "required_when_enabled": (
                document_type in subsidy_required_types_when_enabled
            ),
            "document": subsidy_documents.get(document_type),
        }
        for document_type, label in SubsidyDocument.DocumentType.choices
        if document_type != SubsidyDocument.DocumentType.OTHER
        and (
            document_type
            not in {
                SubsidyDocument.DocumentType.OWNER_DECLARATION,
                SubsidyDocument.DocumentType.OLD_OWNER_BANKBOOK,
            }
            or document_type in subsidy_required_types
            or document_type in subsidy_required_types_when_enabled
            or document_type in subsidy_documents
        )
    ]
    operations_profile = getattr(order, "operations", None)
    balance_payment = next(
        (
            payment
            for payment in order.payment_records.all()
            if payment.system_key == "balance"
        ),
        None,
    )
    if balance_payment is None:
        sync_order_operations(order.pk)
        balance_payment = PaymentRecord.objects.filter(
            order=order,
            system_key="balance",
        ).first()
    balance_ready_for_delivery = bool(
        order.source_type == SalesOrder.SourceType.DEALER
        or (balance_payment is not None and balance_payment.is_settled)
    )
    registration_missing = order.missing_registration_requirements()
    subsidy_missing = order.missing_subsidy_requirements()
    next_actions = build_order_next_actions(
        order,
        registration_missing=registration_missing,
        subsidy_missing=subsidy_missing,
    )
    valid_tabs = {
        "order",
        "allocation",
        "subsidy",
        "registration",
        "delivery",
        "history",
    }
    requested_tab = request.GET.get("tab", "")
    active_tab = requested_tab if requested_tab in valid_tabs else "order"
    return render(
        request,
        "sales/order_detail.html",
        {
            "order": order,
            "contract_form": SignedContractForm(instance=order),
            "privacy_consent_form": PrivacyConsentForm(instance=order),
            "allocation_form": AllocationForm(order),
            "reallocation_form": ReallocationForm(order),
            "registration_form": RegistrationStageForm(instance=order),
            "registration_document_rows": registration_document_rows,
            "other_insurance_documents": order.registration_documents.filter(
                document_type=RegistrationDocument.DocumentType.OTHER_INSURANCE
            ),
            "registration_missing": registration_missing,
            "subsidy_document_rows": subsidy_document_rows,
            "other_subsidy_documents": order.subsidy_documents.filter(
                document_type=SubsidyDocument.DocumentType.OTHER
            ),
            "subsidy_missing": subsidy_missing,
            "subsidy_form": SubsidyDataForm(instance=order),
            "subsidy_item_formset": SubsidyItemFormSet(instance=order, prefix="subsidy_items"),
            "change_cards": build_order_change_cards(order.changes.all()),
            "operations_profile": operations_profile,
            "next_actions": next_actions,
            "active_tab": active_tab,
            "delivery_record": getattr(order, "delivery_record", None),
            "delivery_form": DeliveryCompletionForm(order),
            "balance_payment": balance_payment,
            "balance_payment_form": (
                DeliveryPaymentForm(instance=balance_payment)
                if balance_payment is not None
                else None
            ),
            "balance_ready_for_delivery": balance_ready_for_delivery,
            "cancellation_form": CancellationRequestForm(),
            "refund_form": RefundCompletionForm(order),
            "positioned_templates": PositionedPrintTemplate.objects.filter(active=True).order_by(
                "document_type", "-version"
            ),
        },
    )


def _operations_snapshot(profile):
    values = {}
    for field in profile._meta.fields:
        if field.name in {
            "id", "order", "created_at", "updated_at", "updated_by",
            "manual_financial_fields",
            "vehicle_control_password_encrypted", "battery_password_encrypted",
        }:
            continue
        value = getattr(profile, field.name)
        values[field.name] = "" if value is None else str(value)
    return values


@login_required
def order_operations(request, pk):
    order = get_object_or_404(
        SalesOrder.objects.select_related(
            "vehicle_model", "color", "allocated_vehicle"
        ),
        pk=pk,
    )
    sync_order_operations(order.pk)
    profile = OrderOperationsProfile.objects.get(order=order)
    financial_before = {
        field_name: getattr(profile, field_name)
        for field_name in profile.MANUAL_PROTECTABLE_FINANCIAL_FIELDS
    }
    before = _operations_snapshot(profile)
    form = OrderOperationsForm(
        request.POST or None,
        instance=profile,
        prefix="operations",
    )
    payment_formset = PaymentRecordFormSet(
        request.POST or None,
        request.FILES or None,
        instance=order,
        prefix="payments",
    )
    previous_payment_proofs = list(
        order.payment_records.exclude(proof="").values_list("proof", flat=True)
    )
    if request.method == "POST" and form.is_valid() and payment_formset.is_valid():
        with transaction.atomic():
            profile = form.save(commit=False)
            protected_fields = set(profile.manual_financial_fields or [])
            protected_fields.update(
                field_name
                for field_name, original_value in financial_before.items()
                if form.cleaned_data.get(field_name) != original_value
            )
            profile.manual_financial_fields = sorted(protected_fields)
            vehicle_secret = form.cleaned_data.get("vehicle_control_password")
            battery_secret = form.cleaned_data.get("battery_password")
            if vehicle_secret:
                profile.vehicle_control_password_encrypted = encrypt_secret(
                    vehicle_secret
                )
            if battery_secret:
                profile.battery_password_encrypted = encrypt_secret(battery_secret)
            profile.updated_by = _editing_name(request.user)
            profile.save()
            payments = payment_formset.save()
            for previous_proof in previous_payment_proofs:
                _schedule_model_file_cleanup(
                    PaymentRecord,
                    "proof",
                    previous_proof,
                )
            now = timezone.now()
            for payment in payments:
                if payment.confirmed and not payment.confirmed_at:
                    payment.confirmed_at = now
                    payment.confirmed_by = _editing_name(request.user)
                    payment.save(
                        update_fields=[
                            "confirmed_at", "confirmed_by", "updated_at"
                        ]
                    )
            card_totals = order.payment_records.aggregate(
                income=Sum("card_fee_charged"),
                expense=Sum("bank_card_fee"),
            )
            profile.card_fee_income = card_totals["income"] or 0
            profile.card_fee_expense = card_totals["expense"] or 0
            profile.save(
                update_fields=["card_fee_income", "card_fee_expense", "updated_at"]
            )
            after = _operations_snapshot(profile)
            changes = {
                key: {"before": before.get(key, ""), "after": value}
                for key, value in after.items()
                if before.get(key, "") != value
            }
            OrderChange.objects.create(
                order=order,
                reason=form.cleaned_data.get("change_reason")
                or "更新營運與對帳資料",
                changes=changes,
                actor_name=_editing_name(request.user),
            )
            OrderEvent.objects.create(
                order=order,
                event_type="operations_updated",
                description=f"更新營運與對帳資料（{len(changes)} 個欄位）",
                actor_name=_editing_name(request.user),
            )
        messages.success(request, "營運、收款及損益資料已更新。")
        return redirect("order_operations", pk=order.pk)
    return render(
        request,
        "sales/order_operations.html",
        {
            "order": order,
            "profile": profile,
            "form": form,
            "payment_formset": payment_formset,
            "manual_financial_fields": profile.manual_financial_fields or [],
            "is_electric": order.vehicle_model.energy_type
            != VehicleModel.EnergyType.GAS,
            "discount_request_form": DiscountRequestForm(
                initial={"amount": order.discount_requested_amount or None, "reason": order.discount_reason}
            ),
            "discount_decision_form": DiscountDecisionForm(initial={"decision": "approve"}),
        },
    )


@login_required
@transaction.atomic
def order_discount_request(request, pk):
    order = get_object_or_404(SalesOrder.objects.select_for_update(), pk=pk)
    if request.method != "POST" or not order.is_editable:
        messages.error(request, "已交付、完成或取消的訂單不可再申請折扣。")
        return redirect("order_operations", pk=pk)
    form = DiscountRequestForm(request.POST)
    if not form.is_valid():
        messages.error(request, "折扣申請未送出：" + " ".join(error for errors in form.errors.values() for error in errors))
        return redirect("order_operations", pk=pk)
    order.discount_requested_amount = form.cleaned_data["amount"]
    order.discount_reason = form.cleaned_data["reason"]
    order.discount_status = SalesOrder.DiscountStatus.PENDING
    order.discount_requested_at = timezone.now()
    order.discount_requested_by = _editing_name(request.user)
    order.discount_decided_at = None
    order.discount_decided_by = ""
    order.discount_decision_note = ""
    order.save(
        update_fields=[
            "discount_requested_amount", "discount_reason", "discount_status",
            "discount_requested_at", "discount_requested_by", "discount_decided_at",
            "discount_decided_by", "discount_decision_note", "updated_at",
        ]
    )
    OrderEvent.objects.create(
        order=order,
        event_type="discount_requested",
        description=f"申請內部折扣 {order.discount_requested_amount:,.0f} 元：{order.discount_reason}",
        actor_name=_editing_name(request.user),
    )
    messages.success(request, "折扣申請已送出；核准前不會改變應收金額。")
    return redirect("order_operations", pk=pk)


@login_required
@transaction.atomic
def order_discount_decide(request, pk):
    order = get_object_or_404(SalesOrder.objects.select_for_update(), pk=pk)
    if request.method != "POST" or order.discount_status != SalesOrder.DiscountStatus.PENDING:
        messages.error(request, "目前沒有待確認的折扣申請。")
        return redirect("order_operations", pk=pk)
    form = DiscountDecisionForm(request.POST)
    if not form.is_valid():
        messages.error(request, "折扣確認失敗，請重新選擇處理結果。")
        return redirect("order_operations", pk=pk)
    old_calculated = order.calculated_balance
    old_expected = order.calculate_balance()
    was_automatic = order.actual_balance in {old_calculated, old_expected}
    approved = form.cleaned_data["decision"] == "approve"
    before_amount = order.approved_discount_amount
    order.approved_discount_amount = order.discount_requested_amount if approved else before_amount
    order.discount_status = SalesOrder.DiscountStatus.APPROVED if approved else SalesOrder.DiscountStatus.REJECTED
    order.discount_decided_at = timezone.now()
    order.discount_decided_by = _editing_name(request.user)
    order.discount_decision_note = form.cleaned_data["note"]
    order.calculated_balance = order.calculate_balance()
    if was_automatic:
        order.actual_balance = order.calculated_balance
    elif order.actual_balance != order.calculated_balance:
        order.balance_adjustment_reason = (
            f"{order.balance_adjustment_reason}；" if order.balance_adjustment_reason else ""
        ) + "折扣核准後保留原人工尾款"
    order.save(
        update_fields=[
            "approved_discount_amount", "discount_status", "discount_decided_at",
            "discount_decided_by", "discount_decision_note", "calculated_balance",
            "actual_balance", "updated_at",
            "balance_adjustment_reason",
        ]
    )
    sync_order_operations(order.pk)
    OrderChange.objects.create(
        order=order,
        reason=f"折扣申請{'核准' if approved else '不採用'}",
        changes={
            "approved_discount_amount": {
                "before": str(before_amount),
                "after": str(order.approved_discount_amount),
            }
        },
        actor_name=_editing_name(request.user),
    )
    OrderEvent.objects.create(
        order=order,
        event_type="discount_decided",
        description=f"折扣申請{'已核准並套用' if approved else '未採用'}：{order.discount_requested_amount:,.0f} 元",
        actor_name=_editing_name(request.user),
    )
    messages.success(request, "折扣已核准並重算應收。" if approved else "折扣申請已標記為不採用。")
    return redirect("order_operations", pk=pk)


@login_required
def order_secret_reveal(request, pk):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "僅接受 POST。"}, status=405)
    order = get_object_or_404(SalesOrder, pk=pk)
    profile = get_object_or_404(OrderOperationsProfile, order=order)
    field = request.POST.get("field")
    encrypted_fields = {
        "vehicle_control_password": profile.vehicle_control_password_encrypted,
        "battery_password": profile.battery_password_encrypted,
    }
    if field not in encrypted_fields:
        return JsonResponse({"ok": False, "error": "不支援的欄位。"}, status=400)
    value = decrypt_secret(encrypted_fields[field])
    OrderEvent.objects.create(
        order=order,
        event_type="secret_viewed",
        description=f"查看{'車控' if field == 'vehicle_control_password' else '電池合約'}密碼",
        actor_name=_editing_name(request.user),
    )
    return JsonResponse({"ok": True, "value": value})


def _order_snapshot(order):
    excluded = {
        "id",
        "created_at",
        "updated_at",
        "revision",
        "editing_session",
        "editing_by",
        "editing_at",
        "calculated_balance",
        "old_owner_ocr_name",
        "old_owner_ocr_id_number",
    }
    snapshot = {}
    for field in order._meta.concrete_fields:
        if field.name in excluded:
            continue
        value = field.value_from_object(order)
        snapshot[field.verbose_name] = str(value or "")
    snapshot["配件"] = [
        {
            "名稱": line.name,
            "數量": line.quantity,
            "類型": line.get_line_type_display(),
            "售價": str(line.amount),
            "工資": str(line.labor_fee),
            "總價": str(line.display_total),
            "備註": line.note,
        }
        for line in order.accessories.all()
    ]
    snapshot["其他費用"] = [
        {"項目": line.name, "金額": str(line.amount)}
        for line in order.other_fees.all()
    ]
    return snapshot


def _snapshot_changes(before, after):
    return {
        key: {"before": before.get(key), "after": after.get(key)}
        for key in sorted(set(before) | set(after))
        if before.get(key) != after.get(key)
    }


def _schedule_replaced_identity_file_cleanup(order, previous_names):
    """Delete replaced identity photos only after the order transaction commits."""
    current_names = {
        field_name: getattr(getattr(order, field_name), "name", "")
        for field_name in ("id_front", "id_back")
    }
    for field_name, previous_name in previous_names.items():
        if not previous_name or previous_name == current_names.get(field_name):
            continue
        storage = SalesOrder._meta.get_field(field_name).storage

        def delete_if_unreferenced(name=previous_name, file_storage=storage):
            order_reference = SalesOrder.objects.filter(
                Q(id_front=name) | Q(id_back=name)
            ).exists()
            draft_reference = OrderDraft.objects.filter(
                Q(id_front=name) | Q(id_back=name)
            ).exists()
            if not order_reference and not draft_reference:
                file_storage.delete(name)

        transaction.on_commit(delete_if_unreferenced)


def _schedule_model_file_cleanup(model, field_name, previous_name):
    """Delete a replaced file after commit when no row still references it."""
    if not previous_name:
        return
    storage = model._meta.get_field(field_name).storage

    def delete_if_unreferenced():
        if not model._default_manager.filter(
            **{field_name: previous_name}
        ).exists():
            storage.delete(previous_name)

    transaction.on_commit(delete_if_unreferenced)


@login_required
@transaction.atomic
def order_edit(request, pk):
    order = get_object_or_404(
        SalesOrder.objects.select_for_update(),
        pk=pk,
    )
    if not order.is_editable:
        messages.error(request, "此訂單已交車、完成或取消，內容已鎖定。")
        return redirect("order_detail", pk=pk)
    if not _claim_edit_lock(order, request, ORDER_PRESENCE_TIMEOUT):
        messages.error(
            request,
            f"此訂單目前由 {order.editing_by or '其他人員'} 編輯，暫時只能查看。",
        )
        return redirect("order_detail", pk=pk)

    if request.method == "POST":
        try:
            submitted_revision = int(request.POST.get("_order_revision", 0))
        except (TypeError, ValueError):
            submitted_revision = 0
        if submitted_revision != order.revision:
            messages.error(request, "此訂單已被其他人更新，請重新載入後再修改。")
            return redirect("order_edit", pk=pk)

        before = _order_snapshot(order)
        previous_identity_names = {
            field_name: getattr(getattr(order, field_name), "name", "")
            for field_name in ("id_front", "id_back")
        }
        previous_vehicle_model_id = order.vehicle_model_id
        balance_was_automatic = order.actual_balance == order.calculated_balance
        previous_actual_balance = order.actual_balance
        form = OrderEditForm(request.POST, request.FILES, instance=order)
        formset = AccessoryFormSet(request.POST, instance=order)
        fee_formset = OtherFeeFormSet(
            request.POST, instance=order, prefix="other_fees"
        )
        if form.is_valid() and formset.is_valid() and fee_formset.is_valid():
            order = form.save(commit=False)
            if (
                order.actual_balance != order.calculate_balance()
                and not order.balance_adjustment_reason
            ):
                order.balance_adjustment_reason = form.cleaned_data["change_reason"]
            order.revision += 1
            order.editing_session = ""
            order.editing_by = ""
            order.editing_at = None
            order.save()
            _schedule_replaced_identity_file_cleanup(
                order,
                previous_identity_names,
            )
            apply_order_price_snapshot(
                order,
                force=(
                    previous_vehicle_model_id != order.vehicle_model_id
                    or not order.price_snapshot
                ),
            )
            apply_order_installment_snapshot(order)
            formset.save()
            fee_formset.save()
            order._prefetched_objects_cache = {}
            order.calculated_balance = order.calculate_balance()
            order.actual_balance = (
                order.calculated_balance
                if balance_was_automatic
                else previous_actual_balance
            )
            order.save(
                update_fields=[
                    "calculated_balance",
                    "actual_balance",
                    "revision",
                    "editing_session",
                    "editing_by",
                    "editing_at",
                    "updated_at",
                ]
            )
            after = _order_snapshot(order)
            changes = _snapshot_changes(before, after)
            reason = form.cleaned_data["change_reason"]
            OrderChange.objects.create(
                order=order,
                reason=reason,
                changes=changes,
                actor_name=_editing_name(request.user),
            )
            OrderEvent.objects.create(
                order=order,
                event_type="updated",
                description=f"修改訂單：{reason}（{len(changes)} 個項目）",
                actor_name=_editing_name(request.user),
            )
            messages.success(request, "訂單內容已更新，變更紀錄已保存。")
            return redirect("order_detail", pk=pk)
    else:
        form = OrderEditForm(instance=order)
        formset = AccessoryFormSet(instance=order)
        fee_formset = OtherFeeFormSet(instance=order, prefix="other_fees")

    return render(
        request,
        "sales/order_form.html",
        {
            "form": form,
            "formset": formset,
            "fee_formset": fee_formset,
            "editing_order": order,
            "document_source": order,
            "document_model": "order",
            "vehicle_rate_data": _vehicle_rate_data(),
            "accessory_product_data": _accessory_product_data(),
        },
    )


@login_required
@transaction.atomic
def order_edit_presence(request, pk):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "僅接受 POST。"}, status=405)
    if not request.session.session_key:
        request.session.create()
    order = get_object_or_404(SalesOrder.objects.select_for_update(), pk=pk)
    session_key = request.session.session_key
    now = timezone.now()
    mine = order.editing_session == session_key
    active = bool(
        order.editing_session
        and order.editing_at
        and order.editing_at >= now - ORDER_PRESENCE_TIMEOUT
    )
    if request.POST.get("action") == "release":
        if mine:
            order.editing_session = ""
            order.editing_by = ""
            order.editing_at = None
            SalesOrder.objects.filter(pk=order.pk).update(
                editing_session="", editing_by="", editing_at=None
            )
        return JsonResponse({"ok": True, "active": False})
    if active and not mine:
        return JsonResponse(
            {
                "ok": True,
                "active": True,
                "mine": False,
                "editing_by": order.editing_by or "其他人員",
            }
        )
    editing_by = _editing_name(request.user)
    SalesOrder.objects.filter(pk=order.pk).update(
        editing_session=session_key, editing_by=editing_by, editing_at=now
    )
    return JsonResponse(
        {"ok": True, "active": True, "mine": True, "editing_by": editing_by}
    )


@login_required
def contract_print(request, pk):
    order = get_object_or_404(
        SalesOrder.objects.select_related(
            "source", "vehicle_model", "color"
        ).prefetch_related("accessories", "other_fees"),
        pk=pk,
    )
    pdf = build_order_contract_pdf(order)
    response = FileResponse(
        BytesIO(pdf),
        content_type="application/pdf",
        filename=f"{order.number}.pdf",
    )
    response["Content-Disposition"] = (
        f'inline; filename="{order.number}.pdf"; '
        f"filename*=UTF-8''{order.number}%E8%A8%82%E8%B3%BC%E5%96%AE.pdf"
    )
    return _protect_private_response(response)


@login_required
def privacy_consent_print(request, pk):
    order = get_object_or_404(SalesOrder, pk=pk)
    pdf = build_privacy_consent_pdf(order)
    response = FileResponse(
        BytesIO(pdf),
        content_type="application/pdf",
    )
    response["Content-Disposition"] = (
        f'inline; filename="{order.number}-privacy-consent.pdf"; '
        f"filename*=UTF-8''{order.number}%E5%80%8B%E8%B3%87%E5%90%8C%E6%84%8F%E6%9B%B8.pdf"
    )
    return _protect_private_response(response)


@login_required
def order_documents_print(request, pk):
    from pypdf import PdfReader, PdfWriter

    order = get_object_or_404(
        SalesOrder.objects.select_related(
            "source", "vehicle_model", "color"
        ).prefetch_related("accessories", "other_fees"),
        pk=pk,
    )
    writer = PdfWriter()
    for content in (
        build_order_contract_pdf(order),
        build_privacy_consent_pdf(order),
    ):
        reader = PdfReader(BytesIO(content))
        for page in reader.pages:
            writer.add_page(page)
    output = BytesIO()
    writer.write(output)
    output.seek(0)
    response = FileResponse(output, content_type="application/pdf")
    response["Content-Disposition"] = (
        f'inline; filename="{order.number}-documents.pdf"; '
        f"filename*=UTF-8''{order.number}%E7%B0%BD%E7%BD%B2%E6%96%87%E4%BB%B6.pdf"
    )
    return _protect_private_response(response)


@login_required
def identity_documents_print(request, pk):
    order = get_object_or_404(SalesOrder, pk=pk)
    if request.method != "POST":
        return redirect(f"{reverse('order_detail', args=[pk])}?tab=order")
    purpose = request.POST.get("purpose", "")
    requested_sides = set(request.POST.getlist("sides"))
    side_fields = [
        field_name
        for field_name in ("id_front", "id_back")
        if field_name in requested_sides
    ]
    try:
        pdf = build_identity_document_pdf(
            order,
            purpose,
            side_fields,
            timezone.localdate(),
        )
    except (OSError, ValueError) as exc:
        messages.error(request, f"證件文件未產生：{exc}")
        return redirect(f"{reverse('order_detail', args=[pk])}?tab=order")

    purpose_label = IDENTITY_DOCUMENT_PURPOSES[purpose]
    side_labels = ["正面" if value == "id_front" else "反面" for value in side_fields]
    OrderEvent.objects.create(
        order=order,
        event_type="identity_document_printed",
        description=f"已產生證件浮水印文件：{purpose_label}（{'、'.join(side_labels)}）。",
        actor_name=_editing_name(request.user),
    )
    response = FileResponse(BytesIO(pdf), content_type="application/pdf")
    response["Content-Disposition"] = (
        f'inline; filename="{order.number}-identity.pdf"; '
        f"filename*=UTF-8''{order.number}%E8%AD%89%E4%BB%B6%E5%BD%B1%E6%9C%AC.pdf"
    )
    return _protect_private_response(response)


@login_required
@transaction.atomic
def contract_upload(request, pk):
    order = get_object_or_404(SalesOrder, pk=pk)
    signed_documents_url = _order_detail_section_url(pk, "signed-documents")
    if request.method != "POST":
        return redirect(signed_documents_url)
    previous_file = getattr(order.signed_contract, "name", "")
    form = SignedContractForm(request.POST, request.FILES, instance=order)
    form_is_valid = form.is_valid()
    if "signed_contract" not in request.FILES:
        form.add_error("signed_contract", "請先選擇要上傳的合約檔案。")
        form_is_valid = False
    if form_is_valid:
        order = form.save(commit=False)
        order.signed_contract_uploaded_at = timezone.now()
        order.save()
        _schedule_model_file_cleanup(
            SalesOrder,
            "signed_contract",
            previous_file,
        )
        OrderEvent.objects.create(
            order=order,
            event_type="contract_uploaded",
            description="已上傳訂購合約附件。",
            actor_name=request.user.get_username(),
        )
        return _document_upload_response(
            request,
            order_pk=pk,
            tab="order",
            section="signed-documents",
            ok=True,
            message="訂購合約附件已上傳。",
        )
    return _document_upload_response(
        request,
        order_pk=pk,
        tab="order",
        section="signed-documents",
        ok=False,
        message="合約上傳失敗："
        + (_form_error_text(form) or "請重新選擇檔案。"),
        status=400,
    )


@login_required
@transaction.atomic
def privacy_consent_upload(request, pk):
    order = get_object_or_404(SalesOrder, pk=pk)
    signed_documents_url = _order_detail_section_url(pk, "signed-documents")
    if request.method != "POST":
        return redirect(signed_documents_url)
    previous_file = getattr(order.privacy_consent, "name", "")
    form = PrivacyConsentForm(request.POST, request.FILES, instance=order)
    form_is_valid = form.is_valid()
    if "privacy_consent" not in request.FILES:
        form.add_error("privacy_consent", "請先選擇要上傳的同意書檔案。")
        form_is_valid = False
    if form_is_valid:
        order = form.save(commit=False)
        order.privacy_consent_uploaded_at = timezone.now()
        order.save()
        _schedule_model_file_cleanup(
            SalesOrder,
            "privacy_consent",
            previous_file,
        )
        OrderEvent.objects.create(
            order=order,
            event_type="privacy_consent_uploaded",
            description="已上傳個資同意書附件。",
            actor_name=request.user.get_username(),
        )
        return _document_upload_response(
            request,
            order_pk=pk,
            tab="order",
            section="signed-documents",
            ok=True,
            message="個資同意書附件已上傳。",
        )
    return _document_upload_response(
        request,
        order_pk=pk,
        tab="order",
        section="signed-documents",
        ok=False,
        message="個資同意書上傳失敗："
        + (_form_error_text(form) or "請重新選擇檔案。"),
        status=400,
    )


@login_required
@transaction.atomic
def allocate_vehicle(request, pk):
    order = get_object_or_404(SalesOrder.objects.select_for_update(), pk=pk)
    if request.method != "POST":
        return redirect("order_detail", pk=pk)
    form = AllocationForm(order, request.POST)
    if form.is_valid():
        try:
            order.allocate(form.cleaned_data["vehicle"])
        except ValidationError as exc:
            messages.error(request, " ".join(exc.messages))
        else:
            OrderEvent.objects.create(
                order=order,
                event_type="allocated",
                description=f"已配車：{order.allocated_vehicle}",
                actor_name=request.user.get_username(),
            )
            messages.success(request, "配車完成，實體車輛已鎖定。")
    else:
        messages.error(request, "請選擇可用的實體車輛。")
    return redirect("order_detail", pk=pk)


@login_required
@transaction.atomic
def reallocate_vehicle(request, pk):
    order = get_object_or_404(
        SalesOrder.objects.select_for_update(),
        pk=pk,
    )
    detail_url = f"{reverse('order_detail', args=[pk])}?tab=allocation"
    if request.method != "POST":
        return redirect(detail_url)
    if order.has_registration_started:
        messages.error(
            request,
            "已有車牌號碼、領牌完成紀錄或領牌文件，為避免車輛與文件對錯，無法直接改配。",
        )
        return redirect(detail_url)

    form = ReallocationForm(order, request.POST)
    if not form.is_valid():
        messages.error(
            request,
            "改配未完成：" + " ".join(
                error
                for errors in form.errors.values()
                for error in errors
            ),
        )
        return redirect(detail_url)
    try:
        original, replacement = order.reallocate(
            form.cleaned_data["vehicle"]
        )
    except ValidationError as exc:
        messages.error(request, " ".join(exc.messages))
        return redirect(detail_url)

    reason = form.cleaned_data["reason"]
    OrderChange.objects.create(
        order=order,
        reason=reason,
        changes={
            "已配車輛": {
                "before": str(original),
                "after": str(replacement),
            }
        },
        actor_name=_editing_name(request.user),
    )
    OrderEvent.objects.create(
        order=order,
        event_type="reallocated",
        description=f"改配車輛：{original} → {replacement}；原因：{reason}",
        actor_name=_editing_name(request.user),
    )
    messages.success(request, f"改配完成，新配車輛為 {replacement}。")
    return redirect(detail_url)


@login_required
@transaction.atomic
def registration_save(request, pk):
    order = get_object_or_404(SalesOrder.objects.select_for_update(), pk=pk)
    if request.method != "POST":
        return redirect("order_detail", pk=pk)
    if order.is_registration_complete:
        messages.error(request, "此訂單已完成領牌，領牌資料已鎖定。")
        return redirect("order_detail", pk=pk)
    if not order.allocated_vehicle_id:
        messages.error(request, "請先完成配車，再填寫領牌資料。")
        return redirect("order_detail", pk=pk)
    form = RegistrationStageForm(request.POST, instance=order)
    if form.is_valid():
        order = form.save()
        cost_profile = apply_order_settlement_cost(
            order,
            _editing_name(request.user),
        )
        apply_order_incentive_rule(
            order,
            _editing_name(request.user),
        )
        apply_order_dealer_commission(order)
        OrderEvent.objects.create(
            order=order,
            event_type="registration_data_updated",
            description=(
                f"已更新領牌資料：{order.registration_date}／"
                f"{order.registration_county}／{order.final_plate_number}"
            ),
            actor_name=_editing_name(request.user),
        )
        if cost_profile.vehicle_cost_rule_id:
            messages.success(
                request,
                "領牌資料已保存，已依領牌日期與縣市帶入"
                f"代銷結算成本 ${cost_profile.vehicle_cost:,.0f}。",
            )
        else:
            messages.warning(
                request,
                "領牌資料已保存，但目前找不到適用的代銷結算成本規則；"
                "請先至車型資料的成本規則補建，否則無法完成領牌。",
            )
    else:
        messages.error(
            request,
            "領牌資料未保存：" + " ".join(
                error
                for errors in form.errors.values()
                for error in errors
            ),
        )
    return redirect("order_detail", pk=pk)


@login_required
@transaction.atomic
def registration_document_upload(request, pk):
    order = get_object_or_404(SalesOrder.objects.select_for_update(), pk=pk)
    if request.method != "POST":
        return redirect("order_detail", pk=pk)
    if not order.allocated_vehicle_id:
        return _document_upload_response(
            request,
            order_pk=pk,
            tab="registration",
            ok=False,
            message="請先完成配車，再上傳領牌文件。",
            status=400,
        )
    if order.is_registration_complete or order.status in {
        SalesOrder.Status.COMPLETED,
        SalesOrder.Status.CANCELLED,
    }:
        return _document_upload_response(
            request,
            order_pk=pk,
            tab="registration",
            ok=False,
            message="此訂單的領牌階段已完成，無法修改文件。",
            status=409,
        )
    form = RegistrationDocumentUploadForm(request.POST, request.FILES)
    if not form.is_valid():
        return _document_upload_response(
            request,
            order_pk=pk,
            tab="registration",
            ok=False,
            message="文件上傳失敗：" + " ".join(
                error
                for errors in form.errors.values()
                for error in errors
            ),
            status=400,
        )

    document = form.save(commit=False)
    document.order = order
    document.uploaded_by = _editing_name(request.user)
    if document.document_type != RegistrationDocument.DocumentType.OTHER_INSURANCE:
        existing = order.registration_documents.filter(
            document_type=document.document_type
        ).first()
        if existing:
            previous_file = getattr(existing.file, "name", "")
            existing.name = document.name
            existing.file = document.file
            existing.uploaded_by = document.uploaded_by
            existing.save(
                update_fields=["name", "file", "uploaded_by", "updated_at"]
            )
            _schedule_model_file_cleanup(
                RegistrationDocument,
                "file",
                previous_file,
            )
            document = existing
        else:
            document.save()
    else:
        document.save()
    OrderEvent.objects.create(
        order=order,
        event_type="registration_document_uploaded",
        description=f"已上傳領牌文件：{document.display_name}",
        actor_name=_editing_name(request.user),
    )
    return _document_upload_response(
        request,
        order_pk=pk,
        tab="registration",
        ok=True,
        message=f"{document.display_name}已上傳。",
    )


@login_required
@transaction.atomic
def registration_document_delete(request, pk, document_pk):
    order = get_object_or_404(SalesOrder.objects.select_for_update(), pk=pk)
    if request.method != "POST":
        return redirect("order_detail", pk=pk)
    if order.is_registration_complete or order.status in {
        SalesOrder.Status.COMPLETED,
        SalesOrder.Status.CANCELLED,
    }:
        messages.error(request, "此訂單的領牌階段已完成，無法刪除文件。")
        return redirect("order_detail", pk=pk)
    document = get_object_or_404(
        RegistrationDocument,
        pk=document_pk,
        order=order,
    )
    display_name = document.display_name
    previous_file = getattr(document.file, "name", "")
    document.delete()
    _schedule_model_file_cleanup(
        RegistrationDocument,
        "file",
        previous_file,
    )
    OrderEvent.objects.create(
        order=order,
        event_type="registration_document_deleted",
        description=f"已刪除領牌文件：{display_name}",
        actor_name=_editing_name(request.user),
    )
    messages.success(request, f"{display_name}已刪除。")
    return redirect("order_detail", pk=pk)


@login_required
@transaction.atomic
def registration_complete(request, pk):
    order = get_object_or_404(SalesOrder.objects.select_for_update(), pk=pk)
    if request.method != "POST":
        return redirect("order_detail", pk=pk)
    if order.is_registration_complete:
        messages.info(request, "此訂單已完成領牌。")
        return redirect("order_detail", pk=pk)
    rule = resolve_settlement_cost(
        order.vehicle_model_id,
        order.registration_date,
    )
    if not rule:
        messages.error(
            request,
            "找不到符合車型及領牌日期的代銷結算成本規則，"
            "請先至車型資料補建後再完成領牌。",
        )
        return redirect("order_detail", pk=pk)
    try:
        order.complete_registration(_editing_name(request.user))
    except ValidationError as exc:
        messages.error(request, " ".join(exc.messages))
    else:
        apply_order_settlement_cost(
            order,
            _editing_name(request.user),
            lock=True,
        )
        apply_order_incentive_rule(
            order,
            _editing_name(request.user),
            lock=True,
        )
        apply_order_dealer_commission(order, lock=True)
        OrderEvent.objects.create(
            order=order,
            event_type="registration_completed",
            description=(
                f"領牌完成：{order.registration_date}／"
                f"{order.final_plate_number}"
            ),
            actor_name=_editing_name(request.user),
        )
        messages.success(request, "領牌階段已完成，訂單進入待交付。")
    return redirect("order_detail", pk=pk)


@login_required
@transaction.atomic
def delivery_complete(request, pk):
    # PostgreSQL 不允許 FOR UPDATE 套在 nullable OUTER JOIN；只鎖訂單本身，
    # 實體車輛會在 SalesOrder.complete_delivery() 內另行鎖定。
    order = get_object_or_404(SalesOrder.objects.select_for_update(), pk=pk)
    detail_url = f"{reverse('order_detail', args=[pk])}?tab=delivery"
    if request.method != "POST":
        return redirect(detail_url)
    if order.is_delivered or DeliveryRecord.objects.filter(order=order).exists():
        messages.info(request, "此訂單已完成交付，不需要重複送出。")
        return redirect(detail_url)
    form = DeliveryCompletionForm(order, request.POST, request.FILES)
    if not form.is_valid():
        messages.error(
            request,
            "交付未完成：" + " ".join(
                error
                for errors in form.errors.values()
                for error in errors
            ),
        )
        return redirect(detail_url)
    try:
        order, record = form.save(_editing_name(request.user))
    except ValidationError as exc:
        messages.error(request, "交付未完成：" + " ".join(exc.messages))
        return redirect(detail_url)

    VehicleInventoryHistory.objects.create(
        vehicle=order.allocated_vehicle,
        event_type=VehicleInventoryHistory.EventType.UPDATED,
        actor_name=_editing_name(request.user),
        reason=f"訂單 {order.number} 完成交付",
        changes={"庫存狀態": {"before": "已預留", "after": "已交車"}},
        status_snapshot=order.allocated_vehicle.status,
        location_store_snapshot=order.allocated_vehicle.location_store,
        location_label_snapshot=order.allocated_vehicle.actual_location_label,
        condition_note_snapshot=record.vehicle_condition_note,
        condition_resolution_snapshot=record.damage_note,
        condition_photo_snapshot=record.handover_photo,
    )
    OrderEvent.objects.create(
        order=order,
        event_type="delivery_completed",
        description=(
            f"完成交付：{order.get_delivery_method_display()}／"
            f"{record.handover_location}／收車人 {record.recipient_name}"
        ),
        actor_name=_editing_name(request.user),
    )
    if order.status == SalesOrder.Status.DELIVERED_DOCS_PENDING:
        messages.success(request, "車輛已交付；合作車行領牌文件與尾款將持續提醒。")
    else:
        messages.success(request, "車輛交付完成，訂單已結案。")
    return redirect(detail_url)


@login_required
@transaction.atomic
def delivery_payment_update(request, pk):
    order = get_object_or_404(SalesOrder.objects.select_for_update(), pk=pk)
    detail_url = f"{reverse('order_detail', args=[pk])}?tab=delivery#delivery-payment"
    if request.method != "POST":
        return redirect(detail_url)
    if order.status in {
        SalesOrder.Status.CANCEL_REFUND_PENDING,
        SalesOrder.Status.CANCELLED,
    }:
        messages.error(request, "此訂單已進入取消流程，不能登記尾款。")
        return redirect(detail_url)
    if order.is_delivered and order.source_type != SalesOrder.SourceType.DEALER:
        messages.error(request, "一般訂單交付後，請至營運與對帳進行收款更正。")
        return redirect(detail_url)

    sync_order_operations(order.pk)
    payment = get_object_or_404(
        PaymentRecord.objects.select_for_update(),
        order=order,
        system_key="balance",
    )
    previous_proof = getattr(payment.proof, "name", "")
    before = {
        "received_amount": payment.received_amount,
        "received_on": payment.received_on,
        "payment_method": payment.payment_method,
        "receiving_account": payment.receiving_account,
        "confirmed": payment.confirmed,
        "note": payment.note,
        "proof": previous_proof,
    }
    form = DeliveryPaymentForm(request.POST, request.FILES, instance=payment)
    if not form.is_valid():
        messages.error(
            request,
            "尾款資料未保存："
            + " ".join(
                error for errors in form.errors.values() for error in errors
            ),
        )
        return redirect(detail_url)

    payment = form.save(_editing_name(request.user))
    current_proof = getattr(payment.proof, "name", "")
    if previous_proof and previous_proof != current_proof:
        _schedule_model_file_cleanup(PaymentRecord, "proof", previous_proof)
    after = {
        "received_amount": payment.received_amount,
        "received_on": payment.received_on,
        "payment_method": payment.payment_method,
        "receiving_account": payment.receiving_account,
        "confirmed": payment.confirmed,
        "note": payment.note,
        "proof": current_proof,
    }
    changed_labels = {
        "received_amount": "實收金額",
        "received_on": "收款日期",
        "payment_method": "收款方式",
        "receiving_account": "收款帳戶",
        "confirmed": "收款確認",
        "note": "收款備註",
        "proof": "收款證明",
    }
    changes = {
        changed_labels[key]: {
            "before": str(before[key] or ""),
            "after": str(value or ""),
        }
        for key, value in after.items()
        if before[key] != value
    }
    if changes:
        OrderChange.objects.create(
            order=order,
            reason="更新交付尾款",
            changes=changes,
            actor_name=_editing_name(request.user),
        )
    OrderEvent.objects.create(
        order=order,
        event_type="delivery_payment_updated",
        description=(
            f"更新{payment.item_name}：實收 ${payment.received_amount:,.0f}／"
            f"{'已確認收清' if payment.is_settled else '尚未收清'}"
        ),
        actor_name=_editing_name(request.user),
    )
    if payment.is_settled:
        messages.success(request, f"{payment.item_name}已確認收清，可以繼續交付。")
    elif payment.received_amount:
        messages.warning(
            request,
            f"已保存部分收款，尚差 {payment.outstanding_amount:,.0f} 元。",
        )
    else:
        messages.success(request, "尾款資料已保存。")
    return redirect(detail_url)


@login_required
@transaction.atomic
def cancellation_request(request, pk):
    order = get_object_or_404(SalesOrder.objects.select_for_update(), pk=pk)
    detail_url = f"{reverse('order_detail', args=[pk])}?tab=order"
    if request.method != "POST":
        return redirect(detail_url)
    form = CancellationRequestForm(request.POST)
    if not form.is_valid():
        messages.error(
            request,
            "取消未登記：" + " ".join(
                error for errors in form.errors.values() for error in errors
            ),
        )
        return redirect(detail_url)
    released_identifier = (
        order.allocated_vehicle.identifier if order.allocated_vehicle_id else ""
    )
    try:
        order.request_cancellation(
            _editing_name(request.user),
            form.cleaned_data["reason"],
            form.cleaned_data["note"],
        )
    except ValidationError as exc:
        messages.error(request, "取消未登記：" + " ".join(exc.messages))
        return redirect(detail_url)
    description = f"登記取消：{order.cancellation_reason}"
    if released_identifier:
        description += f"；已解除配車 {released_identifier}"
    if order.status == SalesOrder.Status.CANCELLED:
        description += "；本單未收訂金，已完成取消"
    OrderEvent.objects.create(
        order=order,
        event_type="cancellation_requested",
        description=description,
        actor_name=_editing_name(request.user),
    )
    if order.status == SalesOrder.Status.CANCEL_REFUND_PENDING:
        messages.warning(
            request,
            f"已登記取消，必須全額退還訂金 ${order.deposit_amount:,.0f} 後才會完成取消。",
        )
    else:
        messages.success(request, "訂單已取消。")
    return redirect(detail_url)


@login_required
@transaction.atomic
def refund_complete(request, pk):
    order = get_object_or_404(SalesOrder.objects.select_for_update(), pk=pk)
    detail_url = f"{reverse('order_detail', args=[pk])}?tab=order"
    if request.method != "POST":
        return redirect(detail_url)
    form = RefundCompletionForm(order, request.POST, request.FILES)
    if not form.is_valid():
        messages.error(
            request,
            "退款未完成：" + " ".join(
                error for errors in form.errors.values() for error in errors
            ),
        )
        return redirect(detail_url)
    try:
        order.complete_refund(
            _editing_name(request.user),
            form.cleaned_data["amount"],
            form.cleaned_data["completed_on"],
            form.cleaned_data["method"],
            form.cleaned_data["reference"],
            form.cleaned_data.get("proof"),
        )
    except ValidationError as exc:
        messages.error(request, "退款未完成：" + " ".join(exc.messages))
        return redirect(detail_url)
    OrderEvent.objects.create(
        order=order,
        event_type="refund_completed",
        description=(
            f"訂金已全額退款 ${order.refund_amount:,.0f}／"
            f"{order.get_refund_method_display()}／{order.refund_completed_on}；訂單完成取消"
        ),
        actor_name=_editing_name(request.user),
    )
    messages.success(request, "訂金已全額退款，訂單已取消。")
    return redirect(detail_url)


@login_required
def registration_document_file(request, document_pk):
    document = get_object_or_404(RegistrationDocument, pk=document_pk)
    if not document.file:
        raise Http404
    response = FileResponse(
        document.file.open("rb"),
        as_attachment=False,
        filename=Path(document.file.name).name,
    )
    return _protect_private_response(response)


def _recognize_old_owner_documents(order, actor_name):
    documents = {
        document.document_type: document
        for document in order.subsidy_documents.filter(
            document_type__in={
                SubsidyDocument.DocumentType.OLD_OWNER_ID_FRONT,
                SubsidyDocument.DocumentType.OLD_OWNER_ID_BACK,
            }
        )
    }
    front = documents.get(SubsidyDocument.DocumentType.OLD_OWNER_ID_FRONT)
    back = documents.get(SubsidyDocument.DocumentType.OLD_OWNER_ID_BACK)
    if not front or not back:
        return None
    image_extensions = {".jpg", ".jpeg", ".png", ".webp"}
    if (
        Path(front.file.name).suffix.lower() not in image_extensions
        or Path(back.file.name).suffix.lower() not in image_extensions
    ):
        return "證件已保存；OCR 僅支援 JPG、PNG 或 WebP，請人工填寫舊車主資料。"

    try:
        with front.file.open("rb") as front_file, back.file.open("rb") as back_file:
            result = recognize_id_card(front_file.read(), back_file.read())
    except IdOcrError as exc:
        return f"證件已保存；OCR 未完成：{exc}"
    except Exception:
        logger.exception("舊車主身分證 OCR 發生未預期錯誤")
        return "證件已保存；辨識服務暫時無法使用，請稍後重傳或人工填寫。"

    fields = result.get("fields", {})
    recognized_name = (fields.get("name") or "").strip()
    recognized_id = (fields.get("id_number") or "").strip().upper()
    before = _order_snapshot(order)
    conflicts = []
    changed_fields = []
    for field_name, candidate_field, recognized_value in (
        ("old_owner_name", "old_owner_ocr_name", recognized_name),
        ("old_owner_id_number", "old_owner_ocr_id_number", recognized_id),
    ):
        current_value = getattr(order, field_name)
        if not recognized_value:
            continue
        if not current_value:
            setattr(order, field_name, recognized_value)
            setattr(order, candidate_field, "")
            changed_fields.extend([field_name, candidate_field])
        elif current_value != recognized_value:
            setattr(order, candidate_field, recognized_value)
            changed_fields.append(candidate_field)
            conflicts.append(field_name)
        elif getattr(order, candidate_field):
            setattr(order, candidate_field, "")
            changed_fields.append(candidate_field)

    if changed_fields:
        order.revision += 1
        order.save(
            update_fields=[*set(changed_fields), "revision", "updated_at"]
        )
        changes = _snapshot_changes(before, _order_snapshot(order))
        if changes:
            OrderChange.objects.create(
                order=order,
                reason="舊車主證件 OCR 自動辨識",
                changes=changes,
                actor_name=actor_name,
            )
    warning_text = " ".join(result.get("warnings", []))
    if conflicts:
        return "辨識完成，但結果與目前資料不同，請選擇要採用的內容。"
    if recognized_name or recognized_id:
        return f"已自動帶入舊車主姓名與身分證字號。{warning_text}".strip()
    return f"未辨識到舊車主姓名或身分證字號，請人工填寫。{warning_text}".strip()


@login_required
@transaction.atomic
def subsidy_toggle(request, pk):
    order = get_object_or_404(SalesOrder.objects.select_for_update(), pk=pk)
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "僅接受 POST。"}, status=405)
    if not order.can_manage_subsidy:
        return JsonResponse(
            {"ok": False, "error": "已取消訂單無法修改補助資料。"},
            status=409,
        )
    enabled_value = request.POST.get("enabled")
    if enabled_value not in {"0", "1"}:
        return JsonResponse(
            {"ok": False, "error": "補助狀態格式不正確。"}, status=400
        )
    try:
        submitted_revision = int(request.POST.get("_order_revision", 0))
    except (TypeError, ValueError):
        submitted_revision = 0
    if submitted_revision != order.revision:
        return JsonResponse(
            {
                "ok": False,
                "conflict": True,
                "revision": order.revision,
                "error": "此訂單已被其他人更新，請重新載入後再操作補助。",
            },
            status=409,
        )

    enabled = enabled_value == "1"
    message = (
        "補助申請已啟用，可以直接上傳文件。"
        if enabled
        else "補助申請已關閉，既有文件仍會保留。"
    )
    if order.is_trade_in_subsidy == enabled:
        return JsonResponse(
            {
                "ok": True,
                "enabled": enabled,
                "revision": order.revision,
                "message": message,
            }
        )

    before = _order_snapshot(order)
    order.is_trade_in_subsidy = enabled
    order.revision += 1
    order.save(update_fields=["is_trade_in_subsidy", "revision", "updated_at"])
    reason = "啟用補助申請" if enabled else "關閉補助申請"
    changes = _snapshot_changes(before, _order_snapshot(order))
    OrderChange.objects.create(
        order=order,
        reason=reason,
        changes=changes,
        actor_name=_editing_name(request.user),
    )
    OrderEvent.objects.create(
        order=order,
        event_type="subsidy_enabled" if enabled else "subsidy_disabled",
        description=reason,
        actor_name=_editing_name(request.user),
    )
    return JsonResponse(
        {
            "ok": True,
            "enabled": enabled,
            "revision": order.revision,
            "message": message,
        }
    )


@login_required
@transaction.atomic
def subsidy_document_upload(request, pk):
    order = get_object_or_404(SalesOrder.objects.select_for_update(), pk=pk)
    if request.method != "POST":
        return redirect("order_detail", pk=pk)
    if not order.can_manage_subsidy:
        return _document_upload_response(
            request,
            order_pk=pk,
            tab="subsidy",
            ok=False,
            message="已取消訂單無法修改補助文件。",
            status=409,
        )
    if not order.is_trade_in_subsidy:
        return _document_upload_response(
            request,
            order_pk=pk,
            tab="subsidy",
            ok=False,
            message="此訂單未勾選汰舊／政府補助。",
            status=400,
        )
    form = SubsidyDocumentUploadForm(request.POST, request.FILES)
    if not form.is_valid():
        return _document_upload_response(
            request,
            order_pk=pk,
            tab="subsidy",
            ok=False,
            message="補助文件上傳失敗：" + " ".join(
                error
                for errors in form.errors.values()
                for error in errors
            ),
            status=400,
        )

    document = form.save(commit=False)
    document.order = order
    document.uploaded_by = _editing_name(request.user)
    existing = None
    if document.document_type != SubsidyDocument.DocumentType.OTHER:
        existing = order.subsidy_documents.filter(
            document_type=document.document_type
        ).first()
    if existing:
        previous_file = getattr(existing.file, "name", "")
        existing.file = document.file
        existing.name = document.name
        existing.note = document.note
        existing.uploaded_by = document.uploaded_by
        existing.save(
            update_fields=[
                "file",
                "name",
                "note",
                "uploaded_by",
                "updated_at",
            ]
        )
        _schedule_model_file_cleanup(
            SubsidyDocument,
            "file",
            previous_file,
        )
        document = existing
    else:
        document.save()
    ocr_message = None
    if document.document_type in {
        SubsidyDocument.DocumentType.OLD_OWNER_ID_FRONT,
        SubsidyDocument.DocumentType.OLD_OWNER_ID_BACK,
    }:
        ocr_message = _recognize_old_owner_documents(
            order, _editing_name(request.user)
        )
    OrderEvent.objects.create(
        order=order,
        event_type="subsidy_document_uploaded",
        description=(
            f"已上傳補助文件："
            f"{document.name or document.get_document_type_display()}"
        ),
        actor_name=_editing_name(request.user),
    )
    return _document_upload_response(
        request,
        order_pk=pk,
        tab="subsidy",
        ok=True,
        message=(
            ocr_message
            or f"{document.name or document.get_document_type_display()}已上傳。"
        ),
    )


@login_required
@transaction.atomic
def subsidy_document_delete(request, pk, document_pk):
    order = get_object_or_404(SalesOrder.objects.select_for_update(), pk=pk)
    if request.method != "POST":
        return redirect("order_detail", pk=pk)
    if not order.can_manage_subsidy:
        messages.error(request, "已取消訂單無法修改補助文件。")
        return redirect("order_detail", pk=pk)
    document = get_object_or_404(
        SubsidyDocument,
        pk=document_pk,
        order=order,
    )
    display_name = document.name or document.get_document_type_display()
    previous_file = getattr(document.file, "name", "")
    document.delete()
    _schedule_model_file_cleanup(
        SubsidyDocument,
        "file",
        previous_file,
    )
    OrderEvent.objects.create(
        order=order,
        event_type="subsidy_document_deleted",
        description=f"已刪除補助文件：{display_name}",
        actor_name=_editing_name(request.user),
    )
    messages.success(request, f"{display_name}已刪除。")
    return redirect("order_detail", pk=pk)


@login_required
def subsidy_document_file(request, document_pk):
    document = get_object_or_404(SubsidyDocument, pk=document_pk)
    if not document.file:
        raise Http404
    response = FileResponse(
        document.file.open("rb"),
        as_attachment=False,
        filename=Path(document.file.name).name,
    )
    return _protect_private_response(response)


@login_required
@transaction.atomic
def subsidy_data_update(request, pk):
    order = get_object_or_404(SalesOrder.objects.select_for_update(), pk=pk)
    detail_url = f"{reverse('order_detail', args=[pk])}?tab=subsidy"
    if request.method != "POST":
        return redirect(detail_url)
    if not order.can_manage_subsidy:
        messages.error(request, "已取消訂單無法修改補助資料。")
        return redirect(detail_url)
    try:
        submitted_revision = int(request.POST.get("_order_revision", 0))
    except (TypeError, ValueError):
        submitted_revision = 0
    if submitted_revision != order.revision:
        messages.error(request, "此訂單已被其他人更新，請重新確認補助資料。")
        return redirect(detail_url)

    before = _order_snapshot(order)
    balance_was_automatic = order.actual_balance == order.calculated_balance
    previous_actual_balance = order.actual_balance
    form = SubsidyDataForm(request.POST, instance=order)
    items_submitted = "subsidy_items-TOTAL_FORMS" in request.POST
    item_formset = SubsidyItemFormSet(
        request.POST if items_submitted else None,
        instance=order,
        prefix="subsidy_items",
    )
    if not form.is_valid() or (items_submitted and not item_formset.is_valid()):
        error_messages = [str(error) for errors in form.errors.values() for error in errors]
        if items_submitted:
            error_messages.extend(str(error) for error in item_formset.non_form_errors())
            for row_errors in item_formset.errors:
                error_messages.extend(
                    str(error) for errors in row_errors.values() for error in errors
                )
        messages.error(
            request,
            "補助資料未保存：" + " ".join(error_messages),
        )
        return redirect(detail_url)

    order = form.save(commit=False)
    order.revision += 1
    order.calculated_balance = order.calculate_balance()
    if balance_was_automatic:
        order.actual_balance = order.calculated_balance
    else:
        order.actual_balance = previous_actual_balance
        if order.actual_balance != order.calculated_balance:
            order.balance_adjustment_reason = form.cleaned_data["change_reason"]
    order.save()
    if items_submitted:
        item_formset.instance = order
        item_formset.save()
    profile, _ = OrderOperationsProfile.objects.get_or_create(order=order)
    profile.subsidy_amount = order.subsidy_total
    profile.subsidy_applied_on = order.subsidy_last_applied_on
    profile.updated_by = _editing_name(request.user)
    profile.save(
        update_fields=["subsidy_amount", "subsidy_applied_on", "updated_by", "updated_at"]
    )

    after = _order_snapshot(order)
    changes = _snapshot_changes(before, after)
    reason = form.cleaned_data["change_reason"]
    OrderChange.objects.create(
        order=order,
        reason=reason,
        changes=changes,
        actor_name=_editing_name(request.user),
    )
    OrderEvent.objects.create(
        order=order,
        event_type="subsidy_data_updated",
        description=f"修改補助資料：{reason}（{len(changes)} 個項目）",
        actor_name=_editing_name(request.user),
    )
    messages.success(request, "補助資料已更新，尾款與變更紀錄已同步。")
    return redirect(detail_url)


@login_required
@transaction.atomic
def subsidy_ocr_decision(request, pk):
    order = get_object_or_404(SalesOrder.objects.select_for_update(), pk=pk)
    detail_url = f"{reverse('order_detail', args=[pk])}?tab=subsidy"
    if request.method != "POST" or not order.can_manage_subsidy:
        return redirect(detail_url)
    decision = request.POST.get("decision")
    if decision not in {"apply", "keep"}:
        messages.error(request, "無效的 OCR 資料處理方式。")
        return redirect(detail_url)

    before = _order_snapshot(order)
    if decision == "apply":
        if order.old_owner_ocr_name:
            order.old_owner_name = order.old_owner_ocr_name
        if order.old_owner_ocr_id_number:
            order.old_owner_id_number = order.old_owner_ocr_id_number
    order.old_owner_ocr_name = ""
    order.old_owner_ocr_id_number = ""
    order.revision += 1
    order.save(
        update_fields=[
            "old_owner_name",
            "old_owner_id_number",
            "old_owner_ocr_name",
            "old_owner_ocr_id_number",
            "revision",
            "updated_at",
        ]
    )
    changes = _snapshot_changes(before, _order_snapshot(order))
    reason = (
        "採用舊車主證件 OCR 結果"
        if decision == "apply"
        else "保留人工輸入的舊車主資料"
    )
    if changes:
        OrderChange.objects.create(
            order=order,
            reason=reason,
            changes=changes,
            actor_name=_editing_name(request.user),
        )
    OrderEvent.objects.create(
        order=order,
        event_type="old_owner_ocr_decided",
        description=reason,
        actor_name=_editing_name(request.user),
    )
    messages.success(
        request,
        "已採用辨識結果。" if decision == "apply" else "已保留目前內容。",
    )
    return redirect(detail_url)


@login_required
def inventory_list(request):
    current_statuses = (
        VehicleInventory.Status.AVAILABLE,
        VehicleInventory.Status.RESERVED,
        VehicleInventory.Status.TRANSFER_PENDING,
        VehicleInventory.Status.IN_TRANSFER,
        VehicleInventory.Status.DELIVERY_PENDING,
        VehicleInventory.Status.CONDITION_ISSUE,
    )
    historical_statuses = (
        VehicleInventory.Status.DELIVERED,
        VehicleInventory.Status.SOLD,
        VehicleInventory.Status.INACTIVE,
    )
    requested_statuses = list(
        dict.fromkeys(value for value in request.GET.getlist("status") if value)
    )
    scope = request.GET.get("scope", "")
    if scope not in {"current", "history"}:
        scope = (
            "history"
            if requested_statuses
            and all(value in historical_statuses for value in requested_statuses)
            else "current"
        )
    scope_statuses = historical_statuses if scope == "history" else current_statuses

    vehicles = VehicleInventory.objects.select_related(
        "vehicle_model", "vehicle_model__family", "color", "current_dealer"
    ).filter(status__in=scope_statuses)
    keyword = request.GET.get("q", "").strip()
    requested_family_ids = list(
        dict.fromkeys(
            int(value)
            for value in request.GET.getlist("vehicle_family")
            if value.isdigit()
        )
    )
    requested_colors = list(
        dict.fromkeys(
            value.strip()
            for value in request.GET.getlist("color")
            if value.strip()
        )
    )
    requested_locations = list(
        dict.fromkeys(value for value in request.GET.getlist("location") if value)
    )
    sort = request.GET.get("sort", "received_desc")
    selected_statuses = []
    status_values = set()
    for requested_status in requested_statuses:
        if requested_status == "transfer" and scope == "current":
            selected_statuses.append(requested_status)
            status_values.update(
                {
                    VehicleInventory.Status.TRANSFER_PENDING,
                    VehicleInventory.Status.IN_TRANSFER,
                }
            )
        elif requested_status in scope_statuses:
            selected_statuses.append(requested_status)
            status_values.add(requested_status)
    if status_values:
        vehicles = vehicles.filter(status__in=status_values)

    selected_families = list(
        VehicleModelFamily.objects.filter(
            pk__in=requested_family_ids,
            active=True,
        )
    )
    selected_family_keys = {
        (family.brand.strip().casefold(), family.name.strip().casefold())
        for family in selected_families
    }
    if selected_family_keys:
        family_query = Q()
        for brand_key, family_name_key in selected_family_keys:
            family_query |= Q(
                vehicle_model__family__brand__iexact=brand_key,
                vehicle_model__family__name__iexact=family_name_key,
            )
        vehicles = vehicles.filter(family_query)
    color_rows = list(
        VehicleColor.objects.filter(active=True)
        .values("id", "name")
        .order_by("name", "id")
    )
    color_groups = {}
    color_choices = []
    for color_row in color_rows:
        color_name = color_row["name"].strip()
        color_key = color_name.casefold()
        if not color_key:
            continue
        if color_key not in color_groups:
            color_groups[color_key] = {
                "value": color_name,
                "label": color_name,
                "ids": [],
            }
            color_choices.append(color_groups[color_key])
        color_groups[color_key]["ids"].append(color_row["id"])
    selected_colors = []
    selected_color_ids = set()
    for requested_color in requested_colors:
        if requested_color.isdigit():
            legacy_color = next(
                (row for row in color_rows if row["id"] == int(requested_color)),
                None,
            )
            color_key = legacy_color["name"].strip().casefold() if legacy_color else ""
        else:
            color_key = requested_color.casefold()
        selected_color_group = color_groups.get(color_key)
        if selected_color_group:
            selected_colors.append(selected_color_group["value"])
            selected_color_ids.update(selected_color_group["ids"])
    selected_colors = list(dict.fromkeys(selected_colors))
    if selected_color_ids:
        vehicles = vehicles.filter(color_id__in=selected_color_ids)

    selected_locations = []
    selected_dealer_ids = []
    store_selected = "store" in requested_locations
    if store_selected:
        selected_locations.append("store")
    for requested_location in requested_locations:
        dealer_id = requested_location.removeprefix("dealer-")
        if requested_location.startswith("dealer-") and dealer_id.isdigit():
            selected_locations.append(requested_location)
            selected_dealer_ids.append(int(dealer_id))
    selected_dealer_ids = list(dict.fromkeys(selected_dealer_ids))
    if store_selected or selected_dealer_ids:
        location_query = Q(current_dealer__isnull=True) if store_selected else Q()
        if selected_dealer_ids:
            location_query |= Q(current_dealer_id__in=selected_dealer_ids)
        vehicles = vehicles.filter(location_query)
    selected_dealer_id = (
        str(selected_dealer_ids[0]) if selected_dealer_ids else ""
    )
    if keyword:
        matching_statuses = [
            value
            for value, label in VehicleInventory.Status.choices
            if keyword.casefold() in label.casefold()
        ]
        query = (
            Q(engine_number__icontains=keyword)
            | Q(frame_number__icontains=keyword)
            | vehicle_brand_search_q(keyword, "vehicle_model__brand")
            | Q(vehicle_model__name__icontains=keyword)
            | Q(vehicle_model__model_number__icontains=keyword)
            | Q(vehicle_model__factory_model_codes__code__icontains=keyword)
            | Q(color__name__icontains=keyword)
            | Q(current_dealer__name__icontains=keyword)
            | Q(condition_note__icontains=keyword)
        )
        if matching_statuses:
            query |= Q(status__in=matching_statuses)
        if "本店" in keyword:
            query |= Q(current_dealer__isnull=True)
        vehicles = vehicles.filter(query).distinct()
    sort_options = {
        "received_desc": ("-received_on", "-id"),
        "received_asc": ("received_on", "id"),
        "model": ("vehicle_model__name", "color__name", "-received_on"),
        "color": ("color__name", "vehicle_model__name", "-received_on"),
        "identifier": ("engine_number", "frame_number", "-received_on"),
        "status": ("status", "-received_on"),
        "location": ("current_dealer__name", "vehicle_model__name"),
    }
    if sort not in sort_options:
        sort = "received_desc"
    vehicles = vehicles.order_by(*sort_options[sort])
    paginator = Paginator(vehicles, 100)
    page = paginator.get_page(request.GET.get("page"))
    filter_params = request.GET.copy()
    filter_params.pop("page", None)
    inventory_counts = VehicleInventory.objects.aggregate(
        current=Count("id", filter=Q(status__in=current_statuses)),
        history=Count("id", filter=Q(status__in=historical_statuses)),
    )
    family_rows = VehicleModelFamily.objects.filter(
        active=True,
        versions__active=True,
    ).distinct().order_by("brand", "name", "id")
    vehicle_family_choices = []
    family_choice_keys = set()
    selected_family_choice_ids = []
    for family in family_rows:
        choice_key = (family.brand.strip().casefold(), family.name.strip().casefold())
        if choice_key not in family_choice_keys:
            family_choice_keys.add(choice_key)
            vehicle_family_choices.append(family)
            if choice_key in selected_family_keys:
                selected_family_choice_ids.append(family.pk)
    return render(
        request,
        "sales/inventory_list.html",
        {
            "vehicles": page.object_list,
            "page_obj": page,
            "statuses": [
                (value, label)
                for value, label in VehicleInventory.Status.choices
                if value in scope_statuses
            ],
            "inventory_counts": inventory_counts,
            "vehicle_families": vehicle_family_choices,
            "colors": color_choices,
            "dealers": SalesSource.objects.filter(
                active=True,
                source_type=SalesSource.SourceType.DEALER,
            ).order_by("name"),
            "filter_query": filter_params.urlencode(),
            "selected": {
                "q": keyword,
                "status": selected_statuses[0] if selected_statuses else "",
                "statuses": selected_statuses,
                "vehicle_family": (
                    str(selected_family_choice_ids[0])
                    if selected_family_choice_ids
                    else ""
                ),
                "vehicle_family_ids": selected_family_choice_ids,
                "color": selected_colors[0] if selected_colors else "",
                "colors": selected_colors,
                "location": selected_locations[0] if selected_locations else "",
                "locations": selected_locations,
                "store_selected": store_selected,
                "dealer_ids": selected_dealer_ids,
                "current_dealer": selected_dealer_id,
                "sort": sort,
                "scope": scope,
            },
        },
    )


@login_required
def vehicle_model_list(request):
    keyword = request.GET.get("q", "").strip()
    energy_type = request.GET.get("energy_type", "")
    active = request.GET.get("active", "")
    today = timezone.localdate()
    current_prices = VehiclePriceVersion.objects.filter(
        vehicle_model_id=OuterRef("pk"),
        active=True,
        effective_from__lte=today,
    ).filter(Q(effective_to__isnull=True) | Q(effective_to__gte=today))
    models = VehicleModel.objects.annotate(
        available_count=Count(
            "vehicleinventory",
            filter=Q(vehicleinventory__status=VehicleInventory.Status.AVAILABLE),
            distinct=True,
        ),
        active_color_count=Count(
            "colors",
            filter=Q(colors__active=True),
            distinct=True,
        ),
        current_suggested_price=Subquery(
            current_prices.order_by("-effective_from", "-id").values(
                "suggested_price"
            )[:1],
            output_field=DecimalField(max_digits=12, decimal_places=0),
        ),
    )
    matched_model_ids = set()
    if keyword:
        match_query = (
            vehicle_brand_search_q(keyword)
            | Q(family__name__icontains=keyword)
            | Q(name__icontains=keyword)
            | Q(model_number__icontains=keyword)
            | Q(factory_model_codes__code__icontains=keyword)
            | Q(model_code__icontains=keyword)
            | Q(colors__name__icontains=keyword)
        )
        if keyword.isdigit() and len(keyword) == 4:
            match_query |= Q(model_year=int(keyword))
        matched = models.filter(match_query).distinct()
        matched_model_ids = set(matched.values_list("pk", flat=True))
        matched_family_ids = list(
            matched.exclude(family_id__isnull=True)
            .values_list("family_id", flat=True)
            .distinct()
        )
        unmatched_family_model_ids = list(
            matched.filter(family_id__isnull=True).values_list("pk", flat=True)
        )
        models = models.filter(
            Q(family_id__in=matched_family_ids)
            | Q(pk__in=unmatched_family_model_ids)
        ).distinct()
    valid_energy_types = {
        value for value, _label in VehicleModel.EnergyType.choices
    }
    if energy_type in valid_energy_types:
        models = models.filter(energy_type=energy_type)
    if active == "yes":
        models = models.filter(active=True)
    elif active == "no":
        models = models.filter(active=False)
    models = list(
        models.select_related("family")
        .prefetch_related("factory_model_codes")
        .order_by("brand", "family__name", "name", "-model_year", "model_code")
    )

    brand_records = {
        brand.name.casefold(): brand
        for brand in VehicleBrand.objects.select_related("parent").all()
    }
    grouped_models = {}
    for model in models:
        brand_record = brand_records.get(model.brand.casefold())
        root_brand = (
            brand_record.parent
            if brand_record and brand_record.parent_id
            else brand_record
        )
        root_name = root_brand.name if root_brand else model.brand
        root_key = root_name.casefold()
        group = grouped_models.setdefault(
            root_key,
            {
                "name": root_name,
                "display_order": root_brand.display_order if root_brand else 999,
                "models": [],
                "families": {},
                "child_brands": set(),
                "active_count": 0,
            },
        )
        model.child_brand_label = (
            model.brand if model.brand.casefold() != root_key else ""
        )
        model.brand_display_order = (
            brand_record.display_order if brand_record else 999
        )
        family_name = model.family.name if model.family_id else model.name
        family_key = model.family_id or f"legacy:{model.brand.casefold()}:{family_name.casefold()}"
        family = group["families"].setdefault(
            family_key,
            {
                "key": str(family_key),
                "name": family_name,
                "models": [],
                "child_brand_label": model.child_brand_label,
                "brand_display_order": model.brand_display_order,
            },
        )
        family["models"].append(model)
        if model.child_brand_label:
            group["child_brands"].add(model.child_brand_label)

    vehicle_model_groups = sorted(
        grouped_models.values(),
        key=lambda group: (group["display_order"], group["name"].casefold()),
    )
    filters_applied = bool(keyword or energy_type or active)
    for group in vehicle_model_groups:
        for family in group["families"].values():
            family["models"].sort(
                key=lambda model: (
                    not model.active,
                    -(model.model_year or 0),
                    model.model_code,
                    model.model_number.casefold(),
                    model.pk,
                )
            )

        def family_power_sort_key(family):
            current_model = family["models"][0]
            is_electric = current_model.energy_type in {
                VehicleModel.EnergyType.ELECTRIC,
                VehicleModel.EnergyType.LIGHT_ELECTRIC,
                VehicleModel.EnergyType.MICRO_ELECTRIC,
            }
            if is_electric:
                energy_rank = 0
                power_value = current_model.motor_power_kw
            elif current_model.energy_type == VehicleModel.EnergyType.GAS:
                energy_rank = 1
                power_value = current_model.displacement_cc
            else:
                energy_rank = 2
                power_value = None
            return (
                not any(model.active for model in family["models"]),
                energy_rank,
                power_value is None,
                power_value or 0,
                bool(family["child_brand_label"]),
                family["brand_display_order"],
                family["name"].casefold(),
            )

        families = sorted(
            group["families"].values(),
            key=family_power_sort_key,
        )
        group["models"] = []
        inactive_section_started = False
        for family in families:
            family["is_active"] = any(model.active for model in family["models"])
            family["is_first_inactive"] = (
                not family["is_active"] and not inactive_section_started
            )
            if not family["is_active"]:
                inactive_section_started = True
            factory_codes = {
                code.code
                for model in family["models"]
                for code in model.factory_model_codes.all()
                if code.active
            }
            if not factory_codes:
                factory_codes = {
                    model.model_number
                    for model in family["models"]
                    if model.model_number
                }
            family["factory_codes"] = sorted(factory_codes, key=str.casefold)
            family["version_count"] = len(family["models"])
            family["active_count"] = sum(model.active for model in family["models"])
            family["code_count"] = len(family["factory_codes"])
            family["available_count"] = sum(
                model.available_count for model in family["models"]
            )
            family["active_color_count"] = sum(
                model.active_color_count for model in family["models"]
            )
            family["active_year_count"] = len(
                {
                    model.model_year
                    for model in family["models"]
                    if model.active and model.model_year
                }
            )
            family["energy_labels"] = sorted(
                {model.get_energy_type_display() for model in family["models"]},
                key=str.casefold,
            )
            current_prices = sorted(
                {
                    model.current_suggested_price
                    for model in family["models"]
                    if model.current_suggested_price is not None
                }
            )
            family["current_price_low"] = current_prices[0] if current_prices else None
            family["current_price_high"] = current_prices[-1] if current_prices else None
            latest_model = family["models"][0]
            matched_family_models = [
                model
                for model in family["models"]
                if model.pk in matched_model_ids
            ]
            family["display_model"] = (
                matched_family_models[0]
                if keyword and matched_family_models
                else latest_model
            )
            family["display_is_search_match"] = bool(
                keyword and family["display_model"].pk != latest_model.pk
            )
            group["active_count"] += int(bool(family["active_count"]))
            for family_index, model in enumerate(family["models"]):
                model.show_family_cell = family_index == 0
                model.family_rowspan = family["version_count"]
                model.family_display_name = family["name"]
                model.family_version_count = family["version_count"]
                model.family_code_count = family["code_count"]
                model.family_factory_codes = family["factory_codes"]
                group["models"].append(model)
        group["families"] = families
        group["child_brands"] = sorted(
            group["child_brands"], key=str.casefold
        )
        group["total_count"] = len(families)
        group["version_count"] = len(group["models"])

    vehicle_model_count = sum(group["total_count"] for group in vehicle_model_groups)

    return render(
        request,
        "sales/vehicle_model_list.html",
        {
            "vehicle_model_groups": vehicle_model_groups,
            "vehicle_model_count": vehicle_model_count,
            "vehicle_model_version_count": len(models),
            "energy_types": VehicleModel.EnergyType.choices,
            "filters_applied": filters_applied,
            "selected": {
                "q": keyword,
                "energy_type": energy_type,
                "active": active,
            },
        },
    )


@login_required
def accessory_product_list(request):
    keyword = request.GET.get("q", "").strip()
    active = request.GET.get("active", "")
    products = AccessoryProduct.objects.annotate(
        usage_count=Count("order_lines", distinct=True)
    )
    if keyword:
        products = products.filter(
            Q(name__icontains=keyword) | Q(note__icontains=keyword)
        )
    if active == "yes":
        products = products.filter(active=True)
    elif active == "no":
        products = products.filter(active=False)
    products = products.order_by("name")
    page = Paginator(products, 100).get_page(request.GET.get("page"))
    return render(
        request,
        "sales/accessory_product_list.html",
        {
            "products": page.object_list,
            "page_obj": page,
            "selected": {"q": keyword, "active": active},
        },
    )


def _accessory_product_form_view(request, instance=None):
    form = AccessoryProductForm(request.POST or None, instance=instance)
    if request.method == "POST" and form.is_valid():
        product = form.save()
        messages.success(
            request,
            f"已{'更新' if instance else '新增'}配件：{product.name}。",
        )
        return redirect("accessory_product_list")
    return render(
        request,
        "sales/accessory_product_form.html",
        {"form": form, "editing_product": instance},
    )


@login_required
def accessory_product_create(request):
    return _accessory_product_form_view(request)


@login_required
def accessory_product_edit(request, pk):
    return _accessory_product_form_view(
        request,
        get_object_or_404(AccessoryProduct, pk=pk),
    )


def _coalesce_vehicle_color_post(post_data, instance):
    if not post_data or not instance:
        return post_data, (), ()
    data = post_data.copy()
    try:
        total_forms = int(data.get("colors-TOTAL_FORMS", 0))
    except (TypeError, ValueError):
        return data, (), ()
    original_names = dict(instance.colors.values_list("pk", "name"))
    grouped = {}
    for index in range(total_forms):
        name = (data.get(f"colors-{index}-name") or "").strip()
        if not name:
            continue
        try:
            color_id = int(data.get(f"colors-{index}-id") or 0) or None
        except (TypeError, ValueError):
            color_id = None
        grouped.setdefault(name.casefold(), []).append(
            {"index": index, "id": color_id, "name": name}
        )

    merged_names = []
    preserved_names = []
    for entries in grouped.values():
        if len(entries) < 2:
            continue
        canonical = next(
            (
                entry
                for entry in entries
                if entry["id"]
                and original_names.get(entry["id"], "").strip().casefold()
                == entry["name"].casefold()
            ),
            entries[0],
        )
        wants_active = any(
            f"colors-{entry['index']}-active" in data
            and f"colors-{entry['index']}-DELETE" not in data
            for entry in entries
        )
        canonical_index = canonical["index"]
        data.pop(f"colors-{canonical_index}-DELETE", None)
        if wants_active:
            data[f"colors-{canonical_index}-active"] = "on"

        for duplicate in entries:
            if duplicate is canonical:
                continue
            duplicate_index = duplicate["index"]
            if duplicate["id"]:
                original_name = original_names.get(duplicate["id"], "").strip()
                data[f"colors-{duplicate_index}-name"] = original_name
                data.pop(f"colors-{duplicate_index}-active", None)
                data.pop(f"colors-{duplicate_index}-DELETE", None)
                if original_name:
                    preserved_names.append(original_name)
            else:
                data[f"colors-{duplicate_index}-DELETE"] = "on"
        merged_names.append(canonical["name"])
    return (
        data,
        tuple(dict.fromkeys(merged_names)),
        tuple(dict.fromkeys(preserved_names)),
    )


def _vehicle_model_form_view(request, instance=None):
    is_editing = instance is not None
    today = timezone.localdate()
    installment_plan_versions = []
    current_installment_plan = None
    upcoming_installment_plan = None
    installment_option_count = 0
    if is_editing:
        installment_plan_versions = list(
            instance.installment_plan_versions.prefetch_related(
                "options__company"
            ).all()
        )
        installment_option_count = sum(
            len(version.options.all()) for version in installment_plan_versions
        )
        current_installment_plan = next(
            (
                version
                for version in installment_plan_versions
                if version.active
                and version.effective_from <= today
                and (version.effective_to is None or version.effective_to >= today)
            ),
            None,
        )
        upcoming_installment_plan = next(
            iter(
                sorted(
                    (
                        version
                        for version in installment_plan_versions
                        if version.active and version.effective_from > today
                    ),
                    key=lambda version: (version.effective_from, version.pk),
                )
            ),
            None,
        )
    action = request.POST.get("action", "save_model")
    move_form = (
        VehicleModelFamilyMoveForm(
            request.POST if request.method == "POST" and action == "move_family" else None,
            vehicle_model=instance,
        )
        if is_editing
        else None
    )
    merge_form = (
        VehicleModelVersionMergeForm(
            (
                request.POST
                if request.method == "POST" and action == "merge_version"
                else None
            ),
            vehicle_model=instance,
        )
        if is_editing
        else None
    )
    year_correction_form = (
        VehicleModelYearCorrectionForm(
            request.POST
            if request.method == "POST" and action == "correct_year"
            else None,
            vehicle_model=instance,
        )
        if is_editing
        else None
    )
    if request.method == "POST" and action == "correct_year" and is_editing:
        if year_correction_form.is_valid():
            try:
                corrected_model, original_year = correct_vehicle_model_year(
                    vehicle_model_id=instance.pk,
                    model_year=year_correction_form.cleaned_data["model_year"],
                )
            except ValidationError as exc:
                year_correction_form.add_error(None, exc)
            else:
                original_year_label = original_year or "未設定"
                messages.success(
                    request,
                    f"年份已由 {original_year_label} 修正為 {corrected_model.model_year}。",
                )
                return redirect(
                    f"{reverse('vehicle_model_edit', args=[corrected_model.pk])}#data-correction"
                )
    if request.method == "POST" and action == "merge_version" and is_editing:
        if merge_form.is_valid():
            target_model = merge_form.cleaned_data["target_model"]
            try:
                merged_model = merge_vehicle_model_versions(
                    source_model_id=instance.pk,
                    target_model_id=target_model.pk,
                )
            except ValidationError as exc:
                merge_form.add_error(None, exc)
            else:
                messages.success(
                    request,
                    f"已合併 {merged_model.model_year} 年的重複年式；訂單、庫存、顏色與商務設定均已保留。",
                )
                return redirect(
                    f"{reverse('vehicle_model_edit', args=[merged_model.pk])}#data-correction"
                )
    if request.method == "POST" and action == "move_family" and is_editing:
        if move_form.is_valid():
            try:
                moved_model, source_removed, merged = move_vehicle_model_to_family(
                    vehicle_model_id=instance.pk,
                    target_family_id=move_form.cleaned_data["target_family"].pk,
                )
            except ValidationError as exc:
                move_form.add_error(None, exc)
            else:
                cleanup_note = "，原本空白的錯誤機種也已清除" if source_removed else ""
                action_label = "合併至" if merged else "移至"
                messages.success(
                    request,
                    f"已將 {moved_model.model_year or '待補'} 年式{action_label}「{moved_model.family.name}」；"
                    f"訂單、庫存、顏色與商務設定均已保留{cleanup_note}。",
                )
                return redirect(
                    f"{reverse('vehicle_model_edit', args=[moved_model.pk])}#data-correction"
                )
    if request.method == "POST" and action == "delete_model" and is_editing:
        try:
            delete_unused_vehicle_model(vehicle_model_id=instance.pk)
        except ValidationError as exc:
            messages.error(request, "; ".join(exc.messages))
        else:
            messages.success(request, "未使用的年式／規格已永久刪除。")
            return redirect("vehicle_model_list")
    model_post = request.POST if request.method == "POST" and action == "save_model" else None
    model_post, merged_color_names, preserved_history_color_names = (
        _coalesce_vehicle_color_post(model_post, instance)
    )
    form = VehicleModelMasterForm(model_post, instance=instance)
    color_formset = VehicleColorMasterFormSet(
        model_post,
        instance=instance,
        prefix="colors",
    )
    # 新增車型時預設提供一個顏色欄位；編輯時只顯示既有顏色，
    # 需要更多顏色再由使用者按「新增顏色」。
    if is_editing:
        color_formset.extra = 0
    used_color_ids = set()
    if is_editing:
        used_color_ids = set(
            instance.colors.filter(
                Q(vehicleinventory__isnull=False) | Q(salesorder__isnull=False)
            )
            .values_list("pk", flat=True)
            .distinct()
        )
    for color_row_form in color_formset.forms:
        color_row_form.is_used = bool(
            color_row_form.instance.pk
            and color_row_form.instance.pk in used_color_ids
        )

    if (
        request.method == "POST"
        and action == "save_model"
        and form.is_valid()
        and color_formset.is_valid()
    ):
        renamed_family = None
        try:
            with transaction.atomic():
                if is_editing and instance.family_id:
                    desired_name = form.cleaned_data["name"]
                    if desired_name != instance.family.name:
                        renamed_family = rename_vehicle_model_family(
                            family_id=instance.family_id,
                            new_name=desired_name,
                        )
                vehicle_model = form.save()
                color_formset.instance = vehicle_model
                color_formset.save()
        except ValidationError as exc:
            form.add_error("name", exc)
        else:
            if merged_color_names:
                merged = "、".join(merged_color_names)
                preserved = "、".join(preserved_history_color_names)
                detail = f"；原顏色「{preserved}」已保留為停用歷史資料" if preserved else ""
                messages.info(
                    request,
                    f"偵測到同名顏色，已沿用並重新啟用既有的「{merged}」{detail}。",
                )
            if renamed_family:
                family, original_name, version_count = renamed_family
                messages.success(
                    request,
                    f"機種已由「{original_name}」改為「{family.name}」，共更新 {version_count} 個年式；訂單、庫存及商務設定均已保留。",
                )
            else:
                messages.success(
                    request,
                    "年式／規格已儲存，可以繼續維護售價、傭金或分期方案。",
                )
            return redirect(
                f"{reverse('vehicle_model_edit', args=[vehicle_model.pk])}#business-settings"
            )
    return render(
        request,
        "sales/vehicle_model_form.html",
        {
            "form": form,
            "color_formset": color_formset,
            "vehicle_model": instance,
            "is_editing": is_editing,
            "family_version_count": (
                instance.family.versions.count()
                if is_editing and instance.family_id
                else 0
            ),
            "price_version_count": (
                instance.price_versions.count() if is_editing else 0
            ),
            "current_price_version": (
                instance.price_versions.filter(
                    active=True,
                    effective_from__lte=today,
                )
                .filter(
                    Q(effective_to__isnull=True)
                    | Q(effective_to__gte=today)
                )
                .order_by("-effective_from", "-id")
                .first()
                if is_editing
                else None
            ),
            "settlement_rule_count": (
                instance.settlement_cost_rules.count() if is_editing else 0
            ),
            "incentive_rule_count": (
                instance.incentive_rules.count() if is_editing else 0
            ),
            "installment_plan_version_count": len(installment_plan_versions),
            "installment_option_count": installment_option_count,
            "current_installment_plan": current_installment_plan,
            "current_installment_options": (
                list(current_installment_plan.options.all())
                if current_installment_plan
                else []
            ),
            "upcoming_installment_plan": upcoming_installment_plan,
            "upcoming_installment_options": (
                list(upcoming_installment_plan.options.all())
                if upcoming_installment_plan
                else []
            ),
            "move_form": move_form,
            "merge_form": merge_form,
            "year_correction_form": year_correction_form,
            "has_merge_targets": (
                merge_form.fields["target_model"].queryset.exists()
                if merge_form
                else False
            ),
            "has_move_targets": (
                move_form.fields["target_family"].queryset.exists()
                if move_form
                else False
            ),
            "relation_summary": (
                vehicle_model_relation_summary(instance) if is_editing else {}
            ),
            "delete_blockers": (
                vehicle_model_delete_blockers(instance) if is_editing else []
            ),
        },
    )


@login_required
def vehicle_model_create(request):
    return _vehicle_model_form_view(request)


@login_required
def vehicle_model_edit(request, pk):
    return _vehicle_model_form_view(
        request,
        get_object_or_404(VehicleModel, pk=pk),
    )


@login_required
def settlement_cost_rule_list(request):
    keyword = request.GET.get("q", "").strip()
    rules = VehicleSettlementCostRule.objects.select_related("vehicle_model")
    if keyword:
        rules = rules.filter(
            vehicle_brand_search_q(keyword, "vehicle_model__brand")
            | Q(vehicle_model__family__name__icontains=keyword)
            | Q(vehicle_model__name__icontains=keyword)
            | Q(vehicle_model__model_number__icontains=keyword)
            | Q(vehicle_model__factory_model_codes__code__icontains=keyword)
            | Q(note__icontains=keyword)
        ).distinct()
    rules = rules.order_by(
        "vehicle_model__brand",
        "vehicle_model__name",
        "-effective_from",
    )
    paginator = Paginator(rules, 100)
    page = paginator.get_page(request.GET.get("page"))
    return render(
        request,
        "sales/settlement_cost_rule_list.html",
        {
            "rules": page.object_list,
            "page_obj": page,
            "selected": {"q": keyword},
            "today": timezone.localdate(),
        },
    )


def _settlement_cost_rule_form_view(request, instance=None):
    form = VehicleSettlementCostRuleForm(request.POST or None, instance=instance)
    if request.method == "POST" and form.is_valid():
        rule = form.save()
        messages.success(
            request,
            f"已{'更新' if instance else '建立'}成本規則：{rule}",
        )
        return redirect("settlement_cost_rule_list")
    return render(
        request,
        "sales/settlement_cost_rule_form.html",
        {
            "form": form,
            "rule": instance,
            "is_editing": instance is not None,
        },
    )


@login_required
def settlement_cost_rule_create(request):
    return _settlement_cost_rule_form_view(request)


@login_required
def settlement_cost_rule_edit(request, pk):
    return _settlement_cost_rule_form_view(
        request,
        get_object_or_404(VehicleSettlementCostRule, pk=pk),
    )


@login_required
@transaction.atomic
def settlement_cost_rule_delete(request, pk):
    if request.method != "POST":
        return redirect("settlement_cost_rule_list")
    rule = get_object_or_404(
        VehicleSettlementCostRule.objects.select_for_update(),
        pk=pk,
    )
    if rule.order_snapshots.exists():
        messages.error(
            request,
            "此成本規則已被領牌訂單採用，為保留財務依據不能刪除；"
            "請改為停用。",
        )
    else:
        label = str(rule)
        rule.delete()
        messages.success(request, f"已刪除未使用的成本規則：{label}")
    return redirect("settlement_cost_rule_list")


@login_required
def incentive_rule_list(request):
    keyword = request.GET.get("q", "").strip()
    rules = VehicleIncentiveRule.objects.select_related("vehicle_model")
    if keyword:
        rules = rules.filter(
            vehicle_brand_search_q(keyword, "vehicle_model__brand")
            | Q(vehicle_model__family__name__icontains=keyword)
            | Q(vehicle_model__name__icontains=keyword)
            | Q(vehicle_model__model_number__icontains=keyword)
            | Q(vehicle_model__factory_model_codes__code__icontains=keyword)
            | Q(note__icontains=keyword)
        ).distinct()
    rules = rules.order_by(
        "vehicle_model__brand",
        "vehicle_model__name",
        "-effective_from",
    )
    paginator = Paginator(rules, 100)
    page = paginator.get_page(request.GET.get("page"))
    return render(
        request,
        "sales/incentive_rule_list.html",
        {
            "rules": page.object_list,
            "page_obj": page,
            "selected": {"q": keyword},
        },
    )


def _incentive_rule_form_view(request, instance=None):
    form = VehicleIncentiveRuleForm(request.POST or None, instance=instance)
    if request.method == "POST" and form.is_valid():
        with transaction.atomic():
            rule = form.save()
        messages.success(
            request,
            f"已{'更新' if instance else '建立'}獎勵補助版本：{rule}",
        )
        return redirect("incentive_rule_list")
    return render(
        request,
        "sales/incentive_rule_form.html",
        {
            "form": form,
            "rule": instance,
            "is_editing": instance is not None,
        },
    )


@login_required
def incentive_rule_create(request):
    return _incentive_rule_form_view(request)


@login_required
def incentive_rule_edit(request, pk):
    return _incentive_rule_form_view(
        request,
        get_object_or_404(VehicleIncentiveRule, pk=pk),
    )


@login_required
@transaction.atomic
def incentive_rule_delete(request, pk):
    if request.method != "POST":
        return redirect("incentive_rule_list")
    rule = get_object_or_404(
        VehicleIncentiveRule.objects.select_for_update(),
        pk=pk,
    )
    if rule.order_snapshots.exists():
        messages.error(request, "此版本已被訂單採用，為保留歷史快照不能刪除；請改為停用。")
    else:
        rule.delete()
        messages.success(request, "獎勵補助版本已刪除。")
    return redirect("incentive_rule_list")


def bad_request(request, exception=None):
    return render(request, "errors/400.html", status=400)


def permission_denied(request, exception=None):
    return render(request, "errors/403.html", status=403)


def page_not_found(request, exception=None):
    return render(request, "errors/404.html", status=404)


def server_error(request):
    return render(
        request,
        "errors/500.html",
        {"request_id": getattr(request, "request_id", "")},
        status=500,
    )


@login_required
def vehicle_model_price_versions(request, model_pk):
    vehicle_model = get_object_or_404(VehicleModel, pk=model_pk)
    versions = vehicle_model.price_versions.order_by("-effective_from", "-id")
    requested_id = request.POST.get("version_id") or request.GET.get("edit")
    editing = (
        get_object_or_404(versions, pk=requested_id) if requested_id else None
    )
    if request.method == "POST" and request.POST.get("action") == "delete":
        if editing.orders.exists():
            messages.error(
                request,
                "此售價版本已被訂單採用，為保留歷史價格不能刪除；請改為停用。",
            )
        else:
            editing.delete()
            messages.success(request, "售價版本已刪除。")
        return redirect("vehicle_model_price_versions", model_pk=vehicle_model.pk)
    form = VehiclePriceVersionForm(
        request.POST or None,
        instance=editing or VehiclePriceVersion(vehicle_model=vehicle_model),
    )
    if request.method == "POST" and request.POST.get("action") != "delete" and form.is_valid():
        version = form.save(commit=False)
        version.vehicle_model = vehicle_model
        version.save()
        messages.success(request, f"已{'更新' if editing else '新增'}售價版本。")
        return redirect("vehicle_model_price_versions", model_pk=vehicle_model.pk)
    today = timezone.localdate()
    current_version = resolve_vehicle_price_version(vehicle_model.pk, today)
    version_rows = list(versions)
    for version in version_rows:
        if not version.active:
            version.display_state = "disabled"
            version.display_state_label = "已停用"
        elif version.effective_from > today:
            version.display_state = "future"
            version.display_state_label = "尚未生效"
        elif version.effective_to and version.effective_to < today:
            version.display_state = "expired"
            version.display_state_label = "已到期"
        elif current_version and version.pk == current_version.pk:
            version.display_state = "current"
            version.display_state_label = "目前採用"
        else:
            version.display_state = "superseded"
            version.display_state_label = "已被新版取代"
    fallback_version = None
    editing_is_current = bool(
        editing and current_version and editing.pk == current_version.pk
    )
    if editing_is_current:
        fallback_version = resolve_vehicle_price_version(
            vehicle_model.pk,
            today,
            exclude_version_id=editing.pk,
        )
    return render(
        request,
        "sales/vehicle_model_price_versions.html",
        {
            "vehicle_model": vehicle_model,
            "versions": version_rows,
            "form": form,
            "editing": editing,
            "current_version": current_version,
            "editing_is_current": editing_is_current,
            "fallback_version": fallback_version,
            "today": today,
        },
    )


@login_required
def vehicle_model_commission(request, model_pk):
    vehicle_model = get_object_or_404(VehicleModel, pk=model_pk)
    form = VehicleModelCommissionForm(
        request.POST or None,
        initial={"base_dealer_commission": vehicle_model.base_dealer_commission},
    )
    if request.method == "POST" and form.is_valid():
        vehicle_model.base_dealer_commission = form.cleaned_data[
            "base_dealer_commission"
        ]
        vehicle_model.save(update_fields=["base_dealer_commission", "updated_at"])
        messages.success(request, "車行基礎傭金已更新。")
        return redirect("vehicle_model_edit", pk=vehicle_model.pk)
    return render(
        request,
        "sales/vehicle_model_commission.html",
        {"vehicle_model": vehicle_model, "form": form},
    )


@login_required
def inventory_create(request):
    if request.method == "POST":
        form = VehicleInventoryForm(request.POST, request.FILES)
        if form.is_valid():
            with transaction.atomic():
                vehicle = form.save()
                _create_inventory_history(
                    vehicle,
                    actor_name=_editing_name(request.user),
                    event_type=VehicleInventoryHistory.EventType.CREATED,
                )
            messages.success(request, f"已建立庫存車輛：{vehicle.identifier}")
            return redirect("inventory_list")
    else:
        form = VehicleInventoryForm()
    return render(
        request,
        "sales/inventory_form.html",
        {"form": form, "is_editing": False},
    )


@login_required
def inventory_quick_create(request):
    formset = QuickInventoryEntryFormSet(
        request.POST or None,
        prefix="vehicles",
    )
    if request.method == "POST" and formset.is_valid():
        store = (
            Store.objects.filter(active=True, code__iexact="HQ").first()
            or Store.objects.filter(active=True).order_by("id").first()
        )
        if not store:
            formset._non_form_errors = formset.error_class(
                ["目前沒有啟用中的存放地點，請先建立本店資料。"]
            )
        else:
            try:
                with transaction.atomic():
                    created = []
                    for form in formset:
                        if not form.cleaned_data or form.cleaned_data.get("DELETE"):
                            continue
                        model = form.cleaned_data["vehicle_model"]
                        identifier = form.cleaned_data["identifier"]
                        vehicle = VehicleInventory(
                            vehicle_model=model,
                            color=form.cleaned_data["color"],
                            ownership_store=store,
                            location_store=store,
                            received_on=form.cleaned_data["received_on"],
                            manufactured_year_month=form.cleaned_data.get(
                                "manufactured_year_month", ""
                            ),
                            condition_note=form.cleaned_data.get("condition_note", ""),
                        )
                        if model.energy_type == VehicleModel.EnergyType.GAS:
                            vehicle.engine_number = identifier
                        else:
                            vehicle.frame_number = identifier
                        vehicle.save()
                        _create_inventory_history(
                            vehicle,
                            actor_name=_editing_name(request.user),
                            event_type=VehicleInventoryHistory.EventType.CREATED,
                        )
                        created.append(vehicle)
            except (IntegrityError, ValidationError):
                formset._non_form_errors = formset.error_class(
                    ["資料在送出期間發生變動，可能已有相同號碼，請檢查標示內容後再試一次。"]
                )
            else:
                messages.success(request, f"已完成快速進車，共建立 {len(created)} 台車輛。")
                return redirect("inventory_list")
    return render(
        request,
        "sales/inventory_quick_form.html",
        {
            "formset": formset,
            "energy_types": {
                str(model.pk): model.energy_type
                for model in VehicleModel.objects.filter(active=True)
            },
        },
    )


@login_required
def inventory_edit(request, pk):
    vehicle = get_object_or_404(
        VehicleInventory.objects.select_related(
            "vehicle_model", "color", "ownership_store", "location_store", "current_dealer"
        ),
        pk=pk,
    )
    if request.method == "POST":
        with transaction.atomic():
            vehicle = VehicleInventory.objects.select_for_update().select_related(
                "vehicle_model", "color", "location_store", "current_dealer"
            ).get(pk=pk)
            before = _inventory_values(vehicle)
            before_location_id = vehicle.current_dealer_id
            before_location_label = vehicle.actual_location_label
            form = VehicleInventoryForm(
                request.POST,
                request.FILES,
                instance=vehicle,
            )
            if form.is_valid():
                vehicle = form.save()
                after = _inventory_values(vehicle)
                changes = {
                    field_name: {
                        "label": INVENTORY_HISTORY_FIELDS[field_name],
                        "before": before[field_name][1],
                        "after": after[field_name][1],
                    }
                    for field_name in INVENTORY_HISTORY_FIELDS
                    if before[field_name][0] != after[field_name][0]
                }
                reason = form.cleaned_data.get("change_reason", "").strip()
                if changes or reason:
                    is_transfer = before_location_id != vehicle.current_dealer_id
                    _create_inventory_history(
                        vehicle,
                        actor_name=_editing_name(request.user),
                        event_type=(
                            VehicleInventoryHistory.EventType.TRANSFERRED
                            if is_transfer
                            else VehicleInventoryHistory.EventType.UPDATED
                        ),
                        reason=reason,
                        changes=changes,
                        from_location_label=(before_location_label if is_transfer else ""),
                        to_location_label=(vehicle.actual_location_label if is_transfer else ""),
                    )
                messages.success(request, f"已更新庫存車輛：{vehicle.identifier}")
                return redirect("inventory_list")
    else:
        form = VehicleInventoryForm(instance=vehicle)
    inventory_history = vehicle.history_entries.select_related(
        "from_location", "to_location", "location_store_snapshot"
    )
    return render(
        request,
        "sales/inventory_form.html",
        {
            "form": form,
            "vehicle": vehicle,
            "is_editing": True,
            "core_fields_locked": form.core_fields_locked,
            "final_fields_locked": form.final_fields_locked,
            "inventory_history": inventory_history,
            "transfer_history": inventory_history.filter(
                event_type=VehicleInventoryHistory.EventType.TRANSFERRED
            ),
        },
    )


@login_required
def vehicle_colors(request):
    model_id = request.GET.get("model")
    colors = VehicleColor.objects.filter(
        vehicle_model_id=model_id, active=True
    ).values("id", "name")
    return JsonResponse(
        {"results": list(colors)}
    )


@login_required
def sales_sources(request):
    source_type = request.GET.get("type")
    sources = SalesSource.objects.filter(
        source_type=source_type, active=True
    ).values("id", "name")
    return JsonResponse({"results": list(sources)})


@login_required
def installment_plan_options(request):
    model_id = request.GET.get("vehicle_model")
    raw_date = request.GET.get("order_date")
    try:
        order_date = date.fromisoformat(raw_date) if raw_date else timezone.localdate()
    except ValueError:
        return JsonResponse({"error": "訂單日期格式錯誤。"}, status=400)
    if not model_id or not str(model_id).isdigit():
        return JsonResponse({"options": []})
    version = resolve_installment_plan_version(int(model_id), order_date)
    if not version:
        return JsonResponse({"options": []})
    return JsonResponse(
        {
            "version": {
                "id": version.pk,
                "effective_from": version.effective_from.isoformat(),
                "effective_to": (
                    version.effective_to.isoformat() if version.effective_to else None
                ),
            },
            "options": [
                installment_option_payload(option)
                for option in version.options.select_related("company").all()
            ],
        }
    )


@login_required
def vehicle_price_options(request):
    model_id = request.GET.get("vehicle_model")
    order_id = request.GET.get("order_id")
    payment_type = request.GET.get("payment_type") or SalesOrder.PaymentType.CASH
    raw_date = request.GET.get("order_date")
    try:
        order_date = date.fromisoformat(raw_date) if raw_date else timezone.localdate()
    except ValueError:
        return JsonResponse({"error": "訂單日期格式錯誤。"}, status=400)
    if not model_id or not str(model_id).isdigit():
        return JsonResponse({"version": None, "recommended_price": None})
    valid_payment_types = {value for value, _label in SalesOrder.PaymentType.choices}
    if payment_type not in valid_payment_types:
        return JsonResponse({"error": "付款方式無效。"}, status=400)
    existing_order = None
    if order_id and str(order_id).isdigit():
        existing_order = SalesOrder.objects.filter(pk=order_id).first()
        if existing_order and not raw_date:
            order_date = existing_order.order_date
    if (
        existing_order
        and existing_order.vehicle_model_id == int(model_id)
        and existing_order.price_snapshot
    ):
        snapshot = existing_order.price_snapshot
        recommended_price, recommended_label = recommended_price_from_snapshot(
            snapshot,
            payment_type,
        )
        return JsonResponse(
            {
                "version": {
                    "id": snapshot.get("version_id"),
                    "effective_from": snapshot.get("effective_from"),
                    "effective_to": snapshot.get("effective_to") or None,
                    "suggested_price": snapshot.get("suggested_price") or None,
                    "suggested_price_includes_registration": snapshot.get(
                        "suggested_price_includes_registration", True
                    ),
                    "cash_price": snapshot.get("cash_price") or None,
                    "is_snapshot": True,
                },
                "recommended_price": (
                    str(recommended_price) if recommended_price is not None else None
                ),
                "recommended_label": recommended_label,
            }
        )
    version = resolve_vehicle_price_version(int(model_id), order_date)
    recommended_price, recommended_label = recommended_vehicle_price(
        version,
        payment_type,
    )
    if not version:
        return JsonResponse({"version": None, "recommended_price": None})
    return JsonResponse(
        {
            "version": {
                "id": version.pk,
                "effective_from": version.effective_from.isoformat(),
                "effective_to": (
                    version.effective_to.isoformat() if version.effective_to else None
                ),
                "suggested_price": (
                    str(version.suggested_price)
                    if version.suggested_price is not None
                    else None
                ),
                "suggested_price_includes_registration": (
                    version.suggested_price_includes_registration
                ),
                "cash_price": (
                    str(version.cash_price)
                    if version.cash_price is not None
                    else None
                ),
            },
            "recommended_price": (
                str(recommended_price) if recommended_price is not None else None
            ),
            "recommended_label": recommended_label,
        }
    )


@login_required
def id_card_ocr(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "僅接受 POST。"}, status=405)
    front = request.FILES.get("front")
    back = request.FILES.get("back")
    if not front or not back:
        return JsonResponse(
            {"ok": False, "error": "請先拍攝證件正面與反面。"},
            status=400,
        )
    document_type = request.POST.get(
        "document_type", IdOcrJob.DocumentType.NATIONAL_ID
    )
    valid_document_types = {value for value, _label in IdOcrJob.DocumentType.choices}
    if document_type not in valid_document_types:
        return JsonResponse(
            {"ok": False, "error": "不支援的證件類型。"},
            status=400,
        )
    try:
        validate_image_upload(front)
        validate_image_upload(back)
    except ValidationError as exc:
        return JsonResponse(
            {"ok": False, "error": " ".join(exc.messages)},
            status=400,
        )
    queue = django_rq.get_queue("ocr")
    if queue.count >= 10:
        return JsonResponse(
            {"ok": False, "error": "目前辨識工作較多，請稍候再試。"},
            status=429,
        )
    photo_token = request.POST.get("photo_token") or uuid.uuid4().hex
    job = IdOcrJob.objects.create(
        created_by=request.user,
        front=front,
        back=back,
        document_type=document_type,
        photo_token=photo_token,
    )
    try:
        queue.enqueue(
            run_id_ocr_job,
            str(job.pk),
            job_timeout=45,
            retry=Retry(max=2, interval=[2, 5]),
            result_ttl=300,
            failure_ttl=86400,
        )
    except Exception:
        logger.exception("無法加入身分證 OCR 背景佇列")
        job.status = IdOcrJob.Status.FAILED
        job.error = "辨識工作無法排入佇列，請稍後再試。"
        job.finished_at = timezone.now()
        job.save(update_fields=["status", "error", "finished_at", "updated_at"])
        delete_id_ocr_job_files(job)
        return JsonResponse({"ok": False, "error": job.error}, status=503)
    return JsonResponse(
        {
            "ok": True,
            "job_id": str(job.pk),
            "photo_token": photo_token,
            "document_type": job.document_type,
            "status": job.status,
        },
        status=202,
    )


@login_required
def id_card_ocr_status(request, job_id):
    job = get_object_or_404(IdOcrJob, pk=job_id, created_by=request.user)
    payload = {
        "ok": True,
        "job_id": str(job.pk),
        "photo_token": job.photo_token,
        "status": job.status,
    }
    if job.status == IdOcrJob.Status.SUCCEEDED:
        payload.update(job.result)
    elif job.status == IdOcrJob.Status.FAILED:
        payload["error"] = job.error
    return JsonResponse(payload)


@login_required
def id_card_ocr_invalidate(request, job_id):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "僅接受 POST。"}, status=405)
    job = get_object_or_404(IdOcrJob, pk=job_id, created_by=request.user)
    if job.status != IdOcrJob.Status.INVALIDATED:
        job.status = IdOcrJob.Status.INVALIDATED
        job.result = {}
        job.error = ""
        job.finished_at = timezone.now()
        job.save(
            update_fields=[
                "status",
                "result",
                "error",
                "finished_at",
                "updated_at",
            ]
        )
        delete_id_ocr_job_files(job)
    return JsonResponse({"ok": True, "status": job.status})


@login_required
def protected_media(request, model_name, pk, field_name):
    allowed = {
        "order": (
            SalesOrder,
            {
                "id_front",
                "id_back",
                "signed_contract",
                "privacy_consent",
                "refund_proof",
            },
        ),
        "draft": (OrderDraft, {"id_front", "id_back"}),
        "vehicle": (VehicleInventory, {"condition_photo"}),
        "vehicle_history": (
            VehicleInventoryHistory,
            {"condition_photo_snapshot"},
        ),
        "payment": (PaymentRecord, {"proof"}),
        "delivery": (DeliveryRecord, {"handover_photo"}),
    }
    if model_name not in allowed or field_name not in allowed[model_name][1]:
        raise Http404
    model = allowed[model_name][0]
    instance = get_object_or_404(model, pk=pk)
    file_field = getattr(instance, field_name)
    if not file_field:
        raise Http404
    response = FileResponse(file_field.open("rb"))
    response["Content-Disposition"] = f'inline; filename="{file_field.name.split("/")[-1]}"'
    return _protect_private_response(response)
